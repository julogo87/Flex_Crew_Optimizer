import pandas as pd
import networkx as nx
import pulp
from datetime import datetime, timedelta
import sys
import os
import io
import contextlib
from bisect import bisect_left
import heapq
import itertools
from collections import Counter
from functools import lru_cache

# Carpeta raíz del proyecto (donde está este script)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

# ============================================================================
#                    SISTEMA DE CONFIGURACIÓN EXTERNA (flex_config_6R.xlsx)
# ============================================================================
def load_config_excel(config_path=None) -> dict:
    """Lee flex_config_6R.xlsx y retorna dict con todos los parámetros."""
    if config_path is None:
        config_path = os.path.join(PROJECT_ROOT, "flex_config_6R.xlsx")
    if not os.path.exists(config_path):
        print(f"  [Config] flex_config_6R.xlsx no encontrado — usando valores por defecto del script")
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(config_path, data_only=True)
        cfg = {}
        if "Hoteles" in wb.sheetnames:
            ws = wb["Hoteles"]
            hotel_map = {}
            for row in ws.iter_rows(min_row=2):
                if row[0].value and row[1].value is not None:
                    hotel_map[str(row[0].value).strip().upper()] = float(row[1].value)
            if hotel_map:
                cfg["HOTEL_RATES_USD"] = hotel_map
        if "Estaciones" in wb.sheetnames:
            ws = wb["Estaciones"]
            for row in ws.iter_rows(min_row=2):
                if row[0].value and row[1].value:
                    key = str(row[0].value).strip()
                    vals = {v.strip().upper() for v in str(row[1].value).split(",") if v.strip()}
                    cfg[key] = vals
        scalar_sheets = ["Viaticos", "Tiempos", "Reglas_Duty", "Config_DH",
                         "Costos", "Solver", "Balance_Bases", "Planta_Pilotos",
                         "Buffers"]
        for sheet in scalar_sheets:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=2):
                    if row[0].value is not None and row[1].value is not None:
                        key = str(row[0].value).strip()
                        cfg[key] = row[1].value
        if "Flota" in wb.sheetnames:
            ws = wb["Flota"]
            tails = [str(row[0].value).strip() for row in ws.iter_rows(min_row=2)
                     if row[0].value and str(row[0].value).strip()]
            if tails:
                cfg["TARGET_TAILS"] = tails
        if "Balance_Bases" in wb.sheetnames:
            ws = wb["Balance_Bases"]
            base_targets = {}
            for row in ws.iter_rows(min_row=2):
                k = str(row[0].value).strip() if row[0].value else ""
                if len(k) == 3 and k.isalpha() and k.isupper():
                    try:
                        base_targets[k] = float(row[1].value)
                    except (TypeError, ValueError):
                        pass
            if base_targets:
                cfg["BASE_TARGETS"] = base_targets
        print(f"  [Config] Cargados {len(cfg)} parámetros desde flex_config_6R.xlsx")
        return cfg
    except Exception as exc:
        print(f"  [Config] Error leyendo flex_config_6R.xlsx: {exc} — usando valores por defecto")
        return {}

_CFG = load_config_excel()

def _cfg(key, default):
    """Accessor seguro al config externo con type coercion al tipo del default."""
    val = _CFG.get(key, default)
    if val is None:
        return default
    try:
        if isinstance(default, bool):
            if isinstance(val, bool):
                return val
            return str(val).strip().lower() in ("true", "1", "si", "yes")
        return type(default)(val)
    except (TypeError, ValueError):
        return default

def _cfg_set(key, default):
    """Accessor para parámetros tipo set."""
    val = _CFG.get(key)
    if val is None:
        return default
    if isinstance(val, set):
        return val
    try:
        return {v.strip().upper() for v in str(val).split(",") if v.strip()}
    except Exception:
        return default

# ============================================================================
#                           PARÁMETROS CONFIGURABLES
# ============================================================================

# ----------------------------------------------------------------------------
# 1. SISTEMA
# ----------------------------------------------------------------------------
APP_NAME = "Flex Optimizer-6R"
APP_VERSION = "1.0"
APP_NAME_FILE = APP_NAME.replace(" ", "_")
VERBOSE_OUTPUT = False  # True para ver detalles de iteraciones, False para interfaz limpia


# ----------------------------------------------------------------------------
# UTILIDADES DE CONSOLA
# ----------------------------------------------------------------------------
class ConsoleUI:
    """Interfaz de consola para mostrar progreso y resultados."""

    @staticmethod
    def clear_console():
        """Limpia la consola."""
        import platform
        if platform.system() == "Windows":
            os.system("cls")
        else:
            os.system("clear")

    @staticmethod
    def clear_line():
        """Limpia la línea actual."""
        print("\r" + " " * 80 + "\r", end="", flush=True)

    @staticmethod
    def header():
        """Muestra el encabezado del sistema."""
        print("\n" + "=" * 60)
        print(f"  {APP_NAME} v{APP_VERSION}")
        print("  Sistema de Optimización de Itinerarios de Tripulación")
        print("=" * 60)

    @staticmethod
    def section(title):
        """Muestra un título de sección."""
        print(f"\n> {title}")
        print("-" * 50)

    @staticmethod
    def status(message, done=False):
        """Muestra un mensaje de estado."""
        icon = "[OK]" if done else "[..]"
        print(f"  {icon} {message}")

    @staticmethod
    def progress(current, total, prefix="Progreso"):
        """Muestra una barra de progreso."""
        pct = int(100 * current / total) if total > 0 else 0
        bar_len = 30
        filled = int(bar_len * current / total) if total > 0 else 0
        bar = "#" * filled + "-" * (bar_len - filled)
        print(f"\r  {prefix}: [{bar}] {pct}% ({current}/{total})", end="", flush=True)
        if current >= total:
            print()

    @staticmethod
    def result(label, value, indent=2):
        """Muestra un resultado."""
        spaces = " " * indent
        print(f"{spaces}- {label}: {value}")

    @staticmethod
    def summary(title, results_dict):
        """Muestra un resumen de resultados."""
        print(f"\n{'=' * 60}")
        print(f"  RESUMEN: {title}")
        print("=" * 60)
        for key, val in results_dict.items():
            print(f"  - {key}: {val}")
        print("=" * 60)

    @staticmethod
    def error(message):
        """Muestra un mensaje de error."""
        print(f"  [ERROR] {message}")

    @staticmethod
    def warning(message):
        """Muestra una advertencia."""
        print(f"  [WARNING] {message}")

    @staticmethod
    def success(message):
        """Muestra un mensaje de éxito."""
        print(f"  [OK] {message}")


ui = ConsoleUI()

# ----------------------------------------------------------------------------
# SUPRESIÓN DE STDOUT Y BARRA DE PROGRESO
# ----------------------------------------------------------------------------
@contextlib.contextmanager
def _suppress_stdout():
    old_stdout = sys.stdout
    sys.stdout = io.StringIO()
    try:
        yield
    finally:
        sys.stdout = old_stdout

class StageProgress:
    def __init__(self, stages):
        self.stages = stages
        self.total = len(stages)
        self.current = 0
    def advance(self, detail=""):
        if self.current < self.total:
            label = self.stages[self.current]
            self.current += 1
            pct = int(100 * self.current / self.total)
            bar_len = 30
            filled = int(bar_len * self.current / self.total)
            bar = "#" * filled + "-" * (bar_len - filled)
            suffix = f" — {detail}" if detail else ""
            print(f"\r  [{bar}] {pct}%  {label}{suffix}          ", end="", flush=True)
    def done(self, message="Completado"):
        bar = "#" * 30
        print(f"\r  [{bar}] 100%  {message}          ")

# ----------------------------------------------------------------------------
# HELPER PARA FORMATO DECIMAL (coma como separador decimal para Excel)
# ----------------------------------------------------------------------------
def fmt_dec(value, decimals=1):
    """Formatea un número con coma como separador decimal para Excel.

    Args:
        value: Valor numérico a formatear
        decimals: Número de decimales (por defecto 1)

    Returns:
        String con el número formateado usando coma como decimal
    """
    try:
        formatted = f"{float(value):.{decimals}f}"
        return formatted.replace(".", ",")
    except (ValueError, TypeError):
        return str(value)

def fmt_pct(value, decimals=1):
    """Formatea un porcentaje con coma como separador decimal.

    Args:
        value: Valor numérico del porcentaje
        decimals: Número de decimales (por defecto 1)

    Returns:
        String con el porcentaje formateado usando coma como decimal
    """
    try:
        formatted = f"{float(value):.{decimals}f}%"
        return formatted.replace(".", ",")
    except (ValueError, TypeError):
        return str(value)

def fmt_money(value, decimals=2, symbol="$"):
    """Formatea un valor monetario con coma como separador decimal.

    Args:
        value: Valor numérico
        decimals: Número de decimales (por defecto 2)
        symbol: Símbolo de moneda (por defecto $)

    Returns:
        String con el valor formateado usando coma como decimal
    """
    try:
        formatted = f"{float(value):.{decimals}f}"
        formatted = formatted.replace(".", ",")
        return f"{symbol}{formatted}"
    except (ValueError, TypeError):
        return str(value)

# ----------------------------------------------------------------------------
# DICCIONARIO GLOBAL PARA RASTREAR RAZONES DE NO COBERTURA
# ----------------------------------------------------------------------------
# Estructura: {flight_id: {"razon": "CODIGO", "descripcion": "Texto explicativo"}}
# Códigos de razón:
#   - "SIN_DUTY_VALIDO": No forma parte de ningún duty legal
#   - "EXCEDE_BLOCK": Combinado con otros vuelos supera límite de block hours
#   - "EXCEDE_DUTY": Supera límite de duty hours
#   - "SIN_CONEXION_DH": No hay forma de llegar desde base con DH
#   - "ESTACION_PROHIBIDA": Vuelo termina en estación sin pernocta
#   - "RESTRICCION_DESCANSO": No cumple requisitos de descanso
#   - "NO_SELECCIONADO": Existe en trip válido pero optimizador no lo eligió
#   - "SIN_TRIP_VIABLE": No se pudo construir un trip que incluya este vuelo
FLIGHT_EXCLUSION_REASONS = {}

def reset_exclusion_reasons():
    """Reinicia el diccionario de razones de exclusión."""
    global FLIGHT_EXCLUSION_REASONS
    FLIGHT_EXCLUSION_REASONS = {}

def set_flight_exclusion(flight_id, razon, descripcion):
    """Registra la razón de exclusión de un vuelo (solo si no tiene una razón previa)."""
    global FLIGHT_EXCLUSION_REASONS
    if flight_id not in FLIGHT_EXCLUSION_REASONS:
        FLIGHT_EXCLUSION_REASONS[flight_id] = {"razon": razon, "descripcion": descripcion}

def get_flight_exclusion(flight_id):
    """Obtiene la razón de exclusión de un vuelo."""
    return FLIGHT_EXCLUSION_REASONS.get(flight_id, {"razon": "", "descripcion": ""})

# ----------------------------------------------------------------------------
# 2. FLOTA (Aviones a incluir)
# ----------------------------------------------------------------------------
TARGET_TAILS = _CFG.get("TARGET_TAILS", ['N330QT', 'N331QT', 'N332QT', 'N334QT', 'N335QT', 'N336QT', 'N342AV','N338QT','N337QT',"N343QT"])

# ----------------------------------------------------------------------------
# 3. ESTACIONES Y BASES
# ----------------------------------------------------------------------------
DEFAULT_BASES = _cfg_set("DEFAULT_BASES", {"NLU"})
PROHIBITED_LAYOVERS = _cfg_set("PROHIBITED_LAYOVERS", {"CCS", "VIX", "GUA", "MAO","FLN","ASU","AGT","GYE","MST","VVI"})
SAFE_HAVENS = _cfg_set("SAFE_HAVENS", {"BOG", "MDE", "MIA", "VCP", "SCL", "GDL","LAX","NLU"})
COLOMBIA_STATIONS = _cfg_set("COLOMBIA_STATIONS", {"MDE", "BOG", "BAQ", "CLO"})
OUTSIDE_AMERICAS_STATIONS = _cfg_set("OUTSIDE_AMERICAS_STATIONS", {"MAD", "AMD", "LGG", "KEF", "MST","ZAZ","DSS"})

# ----------------------------------------------------------------------------
# 4. COSTOS - TARIFAS HOTEL (USD por noche-persona)
# ----------------------------------------------------------------------------
HOTEL_RATES_USD = _CFG.get("HOTEL_RATES_USD", {
    "MIA": 109, "SCL": 100, "VCP": 120, "EZE": 120,
    "MDE": 90, "BOG": 90, "NLU": 90, "MVD": 120, "UIO": 110, "SAL": 100,
})

# ----------------------------------------------------------------------------
# 5B. VIÁTICOS Y PRIMAS (USD)
# ----------------------------------------------------------------------------
VIATICO_CAP_AMERICAS_USD = _cfg("VIATICO_CAP_AMERICAS_USD", 71)
VIATICO_CAP_OUTSIDE_USD = _cfg("VIATICO_CAP_OUTSIDE_USD", 73)
VIATICO_COP_AMERICAS_USD = _cfg("VIATICO_COP_AMERICAS_USD", 46)
VIATICO_COP_OUTSIDE_USD = _cfg("VIATICO_COP_OUTSIDE_USD", 47)
PRIMA_NAV_CAP_AMERICAS_USD = _cfg("PRIMA_NAV_CAP_AMERICAS_USD", 22)
PRIMA_NAV_CAP_OUTSIDE_USD = _cfg("PRIMA_NAV_CAP_OUTSIDE_USD", 28)
PRIMA_NAV_COP_USD = _cfg("PRIMA_NAV_COP_USD", 13)
PRIMA_NAV_POR_VUELO_USD = _cfg("PRIMA_NAV_POR_VUELO_USD", 22)
EXTRA_POSICIONAMIENTO_USD = _cfg("EXTRA_POSICIONAMIENTO_USD", 100)
PRIMA_COMANDO_2P_CAP_USD = _cfg("PRIMA_COMANDO_2P_CAP_USD", 26.0)
PRIMA_COMANDO_3P_CAP_USD = _cfg("PRIMA_COMANDO_3P_CAP_USD", 43.0)
PRIMA_COMANDO_3P_COP_USD = _cfg("PRIMA_COMANDO_3P_COP_USD", 24.0)
PRIMA_COMANDO_4P_CAP1_USD = _cfg("PRIMA_COMANDO_4P_CAP1_USD", 38.0)
PRIMA_COMANDO_4P_CAP2_USD = _cfg("PRIMA_COMANDO_4P_CAP2_USD", 26.0)

# ----------------------------------------------------------------------------
# 5. TIEMPOS Y CONEXIONES
# ----------------------------------------------------------------------------
MIN_CONNECT_MINUTES = _cfg("MIN_CONNECT_MINUTES", 60)
MAX_CONNECT_HOURS = _cfg("MAX_CONNECT_HOURS", 4)
REPORT_MINUTES = _cfg("REPORT_MINUTES", 90)
DEBRIEF_MINUTES = _cfg("DEBRIEF_MINUTES", 0)
BLOCK_BUFFER_MINUTES = _cfg("BLOCK_BUFFER_MINUTES", 0)
DUTY_BUFFER_MINUTES = _cfg("DUTY_BUFFER_MINUTES", 0)
GROUND_TRANSFER_MINUTES_VCP_BOG = _cfg("GROUND_TRANSFER_MINUTES_VCP_BOG", 150)

# 5C. ZONA HORARIA
UTC_OFFSET_HOURS = _cfg("UTC_OFFSET_HOURS", 6)

# 6. LÍMITES DE TRIP Y DESCANSO
MAX_TRIP_DAYS = _cfg("MAX_TRIP_DAYS", 5)
MAX_REST_GAP_HOURS = _cfg("MAX_REST_GAP_HOURS", 48)
OPEN_TOUR_MAX_REST_GAP_HOURS = _cfg("OPEN_TOUR_MAX_REST_GAP_HOURS", 72)
REST_BUFFER_MINUTES = _cfg("REST_BUFFER_MINUTES", 0)
REST_THRESHOLD_HOURS = _cfg("REST_THRESHOLD_HOURS", 8.5)
REST_SHORT_HOURS = _cfg("REST_SHORT_HOURS", 12.0)
REST_LONG_HOURS = _cfg("REST_LONG_HOURS", 24.0)

# ----------------------------------------------------------------------------
# 7. REGLAS DE NOCTURNIDAD
# ----------------------------------------------------------------------------
# Regla 1: Ventana nocturna para MAX_CONSEC_NIGHT_DUTIES (duty que TOCA esta franja)
NIGHT_WINDOW_START_HOUR = _cfg("NIGHT_WINDOW_START_HOUR", 1)
NIGHT_WINDOW_END_HOUR = _cfg("NIGHT_WINDOW_END_HOUR", 5)
MAX_CONSEC_NIGHT_DUTIES = _cfg("MAX_CONSEC_NIGHT_DUTIES", 3)
DUTY_REDUCTION_START_HOUR = _cfg("DUTY_REDUCTION_START_HOUR", 15)
DUTY_REDUCTION_END_HOUR = _cfg("DUTY_REDUCTION_END_HOUR", 3)
DUTY_REDUCTION_MINUTES = _cfg("DUTY_REDUCTION_MINUTES", 60)

# 8. REGLAS DE DUTY
ALLOW_TAIL_CHANGE_IN_DUTY = _cfg("ALLOW_TAIL_CHANGE_IN_DUTY", False)
MAX_DUTY_LEGS = _cfg("MAX_DUTY_LEGS", 4)
ENFORCE_NO_SAME_DAY_DUTY = _cfg("ENFORCE_NO_SAME_DAY_DUTY", False)
ENFORCE_NO_SAME_DAY_AFTER_DH = _cfg("ENFORCE_NO_SAME_DAY_AFTER_DH", False)

# 9. CONFIGURACIÓN DE DH
ALLOW_DH_IN_DUTY = _cfg("ALLOW_DH_IN_DUTY", True)
DH_IN_DUTY_MODE = "BOTH"
ALLOW_DH_ONLY_DUTY = _cfg("ALLOW_DH_ONLY_DUTY", True)
DEFAULT_DH_HOURS = _cfg("DEFAULT_DH_HOURS", 4.0)
MAX_DH_LEGS = _cfg("MAX_DH_LEGS", 2)
MAX_DH_HOURS = _cfg("MAX_DH_HOURS", 12)
FILTER_DH_TO_CARGO_STATIONS = _cfg("FILTER_DH_TO_CARGO_STATIONS", True)
DH_DATE_BUFFER_DAYS = _cfg("DH_DATE_BUFFER_DAYS", 2)
MAX_DH_IN_DUTY_FLIGHTS = _cfg("MAX_DH_IN_DUTY_FLIGHTS", 2)
MAX_DH_HOURS_PER_TRIP = _cfg("MAX_DH_HOURS_PER_TRIP", 20.0)
MAX_DH_RATIO = _cfg("MAX_DH_RATIO", 0.6)

# ----------------------------------------------------------------------------
# 10. OPEN TOURS (Vuelos no cubiertos)
# ----------------------------------------------------------------------------
ENABLE_OPEN_TOURS = _cfg("ENABLE_OPEN_TOURS", True)
OPEN_TOUR_MAX_DAYS = _cfg("OPEN_TOUR_MAX_DAYS", 6)
OPEN_TOUR_EXPORT_EXCEL = _cfg("OPEN_TOUR_EXPORT_EXCEL", True)
OPEN_TOUR_ALLOW_SAME_DAY = _cfg("OPEN_TOUR_ALLOW_SAME_DAY", True)
OPEN_TOUR_MAX_DH_HOURS = OPEN_TOUR_MAX_DAYS * 72
OPEN_TOUR_INCLUDE_DH_FLIGHTS = _cfg("OPEN_TOUR_INCLUDE_DH_FLIGHTS", True)
OPEN_TOUR_MAX_DH_LEGS = _cfg("OPEN_TOUR_MAX_DH_LEGS", 4)
OPEN_TOUR_MAX_DH_HOURS_PER_TRIP = OPEN_TOUR_MAX_DH_HOURS
OPEN_TOUR_MAX_DH_RATIO = _cfg("OPEN_TOUR_MAX_DH_RATIO", 10.0)
POSITIONING_DH_HOURS_BACK = _cfg("POSITIONING_DH_HOURS_BACK", 48)
POSITIONING_DH_SEARCH_WINDOW = _cfg("POSITIONING_DH_SEARCH_WINDOW", 48)

# ----------------------------------------------------------------------------
# 11. TRIPULACIONES
# ----------------------------------------------------------------------------
CREW_RANK = {'2P': 2, '3P': 3, '4P': 4}  # Sencilla, Multiple, Doble


def _ops_for_req(trip_crew, req):
    """Posiciones operacionales para un duty requirement dado el trip_crew."""
    if trip_crew == '4P':
        if req == '4P': return {"CAP1", "CAP2", "COP1", "COP2"}
        if req == '3P': return {"CAP1", "COP1", "COP2"}
        return {"CAP1", "COP1"}
    if trip_crew == '3P':
        if req == '3P': return {"CAP", "COP", "CRP"}
        return {"CAP", "COP"}
    return {"CAP", "COP"}


def _get_pilot_list(trip_crew):
    """Lista de (pilot_id, role) para el trip_crew."""
    if trip_crew == '4P':
        return [("CAP1", "COMANDANTE"), ("CAP2", "CAP"), ("COP1", "COP"), ("COP2", "COP")]
    elif trip_crew == '3P':
        return [("CAP", "COMANDANTE"), ("COP", "COP"), ("CRP", "CRP")]
    else:
        return [("CAP", "COMANDANTE"), ("COP", "COP")]


MAX_DUTY_2P = _cfg("MAX_DUTY_2P", 13)
MAX_BLOCK_2P = _cfg("MAX_BLOCK_2P", 9.0)
MAX_DUTY_3P = _cfg("MAX_DUTY_3P", 15)
MAX_BLOCK_3P = _cfg("MAX_BLOCK_3P", 12)
MAX_DUTY_4P = _cfg("MAX_DUTY_4P", 18)
MAX_BLOCK_4P = _cfg("MAX_BLOCK_4P", 15)

# Filtro de crew type: trips largos deben justificar su crew type
MIN_HIGH_CREW_DUTIES    = _cfg("MIN_HIGH_CREW_DUTIES",    2)   # Min duties que requieran el crew type
MIN_DAYS_FOR_CREW_CHECK = _cfg("MIN_DAYS_FOR_CREW_CHECK", 4)   # Solo trips de N+ días
SKIP_CREW_CHECK_SECOND_PASS = _cfg("SKIP_CREW_CHECK_SECOND_PASS", True)  # No aplicar en 2da pasada
MIN_BLK_PER_DAY_FIRST_PASS  = _cfg("MIN_BLK_PER_DAY_FIRST_PASS", 4.0)   # Block/día mínimo 1ra pasada

# ----------------------------------------------------------------------------
# 12. PLANTA DE PILOTOS (Requerimiento mensual)
# ----------------------------------------------------------------------------
FREE_DAYS_PER_MONTH = _cfg("FREE_DAYS_PER_MONTH", 9)
RESERVE_PCT = _cfg("RESERVE_PCT", 0.0)
TRAINING_PILOTS = _cfg("TRAINING_PILOTS", 0)
VACATION_PILOTS = _cfg("VACATION_PILOTS", 0)
ADMIN_PILOTS = _cfg("ADMIN_PILOTS", 0)
DOCS_PILOTS = _cfg("DOCS_PILOTS", 0)
INCAP_PILOTS = _cfg("INCAP_PILOTS", 0)
UNION_PILOTS = _cfg("UNION_PILOTS", 0)

# 13. BALANCE POR BASE
BASE_TARGETS = _CFG.get("BASE_TARGETS", {'NLU': 1})
ENFORCE_BASE_BALANCE = _cfg("ENFORCE_BASE_BALANCE", True)
BASE_BALANCE_TOLERANCE = _cfg("BASE_BALANCE_TOLERANCE", 0.0)

# 14. OBJETIVO DE OPTIMIZACIÓN
OPTIMIZATION_OBJECTIVE = _cfg("OPTIMIZATION_OBJECTIVE", "PILOTS")
EFFICIENCY_DAY_WEIGHT = 10.0
EFFICIENCY_BLOCK_WEIGHT = 1.0
PILOT_DAY_WEIGHT = 1.0
HYBRID_WEIGHT_EFF = 1.0
HYBRID_WEIGHT_PILOTS = 1.0
HYBRID_WEIGHT_COST = 0.0
PILOTS_EFF_WEIGHT_PILOTS = 10.0
PILOTS_EFF_WEIGHT_EFF = 2.0
OBJECTIVE_TIEBREAKER_COST = 0.0
ENABLE_LEXICOGRAPHIC = _cfg("ENABLE_LEXICOGRAPHIC", True)
PILOT_DAY_TOLERANCE = _cfg("PILOT_DAY_TOLERANCE", 0.0)

# ----------------------------------------------------------------------------
# 15. PENALIZACIONES Y FILTROS
# ----------------------------------------------------------------------------
PENALTY_4P = _cfg("PENALTY_4P", 50)
TRIP_DAY_PENALTY = _cfg("TRIP_DAY_PENALTY", 300)

# 16. RENDIMIENTO / SOLVER
MAX_ITER = _cfg("MAX_ITER", 600000)
MAX_DUTY_PATHS_PER_START = _cfg("MAX_DUTY_PATHS_PER_START", 1500)
SOLVER_TIME_LIMIT_SECONDS = _cfg("SOLVER_TIME_LIMIT_SECONDS", 600)

# 17. SEGUNDA PASADA PARA OPEN TIME CON DH
ENABLE_SECOND_PASS_DH = _cfg("ENABLE_SECOND_PASS_DH", True)
MAX_TRIP_DAYS_SECOND_PASS = _cfg("MAX_TRIP_DAYS_SECOND_PASS", 6)
MAX_DH_RATIO_SECOND_PASS = _cfg("MAX_DH_RATIO_SECOND_PASS", 10.0)
SECOND_PASS_PENALTY = _cfg("SECOND_PASS_PENALTY", 10_000)
SECOND_PASS_DEBUG = _cfg("SECOND_PASS_DEBUG", False)

# 18. TERCERA PASADA Y SOLO DH
ENABLE_THIRD_PASS_PARTIAL = _cfg("ENABLE_THIRD_PASS_PARTIAL", False)
ENABLE_SOLO_DH_PASS = _cfg("ENABLE_SOLO_DH_PASS", True)
SOLO_DH_SEARCH_WINDOW_HOURS = _cfg("SOLO_DH_SEARCH_WINDOW_HOURS", 72)
SOLO_DH_PENALTY = _cfg("SOLO_DH_PENALTY", 2000)

# ==========================================
# 1. MOTOR DE REGLAS (COMPLETO)
# ==========================================

class RuleEngine:
    def __init__(self, dh_table=None, dh_index=None, allowed_crews=None, allow_same_day_duty=None):
        # Configuración de Bases y Aeropuertos
        self.BASES = set(DEFAULT_BASES)
        self.PROHIBITED_LAYOVERS = set(PROHIBITED_LAYOVERS)  # ¡Prohibido dormir aquí!
        self.SAFE_HAVENS = set(SAFE_HAVENS)
        
        # Parámetros de Tiempo
        self.MIN_CONNECT = timedelta(minutes=MIN_CONNECT_MINUTES)
        self.MAX_CONNECT = timedelta(hours=MAX_CONNECT_HOURS)
        self.DEBRIEF = timedelta(minutes=DEBRIEF_MINUTES)
        self.MAX_TRIP_DAYS = MAX_TRIP_DAYS
        self.REPORT_TIME = timedelta(minutes=REPORT_MINUTES)
        self.REST_BUFFER_HOURS = REST_BUFFER_MINUTES / 60.0  # Buffer de descanso en horas
        self.CREW_RANK = dict(CREW_RANK)
        self.DEFAULT_DH = DEFAULT_DH_HOURS
        self.MAX_DH_LEGS = MAX_DH_LEGS
        self.MAX_DH_HOURS = MAX_DH_HOURS
        # Tabla DH (org,dst) -> horas block
        self.DH_TABLE = dh_table or {}
        # Índice de vuelos DH para rescates (time-dependent)
        self.DH_INDEX = dh_index or {}
        self._rescue_cache = {}
        # Tipos de tripulación permitidos (None = todos)
        self.ALLOWED_CREWS = set(allowed_crews) if allowed_crews else None
        # Permitir iniciar duty el mismo día calendario (por defecto según flag global)
        if allow_same_day_duty is None:
            self.ALLOW_SAME_DAY_DUTY = not ENFORCE_NO_SAME_DAY_DUTY
            self.ALLOW_SAME_DAY_AFTER_DH = not ENFORCE_NO_SAME_DAY_AFTER_DH
        else:
            self.ALLOW_SAME_DAY_DUTY = bool(allow_same_day_duty)
            self.ALLOW_SAME_DAY_AFTER_DH = bool(allow_same_day_duty)  # Si permite same day, también después de DH

    def get_dh_time(self, org, dst):
        """Devuelve tiempo estimado de DH entre estaciones comunes (incluye report/debrief)."""
        if (org, dst) in self.DH_TABLE:
            return self.DH_TABLE[(org, dst)]
        if {org, dst} == {'BOG', 'MDE'}: base = 1.0
        elif {org, dst} == {'BOG', 'CLO'}: base = 1.0
        elif {org, dst} == {'BOG', 'BAQ'}: base = 1.0
        elif {org, dst} == {'BOG', 'MIA'}: base = 3.5
        elif {org, dst} == {'MDE', 'MIA'}: base = 3.5
        elif {org, dst} == {'BOG', 'UIO'}: base = 1.5
        elif {org, dst} == {'BOG', 'CCS'}: base = 1.5
        else: base = self.DEFAULT_DH
        # Convertir a duración duty agregando report/debrief
        buffer_h = (self.REPORT_TIME + self.DEBRIEF).total_seconds() / 3600
        return base + buffer_h

    def get_dh_time_required(self, org, dst):
        """DH solo válido si existe en la tabla real."""
        if (org, dst) in self.DH_TABLE:
            return self.DH_TABLE[(org, dst)]
        return None

    def get_positioning_itinerary(self, org, dst, latest_arrival_base_dt, max_legs=None, max_hours=None):
        """Busca el DH MÁS TARDÍO que llegue antes de latest_arrival_base_dt,
        cumpliendo descanso reglamentario mínimo y regla de same-day duty."""
        if max_legs is None:
            max_legs = self.MAX_DH_LEGS
        if max_hours is None:
            max_hours = self.MAX_DH_HOURS

        best_info = None
        best_arrival = None

        # Ventanas de 6h con overlap de 12h para encontrar el DH más cercano al duty
        search_step = 6
        window_hours = 12
        for hours_back in range(search_step, int(max_hours) + search_step, search_step):
            earliest_report = latest_arrival_base_dt - timedelta(hours=hours_back)
            info = self.get_rescue_itinerary(org, dst, earliest_report,
                max_legs=max_legs, max_hours=min(window_hours, POSITIONING_DH_SEARCH_WINDOW))
            if info:
                dh_hours, arr_base_dt, path = info
                if arr_base_dt > latest_arrival_base_dt:
                    continue
                # Validar descanso reglamentario mínimo
                dh_block = sum(l.get("blk", 0.0) for l in path)
                is_at_base = (dst in self.BASES)
                req_rest = self.calculate_required_rest(dh_block, is_base=is_at_base)
                rest_available = (latest_arrival_base_dt - arr_base_dt).total_seconds() / 3600
                if rest_available < req_rest:
                    continue
                # Validar same-day duty
                if not self.ALLOW_SAME_DAY_DUTY:
                    if latest_arrival_base_dt.date() <= arr_base_dt.date():
                        continue
                if not self.ALLOW_SAME_DAY_AFTER_DH:
                    if latest_arrival_base_dt.date() <= arr_base_dt.date():
                        continue
                # Actualizar si es más tardío (más cercano al duty)
                if best_arrival is None or arr_base_dt > best_arrival:
                    best_info = info
                    best_arrival = arr_base_dt

        return best_info

    def get_rescue_itinerary(self, org, dst, earliest_report_base_dt, max_legs=None, max_hours=None):
        """Busca el itinerario DH más temprano (con escalas) desde org a dst.
        earliest_report_base_dt: hora más temprana para reportar (ya cumplido descanso).
        Retorna (dh_hours, arr_base_dt) o None si no hay opción.
        """
        if max_legs is None:
            max_legs = self.MAX_DH_LEGS
        if max_hours is None:
            max_hours = self.MAX_DH_HOURS

        # DEBUG: Mostrar búsqueda específica para SCL/MIA y VCP/GRU-BOG
        debug_this = (org == 'BOG' and dst == 'SCL') or (org == 'MIA' and dst in ('BOG', 'MDE')) or (org in ('VCP', 'GRU') and dst == 'BOG')
        if debug_this and SECOND_PASS_DEBUG:
            print(f"\n    [DEBUG RESCUE] Buscando {org}->{dst}")
            print(f"      earliest_report_base: {earliest_report_base_dt}")
            print(f"      max_legs: {max_legs}, max_hours: {max_hours}")
            print(f"      DH_INDEX presente: {bool(self.DH_INDEX)}")

        if not self.DH_INDEX:
            if debug_this and SECOND_PASS_DEBUG:
                print(f"      [FAIL] DH_INDEX vacío")
            return None

        earliest_report_utc = earliest_report_base_dt + timedelta(hours=UTC_OFFSET_HOURS)
        min_dep_utc = earliest_report_utc + self.REPORT_TIME
        # Si el DH sale desde VCP/GRU, sumar traslado terrestre al min_dep
        if org.upper() in ('VCP', 'GRU'):
            min_dep_utc = min_dep_utc + timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
        key = (org, dst, earliest_report_utc.strftime("%Y-%m-%d %H"), int(max_legs), float(max_hours))
        cached = self._rescue_cache.get(key)
        if cached is not None:
            return cached

        flights_by_org = self.DH_INDEX.get("flights_by_org", {})
        times_by_org = self.DH_INDEX.get("times_by_org", {})

        if debug_this and SECOND_PASS_DEBUG:
            print(f"      Vuelos desde {org}: {len(flights_by_org.get(org, []))}")
            print(f"      min_dep_utc: {min_dep_utc}")
            print(f"      max_dep_utc: {min_dep_utc + timedelta(hours=max_hours)}")

        if not flights_by_org:
            if debug_this and SECOND_PASS_DEBUG:
                print(f"      [FAIL] flights_by_org vacío")
            return None

        dest_set = _equiv_stations(dst)
        max_dep_utc = min_dep_utc + timedelta(hours=max_hours)
        if _has_direct_in_window(flights_by_org, times_by_org, org, dest_set, min_dep_utc, max_dep_utc):
            max_legs = 1

        heap = []
        seq = itertools.count()
        best = {}

        candidates_found = 0
        for start_org in _equiv_stations(org):
            next_list = flights_by_org.get(start_org, [])
            times = times_by_org.get(start_org, [])
            if not next_list or not times:
                if debug_this and SECOND_PASS_DEBUG:
                    print(f"      Sin vuelos desde {start_org}")
                continue
            start_idx = bisect_left(times, min_dep_utc)
            if debug_this and SECOND_PASS_DEBUG:
                print(f"      Buscando desde {start_org}: {len(next_list)} vuelos, start_idx={start_idx}")
                # DIAGNÓSTICO: mostrar rango de fechas en el índice
                if times:
                    print(f"      ÍNDICE {start_org}: primer_fecha={times[0]}, última_fecha={times[-1]}")
                    print(f"      BÚSQUEDA: min_dep_utc={min_dep_utc}, max_dep_utc={max_dep_utc}")
                    if start_idx >= len(times):
                        print(f"      [WARN] start_idx={start_idx} >= len={len(times)}: fecha busqueda DESPUES del indice")
                    elif start_idx == 0 and times[0] > min_dep_utc:
                        print(f"      [WARN] Todos los vuelos del indice son DESPUES de min_dep_utc")
            for i in range(start_idx, len(next_list)):
                nf = next_list[i]
                if nf["dep_utc"] > max_dep_utc:
                    break
                start_dep = nf["dep_utc"]
                block_sum = nf["blk"]
                if not _duty_ok_2p(start_dep, nf["arr_utc"], block_sum, 1, self,
                                    first_flight=nf, last_flight=nf):
                    if debug_this and SECOND_PASS_DEBUG and candidates_found < 3:
                        print(f"        Vuelo {nf.get('org')}-{nf.get('dst')} dep={nf['dep_utc']} rechazado por duty_ok_2p")
                    continue
                candidates_found += 1
                path = [nf]
                heapq.heappush(heap, (nf["arr_utc"], 1, next(seq), nf["dst"], start_dep, block_sum, path))

        if debug_this and SECOND_PASS_DEBUG:
            print(f"      Candidatos iniciales en heap: {candidates_found}")

        while heap:
            arr_utc, legs, _, stn, start_dep, block_sum, path = heapq.heappop(heap)
            if arr_utc > max_dep_utc:
                continue
            k = (stn, legs)
            prev = best.get(k)
            if prev is not None and arr_utc >= prev:
                continue
            best[k] = arr_utc

            if stn in dest_set:
                # Ajustar por traslado terrestre VCP/GRU ↔ BOG
                gt = timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
                effective_start = start_dep - self.REPORT_TIME
                effective_end = arr_utc + self.DEBRIEF
                first_leg = path[0] if path else None
                last_leg = path[-1] if path else None
                if first_leg:
                    fl_org = str(first_leg.get('org', '')).upper()
                    fl_dst = str(first_leg.get('dst', '')).upper()
                    if fl_org in ('VCP', 'GRU') and fl_dst == 'BOG':
                        effective_start = effective_start - gt
                if last_leg:
                    ll_org = str(last_leg.get('org', '')).upper()
                    ll_dst = str(last_leg.get('dst', '')).upper()
                    if ll_org == 'BOG' and ll_dst in ('VCP', 'GRU'):
                        effective_end = effective_end + gt
                dh_hours = (effective_end - effective_start).total_seconds() / 3600
                arr_base_dt = effective_end - timedelta(hours=UTC_OFFSET_HOURS)
                result = (dh_hours, arr_base_dt, path)
                self._rescue_cache[key] = result
                return result

            if legs >= max_legs:
                continue

            min_dep = arr_utc + self.MIN_CONNECT
            max_dep = arr_utc + self.MAX_CONNECT
            for next_org in _equiv_stations(stn):
                next_list = flights_by_org.get(next_org, [])
                times = times_by_org.get(next_org, [])
                if not next_list or not times:
                    continue
                start_idx = bisect_left(times, min_dep)
                for i in range(start_idx, len(next_list)):
                    nf = next_list[i]
                    if nf["dep_utc"] > max_dep:
                        break
                    new_block = block_sum + nf["blk"]
                    new_legs = legs + 1
                    if not _duty_ok_2p(start_dep, nf["arr_utc"], new_block, new_legs, self,
                                        first_flight=path[0], last_flight=nf):
                        continue
                    new_path = path + [nf]
                    heapq.heappush(heap, (nf["arr_utc"], new_legs, next(seq), nf["dst"], start_dep, new_block, new_path))

        if debug_this and SECOND_PASS_DEBUG:
            print(f"      [FAIL] No se encontró ruta {org}->{dst}")
        self._rescue_cache[key] = None
        return None

    def calculate_required_rest(self, block_hours, is_base):
        """
        Calculadora de Descanso Simplificada:
        - < REST_THRESHOLD_HOURS: REST_SHORT_HOURS de descanso
        - >= REST_THRESHOLD_HOURS: REST_LONG_HOURS de descanso

        + REST_BUFFER_MINUTES adicionales configurables
        """
        if block_hours < REST_THRESHOLD_HOURS:
            base_rest = REST_SHORT_HOURS
        else:
            base_rest = REST_LONG_HOURS

        # Aplicar buffer adicional
        return base_rest + self.REST_BUFFER_HOURS

    def validate_daily_duty(self, flights):
        if not flights: return False, None, None
        first, last = flights[0], flights[-1]
        
        # 1. Validación: Mismo Avión (si está habilitado)
        first_tail = flights[0]['tail']
        tails = []
        for f in flights:
            t = f.get('tail')
            if t not in tails:
                tails.append(t)
        tail_change = False
        if not ALLOW_TAIL_CHANGE_IN_DUTY:
            primary_tail = None
            for f in flights:
                if f.get("is_dh"):
                    continue
                primary_tail = f.get("tail")
                break
            if primary_tail is None:
                primary_tail = first_tail
            for f in flights:
                if f.get("is_dh"):
                    continue
                if f.get('tail') != primary_tail:
                    return False, None, None
        else:
            tail_change = any(f.get('tail') != first_tail for f in flights)

        if not _dh_positions_ok(flights):
            return False, None, None

        # 2. Tiempos (hora base)
        duty_start = first['dep_base'] - self.REPORT_TIME
        duty_end = last['arr_base'] + self.DEBRIEF

        # Ajuste traslado terrestre VCP/GRU ↔ BOG para DH
        ground_transfer = timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
        first_org = str(first.get('org', '')).upper()
        first_dst = str(first.get('dst', '')).upper()
        if first.get('is_dh') and first_org in ('VCP', 'GRU') and first_dst == 'BOG':
            duty_start = duty_start - ground_transfer
        last_org = str(last.get('org', '')).upper()
        last_dst = str(last.get('dst', '')).upper()
        if last.get('is_dh') and last_org == 'BOG' and last_dst in ('VCP', 'GRU'):
            duty_end = duty_end + ground_transfer

        duty_dur = (duty_end - duty_start).total_seconds() / 3600
        block_hrs = sum([f['blk_hours'] for f in flights])
        num_sectors = len(flights)
        
        # 3. Límites Legales (Dinámicos)
        start_hour = duty_start.hour
        # Reducción de duty si empieza entre 15:00 y 03:00
        needs_reduction = (start_hour >= DUTY_REDUCTION_START_HOUR or start_hour < DUTY_REDUCTION_END_HOUR)
        reduction = (DUTY_REDUCTION_MINUTES / 60.0) if needs_reduction else 0.0
        
        duty_buffer_h = DUTY_BUFFER_MINUTES / 60.0
        block_buffer_h = BLOCK_BUFFER_MINUTES / 60.0
        limits = {
            '2P': {'max_duty': MAX_DUTY_2P - reduction, 'max_blk': MAX_BLOCK_2P},
            '3P': {'max_duty': MAX_DUTY_3P - reduction, 'max_blk': MAX_BLOCK_3P},
            '4P': {'max_duty': MAX_DUTY_4P - reduction, 'max_blk': MAX_BLOCK_4P}
        }

        daily_crew_req = None
        limit_info = {}

        # Debug para MIA-ASU-VCP
        is_target_duty = 'MIA' in str(first.get('org', '')) and 'VCP' in str(last.get('dst', ''))
        if VERBOSE_OUTPUT and is_target_duty:
            print(f"\n  [DUTY DEBUG] Validating: {first.get('org')}-...-{last.get('dst')}")
            print(f"    Block: {block_hrs:.2f}h, Duty: {duty_dur:.2f}h, Sectors: {num_sectors}")
            print(f"    Allowed crews: {self.ALLOWED_CREWS}")
            print(f"    Needs Reduction (15:00-03:00): {needs_reduction}, Reduction: {reduction}h")

        for c_type in ['2P', '3P', '4P']:
            if self.ALLOWED_CREWS is not None and c_type not in self.ALLOWED_CREWS:
                if VERBOSE_OUTPUT and is_target_duty:
                    print(f"    [{c_type}] SKIPPED - not in allowed crews")
                continue
            rule = limits[c_type]
            limit_duty = rule['max_duty'] if num_sectors <= 6 else (rule['max_duty'] - 2.0)
            max_duty_eff = max(0.0, limit_duty - duty_buffer_h)
            max_blk_eff = max(0.0, rule['max_blk'] - block_buffer_h)

            if VERBOSE_OUTPUT and is_target_duty:
                print(f"    [{c_type}] Max duty: {max_duty_eff:.2f}h, Max block: {max_blk_eff:.2f}h")
                if duty_dur <= max_duty_eff and block_hrs <= max_blk_eff:
                    print(f"    [{c_type}] OK FITS!")
                else:
                    fails = []
                    if duty_dur > max_duty_eff:
                        fails.append(f"duty {duty_dur:.2f}h > {max_duty_eff:.2f}h")
                    if block_hrs > max_blk_eff:
                        fails.append(f"block {block_hrs:.2f}h > {max_blk_eff:.2f}h")
                    print(f"    [{c_type}] FAIL: {', '.join(fails)}")

            if duty_dur <= max_duty_eff and block_hrs <= max_blk_eff:
                daily_crew_req = c_type
                limit_info = {'max_duty': max_duty_eff, 'max_blk': max_blk_eff}
                break

        if not daily_crew_req:
            if VERBOSE_OUTPUT and is_target_duty:
                print(f"    [RESULT] REJECTED - no crew type fits")
            return False, None, None

        if VERBOSE_OUTPUT and is_target_duty:
            print(f"    [RESULT] ACCEPTED as {daily_crew_req}")
            
        night_touch = self._touches_night_window(duty_start, duty_end)
        duty_obj = {
            'id': f"D_{first['id']}_{last['id']}",
            'flights': flights,
            'start_utc': first['dep_utc'] - self.REPORT_TIME,
            'start_base': duty_start,
            'end_base': duty_end,
            'org': first['org'],
            'dst': last['dst'],
            'min_crew': daily_crew_req, 
            'limits': limit_info,
            'block': block_hrs,
            'duty_dur': duty_dur,
            'base_cost': (duty_dur - block_hrs) * 100, # Costo ineficiencia base
            'tail': first_tail,
            'tails': tails,
            'tail_change': tail_change,
            'night_touch': night_touch,
        }
        return True, daily_crew_req, duty_obj

    def _touches_night_window(self, start_dt, end_dt):
        """True si el duty toca la ventana 01:00-05:00 HL en cualquier día."""
        if end_dt <= start_dt:
            return False
        cur = start_dt.date()
        end_date = end_dt.date()
        while cur <= end_date:
            win_start = datetime.combine(cur, datetime.min.time()) + timedelta(hours=NIGHT_WINDOW_START_HOUR)
            win_end = datetime.combine(cur, datetime.min.time()) + timedelta(hours=NIGHT_WINDOW_END_HOUR)
            if start_dt < win_end and end_dt > win_start:
                return True
            cur += timedelta(days=1)
        return False

# ==========================================
# 2. CARGA DE DATOS
# ==========================================

def load_schedule(file_path, tails_filter=None):
    if VERBOSE_OUTPUT:
        print(f"\n--- Cargando archivo: {os.path.basename(file_path)} ---")
    if not os.path.exists(file_path): return pd.DataFrame()

    try:
        try:
            df = pd.read_excel(file_path, sheet_name='Database', engine='openpyxl')
        except:
            df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        if VERBOSE_OUTPUT:
            print(f"Error: {e}")
        return pd.DataFrame()

    if tails_filter:
        df['Tail'] = df['Tail'].astype(str).str.strip()
        df = df[df['Tail'].isin(tails_filter)].copy()

    flight_list = []
    if VERBOSE_OUTPUT:
        print(f"Procesando {len(df)} filas...")

    for idx, row in df.iterrows():
        try:
            if isinstance(row['Day'], datetime): s_day = row['Day'].strftime("%Y-%m-%d")
            else: s_day = str(row['Day']).split(' ')[0]

            def get_time_str(val):
                if isinstance(val, datetime) or hasattr(val, 'strftime'): return val.strftime("%H:%M:%S")
                return str(val).strip()

            dep_utc = datetime.strptime(f"{s_day} {get_time_str(row['Dept Time'])}", "%Y-%m-%d %H:%M:%S")
            arr_utc = datetime.strptime(f"{s_day} {get_time_str(row['Arrv Time'])}", "%Y-%m-%d %H:%M:%S")
            if arr_utc < dep_utc: arr_utc += timedelta(days=1)
            
            # Conversión UTC-5
            dep_base = dep_utc - timedelta(hours=UTC_OFFSET_HOURS)
            arr_base = arr_utc - timedelta(hours=UTC_OFFSET_HOURS)
            blk = (arr_utc - dep_utc).total_seconds() / 3600
            
            flight_list.append({
                'id': f"{row['Flt Desg']}_{idx}", 'org': row['Dept Arp'], 'dst': row['Arvl Arp'],
                'dep_utc': dep_utc, 'arr_utc': arr_utc, 'dep_base': dep_base, 'arr_base': arr_base,
                'tail': str(row['Tail']).strip(), 'blk_hours': blk, 'flt_num': row['Flt Desg'],
                'is_dh': False
            })
        except: continue
    return pd.DataFrame(flight_list)

# ==========================================
# 2b. CARGA ITINERARIO PASAJEROS (DH REAL)
# ==========================================

def _parse_time_str(val):
    if isinstance(val, datetime) or hasattr(val, 'strftime'):
        return val.strftime("%H:%M:%S")
    s = str(val).strip()
    if len(s.split(":")) == 2:
        return s + ":00"
    return s

def _fmt_flt_num(val):
    if val is None:
        return ""
    try:
        if isinstance(val, (int, float)):
            return str(int(val))
        s = str(val).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s
    except Exception:
        return str(val).strip()

@lru_cache(maxsize=128)
def _equiv_stations(stn):
    """Devuelve estaciones equivalentes (cached para rendimiento)."""
    stn = str(stn).strip()
    nearby = {
        "GRU": {"VCP"},
        "VCP": {"GRU"},
        "MEX": {"NLU"},
        "NLU": {"MEX"},
    }
    eq = frozenset({stn} | nearby.get(stn, set()))
    return eq

def _direct_only_dh(rules, org, dst):
    bases = getattr(rules, "BASES", {"BOG", "MDE"})
    return (org in bases) or (dst in bases)

def _has_direct_pair(direct_pairs, org, dst):
    for o in _equiv_stations(org):
        for d in _equiv_stations(dst):
            if (o, d) in direct_pairs:
                return True
    return False

def _has_direct_in_window(flights_by_org, times_by_org, org, dest_set, min_dep_utc, max_dep_utc):
    for o in _equiv_stations(org):
        next_list = flights_by_org.get(o, [])
        times = times_by_org.get(o, [])
        if not next_list or not times:
            continue
        start_idx = bisect_left(times, min_dep_utc)
        for i in range(start_idx, len(next_list)):
            nf = next_list[i]
            if nf["dep_utc"] > max_dep_utc:
                break
            if nf["dst"] in dest_set:
                return True
    return False

def _duty_ok_2p(start_dep_utc, end_arr_utc, block_sum, num_sectors, rules,
                 first_flight=None, last_flight=None):
    duty_start = start_dep_utc - rules.REPORT_TIME
    duty_end = end_arr_utc + rules.DEBRIEF
    # Ajuste traslado terrestre VCP/GRU <-> BOG
    if first_flight and first_flight.get('is_dh'):
        f_org = str(first_flight.get('org', '')).upper()
        f_dst = str(first_flight.get('dst', '')).upper()
        if f_org in ('VCP', 'GRU') and f_dst == 'BOG':
            duty_start = duty_start - timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
    if last_flight and last_flight.get('is_dh'):
        l_org = str(last_flight.get('org', '')).upper()
        l_dst = str(last_flight.get('dst', '')).upper()
        if l_org == 'BOG' and l_dst in ('VCP', 'GRU'):
            duty_end = duty_end + timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
    duty_dur = (duty_end - duty_start).total_seconds() / 3600
    start_hour = duty_start.hour
    reduction = 1.0 if (start_hour >= 20 or start_hour < 8) else 0.0
    duty_buffer_h = DUTY_BUFFER_MINUTES / 60.0
    block_buffer_h = BLOCK_BUFFER_MINUTES / 60.0
    max_duty = 12.5 - reduction
    if num_sectors > 6:
        max_duty -= 2.0
    max_duty = max(0.0, max_duty - duty_buffer_h)
    max_blk = max(0.0, 9.0 - block_buffer_h)
    return duty_dur <= max_duty and block_sum <= max_blk

def _load_dh_flights(file_path, sheet_name=None):
    """Carga vuelos DH desde archivo Excel. Si sheet_name es None, usa la primera hoja."""
    if not os.path.exists(file_path):
        print(f"  [WARNING] Archivo DH no encontrado: {file_path}")
        return []
    try:
        # Siempre usar la primera hoja (índice 0) para mayor flexibilidad
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
        print(f"  [OK] Archivo DH cargado: {os.path.basename(file_path)} ({len(df)} filas)")
    except Exception as e:
        print(f"  [ERROR] No se pudo leer archivo DH: {e}")
        return []

    df.columns = [str(c).strip() for c in df.columns]

    # DEBUG: Mostrar columnas detectadas
    print(f"  [DEBUG] Columnas encontradas en archivo DH: {df.columns.tolist()}")
    if len(df) > 0:
        print(f"  [DEBUG] Primera fila de datos: {df.iloc[0].to_dict()}")

    # Detectar automáticamente el mapeo de columnas basándose en el contenido
    # Caso 1: Formato estándar con columnas correctas
    # Caso 2: Formato con columnas desfasadas (Weekday=org, Dept Sta=dep_time, Dept Time=dst, Arvl Time=arr_time)
    col_mapping = {}

    # Primero intentar detectar el formato basándose en las columnas disponibles
    available_cols = set(df.columns)

    # Detectar nombres de columnas de tiempo (soporta "Dept Time", "Dept Time (UTC)", etc.)
    dep_time_col = None
    arr_time_col = None
    for col in available_cols:
        cl = col.lower()
        if cl.startswith('dept') and 'time' in cl and dep_time_col is None:
            dep_time_col = col
        if cl.startswith('arvl') and 'time' in cl and arr_time_col is None:
            arr_time_col = col
    if dep_time_col is None:
        dep_time_col = 'Dept Time'
    if arr_time_col is None:
        arr_time_col = 'Arvl Time'

    # Verificar si las columnas tienen el contenido esperado
    if 'Dept Sta' in df.columns and len(df) > 0:
        sample_dept_sta = df['Dept Sta'].dropna().iloc[0] if not df['Dept Sta'].dropna().empty else None
        # Si Dept Sta contiene tiempos en lugar de códigos de aeropuerto, las columnas están desfasadas
        if sample_dept_sta is not None and (isinstance(sample_dept_sta, pd.Timestamp) or
            (isinstance(sample_dept_sta, str) and ':' in str(sample_dept_sta))):
            # Columnas desfasadas: usar Weekday como origen, Dept Time como destino
            col_mapping = {
                'org': 'Weekday',
                'dst': dep_time_col,
                'dep_time': 'Dept Sta',
                'arr_time': arr_time_col,
                'day': 'Day',
                'flt_num': 'Flt Num'
            }
            print(f"  [INFO] Detectado formato con columnas desfasadas, ajustando mapeo")
        else:
            # Formato estándar
            col_mapping = {
                'org': 'Dept Sta',
                'dst': 'Arvl Sta',
                'dep_time': dep_time_col,
                'arr_time': arr_time_col,
                'day': 'Day',
                'flt_num': 'Flt Num'
            }
            print(f"  [INFO] Usando formato estándar de columnas (dep_time='{dep_time_col}', arr_time='{arr_time_col}')")
    else:
        # Si no hay 'Dept Sta', intentar con nombres de columnas alternativos
        print(f"  [WARNING] No se encontró columna 'Dept Sta', buscando alternativas...")
        # Intentar detectar columnas por contenido o nombre similar
        for col in available_cols:
            if 'dept' in col.lower() and 'sta' in col.lower():
                print(f"  [INFO] Posible columna de origen: {col}")
            if 'arvl' in col.lower() and 'sta' in col.lower():
                print(f"  [INFO] Posible columna de destino: {col}")

        # Si no se puede detectar el formato, usar el estándar por defecto
        col_mapping = {
            'org': 'Dept Sta',
            'dst': 'Arvl Sta',
            'dep_time': dep_time_col,
            'arr_time': arr_time_col,
            'day': 'Day',
            'flt_num': 'Flt Num'
        }

    print(f"  [DEBUG] Mapeo de columnas: {col_mapping}")

    # Verificar que las columnas mapeadas existan
    required_mapped = {col_mapping.get('day', 'Day'), col_mapping.get('org', 'Dept Sta'),
                       col_mapping.get('dep_time', 'Dept Time'), col_mapping.get('dst', 'Arvl Sta'),
                       col_mapping.get('arr_time', 'Arvl Time')}
    missing_cols = required_mapped - available_cols
    if missing_cols:
        print(f"  [ERROR] Columnas requeridas no encontradas: {missing_cols}")
        print(f"  [ERROR] Columnas disponibles: {df.columns.tolist()}")
        print(f"  [INFO] Mostrando muestra de datos para diagnóstico:")
        print(df.head(2).to_string())
        return []

    # Usar mapeo de columnas
    col_day = col_mapping.get('day', 'Day')
    col_org = col_mapping.get('org', 'Dept Sta')
    col_dst = col_mapping.get('dst', 'Arvl Sta')
    col_dep_time = col_mapping.get('dep_time', 'Dept Time')
    col_arr_time = col_mapping.get('arr_time', 'Arvl Time')
    col_flt_num = col_mapping.get('flt_num', 'Flt Num')

    flights = []
    parse_errors = 0
    invalid_airports = 0
    invalid_block = 0

    for idx, row in df.iterrows():
        try:
            # Parsear la fecha en múltiples formatos posibles
            if isinstance(row[col_day], datetime):
                s_day = row[col_day].strftime("%Y-%m-%d")
            else:
                day_str = str(row[col_day]).split(' ')[0]
                # Intentar parsear diferentes formatos de fecha
                date_obj = None
                for fmt in ["%d-%b-%y", "%d-%B-%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]:
                    try:
                        date_obj = datetime.strptime(day_str, fmt)
                        break
                    except ValueError:
                        continue

                if date_obj is None:
                    # Si no se puede parsear, intentar como timestamp de pandas
                    try:
                        date_obj = pd.to_datetime(day_str)
                    except:
                        raise ValueError(f"No se pudo parsear la fecha: {day_str}")

                s_day = date_obj.strftime("%Y-%m-%d")

            dep_utc = datetime.strptime(f"{s_day} {_parse_time_str(row[col_dep_time])}", "%Y-%m-%d %H:%M:%S")
            arr_utc = datetime.strptime(f"{s_day} {_parse_time_str(row[col_arr_time])}", "%Y-%m-%d %H:%M:%S")
            if arr_utc < dep_utc:
                arr_utc += timedelta(days=1)

            blk = (arr_utc - dep_utc).total_seconds() / 3600
            if blk <= 0:
                invalid_block += 1
                continue

            org = str(row[col_org]).strip()
            dst = str(row[col_dst]).strip()
            flt_num = _fmt_flt_num(row.get(col_flt_num, ''))

            if not org or not dst or len(org) != 3 or len(dst) != 3:
                invalid_airports += 1
                if idx < 3:  # Mostrar primeras 3 filas con problemas
                    print(f"  [WARNING] Fila {idx}: aeropuertos inválidos org='{org}' dst='{dst}'")
                continue

            aln = str(row.get("Aln", "")).strip() or "AV"
            flights.append({
                "org": org, "dst": dst,
                "dep_utc": dep_utc, "arr_utc": arr_utc,
                "blk": blk,
                "flt_num": flt_num,
                "aln": aln,
                "is_dh": True,
            })
        except Exception as e:
            parse_errors += 1
            if idx < 3:  # Mostrar primeras 3 filas con errores
                print(f"  [WARNING] Error parseando fila {idx}: {str(e)}")
            continue

    print(f"  [OK] Vuelos DH parseados: {len(flights)}")
    if parse_errors > 0:
        print(f"  [WARNING] Errores de parseo: {parse_errors} filas")
    if invalid_airports > 0:
        print(f"  [WARNING] Aeropuertos inválidos: {invalid_airports} filas")
    if invalid_block > 0:
        print(f"  [WARNING] Block time inválido: {invalid_block} filas")

    if len(flights) == 0 and len(df) > 0:
        print(f"  [ERROR] No se pudo parsear ningún vuelo DH. Revise el formato del archivo.")
        print(f"  [DEBUG] Muestra de la primera fila:")
        print(f"    Day ({col_day}): {df.iloc[0][col_day]}")
        print(f"    Org ({col_org}): {df.iloc[0][col_org]}")
        print(f"    Dst ({col_dst}): {df.iloc[0][col_dst]}")
        print(f"    Dep Time ({col_dep_time}): {df.iloc[0][col_dep_time]}")
        print(f"    Arr Time ({col_arr_time}): {df.iloc[0][col_arr_time]}")

    return flights

def _dh_flights_to_schedule_df(flights):
    rows = []
    for i, f in enumerate(flights or []):
        try:
            dep_utc = f["dep_utc"]
            arr_utc = f["arr_utc"]
            dep_base = dep_utc - timedelta(hours=UTC_OFFSET_HOURS)
            arr_base = arr_utc - timedelta(hours=UTC_OFFSET_HOURS)
            flt_num = _fmt_flt_num(f.get("flt_num", ""))
            aln = f.get("aln", "AV")
            flt_num = f"{aln}{flt_num}" if flt_num else "DH"
            rows.append({
                "id": f"DH_{i}",
                "org": f.get("org"),
                "dst": f.get("dst"),
                "dep_utc": dep_utc,
                "arr_utc": arr_utc,
                "dep_base": dep_base,
                "arr_base": arr_base,
                "tail": "DH",
                "blk_hours": f.get("blk", 0.0),
                "flt_num": flt_num,
                "is_dh": True,
            })
        except Exception:
            continue
    return pd.DataFrame(rows)

def _filter_dh_flights_for_duty(dh_flights, cargo_df):
    if not dh_flights:
        return dh_flights
    if cargo_df is None or cargo_df.empty:
        return dh_flights
    filtered = list(dh_flights)
    if FILTER_DH_TO_CARGO_STATIONS:
        cargo_stations = set(cargo_df["org"]).union(set(cargo_df["dst"]))
        filtered = [f for f in filtered if f.get("org") in cargo_stations and f.get("dst") in cargo_stations]
    if DH_DATE_BUFFER_DAYS is not None:
        try:
            min_dep = cargo_df["dep_utc"].min() - timedelta(days=DH_DATE_BUFFER_DAYS)
            max_dep = cargo_df["dep_utc"].max() + timedelta(days=DH_DATE_BUFFER_DAYS)
            filtered = [f for f in filtered if min_dep <= f.get("dep_utc") <= max_dep]
        except Exception:
            pass
    if MAX_DH_IN_DUTY_FLIGHTS is not None and MAX_DH_IN_DUTY_FLIGHTS > 0:
        filtered = sorted(filtered, key=lambda x: x.get("dep_utc"))
        filtered = filtered[:MAX_DH_IN_DUTY_FLIGHTS]
    return filtered

def _dh_positions_ok(flights):
    if not ALLOW_DH_IN_DUTY:
        return True
    mode = (DH_IN_DUTY_MODE or "").strip().upper()
    if mode not in {"START", "END", "BOTH", "ANY", "NONE"}:
        mode = "ANY"
    if mode == "ANY":
        return True
    if mode == "NONE":
        return not any(f.get("is_dh") for f in flights)
    # START / END / BOTH
    has_non_dh = any(not f.get("is_dh") for f in flights)
    if not has_non_dh:
        return ALLOW_DH_ONLY_DUTY
    first_non = next(i for i, f in enumerate(flights) if not f.get("is_dh"))
    last_non = max(i for i, f in enumerate(flights) if not f.get("is_dh"))
    if mode == "START":
        return all((not f.get("is_dh") or i <= first_non) for i, f in enumerate(flights))
    if mode == "END":
        return all((not f.get("is_dh") or i >= last_non) for i, f in enumerate(flights))
    if mode == "BOTH":
        return all((not f.get("is_dh") or i <= first_non or i >= last_non) for i, f in enumerate(flights))
    return True

def _build_dh_index(flights):
    flights_by_org = {}
    times_by_org = {}
    for f in flights:
        flights_by_org.setdefault(f["org"], []).append(f)
    for org in flights_by_org:
        flights_by_org[org].sort(key=lambda x: x["dep_utc"])
        times_by_org[org] = [x["dep_utc"] for x in flights_by_org[org]]
    return {"flights_by_org": flights_by_org, "times_by_org": times_by_org}

def _cargo_flights_to_dh(df):
    flights = []
    if df is None or df.empty:
        return flights
    for _, row in df.iterrows():
        try:
            dep_utc = row.get("dep_utc")
            arr_utc = row.get("arr_utc")
            if pd.isna(dep_utc) or pd.isna(arr_utc):
                continue
            org = str(row.get("org", "")).strip()
            dst = str(row.get("dst", "")).strip()
            if not org or not dst:
                continue
            blk = float(row.get("blk_hours", 0) or 0)
            if blk <= 0:
                continue
            flt_num = _fmt_flt_num(row.get("flt_num", ""))
            flights.append({
                "org": org,
                "dst": dst,
                "dep_utc": dep_utc,
                "arr_utc": arr_utc,
                "blk": blk,
                "flt_num": flt_num,
            })
        except Exception:
            continue
    return flights

def load_dh_table(file_path, rules, sheet_name='Vuelos dia por dia_UTC', max_legs=4, flights=None, dh_index=None):
    """Construye tabla DH (org,dst) -> horas usando itinerario de pasajeros en UTC.
    Permite escalas si el duty cumple límites de 2P.
    """
    if flights is None:
        flights = _load_dh_flights(file_path, sheet_name=sheet_name)
    if not flights:
        return {}
    if dh_index is None:
        dh_index = _build_dh_index(flights)

    flights_by_org = dh_index.get("flights_by_org", {})
    flights_by_org_times = dh_index.get("times_by_org", {})
    direct_pairs = {(f["org"], f["dst"]) for f in flights}

    dh_table = {}
    for f in flights:
        start_orgs = _equiv_stations(f["org"])
        for start_org in start_orgs:
            stack = [(f, f["dep_utc"], f["arr_utc"], f["blk"], 1, f)]
            while stack:
                curr, start_dep, end_arr, block_sum, legs, first_f = stack.pop()
                direct_only = _has_direct_pair(direct_pairs, start_org, curr["dst"])
                if _duty_ok_2p(start_dep, end_arr, block_sum, legs, rules,
                               first_flight=first_f, last_flight=curr):
                    if direct_only and legs > 1:
                        continue
                    dh_hours = ((end_arr + rules.DEBRIEF) - (start_dep - rules.REPORT_TIME)).total_seconds() / 3600
                    for dest in _equiv_stations(curr["dst"]):
                        prev = dh_table.get((start_org, dest))
                        if prev is None or dh_hours < prev:
                            dh_table[(start_org, dest)] = dh_hours

                if direct_only or legs >= max_legs:
                    continue

                min_dep = end_arr + rules.MIN_CONNECT
                max_dep = end_arr + rules.MAX_CONNECT
                for next_org in _equiv_stations(curr["dst"]):
                    next_list = flights_by_org.get(next_org, [])
                    if not next_list:
                        continue
                    times = flights_by_org_times.get(next_org, [])
                    if not times:
                        continue
                    start_idx = bisect_left(times, min_dep)
                    for i in range(start_idx, len(next_list)):
                        nf = next_list[i]
                        if nf["dep_utc"] > max_dep:
                            break
                        new_block = block_sum + nf["blk"]
                        new_end = nf["arr_utc"]
                        new_legs = legs + 1
                        if not _duty_ok_2p(start_dep, new_end, new_block, new_legs, rules,
                                           first_flight=first_f, last_flight=nf):
                            continue
                        stack.append((nf, start_dep, new_end, new_block, new_legs, first_f))

    return dh_table

def load_dh_data(file_path, rules, sheet_name=None, max_legs=MAX_DH_LEGS, extra_flights=None):
    flights = _load_dh_flights(file_path, sheet_name=sheet_name)

    # DEBUG: Mostrar vuelos DH cargados
    if VERBOSE_OUTPUT:
        print(f"\n[DEBUG DH LOAD] Vuelos DH cargados desde archivo: {len(flights or [])}")
        if flights:
            orgs = set(f.get('org') for f in flights)
            dsts = set(f.get('dst') for f in flights)
            print(f"  Orígenes encontrados: {sorted(orgs)}")
            print(f"  Destinos encontrados: {sorted(dsts)}")
            # Buscar específicamente vuelos BOG->SCL y MIA->BOG
            bog_scl = [f for f in flights if f.get('org') == 'BOG' and f.get('dst') == 'SCL']
            mia_bog = [f for f in flights if f.get('org') == 'MIA' and f.get('dst') == 'BOG']
            bog_gru = [f for f in flights if f.get('org') == 'BOG' and f.get('dst') == 'GRU']
            print(f"  Vuelos BOG->SCL: {len(bog_scl)}")
            print(f"  Vuelos MIA->BOG: {len(mia_bog)}")
            print(f"  Vuelos BOG->GRU: {len(bog_gru)}")
            if bog_scl:
                print(f"    Ejemplo BOG->SCL: {bog_scl[0].get('flt_num')} dep={bog_scl[0].get('dep_utc')}")

    if extra_flights:
        flights = (flights or []) + list(extra_flights)
    if not flights:
        return {}, {}
    dh_index = _build_dh_index(flights)
    dh_table = load_dh_table(file_path, rules, sheet_name=sheet_name, max_legs=max_legs, flights=flights, dh_index=dh_index)

    # DEBUG: Mostrar rango de fechas en el índice para estaciones críticas
    if SECOND_PASS_DEBUG and dh_index:
        times_by_org = dh_index.get("times_by_org", {})
        critical_orgs = ['MIA', 'BOG', 'SCL', 'VCP', 'GRU']
        print("\n[DEBUG DH INDEX] Rango de fechas por origen:")
        for org in critical_orgs:
            times = times_by_org.get(org, [])
            if times:
                first_dt = times[0]
                last_dt = times[-1]
                print(f"  {org}: {len(times)} vuelos, desde {first_dt.strftime('%Y-%m-%d %H:%M')} UTC hasta {last_dt.strftime('%Y-%m-%d %H:%M')} UTC")
            else:
                print(f"  {org}: 0 vuelos en índice")

    return dh_table, dh_index

# ==========================================
# 3. GENERACIÓN DE TRIPS (LOGICA DE NEGOCIO)
# ==========================================

def generate_trips(flights_df, rules, max_dh_hours_per_trip=None, max_dh_ratio=None, allow_dh_only_duties=False):
    """
    Genera trips a partir de vuelos.

    Args:
        flights_df: DataFrame con vuelos
        rules: RuleEngine con reglas
        max_dh_hours_per_trip: Límite de horas DH por trip (None = usar global)
        max_dh_ratio: Límite ratio DH/block (None = usar global)
    """
    # Usar parámetros globales si no se especifican
    if max_dh_hours_per_trip is None:
        max_dh_hours_per_trip = MAX_DH_HOURS_PER_TRIP
    if max_dh_ratio is None:
        max_dh_ratio = MAX_DH_RATIO

    # Permitir duties solo DH para Open Tours (se restaura al finalizar la fase 1)
    global ALLOW_DH_ONLY_DUTY
    prev_allow_dh_only = ALLOW_DH_ONLY_DUTY
    if allow_dh_only_duties:
        ALLOW_DH_ONLY_DUTY = True
    rest_gap_limit = OPEN_TOUR_MAX_REST_GAP_HOURS if allow_dh_only_duties else MAX_REST_GAP_HOURS

    # --- FASE 1: DUTIES ---
    print("\r  [..] Generando duties...                    ", end="", flush=True)

    # Inicializar todos los vuelos con razón de exclusión por defecto
    reset_exclusion_reasons()
    for _, f in flights_df.iterrows():
        set_flight_exclusion(
            f['id'],
            "SIN_DUTY_VALIDO",
            f"Vuelo {f.get('flt', '')} {f.get('org', '')}-{f.get('dst', '')} no forma parte de ningún duty legal"
        )

    # Conjunto para rastrear vuelos que SÍ están en duties válidos
    flights_in_valid_duties = set()

    if VERBOSE_OUTPUT:
        if ALLOW_TAIL_CHANGE_IN_DUTY and ALLOW_DH_IN_DUTY:
            print(f"\n1. Generando Duties (Cambio de avión y DH permitido: {DH_IN_DUTY_MODE})...")
        elif ALLOW_TAIL_CHANGE_IN_DUTY:
            print("\n1. Generando Duties (Cambio de avión permitido)...")
        elif ALLOW_DH_IN_DUTY:
            print(f"\n1. Generando Duties (DH permitido en duty: {DH_IN_DUTY_MODE})...")
        else:
            print("\n1. Generando Duties (Mismo Avión)...")
    G = nx.DiGraph()
    for _, f in flights_df.iterrows(): G.add_node(f['id'], data=f.to_dict())

    flights_by_org = flights_df.groupby('org')
    for _, f1 in flights_df.iterrows():
        if f1['dst'] in flights_by_org.groups:
            for _, f2 in flights_by_org.get_group(f1['dst']).iterrows():
                if not ALLOW_TAIL_CHANGE_IN_DUTY and f1['tail'] != f2['tail']:
                    continue
                conn = f2['dep_base'] - f1['arr_base']
                if rules.MIN_CONNECT <= conn <= rules.MAX_CONNECT:
                    # Debug logging para conexiones con estaciones clave
                    if VERBOSE_OUTPUT and any(stn in f1['dst'] for stn in ['ASU', 'FLN', 'SCL', 'GRU', 'VCP', 'EZE']):
                        print(f"  [GRAPH] Edge: {f1['org']}-{f1['dst']} -> {f2['org']}-{f2['dst']} (conn: {conn.total_seconds()/3600:.2f}h, tail: {f1['tail']}/{f2['tail']})")
                    G.add_edge(f1['id'], f2['id'])
                else:
                    # Debug: mostrar conexiones rechazadas para estaciones clave
                    if VERBOSE_OUTPUT and any(stn in f1['dst'] for stn in ['ASU', 'FLN']):
                        print(f"  [GRAPH REJECT] {f1['org']}-{f1['dst']} -> {f2['org']}-{f2['dst']} (conn: {conn.total_seconds()/3600:.2f}h, min: {rules.MIN_CONNECT.total_seconds()/3600:.2f}h, max: {rules.MAX_CONNECT.total_seconds()/3600:.2f}h)")

    duties = []
    adj = {n: list(G.successors(n)) for n in G.nodes}
    max_legs = max(1, int(MAX_DUTY_LEGS or 1))
    for start_node in G.nodes:
        f_data = G.nodes[start_node]['data']
        valid, _, d = rules.validate_daily_duty([f_data])
        if valid:
            duties.append(d)
            # Marcar vuelo como parte de duty válido
            flights_in_valid_duties.add(f_data['id'])
        # DFS limitado por número de piernas para evitar explosión combinatoria
        stack = [(start_node, [start_node], [f_data])]
        paths_count = 0
        while stack:
            node, path_ids, path_flights = stack.pop()
            if len(path_flights) >= max_legs:
                continue
            for nxt in adj.get(node, []):
                if nxt in path_ids:
                    continue
                new_flights = path_flights + [G.nodes[nxt]['data']]
                if not _dh_positions_ok(new_flights):
                    continue
                valid, _, d = rules.validate_daily_duty(new_flights)

                # Debug logging para duties con estaciones clave
                route = " -> ".join([f"{f['org']}-{f['dst']}" for f in new_flights])
                is_target_route = any(stn in route for stn in ['ASU', 'FLN', 'SCL-MIA', 'GRU-BOG', 'VCP-BOG', 'EZE-SCL'])
                if is_target_route and VERBOSE_OUTPUT:
                    print(f"  [DUTY VALIDATE] {route}: {'VALID' if valid else 'INVALID'}")
                    if not valid:
                        # Try to understand why invalid
                        print(f"    Tails: {[f['tail'] for f in new_flights]}")
                        print(f"    Block: {sum(f.get('blk_hours', 0) for f in new_flights):.2f}h")
                        # Calcular duty para debug
                        first, last = new_flights[0], new_flights[-1]
                        duty_start = first['dep_base'] - rules.REPORT_TIME
                        duty_end = last['arr_base'] + rules.DEBRIEF
                        duty_dur = (duty_end - duty_start).total_seconds() / 3600
                        print(f"    Duty: {duty_dur:.2f}h (start: {duty_start}, end: {duty_end})")

                if not valid:
                    continue
                duties.append(d)
                # Marcar todos los vuelos del duty como parte de duty válido
                for fl in new_flights:
                    flights_in_valid_duties.add(fl['id'])
                stack.append((nxt, path_ids + [nxt], new_flights))
                paths_count += 1
                if MAX_DUTY_PATHS_PER_START and paths_count >= MAX_DUTY_PATHS_PER_START:
                    stack.clear()
                    break
    
    unique_duties = {d['id']: d for d in duties}.values()
    duties = list(unique_duties)
    if ALLOW_DH_IN_DUTY and not allow_dh_only_duties:
        duties = [d for d in duties if any(not f.get("is_dh") for f in d.get("flights", []))]

    # Bloquear pernocta en estaciones prohibidas: eliminar duties que terminan allí
    # Registrar razón de exclusión para vuelos en duties filtrados
    duties_filtered = []
    for d in duties:
        if d['dst'] in rules.PROHIBITED_LAYOVERS:
            # Marcar vuelos de este duty con razón de estación prohibida
            for fl in d.get('flights', []):
                fid = fl['id']
                # Solo actualizar si el vuelo no tiene otra razón válida
                if FLIGHT_EXCLUSION_REASONS.get(fid, {}).get("razon") == "SIN_DUTY_VALIDO":
                    FLIGHT_EXCLUSION_REASONS[fid] = {
                        "razon": "ESTACION_PROHIBIDA",
                        "descripcion": f"Duty termina en {d['dst']} que es estación sin pernocta permitida"
                    }
        else:
            duties_filtered.append(d)
    duties = duties_filtered

    # Aplicar buffer de seguridad: filtrar duties que excedan límites CON buffer
    duty_buffer_h = DUTY_BUFFER_MINUTES / 60.0
    block_buffer_h = BLOCK_BUFFER_MINUTES / 60.0
    duty_limits_1p = {'2P': 12.5, '3P': 17.0, '4P': 20.0}
    block_limits_1p = {'2P': 9.0, '3P': 14.0, '4P': 18.0}

    def check_duty_within_buffer(d):
        """Verifica que duty esté dentro de límites CON buffer aplicado.
        Retorna (is_valid, razon, descripcion)"""
        blk = d.get('block', 0)
        dty = d.get('duty_dur', d.get('duty', 0))  # duty_dur es el campo correcto
        crew = d.get('min_crew', '2P')
        if isinstance(crew, int):
            crew = {2: '2P', 3: '3P', 4: '4P'}.get(crew, '2P')

        # Considerar reducción de duty si empieza entre 15:00 y 03:00
        start_base = d.get('start_base')
        needs_reduction = False
        if start_base:
            start_hour = start_base.hour
            needs_reduction = (start_hour >= DUTY_REDUCTION_START_HOUR or start_hour < DUTY_REDUCTION_END_HOUR)
        reduction = (DUTY_REDUCTION_MINUTES / 60.0) if needs_reduction else 0.0

        # Considerar reducción por sectores (>6 sectores = -2h duty)
        num_sectors = len(d.get('flights', []))
        sector_reduction = 2.0 if num_sectors > 6 else 0.0

        max_blk = block_limits_1p.get(crew, 9.0) - block_buffer_h
        base_duty_limit = duty_limits_1p.get(crew, 12.5) - reduction - sector_reduction
        max_dty = base_duty_limit - duty_buffer_h

        if blk > max_blk:
            return False, "EXCEDE_BLOCK", f"Block {blk:.1f}h excede límite {max_blk:.1f}h para crew {crew}"
        if dty > max_dty:
            return False, "EXCEDE_DUTY", f"Duty {dty:.1f}h excede límite {max_dty:.1f}h para crew {crew}"
        return True, None, None

    before_buffer = len(duties)
    duties_after_buffer = []
    for d in duties:
        is_valid, razon, descripcion = check_duty_within_buffer(d)
        if is_valid:
            duties_after_buffer.append(d)
        else:
            # Registrar razón de exclusión para vuelos de este duty
            for fl in d.get('flights', []):
                fid = fl['id']
                if FLIGHT_EXCLUSION_REASONS.get(fid, {}).get("razon") == "SIN_DUTY_VALIDO":
                    FLIGHT_EXCLUSION_REASONS[fid] = {"razon": razon, "descripcion": descripcion}
    duties = duties_after_buffer
    if VERBOSE_OUTPUT and (before_buffer - len(duties)) > 0:
        print(f"  -> Filtrados {before_buffer - len(duties)} duties por exceder límites con buffer")

    duties.sort(key=lambda x: x['start_base'])
    # Asignar duty_id único para tracking en segunda pasada
    for idx, d in enumerate(duties):
        d['duty_id'] = idx
    print(f"\r  [OK] Duties generados: {len(duties)}              ", flush=True)
    if VERBOSE_OUTPUT:
        print(f"-> {len(duties)} Duties legales.")

    # Restaurar bandera DH-only
    ALLOW_DH_ONLY_DUTY = prev_allow_dh_only

    # --- FASE 2: TRIPS ---
    print("\r  [..] Construyendo trips...                   ", end="", flush=True)
    if VERBOSE_OUTPUT:
        print("\n2. Construyendo Trips (Con bases prohibidas, DH y descansos)...")
    trips = []
    dh_drop = {"missing_transfer": 0, "missing_rescue": 0}
    
    duties_by_org = {}
    for d in duties:
        if d['org'] not in duties_by_org: duties_by_org[d['org']] = []
        duties_by_org[d['org']].append(d)

    # Inicio de trips: directos desde base o con DH de posicionamiento
    stack = []
    for d in duties:
        if d['org'] in rules.BASES:
            night_consec = 1 if d.get("night_touch") else 0
            stack.append((d, [d], d['org'], None, night_consec))
            continue
        for base in rules.BASES:
            pre_info = rules.get_positioning_itinerary(base, d['org'], d['start_base'])
            if not pre_info:
                continue
            pre_dh_hours, pre_arr_base, pre_legs = pre_info
            pre_start_utc = pre_legs[0]['dep_utc'] - rules.REPORT_TIME
            pre_start_base = pre_start_utc - timedelta(hours=UTC_OFFSET_HOURS)
            # DH como duty: descanso según horas voladas y no iniciar duty el mismo día
            dh_block = sum(l.get("blk", 0.0) for l in pre_legs)
            req_rest_dh = rules.calculate_required_rest(dh_block, is_base=False)
            if (not rules.ALLOW_SAME_DAY_DUTY) and d['start_base'].date() <= pre_arr_base.date():
                continue
            # Usar atributo de rules (no la constante global) para permitir DH→Vuelo mismo día en Open Tours
            if (not rules.ALLOW_SAME_DAY_AFTER_DH) and d['start_base'].date() <= pre_arr_base.date():
                continue
            rest_after_dh = (d['start_base'] - pre_arr_base).total_seconds() / 3600
            if rest_after_dh < req_rest_dh:
                continue
            pre_dh = {
                "from": base,
                "to": d['org'],
                "dh_hours": pre_dh_hours,
                "dh_block": dh_block,
                "start_base": pre_start_base,
                "end_base": pre_arr_base,
                "legs": pre_legs,
            }
            night_consec = 1 if d.get("night_touch") else 0
            stack.append((d, [d], base, pre_dh, night_consec))
    
    iter_c = 0
    
    while stack and iter_c < MAX_ITER:
        iter_c += 1
        curr, chain, trip_origin, pre_dh, night_consec = stack.pop()
        
        first_day = chain[0]['start_base'].date()
        if pre_dh and pre_dh.get("start_base"):
            pre_day = pre_dh["start_base"].date()
            if pre_day < first_day:
                first_day = pre_day
        last_day = curr['end_base'].date()
        d_span = (last_day - first_day).days + 1
        if d_span > rules.MAX_TRIP_DAYS: continue

        # ¿Llegamos a base prohibida?
        is_prohibited = (curr['dst'] in rules.PROHIBITED_LAYOVERS)

        # --- A. TRIP COMPLETO (Cerrado en Base de Origen) ---
        if curr['dst'] == trip_origin and not is_prohibited:
            total_dh = 0
            dh_pen = 0
            dh_ok = True
            for i in range(len(chain)-1):
                if chain[i]['dst'] != chain[i+1]['org']:
                    dh = rules.get_dh_time_required(chain[i]['dst'], chain[i+1]['org'])
                    if dh is None:
                        dh_ok = False
                        break
                    total_dh += dh
                    dh_pen += 2000
            if not dh_ok:
                dh_drop["missing_transfer"] += 1
                continue
            if pre_dh:
                total_dh += float(pre_dh.get("dh_hours", 0) or 0)

            # Filtrar trips con exceso de DH
            if total_dh > max_dh_hours_per_trip:
                continue
            if total_dh > 0:
                dh_ratio = total_dh / max(0.1, sum(d['block'] for d in chain))
                if dh_ratio > max_dh_ratio:
                    continue

            # Regla especial: si MAX_TRIP_DAYS=6, el dÃ­a 6 solo puede ser DH de rescate
            if rules.MAX_TRIP_DAYS == 6 and d_span >= rules.MAX_TRIP_DAYS:
                continue

            ranks = [rules.CREW_RANK[d['min_crew']] for d in chain]
            trip_crew = {2:'2P', 3:'3P', 4:'4P'}[max(ranks)]

            # Filtro crew type: trips largos deben justificar su crew type
            if trip_crew in ('3P', '4P') and d_span >= MIN_DAYS_FOR_CREW_CHECK:
                high_crew_count = sum(1 for d in chain if d['min_crew'] == trip_crew)
                if high_crew_count < MIN_HIGH_CREW_DUTIES:
                    continue

            # Filtro eficiencia mínima 1ra pasada
            _trip_blk = sum(d['block'] for d in chain)
            if d_span > 0 and (_trip_blk / d_span) < MIN_BLK_PER_DAY_FIRST_PASS:
                continue

            base_cost = sum(d['base_cost'] for d in chain)
            crew_cost = (0 if trip_crew == '2P' else 500) * d_span
            crew_pen = PENALTY_4P if trip_crew == '4P' else 0

            trips.append({
                'id': len(trips), 'chain': chain,
                'cost': base_cost + crew_cost + (d_span*TRIP_DAY_PENALTY) + dh_pen + crew_pen,
                'block': sum(d['block'] for d in chain), 'days': d_span,
                'trip_crew': trip_crew, 'base': trip_origin,
                'flights_covered': set(f['id'] for d in chain for f in d['flights']),
                'has_rescue_dh': False, 'total_dh_hours': total_dh,
                'pre_dh': pre_dh
            })
            continue

        # --- B. EXTENSIÓN (Si no estamos en base prohibida) ---
        if not is_prohibited and d_span < rules.MAX_TRIP_DAYS:
            candidates = duties_by_org.get(curr['dst'], [])
            for next_d in candidates:
                # Regla: no iniciar duty el mismo día calendario del duty anterior
                if not rules.ALLOW_SAME_DAY_DUTY and next_d['start_base'].date() <= curr['end_base'].date():
                    continue

                # Regla Descanso (incluye DH como duty si hay transferencia)
                eff_block = curr['block']
                req_rest = rules.calculate_required_rest(eff_block, is_base=False)
                available_rest = (next_d['start_base'] - curr['end_base']).total_seconds() / 3600

                required_gap = req_rest
                if curr['dst'] != next_d['org']:
                    dh = rules.get_dh_time_required(curr['dst'], next_d['org'])
                    if dh is None:
                        continue
                    buffer_h = (rules.REPORT_TIME + rules.DEBRIEF).total_seconds() / 3600
                    dh_block = max(0.0, dh - buffer_h)
                    req_rest_dh = rules.calculate_required_rest(dh_block, is_base=False)
                    required_gap = req_rest + dh + req_rest_dh
                    if not rules.ALLOW_SAME_DAY_AFTER_DH:
                        dh_arrival = curr['end_base'] + timedelta(hours=dh)
                        if next_d['start_base'].date() <= dh_arrival.date():
                            continue

                # Máximo 3 jornadas consecutivas con ventana 01:00-05:00
                if next_d.get("night_touch"):
                    next_consec = (night_consec + 1) if curr.get("night_touch") else 1
                else:
                    next_consec = 0
                if next_consec > MAX_CONSEC_NIGHT_DUTIES:
                    continue

                # Verificar pernocta en estación prohibida
                # Si el siguiente duty empieza en otro día, hay pernocta en curr['dst']
                if next_d['start_base'].date() > curr['end_base'].date():
                    if curr['dst'] in rules.PROHIBITED_LAYOVERS:
                        continue  # No permitir pernocta en estación prohibida

                if available_rest >= required_gap:
                    if (rest_gap_limit is None) or (available_rest <= rest_gap_limit):
                        stack.append((next_d, chain + [next_d], trip_origin, pre_dh, next_consec))

        # --- C. RESCATE (Forzar DH a Base) ---
        req_rest_before = rules.calculate_required_rest(curr['block'], is_base=False)
        min_report = curr['end_base'] + timedelta(hours=req_rest_before)
        if not rules.ALLOW_SAME_DAY_DUTY:
            next_day_start = datetime.combine(curr['end_base'].date() + timedelta(days=1), datetime.min.time())
            rescue_report_base = max(min_report, next_day_start)
        else:
            rescue_report_base = min_report
        rescue_info = rules.get_rescue_itinerary(curr['dst'], trip_origin, rescue_report_base)
        rescue_legs = None
        if rescue_info:
            dh_hours, arrival_at_base, rescue_legs = rescue_info
        else:
            dh_drop["missing_rescue"] += 1
            continue
        final_d_span = (arrival_at_base.date() - first_day).days + 1
        
        if final_d_span <= rules.MAX_TRIP_DAYS:
            # Regla especial: si MAX_TRIP_DAYS=6, el dÃ­a 6 solo puede ser DH de rescate
            if rules.MAX_TRIP_DAYS == 6 and final_d_span == rules.MAX_TRIP_DAYS:
                if not rescue_legs:
                    continue
                if curr['end_base'].date() >= arrival_at_base.date():
                    continue
            # Calcular Rest necesario post-rescate
            buffer_h = (rules.REPORT_TIME + rules.DEBRIEF).total_seconds() / 3600
            if rescue_legs:
                dh_block = sum(l.get("blk", 0.0) for l in rescue_legs)
            else:
                dh_block = max(0.0, dh_hours - buffer_h)
            eff_block = curr['block'] + dh_block 
            req_rest_after = rules.calculate_required_rest(eff_block, is_base=True)

            total_dh = dh_hours
            dh_ok = True
            for i in range(len(chain)-1):
                if chain[i]['dst'] != chain[i+1]['org']:
                    dh = rules.get_dh_time_required(chain[i]['dst'], chain[i+1]['org'])
                    if dh is None:
                        dh_ok = False
                        break
                    total_dh += dh
            if not dh_ok:
                dh_drop["missing_transfer"] += 1
                continue
            if pre_dh:
                total_dh += float(pre_dh.get("dh_hours", 0) or 0)

            # Filtrar trips con exceso de DH
            if total_dh > max_dh_hours_per_trip:
                continue
            if total_dh > 0:
                dh_ratio = total_dh / max(0.1, sum(d['block'] for d in chain))
                if dh_ratio > max_dh_ratio:
                    continue

            ranks = [rules.CREW_RANK[d['min_crew']] for d in chain]
            trip_crew = {2:'2P', 3:'3P', 4:'4P'}[max(ranks)]

            # Filtro crew type: trips largos deben justificar su crew type
            if trip_crew in ('3P', '4P') and final_d_span >= MIN_DAYS_FOR_CREW_CHECK:
                high_crew_count = sum(1 for d in chain if d['min_crew'] == trip_crew)
                if high_crew_count < MIN_HIGH_CREW_DUTIES:
                    continue

            base_cost = sum(d['base_cost'] for d in chain)
            crew_cost = (0 if trip_crew == '2P' else 500) * final_d_span
            crew_pen = PENALTY_4P if trip_crew == '4P' else 0

            # Penalizaciones: DH + Prohibida + Rescate
            pen = 3000 + (dh_hours * 200) + (5000 if is_prohibited else 0)

            trips.append({
                'id': len(trips), 'chain': chain,
                'cost': base_cost + crew_cost + (final_d_span*TRIP_DAY_PENALTY) + pen + crew_pen,
                'block': sum(d['block'] for d in chain), 'days': final_d_span,
                'trip_crew': trip_crew, 'base': trip_origin,
                'flights_covered': set(f['id'] for d in chain for f in d['flights']),
                'has_rescue_dh': True, 'rescue_from': curr['dst'],
                'total_dh_hours': total_dh, 'final_rest_req': req_rest_after,
                'rescue_dh_hours': dh_hours, 'rescue_arr_base': arrival_at_base,
                'rescue_legs': rescue_legs, 'pre_dh': pre_dh
            })

    print(f"\r  [OK] Trips generados (1ra pasada): {len(trips)}       ", flush=True)
    if VERBOSE_OUTPUT:
        print(f"-> {len(trips)} Trips generados (primera pasada).")
        print(f"-> Trips descartados por DH: transfer={dh_drop['missing_transfer']}, rescue={dh_drop['missing_rescue']}")

    # SEGUNDA PASADA: Identificar duties sin cubrir y crear trips con DH
    if ENABLE_SECOND_PASS_DH:
        # Identificar duties cubiertos en primera pasada
        all_duty_ids = set(range(len(duties)))
        covered_duty_ids = set()
        for trip in trips:
            for duty in trip['chain']:
                duty_id = duty.get('duty_id')
                if duty_id is not None:
                    covered_duty_ids.add(duty_id)

        uncovered_duty_ids = all_duty_ids - covered_duty_ids
        uncovered_duties = [duties[i] for i in uncovered_duty_ids if i < len(duties)]

        if VERBOSE_OUTPUT:
            print(f"\n=== PRIMERA PASADA COMPLETADA ===")
            print(f"Trips generados: {len(trips)}")
            print(f"Duties cubiertos: {len(covered_duty_ids)} / {len(duties)}")
            print(f"Duties sin cubrir: {len(uncovered_duties)}")

        # Ejecutar segunda pasada si hay duties sin cubrir
        if uncovered_duties:
            print(f"\r  [..] Segunda pasada ({len(uncovered_duties)} duties)...   ", end="", flush=True)
            if VERBOSE_OUTPUT:
                print(f"\n=== INICIANDO SEGUNDA PASADA ===")

            # Crear rules con permisos relajados para Open Tours (segunda pasada)
            # Permitir DH + vuelo el mismo día para cubrir vuelos como SCL-MIA, VCP-BOG, etc.
            rules_second_pass = RuleEngine(
                dh_table=rules.DH_TABLE,
                dh_index=rules.DH_INDEX,
                allowed_crews=rules.ALLOWED_CREWS,
                allow_same_day_duty=OPEN_TOUR_ALLOW_SAME_DAY  # True para permitir DH+vuelo mismo día
            )
            # Ajustar límites de DH para segunda pasada (más permisivos)
            rules_second_pass.MAX_DH_LEGS = OPEN_TOUR_MAX_DH_LEGS  # 4 escalas para rutas complejas (BOG→GRU→VCP)
            rules_second_pass.MAX_DH_HOURS = OPEN_TOUR_MAX_DH_HOURS  # Ventana extendida para DH 1-2 días antes

            new_trips, covered_flights = cover_open_time_with_dh(
                uncovered_duties,
                rules_second_pass,
                trips,
                max_trip_days=MAX_TRIP_DAYS_SECOND_PASS,
                verbose=VERBOSE_OUTPUT
            )

            trips.extend(new_trips)
            print(f"\r  [OK] Segunda pasada: +{len(new_trips)} trips          ", flush=True)

            if VERBOSE_OUTPUT:
                print(f"=== SEGUNDA PASADA COMPLETADA ===")
                print(f"Trips adicionales: {len(new_trips)}")
                print(f"Vuelos cubiertos adicionales: {len(covered_flights)}")
                print(f"Total trips: {len(trips)}")

            # TERCERA PASADA: Crear trips parciales para duties que aún no tienen cobertura
            if ENABLE_THIRD_PASS_PARTIAL:
                # Identificar duties cubiertos después de segunda pasada
                covered_duty_ids_after_2nd = set()
                for trip in trips:
                    for duty in trip['chain']:
                        duty_id = duty.get('duty_id')
                        if duty_id is not None:
                            covered_duty_ids_after_2nd.add(duty_id)

                still_uncovered_duty_ids = all_duty_ids - covered_duty_ids_after_2nd
                still_uncovered_duties = [duties[i] for i in still_uncovered_duty_ids if i < len(duties)]

                if still_uncovered_duties:
                    print(f"\r  [..] Tercera pasada - trips parciales ({len(still_uncovered_duties)} duties)...   ", end="", flush=True)

                    partial_trips, partial_covered = create_partial_trips(
                        still_uncovered_duties,
                        rules_second_pass,
                        trips,
                        max_trip_days=MAX_TRIP_DAYS_SECOND_PASS,
                        verbose=VERBOSE_OUTPUT
                    )

                    trips.extend(partial_trips)
                    print(f"\r  [OK] Tercera pasada: +{len(partial_trips)} trips parciales   ", flush=True)

                    if VERBOSE_OUTPUT:
                        print(f"=== TERCERA PASADA COMPLETADA ===")
                        print(f"Trips parciales: {len(partial_trips)}")
                        print(f"Vuelos cubiertos adicionales: {len(partial_covered)}")
                        print(f"Total trips: {len(trips)}")

    # Actualizar razones de exclusión para vuelos en trips
    # Vuelos que están en algún trip: cambiar razón a NO_SELECCIONADO (el optimizador decidirá)
    flights_in_trips = set()
    for t in trips:
        for fid in t.get('flights_covered', []):
            flights_in_trips.add(fid)

    for fid in flights_in_trips:
        FLIGHT_EXCLUSION_REASONS[fid] = {
            "razon": "NO_SELECCIONADO",
            "descripcion": "Vuelo existe en trip válido pero no fue seleccionado por el optimizador"
        }

    # Vuelos en duties válidos pero no en trips → SIN_TRIP_VIABLE
    for fid in flights_in_valid_duties:
        if fid not in flights_in_trips:
            if FLIGHT_EXCLUSION_REASONS.get(fid, {}).get("razon") == "SIN_DUTY_VALIDO":
                FLIGHT_EXCLUSION_REASONS[fid] = {
                    "razon": "SIN_TRIP_VIABLE",
                    "descripcion": "Vuelo forma parte de duty válido pero no se pudo construir un trip que regrese a base"
                }

    return trips


# ==========================================
# 3B. SEGUNDA PASADA - CUBRIR OPEN TIME CON DH
# ==========================================

def cover_open_time_with_dh(uncovered_duties, rules, trips, max_trip_days=4, verbose=False):
    """
    Segunda pasada: intenta cubrir vuelos open time creando trips con DHs de la base de datos.

    Lógica:
    PASO A: CONSTRUIR CADENAS DE DUTIES CONECTADOS
    1. Para cada duty sin cubrir, intentar extenderlo:
       - Buscar otros duties open que conectan (duty.dst == next.org)
       - Validar conexión (timing, tail si aplica)
       - Crear cadenas de 1, 2, 3+ duties

    PASO B: CREAR TRIPS CON DHs
    2. Para cada cadena de duties
    3. Para cada base (BOG, MDE)
       a) Buscar DH positioning: base → primer_duty_org
       b) Si no existe en BD, continuar
       c) Buscar DH rescue: último_duty_dst → base
       d) Si no existe en BD, continuar
       e) Validar trip resultante (días, DH ratio, descansos)
       f) Si es legal, crear trip y agregarlo

    Returns:
        (trips_agregados, vuelos_cubiertos)
    """
    new_trips = []
    covered_flights = set()
    used_duty_ids = set()  # Para no reusar duties ya combinados

    # Contadores de debug para saber por qué se rechazan trips
    rejection_reasons = {
        'no_positioning_dh': 0,
        'same_day_duty_positioning': 0,
        'insufficient_rest_after_dh': 0,
        'insufficient_rest_between_duties': 0,  # REST insuficiente entre duties consecutivos en cadena
        'no_rescue_dh': 0,
        'exceeds_max_days': 0,
        'exceeds_max_dh_hours': 0,
        'exceeds_dh_ratio': 0,
        'prohibited_layover': 0,
        'exceeds_block_buffer': 0,
        'exceeds_duty_buffer': 0,
    }

    # Calcular límites con buffer para validación
    duty_buffer_h = DUTY_BUFFER_MINUTES / 60.0
    block_buffer_h = BLOCK_BUFFER_MINUTES / 60.0

    # Límites base por tipo de crew (se reducirán con buffer)
    duty_limits = {'2P': 12.5, '3P': 17.0, '4P': 20.0}
    block_limits = {'2P': 9.0, '3P': 14.0, '4P': 18.0}

    # NOTA: Ya NO filtramos duties con destino en estaciones prohibidas aquí.
    # Permitimos que estos duties se combinen en cadenas (ej: MIA-FLN + FLN-VCP).
    # Las cadenas que TERMINAN en estaciones prohibidas se filtrarán después de construirlas.
    # Los duties intermedios en estaciones prohibidas se validan durante la construcción de cadenas
    # para asegurar que NO haya pernocta en esa estación (solo conexiones mismo día).
    original_count = len(uncovered_duties)
    prohibited_duties = [d for d in uncovered_duties if d['dst'] in rules.PROHIBITED_LAYOVERS]
    if len(prohibited_duties) > 0 and verbose:
        print(f"  [2nd Pass] {len(prohibited_duties)} duties terminan en estación prohibida (se intentarán combinar en cadenas)")

    # Filtrar duties que exceden límites de block/duty CON buffer
    def duty_exceeds_buffer(duty):
        """Verifica si un duty excede los límites de block/duty con buffer aplicado."""
        block_hrs = duty.get('block', 0)
        duty_hrs = duty.get('duty_dur', duty.get('duty', 0))
        crew = duty.get('min_crew', '2P')
        if isinstance(crew, int):
            crew = {2: '2P', 3: '3P', 4: '4P'}.get(crew, '2P')

        # Considerar reducción de duty si empieza entre 15:00 y 03:00
        start_base = duty.get('start_base')
        needs_reduction = False
        if start_base:
            start_hour = start_base.hour
            needs_reduction = (start_hour >= DUTY_REDUCTION_START_HOUR or start_hour < DUTY_REDUCTION_END_HOUR)
        reduction = (DUTY_REDUCTION_MINUTES / 60.0) if needs_reduction else 0.0

        # Considerar reducción por sectores (>6 sectores = -2h duty)
        num_sectors = len(duty.get('flights', []))
        sector_reduction = 2.0 if num_sectors > 6 else 0.0

        max_block = block_limits.get(crew, 9.0) - block_buffer_h
        base_duty_limit = duty_limits.get(crew, 12.5) - reduction - sector_reduction
        max_duty = base_duty_limit - duty_buffer_h

        return block_hrs > max_block or duty_hrs > max_duty

    before_buffer_filter = len(uncovered_duties)
    uncovered_duties = [d for d in uncovered_duties if not duty_exceeds_buffer(d)]
    buffer_filtered = before_buffer_filter - len(uncovered_duties)
    if buffer_filtered > 0 and verbose:
        print(f"  [2nd Pass] Filtrados {buffer_filtered} duties que exceden límites con buffer")

    # DEBUG: Mostrar vuelos DH disponibles para estaciones clave
    if SECOND_PASS_DEBUG and verbose:
        dh_index = rules.DH_INDEX or {}
        flights_by_org = dh_index.get("flights_by_org", {})
        times_by_org = dh_index.get("times_by_org", {})
        print(f"\n[DEBUG DH INDEX] Vuelos DH disponibles (rango de fechas):")
        for stn in ['BOG', 'MDE', 'SCL', 'MIA', 'VCP', 'GRU']:
            flights = flights_by_org.get(stn, [])
            times = times_by_org.get(stn, [])
            if flights and times:
                dsts = set(f.get('dst') for f in flights[:50])  # Primeros 50 para no saturar
                first_date = times[0].strftime('%Y-%m-%d')
                last_date = times[-1].strftime('%Y-%m-%d')
                print(f"  {stn} -> {sorted(dsts)} ({len(flights)} vuelos, fechas: {first_date} a {last_date})")
            else:
                print(f"  {stn} -> (sin vuelos DH)")

        # Mostrar los duties que se intentan cubrir con sus fechas
        print(f"\n[DEBUG] Duties sin cubrir que se intentarán procesar:")
        for duty in uncovered_duties[:10]:  # Primeros 10
            org = duty.get('org')
            dst = duty.get('dst')
            start_base = duty.get('start_base')
            end_base = duty.get('end_base')
            duty_id = duty.get('duty_id')
            if start_base and end_base:
                print(f"  Duty #{duty_id}: {org}-{dst} | {start_base.strftime('%Y-%m-%d %H:%M')} a {end_base.strftime('%Y-%m-%d %H:%M')} (base time)")

    # PASO A: Construir cadenas de duties conectados
    duty_chains = []
    # Track duties that are successfully in chains (not rejected)
    duties_in_valid_chains = set()

    # Primero: intentar combinar duties
    for duty in uncovered_duties:
        if duty.get('duty_id') in used_duty_ids:
            continue  # Ya usado en otra cadena

        # Iniciar cadena con este duty
        chain = [duty]
        used_duty_ids.add(duty.get('duty_id'))

        # Intentar extender la cadena hacia adelante
        current_dst = duty['dst']
        current_end_base = duty['end_base']

        while True:
            # Buscar siguiente duty que conecta
            next_duty = None
            for candidate in uncovered_duties:
                if candidate.get('duty_id') in used_duty_ids:
                    continue

                # ¿Conecta?
                if candidate['org'] != current_dst:
                    continue

                # Validar timing de conexión
                connection_time = (candidate['start_base'] - current_end_base).total_seconds() / 3600
                if connection_time < (rules.MIN_CONNECT.total_seconds() / 3600):
                    continue  # Conexión muy rápida

                # Determinar si hay pernocta (días diferentes)
                is_overnight = candidate['start_base'].date() > current_end_base.date()

                if is_overnight:
                    # HAY PERNOCTA - validar descanso mínimo y estación
                    current_block = chain[-1].get('block', 0) if chain else 0
                    required_rest = rules.calculate_required_rest(current_block, is_base=False)
                    actual_rest = connection_time
                    if actual_rest < required_rest:
                        # No hay suficiente descanso entre duties
                        rejection_reasons['insufficient_rest_between_duties'] += 1
                        if verbose:
                            print(f"    [CHAIN REJECT] Insufficient rest between duties: {actual_rest:.1f}h < {required_rest:.1f}h required")
                        continue

                    # Verificar que la estación de pernocta NO esté prohibida
                    if current_dst in rules.PROHIBITED_LAYOVERS:
                        if verbose:
                            print(f"    [CHAIN REJECT] Overnight in prohibited station: {current_dst}")
                        continue  # No permitir pernocta en estación prohibida
                else:
                    # MISMO DÍA - aplicar MAX_CONNECT para conexiones sin pernocta
                    if connection_time > (rules.MAX_CONNECT.total_seconds() / 3600):
                        continue  # Conexión muy lenta para mismo día

                # Validar tail si ALLOW_TAIL_CHANGE_IN_DUTY = False
                # NOTA: Entre duties diferentes SÍ se permite cambio de tail
                # Solo se restringe dentro del mismo duty

                # NUEVO: Verificar que la cadena extendida no exceda límites de block
                # Si excede, no agregar el candidato a esta cadena (dejarlo disponible para standalone)
                candidate_chain = chain + [candidate]
                candidate_block = sum(d.get('block', 0) for d in candidate_chain)
                # Determinar crew máximo de la cadena candidata
                cand_max_crew = max(d.get('req_crew', d.get('min_crew', 2)) for d in candidate_chain)
                if isinstance(cand_max_crew, int):
                    cand_crew = {2: '2P', 3: '3P', 4: '4P'}.get(cand_max_crew, '2P')
                else:
                    cand_crew = cand_max_crew
                cand_max_block = block_limits.get(cand_crew, 9.0) - block_buffer_h
                if candidate_block > cand_max_block:
                    # La cadena extendida excedería límites - no agregar este candidato
                    # Dejarlo disponible para ser procesado como cadena standalone
                    continue

                # Candidato válido
                next_duty = candidate
                break

            if not next_duty:
                break  # No hay más extensiones

            # Agregar a la cadena
            chain.append(next_duty)
            used_duty_ids.add(next_duty.get('duty_id'))
            current_dst = next_duty['dst']
            current_end_base = next_duty['end_base']

        duty_chains.append(chain)

    # Agregar duties que no se combinaron (cadenas de 1 duty)
    for duty in uncovered_duties:
        if duty.get('duty_id') not in used_duty_ids:
            duty_chains.append([duty])
            used_duty_ids.add(duty.get('duty_id'))

    if verbose:
        chains_multi = [c for c in duty_chains if len(c) > 1]
        chains_single = [c for c in duty_chains if len(c) == 1]
        print(f"  Cadenas de duties: {len(chains_multi)} multi-duty, {len(chains_single)} single-duty")

    # Filtrar cadenas que TERMINAN en estaciones prohibidas (no pueden pernoctar ahí antes del DH de rescate)
    # Pero permitimos cadenas con estaciones prohibidas INTERMEDIAS (conexión mismo día)
    before_filter = len(duty_chains)
    duty_chains = [chain for chain in duty_chains if chain[-1]['dst'] not in rules.PROHIBITED_LAYOVERS]
    filtered_chains = before_filter - len(duty_chains)
    if filtered_chains > 0 and verbose:
        print(f"  [2nd Pass] Filtradas {filtered_chains} cadenas que terminan en estación prohibida")

    # PASO B: Crear trips con DHs para cada cadena
    for chain in duty_chains:
        # Información de la cadena
        first_duty = chain[0]
        last_duty = chain[-1]

        chain_org = first_duty['org']
        chain_dst = last_duty['dst']
        chain_start_base = first_duty['start_base']
        chain_end_base = last_duty['end_base']

        # Calcular totales de la cadena
        chain_block = sum(d.get('block', 0) for d in chain)
        chain_duty_hours = sum(d.get('duty', 0) for d in chain)
        chain_flights = []
        for d in chain:
            chain_flights.extend(d.get('flights', []))

        # Determinar crew requerido de la cadena
        max_crew = max(d.get('req_crew', d.get('min_crew', 2)) for d in chain)
        if isinstance(max_crew, str):
            chain_crew = max_crew
        else:
            chain_crew = {2: '2P', 3: '3P', 4: '4P'}.get(max_crew, '2P')

        # VALIDAR límites de block/duty para la CADENA COMBINADA
        # Considerar reducción de duty si la cadena empieza entre 15:00 y 03:00
        chain_start_hour = chain_start_base.hour
        chain_needs_reduction = (chain_start_hour >= DUTY_REDUCTION_START_HOUR or chain_start_hour < DUTY_REDUCTION_END_HOUR)
        chain_reduction = (DUTY_REDUCTION_MINUTES / 60.0) if chain_needs_reduction else 0.0

        # Considerar reducción por sectores (>6 sectores = -2h duty)
        total_sectors = sum(len(d.get('flights', [])) for d in chain)
        chain_sector_reduction = 2.0 if total_sectors > 6 else 0.0

        # Calcular límites máximos para la cadena
        chain_max_block = block_limits.get(chain_crew, 9.0) - block_buffer_h
        chain_base_duty_limit = duty_limits.get(chain_crew, 12.5) - chain_reduction - chain_sector_reduction
        chain_max_duty = chain_base_duty_limit - duty_buffer_h

        # Verificar si la cadena excede límites
        chain_exceeds_block = chain_block > chain_max_block
        chain_exceeds_duty = chain_duty_hours > chain_max_duty

        if chain_exceeds_block or chain_exceeds_duty:
            rejection_reasons['exceeds_block_buffer'] += 1 if chain_exceeds_block else 0
            rejection_reasons['exceeds_duty_buffer'] += 1 if chain_exceeds_duty else 0
            if verbose:
                if chain_exceeds_block:
                    print(f"    [CHAIN SKIP] Block {chain_block:.1f}h > max {chain_max_block:.1f}h for {chain_crew}")
                if chain_exceeds_duty:
                    print(f"    [CHAIN SKIP] Duty {chain_duty_hours:.1f}h > max {chain_max_duty:.1f}h for {chain_crew}")
            continue  # Saltar esta cadena

        # Debug: Log cadenas específicas que involucran MIA, ASU, VCP
        chain_route = " -> ".join([f"{d['org']}-{d['dst']}" for d in chain])
        is_target_chain = any(stn in chain_route for stn in ['MIA', 'ASU', 'VCP', 'FLN', 'GRU'])

        if SECOND_PASS_DEBUG and verbose and is_target_chain:
            print(f"\n  [DEBUG CHAIN] Procesando: {chain_route}")
            print(f"    Start: {chain_start_base.strftime('%Y-%m-%d %H:%M')} (base time)")
            print(f"    End: {chain_end_base.strftime('%Y-%m-%d %H:%M')} (base time)")
            print(f"    Block: {chain_block:.2f}h")

        for base in rules.BASES:
            # PASO 1: Buscar DH positioning (base → chain_org)
            # Para segunda pasada, expandir ventana de búsqueda para permitir trips multi-día
            # Buscar hasta MAX_TRIP_DAYS_SECOND_PASS días hacia atrás
            expanded_search_hours = max_trip_days * 24  # e.g., 6 días = 144 horas

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                print(f"    [SEARCH] Positioning DH: {base} -> {chain_org}, arrive by {chain_start_base.strftime('%Y-%m-%d %H:%M')}")
                print(f"      Search window: {expanded_search_hours}h backward (vs normal {rules.MAX_DH_HOURS}h)")

            # Si ALLOW_SAME_DAY_AFTER_DH es False, buscar DH que llegue el día anterior
            if not rules.ALLOW_SAME_DAY_AFTER_DH:
                # Limitar búsqueda para que el DH llegue antes del inicio del día del duty
                latest_arrival_for_search = datetime.combine(
                    chain_start_base.date(),
                    datetime.min.time()
                )
            else:
                latest_arrival_for_search = chain_start_base

            pre_info = rules.get_positioning_itinerary(
                base, chain_org, latest_arrival_for_search,
                max_legs=rules.MAX_DH_LEGS,
                max_hours=expanded_search_hours  # Usar ventana expandida
            )
            if not pre_info:
                rejection_reasons['no_positioning_dh'] += 1
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"      [REJECT] No positioning DH found (search until {latest_arrival_for_search})")
                continue  # No hay DH en BD

            pre_dh_hours, pre_arr_base, pre_legs = pre_info
            pre_start_utc = pre_legs[0]['dep_utc'] - rules.REPORT_TIME
            pre_start_base = pre_start_utc - timedelta(hours=UTC_OFFSET_HOURS)

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                dh_route = " -> ".join([f"{leg['org']}-{leg['dst']}" for leg in pre_legs])
                print(f"      [FOUND] Positioning DH: {dh_route}")
                print(f"        Departs: {pre_start_base.strftime('%Y-%m-%d %H:%M')} (base time)")
                print(f"        Arrives: {pre_arr_base.strftime('%Y-%m-%d %H:%M')} (base time)")
                print(f"        Duration: {pre_dh_hours:.2f}h")

            # Validar descanso post-DH
            dh_block = sum(l.get("blk", 0.0) for l in pre_legs)
            req_rest_dh = rules.calculate_required_rest(dh_block, is_base=False)

            # Validar pernoctas en estaciones prohibidas durante el DH de posicionamiento
            # Verificar si hay pernoctas en escalas intermedias del DH
            dh_has_prohibited_layover = False
            for i, leg in enumerate(pre_legs):
                leg_dst = leg.get('dst', '')
                leg_arr = leg.get('arr_utc')
                # Verificar si hay siguiente leg y si hay pernocta entre ellos
                if i < len(pre_legs) - 1:
                    next_leg = pre_legs[i + 1]
                    next_dep = next_leg.get('dep_utc')
                    if leg_arr and next_dep and leg_arr.date() < next_dep.date():
                        # Hay pernocta en leg_dst
                        if leg_dst in rules.PROHIBITED_LAYOVERS:
                            dh_has_prohibited_layover = True
                            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                                print(f"      [REJECT] DH layover in prohibited station: {leg_dst}")
                            break

            if dh_has_prohibited_layover:
                rejection_reasons['prohibited_layover'] += 1
                continue

            # Validar pernocta en estación prohibida después del DH de posicionamiento
            # Si el DH llega en un día diferente al inicio del duty, hay pernocta en chain_org
            if pre_arr_base.date() < chain_start_base.date():
                if chain_org in rules.PROHIBITED_LAYOVERS:
                    rejection_reasons['prohibited_layover'] += 1
                    if SECOND_PASS_DEBUG and verbose and is_target_chain:
                        print(f"      [REJECT] Layover in prohibited station: {chain_org}")
                    continue

            # Usar regla desde RuleEngine en lugar de variable global
            if not rules.ALLOW_SAME_DAY_AFTER_DH:
                if chain_start_base.date() <= pre_arr_base.date():
                    rejection_reasons['same_day_duty_positioning'] += 1
                    if SECOND_PASS_DEBUG and verbose and is_target_chain:
                        print(f"      [REJECT] Same-day duty: DH arrives {pre_arr_base.date()}, duty starts {chain_start_base.date()}")
                    continue

            rest_after_dh = (chain_start_base - pre_arr_base).total_seconds() / 3600
            if rest_after_dh < req_rest_dh:
                rejection_reasons['insufficient_rest_after_dh'] += 1
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"      [REJECT] Insufficient rest: {rest_after_dh:.2f}h < {req_rest_dh:.2f}h required")
                continue

            # PASO 2: Buscar DH rescue (chain_dst → base)
            # Si el duty ya termina en la base, NO necesita DH de rescate
            # Normalizar para comparación (strip y upper)
            chain_dst_norm = str(chain_dst).strip().upper()
            base_norm = str(base).strip().upper()

            if chain_dst_norm == base_norm:
                # El duty ya termina en la base, crear trip sin DH de rescate
                first_day = pre_start_base.date()
                last_day = chain_end_base.date()
                trip_days = (last_day - first_day).days + 1

                if trip_days > max_trip_days:
                    rejection_reasons['exceeds_max_days'] += 1
                    continue

                # Solo tiene DH de posicionamiento, no de rescate
                total_dh_hours = pre_dh_hours
                dh_ratio = total_dh_hours / chain_block if chain_block > 0 else 0

                if total_dh_hours > MAX_DH_HOURS_PER_TRIP:
                    rejection_reasons['exceeds_max_dh_hours'] += 1
                    continue

                if dh_ratio > MAX_DH_RATIO_SECOND_PASS:
                    rejection_reasons['exceeds_dh_ratio'] += 1
                    continue

                # Determinar crew requerido
                max_crew = max(d.get('req_crew', 2) for d in chain)
                trip_crew = {2: '2P', 3: '3P', 4: '4P'}.get(max_crew, '2P')

                # Filtro crew type en 2da pasada (si no está desactivado)
                if not SKIP_CREW_CHECK_SECOND_PASS:
                    if trip_crew in ('3P', '4P') and trip_days >= MIN_DAYS_FOR_CREW_CHECK:
                        high_crew_count = sum(1 for d in chain if d.get('min_crew', '2P') == trip_crew)
                        if high_crew_count < MIN_HIGH_CREW_DUTIES:
                            continue

                # Crear trip sin DH de rescate
                pre_dh = {
                    "from": base,
                    "to": chain_org,
                    "dh_hours": pre_dh_hours,
                    "dh_block": dh_block,
                    "start_base": pre_start_base,
                    "end_base": pre_arr_base,
                    "legs": pre_legs,
                    "is_second_pass": True,
                }

                trip = {
                    'id': len(trips) + len(new_trips),
                    'base': base,
                    'chain': chain,
                    'block': chain_block,
                    'duty': chain_duty_hours,
                    'trip_crew': trip_crew,
                    'days': trip_days,
                    'total_dh_hours': total_dh_hours,
                    'dh_ratio': dh_ratio,
                    'pre_dh': pre_dh,
                    'rescue_dh': None,  # No hay DH de rescate
                    'has_rescue_dh': False,
                    'is_second_pass': True,
                    'flights_covered': set(f['id'] for d in chain for f in d.get('flights', [])),
                }

                # Calcular costo
                base_cost = sum(d.get('base_cost', 0) for d in chain)
                crew_cost = (0 if trip_crew == '2P' else 500) * trip_days
                dh_penalty = int(total_dh_hours * 1000)
                days_penalty = trip_days * TRIP_DAY_PENALTY
                second_pass_penalty = SECOND_PASS_PENALTY
                trip['cost'] = base_cost + crew_cost + dh_penalty + days_penalty + second_pass_penalty

                new_trips.append(trip)
                for d in chain:
                    for f in d.get('flights', []):
                        covered_flights.add(f['id'])

                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"  [2nd Pass] Trip #{trip['id']}: {base} -> DH -> {chain_org}-{chain_dst} (NO RESCUE DH NEEDED)")

                continue  # Ya creamos el trip, pasar al siguiente

            req_rest_after_duty = rules.calculate_required_rest(chain_block, is_base=False)
            min_report = chain_end_base + timedelta(hours=req_rest_after_duty)

            # Usar regla desde RuleEngine en lugar de variable global
            if not rules.ALLOW_SAME_DAY_DUTY:
                next_day = datetime.combine(
                    chain_end_base.date() + timedelta(days=1),
                    datetime.min.time()
                )
                rescue_report_base = max(min_report, next_day)
            else:
                rescue_report_base = min_report

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                print(f"    [SEARCH] Rescue DH: {chain_dst} -> {base}, depart after {rescue_report_base.strftime('%Y-%m-%d %H:%M')}")
                print(f"      Required rest after duty: {req_rest_after_duty:.2f}h")
                print(f"      Search window: {expanded_search_hours}h forward (vs normal {rules.MAX_DH_HOURS}h)")

            rescue_info = rules.get_rescue_itinerary(
                chain_dst, base, rescue_report_base,
                max_legs=rules.MAX_DH_LEGS,
                max_hours=expanded_search_hours  # Usar ventana expandida
            )
            if not rescue_info:
                rejection_reasons['no_rescue_dh'] += 1
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"      [REJECT] No rescue DH found")
                continue  # No hay DH rescue en BD

            rescue_dh_hours, arrival_at_base, rescue_legs = rescue_info

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                rescue_route = " -> ".join([f"{leg['org']}-{leg['dst']}" for leg in rescue_legs])
                print(f"      [FOUND] Rescue DH: {rescue_route}")
                print(f"        Duration: {rescue_dh_hours:.2f}h")
                print(f"        Arrives at base: {arrival_at_base.strftime('%Y-%m-%d %H:%M')}")

            # Validar pernocta en estación prohibida antes del DH de rescate
            # Si el duty termina en un día diferente al inicio del DH rescate, hay pernocta en chain_dst
            rescue_start_utc = rescue_legs[0]['dep_utc'] - rules.REPORT_TIME if rescue_legs else None
            if rescue_start_utc:
                rescue_start_base = rescue_start_utc - timedelta(hours=UTC_OFFSET_HOURS)
                if chain_end_base.date() < rescue_start_base.date():
                    if chain_dst in rules.PROHIBITED_LAYOVERS:
                        rejection_reasons['prohibited_layover'] += 1
                        if SECOND_PASS_DEBUG and verbose and is_target_chain:
                            print(f"      [REJECT] Layover before rescue in prohibited station: {chain_dst}")
                        continue

            # Validar pernoctas en escalas intermedias del DH de rescate
            rescue_has_prohibited_layover = False
            for i, leg in enumerate(rescue_legs):
                leg_dst = leg.get('dst', '')
                leg_arr = leg.get('arr_utc')
                if i < len(rescue_legs) - 1:
                    next_leg = rescue_legs[i + 1]
                    next_dep = next_leg.get('dep_utc')
                    if leg_arr and next_dep and leg_arr.date() < next_dep.date():
                        if leg_dst in rules.PROHIBITED_LAYOVERS:
                            rescue_has_prohibited_layover = True
                            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                                print(f"      [REJECT] Rescue DH layover in prohibited station: {leg_dst}")
                            break

            if rescue_has_prohibited_layover:
                rejection_reasons['prohibited_layover'] += 1
                continue

            # PASO 3: Validar trip resultante
            first_day = pre_start_base.date()
            last_day = arrival_at_base.date()
            trip_days = (last_day - first_day).days + 1

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                print(f"    [VALIDATE] Trip duration: {trip_days} days (max allowed: {max_trip_days})")

            if trip_days > max_trip_days:
                rejection_reasons['exceeds_max_days'] += 1
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"      [REJECT] Exceeds max days: {trip_days} > {max_trip_days}")
                continue  # Excede días máximos

            # Calcular DH total
            rescue_block = sum(l.get("blk", 0.0) for l in rescue_legs)
            total_dh_hours = pre_dh_hours + rescue_dh_hours

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                print(f"    [VALIDATE] Total DH: {total_dh_hours:.2f}h (max: {MAX_DH_HOURS_PER_TRIP}h)")

            if total_dh_hours > MAX_DH_HOURS_PER_TRIP:
                rejection_reasons['exceeds_max_dh_hours'] += 1
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"      [REJECT] Exceeds max DH hours: {total_dh_hours:.2f} > {MAX_DH_HOURS_PER_TRIP}")
                continue  # Excede DH máximo

            # Calcular ratio DH/Block
            if chain_block > 0:
                dh_ratio = total_dh_hours / chain_block
                if SECOND_PASS_DEBUG and verbose and is_target_chain:
                    print(f"    [VALIDATE] DH/Block ratio: {dh_ratio:.2f} (max: {MAX_DH_RATIO_SECOND_PASS})")

                if dh_ratio > MAX_DH_RATIO_SECOND_PASS:
                    rejection_reasons['exceeds_dh_ratio'] += 1
                    if SECOND_PASS_DEBUG and verbose and is_target_chain:
                        print(f"      [REJECT] Exceeds DH/Block ratio: {dh_ratio:.2f} > {MAX_DH_RATIO_SECOND_PASS}")
                    continue  # Ratio DH/Block muy alto
            else:
                dh_ratio = 0

            # Determinar crew requerido (basado en la cadena completa)
            max_crew = max(d.get('req_crew', 2) for d in chain)
            trip_crew = {2: '2P', 3: '3P', 4: '4P'}.get(max_crew, '2P')

            # Filtro crew type en 2da pasada (si no está desactivado)
            if not SKIP_CREW_CHECK_SECOND_PASS:
                if trip_crew in ('3P', '4P') and trip_days >= MIN_DAYS_FOR_CREW_CHECK:
                    high_crew_count = sum(1 for d in chain if d.get('min_crew', '2P') == trip_crew)
                    if high_crew_count < MIN_HIGH_CREW_DUTIES:
                        continue

            # Calcular descanso requerido post-rescate
            buffer_h = (rules.REPORT_TIME + rules.DEBRIEF).total_seconds() / 3600
            eff_block = chain_block + rescue_block
            req_rest_after = rules.calculate_required_rest(eff_block, is_base=True)

            # PASO 4: Crear trip
            pre_dh = {
                "from": base,
                "to": chain_org,
                "dh_hours": pre_dh_hours,
                "dh_block": dh_block,
                "start_base": pre_start_base,
                "end_base": pre_arr_base,
                "legs": pre_legs,
                "is_second_pass": True,  # Marcar origen
            }

            trip = {
                'id': len(trips) + len(new_trips),
                'base': base,
                'chain': chain,  # Cadena completa de duties
                'block': chain_block,
                'duty': chain_duty_hours,
                'trip_crew': trip_crew,
                'days': trip_days,
                'total_dh_hours': total_dh_hours,
                'dh_ratio': dh_ratio,
                'pre_dh': pre_dh,
                'rescue_legs': rescue_legs,
                'rescue_dh_hours': rescue_dh_hours,
                'req_rest': req_rest_after,
                'arrival_at_base': arrival_at_base,
                'is_second_pass': True,  # Flag especial
                'flights_covered': set(f['id'] for f in chain_flights),  # Vuelos cubiertos
                'chain_length': len(chain),  # Número de duties en la cadena
                'has_rescue_dh': True,
                'rescue_from': chain_dst,
                'rescue_arr_base': arrival_at_base,
                'final_rest_req': req_rest_after,
            }

            # Calcular costo del trip (para solver)
            base_cost = len(chain_flights) * 100_000
            crew_cost = {2: 200_000, 3: 300_000, 4: 400_000}.get(trip_crew, 200_000)
            dh_penalty = total_dh_hours * 1000
            days_penalty = trip_days * TRIP_DAY_PENALTY
            second_pass_penalty = SECOND_PASS_PENALTY  # Penalizar trips de segunda pasada

            # Bonus por combinar múltiples duties
            if len(chain) > 1:
                second_pass_penalty -= (len(chain) - 1) * 2000  # Reducir penalización

            trip['cost'] = base_cost + crew_cost + dh_penalty + days_penalty + second_pass_penalty

            # Agregar trip
            new_trips.append(trip)
            covered_flights.update(f['id'] for f in chain_flights)

            if verbose:
                chain_route = " -> ".join([f"{d['org']}-{d['dst']}" for d in chain])
                print(f"  [2nd Pass] Trip #{trip['id']}: {base} -> DH -> {chain_route} -> DH -> {base} ({len(chain)} duties, {trip_days}D, {total_dh_hours:.1f}h DH)")

            if SECOND_PASS_DEBUG and verbose and is_target_chain:
                print(f"      [SUCCESS] Trip created!")
                print(f"        Trip ID: {trip['id']}")
                print(f"        Cost: {trip['cost']:,}")
                print(f"        Flights covered: {len(chain_flights)}")

            break  # Ya cubrimos esta cadena, no seguir probando bases

    if verbose:
        if new_trips:
            print(f"\n[2nd Pass] Trips adicionales: {len(new_trips)}")
            print(f"[2nd Pass] Vuelos cubiertos: {len(covered_flights)}")

        # Mostrar razones de rechazo
        if SECOND_PASS_DEBUG:
            print(f"\n[DEBUG] Razones de rechazo:")
            print(f"  - Sin DH positioning: {rejection_reasons['no_positioning_dh']}")
            print(f"  - Same-day duty (positioning): {rejection_reasons['same_day_duty_positioning']}")
            print(f"  - Descanso insuficiente post-DH: {rejection_reasons['insufficient_rest_after_dh']}")
            print(f"  - Descanso insuficiente entre duties: {rejection_reasons['insufficient_rest_between_duties']}")
            print(f"  - Sin DH rescue: {rejection_reasons['no_rescue_dh']}")
            print(f"  - Excede días máximos ({max_trip_days}D): {rejection_reasons['exceeds_max_days']}")
            print(f"  - Excede DH máximo ({MAX_DH_HOURS_PER_TRIP}h): {rejection_reasons['exceeds_max_dh_hours']}")
            print(f"  - Excede ratio DH/Block ({MAX_DH_RATIO_SECOND_PASS}): {rejection_reasons['exceeds_dh_ratio']}")
            print(f"  - Cadena excede block con buffer: {rejection_reasons['exceeds_block_buffer']}")
            print(f"  - Cadena excede duty con buffer: {rejection_reasons['exceeds_duty_buffer']}")

    return new_trips, covered_flights


# ==========================================
# 3C. TERCERA PASADA - CREAR TRIPS PARCIALES
# ==========================================

def create_partial_trips(uncovered_duties, rules, trips, max_trip_days=6, verbose=False):
    """
    Tercera pasada: crea trips PARCIALES para duties que no pudieron ser cubiertos
    en primera ni segunda pasada.

    Un trip parcial es aquel que NO regresa a base (no tiene DH de rescate).
    El crew queda en la estación de destino final.

    IMPORTANTE: Esta función SOLO debe ejecutarse DESPUÉS de la segunda pasada.

    Returns:
        (trips_parciales, vuelos_cubiertos)
    """
    new_trips = []
    covered_flights = set()
    used_duty_ids = set()

    if verbose:
        print(f"\n=== INICIANDO TERCERA PASADA (TRIPS PARCIALES) ===")
        print(f"Duties sin cubrir a procesar: {len(uncovered_duties)}")

    # Bases disponibles
    bases = ['BOG', 'MDE']

    # Construir cadenas de duties conectados (igual que segunda pasada)
    all_chains = []

    # Ordenar duties por fecha de inicio
    sorted_duties = sorted(uncovered_duties, key=lambda d: d.get('report_utc', datetime.min))

    for duty in sorted_duties:
        duty_id = duty.get('duty_id')
        if duty_id in used_duty_ids:
            continue

        # Intentar construir cadena a partir de este duty
        chain = [duty]
        used_duty_ids.add(duty_id)
        current_dst = duty.get('dst')
        current_end = duty.get('release_utc')

        # Buscar duties que conectan
        for next_duty in sorted_duties:
            next_id = next_duty.get('duty_id')
            if next_id in used_duty_ids:
                continue

            next_org = next_duty.get('org')
            next_start = next_duty.get('report_utc')

            if next_org == current_dst and next_start and current_end:
                # Validar tiempo de conexión
                gap = (next_start - current_end).total_seconds() / 3600
                if 10 <= gap <= 72:  # Rango válido de conexión
                    chain.append(next_duty)
                    used_duty_ids.add(next_id)
                    current_dst = next_duty.get('dst')
                    current_end = next_duty.get('release_utc')

        all_chains.append(chain)

    if verbose:
        print(f"Cadenas construidas: {len(all_chains)}")

    # Para cada cadena, intentar crear trip parcial desde alguna base
    for chain in all_chains:
        chain_org = chain[0].get('org')
        chain_dst = chain[-1].get('dst')
        chain_flights = [f for d in chain for f in d.get('flights', [])]

        if not chain_flights:
            continue

        # Validar que destino final no esté prohibido
        if chain_dst in rules.PROHIBITED_LAYOVERS:
            if verbose:
                print(f"  [SKIP] Destino {chain_dst} está prohibido para pernocta")
            continue

        # Calcular métricas de la cadena
        chain_block = sum(d.get('block', 0) for d in chain)
        chain_duty_hours = sum(d.get('duty', 0) for d in chain)

        # Buscar base con DH de posicionamiento
        for base in bases:
            if chain_org == base:
                # No necesita DH de posicionamiento
                pre_dh_hours = 0
                pre_legs = []
                dh_block = 0

                first_flight = chain_flights[0] if chain_flights else None
                if first_flight:
                    pre_start_base = first_flight.get('dep_utc', datetime.now()) - rules.REPORT_TIME
                    pre_arr_base = pre_start_base
                else:
                    continue
            else:
                # Buscar DH de posicionamiento
                first_report_utc = chain[0].get('report_utc')
                if not first_report_utc:
                    continue

                pre_dh = rules.get_positioning_itinerary(
                    base, chain_org, first_report_utc,
                    max_legs=rules.MAX_DH_LEGS,
                    max_hours=48  # Ventana amplia
                )

                if not pre_dh:
                    continue

                pre_dh_hours, pre_arr_base, pre_legs = pre_dh
                dh_block = sum(l.get("blk", 0.0) for l in pre_legs)
                pre_start_base = pre_legs[0]['dep_utc'] - rules.REPORT_TIME if pre_legs else first_report_utc

            # Calcular fechas del trip
            chain_end_utc = chain[-1].get('release_utc')
            if not chain_end_utc:
                last_flight = chain_flights[-1] if chain_flights else None
                if last_flight:
                    chain_end_utc = last_flight.get('arr_utc')
                else:
                    continue

            chain_end_base = chain_end_utc - timedelta(hours=UTC_OFFSET_HOURS)  # UTC-5
            arrival_at_base = chain_end_base  # Trip termina en destino final

            first_day = pre_start_base.date() if pre_start_base else chain_end_base.date()
            last_day = arrival_at_base.date()
            trip_days = (last_day - first_day).days + 1

            if trip_days > max_trip_days:
                continue

            # Determinar crew requerido
            max_crew = max(d.get('req_crew', 2) for d in chain)
            trip_crew = {2: '2P', 3: '3P', 4: '4P'}.get(max_crew, '2P')

            # Calcular costo
            base_cost = len(chain_flights) * 100_000
            crew_cost = {2: 200_000, 3: 300_000, 4: 400_000}.get(trip_crew, 200_000)
            dh_penalty = pre_dh_hours * 1000
            days_penalty = trip_days * TRIP_DAY_PENALTY
            partial_penalty = 10_000  # Penalización extra por ser parcial

            pre_dh_info = {
                "from": base,
                "to": chain_org,
                "dh_hours": pre_dh_hours,
                "dh_block": dh_block,
                "start_base": pre_start_base,
                "end_base": pre_arr_base,
                "legs": pre_legs,
                "is_third_pass": True,
            }

            trip = {
                'id': len(trips) + len(new_trips),
                'base': base,
                'chain': chain,
                'block': chain_block,
                'duty': chain_duty_hours,
                'trip_crew': trip_crew,
                'days': trip_days,
                'total_dh_hours': pre_dh_hours,
                'dh_ratio': pre_dh_hours / chain_block if chain_block > 0 else 0,
                'pre_dh': pre_dh_info,
                'rescue_legs': [],  # Sin rescate
                'rescue_dh_hours': 0,
                'arrival_at_base': arrival_at_base,
                'is_second_pass': False,
                'is_third_pass': True,  # Flag de tercera pasada
                'is_partial_trip': True,  # TRIP PARCIAL
                'has_rescue_dh': False,
                'final_station': chain_dst,  # Donde queda el crew
                'flights_covered': set(f['id'] for f in chain_flights),
                'chain_length': len(chain),
                'cost': base_cost + crew_cost + dh_penalty + days_penalty + partial_penalty,
            }

            new_trips.append(trip)
            covered_flights.update(f['id'] for f in chain_flights)

            if verbose:
                chain_route = " -> ".join([f"{d['org']}-{d['dst']}" for d in chain])
                print(f"  [3rd Pass] Trip PARCIAL #{trip['id']}: {base} -> DH -> {chain_route} -> QUEDA EN {chain_dst} ({trip_days}D)")

            break  # Ya cubrimos esta cadena

    if verbose:
        print(f"=== TERCERA PASADA COMPLETADA ===")
        print(f"Trips parciales creados: {len(new_trips)}")
        print(f"Vuelos cubiertos: {len(covered_flights)}")

    return new_trips, covered_flights


# ==========================================
# 3B. PASE INDIVIDUAL SOLO-DH (ÚLTIMO RECURSO)
# ==========================================

def cover_individual_flights_with_dh(solo_flight_ids, df, dh_table, dh_index):
    """
    Último recurso de cobertura: trata cada vuelo sin cubrir en total aislamiento.
    Construye el trip mínimo posible: [DH posicionamiento] + [vuelo] + [DH rescate].

    A diferencia del Open Tours:
    - Sin restricción ALLOW_SAME_DAY_AFTER_DH (el DH puede llegar el mismo día del duty)
    - Sin límite de ratio DH/bloque (el objetivo es cubrir, no optimizar)
    - Ventana de búsqueda hasta SOLO_DH_SEARCH_WINDOW_HOURS (168h = 7 días por defecto)
    - Cada vuelo se trata individualmente (no se encadenan con otros)
    """
    if not solo_flight_ids:
        return [], None

    # RuleEngine ultra-relajado
    rules_solo = RuleEngine(
        dh_table=dh_table,
        dh_index=dh_index,
        allowed_crews={"2P", "3P", "4P"},
        allow_same_day_duty=True,
    )
    rules_solo.ALLOW_SAME_DAY_AFTER_DH = True  # Override post-init
    rules_solo.MAX_DH_LEGS = 4
    rules_solo.MAX_DH_HOURS = SOLO_DH_SEARCH_WINDOW_HOURS
    rules_solo.MAX_TRIP_DAYS = 14

    all_solo_trips = []
    found_count = 0

    for fid in solo_flight_ids:
        single_df = df[df["id"] == fid].copy()
        if single_df.empty:
            continue

        trips_for_flight = generate_trips(
            single_df,
            rules_solo,
            max_dh_hours_per_trip=SOLO_DH_SEARCH_WINDOW_HOURS,
            max_dh_ratio=999.0,
            allow_dh_only_duties=False,
        )

        if trips_for_flight:
            for t in trips_for_flight:
                t["cost"] = t.get("cost", 0) + SOLO_DH_PENALTY
                t["is_solo_dh"] = True
            all_solo_trips.extend(trips_for_flight)
            found_count += 1

    if VERBOSE_OUTPUT:
        print(f"  [Solo-DH] {found_count}/{len(solo_flight_ids)} vuelos con trip candidato generado")

    return all_solo_trips, rules_solo


# ==========================================
# 4. SOLVER Y REPORTES EJECUTIVOS
# ==========================================


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names max 31 chars and cannot contain some symbols."""
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    for b in bad:
        name = name.replace(b, '_')
    return name[:31]

def _autosize_worksheet_columns(ws):
    """Auto-ajusta ancho de columnas basado en el contenido."""
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _create_parameters_sheet(writer):
    """
    Crea una hoja con todos los parámetros de configuración utilizados en el cálculo.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Crear hoja (posición 1, después del Dashboard que está en 0)
    ws = writer.book.create_sheet("Parametros_Calculo", 1)

    # Estilos
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    category_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    category_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    param_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    value_font = Font(name="Calibri", size=10)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:D1')
    ws['A1'] = "PARÁMETROS DE CONFIGURACIÓN DEL CÁLCULO"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25

    # Fecha de generación
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name="Calibri", size=9, italic=True)
    ws['A2'].alignment = center_align

    # Definir parámetros por categoría
    parameters = [
        ("SISTEMA", [
            ("Nombre de la Aplicación", APP_NAME),
            ("Versión", APP_VERSION),
            ("Salida Verbose", VERBOSE_OUTPUT),
        ]),
        ("FLOTA", [
            ("Matrículas Target", ", ".join(TARGET_TAILS)),
        ]),
        ("ESTACIONES Y BASES", [
            ("Bases Operativas", ", ".join(DEFAULT_BASES)),
            ("Pernoctas Prohibidas", ", ".join(PROHIBITED_LAYOVERS)),
            ("Safe Havens (con hotel)", ", ".join(SAFE_HAVENS)),
            ("Estaciones Colombia", ", ".join(COLOMBIA_STATIONS)),
            ("Estaciones Fuera América", ", ".join(OUTSIDE_AMERICAS_STATIONS)),
        ]),
        ("TIEMPOS Y CONEXIONES", [
            ("Conexión Mínima (min)", MIN_CONNECT_MINUTES),
            ("Conexión Máxima (hrs)", MAX_CONNECT_HOURS),
            ("Report Time (min)", REPORT_MINUTES),
            ("Debrief Time (min)", DEBRIEF_MINUTES),
            ("Buffer Block Hours (min)", BLOCK_BUFFER_MINUTES),
            ("Buffer Duty Hours (min)", DUTY_BUFFER_MINUTES),
        ]),
        ("LÍMITES DE TRIP Y DESCANSO", [
            ("Máximo Días por Trip", MAX_TRIP_DAYS),
            ("Descanso Máximo entre Duties (hrs)", MAX_REST_GAP_HOURS),
            ("Descanso Máximo Open Tours (hrs)", OPEN_TOUR_MAX_REST_GAP_HOURS),
            ("Buffer Descanso (min)", REST_BUFFER_MINUTES),
        ]),
        ("VIÁTICOS Y PRIMAS (USD)", [
            ("Viático CAP Américas", VIATICO_CAP_AMERICAS_USD),
            ("Viático CAP Fuera Américas", VIATICO_CAP_OUTSIDE_USD),
            ("Viático COP Américas", VIATICO_COP_AMERICAS_USD),
            ("Viático COP Fuera Américas", VIATICO_COP_OUTSIDE_USD),
            ("Prima Nav CAP Américas", PRIMA_NAV_CAP_AMERICAS_USD),
            ("Prima Nav CAP Fuera Américas", PRIMA_NAV_CAP_OUTSIDE_USD),
            ("Prima Nav COP", PRIMA_NAV_COP_USD),
            ("Prima Nav Por Vuelo", PRIMA_NAV_POR_VUELO_USD),
            ("Extra Posicionamiento", EXTRA_POSICIONAMIENTO_USD),
            ("Prima Comando 2P CAP", PRIMA_COMANDO_2P_CAP_USD),
            ("Prima Comando 3P CAP", PRIMA_COMANDO_3P_CAP_USD),
            ("Prima Comando 3P COP", PRIMA_COMANDO_3P_COP_USD),
            ("Prima Comando 4P CAP1", PRIMA_COMANDO_4P_CAP1_USD),
            ("Prima Comando 4P CAP2", PRIMA_COMANDO_4P_CAP2_USD),
        ]),
        ("REGLAS DE NOCTURNIDAD", [
            ("Inicio Ventana Nocturna (hr)", NIGHT_WINDOW_START_HOUR),
            ("Fin Ventana Nocturna (hr)", NIGHT_WINDOW_END_HOUR),
            ("Máx Duties Nocturnos Consecutivos", MAX_CONSEC_NIGHT_DUTIES),
            ("Inicio Reducción Duty (hr)", DUTY_REDUCTION_START_HOUR),
            ("Fin Reducción Duty (hr)", DUTY_REDUCTION_END_HOUR),
            ("Reducción Duty (min)", DUTY_REDUCTION_MINUTES),
        ]),
        ("REGLAS DE DUTY", [
            ("Permitir Cambio de Avión en Duty", ALLOW_TAIL_CHANGE_IN_DUTY),
            ("Máximo Vuelos por Duty", MAX_DUTY_LEGS),
            ("Prohibir Duty Mismo Día", ENFORCE_NO_SAME_DAY_DUTY),
            ("Prohibir Duty Mismo Día Después DH", ENFORCE_NO_SAME_DAY_AFTER_DH),
        ]),
        ("CONFIGURACIÓN DE DH (DEADHEAD)", [
            ("Permitir DH en Duty", ALLOW_DH_IN_DUTY),
            ("Modo DH en Duty", DH_IN_DUTY_MODE),
            ("Permitir Duty Solo DH", ALLOW_DH_ONLY_DUTY),
            ("DH Estimado por Defecto (hrs)", DEFAULT_DH_HOURS),
            ("Máximo Escalas en DH", MAX_DH_LEGS),
            ("Ventana Máxima DH (hrs)", MAX_DH_HOURS),
            ("Filtrar DH a Estaciones Cargo", FILTER_DH_TO_CARGO_STATIONS),
            ("Buffer Fecha DH (días)", DH_DATE_BUFFER_DAYS),
            ("Máx DH Flights por Duty", MAX_DH_IN_DUTY_FLIGHTS),
            ("Máx DH Horas por Trip", MAX_DH_HOURS_PER_TRIP),
            ("Máx Ratio DH/Block", MAX_DH_RATIO),
        ]),
        ("OPEN TOURS", [
            ("Activar Open Tours", ENABLE_OPEN_TOURS),
            ("Máximo Días Open Tour", OPEN_TOUR_MAX_DAYS),
            ("Exportar Excel Open Tours", OPEN_TOUR_EXPORT_EXCEL),
            ("Permitir Mismo Día", OPEN_TOUR_ALLOW_SAME_DAY),
            ("Máx DH Horas Open Tours", OPEN_TOUR_MAX_DH_HOURS),
            ("Incluir DH Flights", OPEN_TOUR_INCLUDE_DH_FLIGHTS),
            ("Máx Escalas DH", OPEN_TOUR_MAX_DH_LEGS),
            ("Máx DH Horas por Trip", OPEN_TOUR_MAX_DH_HOURS_PER_TRIP),
            ("Máx Ratio DH", OPEN_TOUR_MAX_DH_RATIO),
            ("DH Positioning Horas Atrás", POSITIONING_DH_HOURS_BACK),
            ("Ventana Búsqueda DH (hrs)", POSITIONING_DH_SEARCH_WINDOW),
        ]),
        ("TRIPULACIONES", [
            ("Configuración Tripulación", f"2P={CREW_RANK['2P']}, 3P={CREW_RANK['3P']}, 4P={CREW_RANK['4P']}"),
        ]),
        ("PLANTA DE PILOTOS", [
            ("Días Libres por Mes", FREE_DAYS_PER_MONTH),
            ("Reserva (%)", fmt_pct(RESERVE_PCT*100, 1)),
            ("Pilotos en Entrenamiento", TRAINING_PILOTS),
            ("Pilotos en Vacaciones", VACATION_PILOTS),
            ("Pilotos en Admin", ADMIN_PILOTS),
            ("Pilotos Documentos/Visas", DOCS_PILOTS),
            ("Pilotos Incapacitados", INCAP_PILOTS),
            ("Pilotos Permiso Sindical", UNION_PILOTS),
        ]),
        ("BALANCE POR BASE", [
            ("Targets por Base", ", ".join([f"{k}={int(v*100)}%" for k, v in BASE_TARGETS.items()])),
            ("Aplicar Balanceo", ENFORCE_BASE_BALANCE),
            ("Tolerancia (%)", BASE_BALANCE_TOLERANCE),
        ]),
        ("OPTIMIZACIÓN", [
            ("Objetivo", OPTIMIZATION_OBJECTIVE),
            ("Efficiency: Peso Días", EFFICIENCY_DAY_WEIGHT),
            ("Efficiency: Peso Block", EFFICIENCY_BLOCK_WEIGHT),
            ("Pilots: Peso Días", PILOT_DAY_WEIGHT),
            ("Hybrid: Peso Efficiency", HYBRID_WEIGHT_EFF),
            ("Hybrid: Peso Pilots", HYBRID_WEIGHT_PILOTS),
            ("Hybrid: Peso Costo", HYBRID_WEIGHT_COST),
            ("Pilots_Eff: Peso Pilots", PILOTS_EFF_WEIGHT_PILOTS),
            ("Pilots_Eff: Peso Efficiency", PILOTS_EFF_WEIGHT_EFF),
            ("Desempate Costo", OBJECTIVE_TIEBREAKER_COST),
            ("Lexicográfico Activado", ENABLE_LEXICOGRAPHIC),
            ("Tolerancia Pilot Days", PILOT_DAY_TOLERANCE),
        ]),
        ("PENALIZACIONES", [
            ("Penalización 4P", PENALTY_4P),
            ("Penalización Día Trip", TRIP_DAY_PENALTY),
        ]),
        ("RENDIMIENTO/SOLVER", [
            ("Máximo Iteraciones", MAX_ITER),
            ("Máx Caminos por Inicio", MAX_DUTY_PATHS_PER_START),
            ("Tiempo Límite Solver (seg)", SOLVER_TIME_LIMIT_SECONDS),
        ]),
        ("SEGUNDA PASADA DH", [
            ("Activar Segunda Pasada", ENABLE_SECOND_PASS_DH),
            ("Máx Días Trip", MAX_TRIP_DAYS_SECOND_PASS),
            ("Máx Ratio DH/Block", MAX_DH_RATIO_SECOND_PASS),
            ("Penalización", SECOND_PASS_PENALTY),
            ("Debug Activado", SECOND_PASS_DEBUG),
        ]),
        ("TERCERA PASADA (PARCIALES)", [
            ("Activar Tercera Pasada", ENABLE_THIRD_PASS_PARTIAL),
        ]),
    ]

    # Escribir parámetros
    row = 4
    for category, params in parameters:
        # Categoría
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = category
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = center_align
        cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Parámetros
        for param_name, param_value in params:
            ws[f'A{row}'] = param_name
            ws[f'A{row}'].font = value_font
            ws[f'A{row}'].alignment = left_align
            ws[f'A{row}'].border = border
            ws[f'A{row}'].fill = param_fill

            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = str(param_value)
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = left_align
            ws[f'B{row}'].border = border
            row += 1

        row += 1  # Espacio entre categorías

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def _create_executive_dashboard(writer, kpis, base_dist, crew_dist, man_days,
                               selected_trips, flights_df, kpis_open=None, open_time_rows=None,
                               monthly_kpis=None):
    """
    Crea un dashboard ejecutivo profesional como primera hoja del Excel.
    Incluye sección de KPIs mensuales si monthly_kpis está disponible.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Crear hoja
    ws = writer.book.create_sheet("Dashboard", 0)

    # Estilos - Colores Avianca Cargo (Rojo)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Rojo oscuro
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Rojo Avianca
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    metric_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Rosa claro
    value_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # Rosa muy claro

    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:F1')
    ws['A1'] = "FLEX OPTIMIZER - DASHBOARD EJECUTIVO"
    ws['A1'].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Fecha de generación
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws['A2'].alignment = center_align

    row = 4

    # ============================================================================
    # SECCIÓN 1: KPIs PRINCIPALES
    # ============================================================================
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "MÉTRICAS PRINCIPALES"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Cobertura
    coverage = kpis_open.get("coverage_pct") if kpis_open else kpis.get("coverage_pct", 0)
    ws[f'A{row}'] = "Cobertura de Vuelos"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = round(float(coverage), 1)
    ws[f'B{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'B{row}'].fill = good_fill if coverage >= 90 else warning_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Eficiencia de Red
    network_eff = kpis_open.get("avg_network_eff") if kpis_open else kpis.get("avg_network_eff", 0)
    ws[f'A{row}'] = "Eficiencia de Red (Blk/Día)"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = round(float(network_eff), 2)
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Total Rotaciones
    total_trips = kpis_open.get("total_trips") if kpis_open else kpis.get("total_trips", 0)
    ws[f'A{row}'] = "Total Rotaciones"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = total_trips
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Total Vuelos en Itinerario
    total_flights = kpis_open.get("total_flights") if kpis_open else kpis.get("total_flights", len(flights_df))
    ws[f'A{row}'] = "Total Vuelos Itinerario"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = total_flights
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Vuelos Cubiertos
    flights_covered = kpis_open.get("flights_covered") if kpis_open else kpis.get("flights_covered", 0)
    ws[f'A{row}'] = "Vuelos Cubiertos"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = flights_covered
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = good_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Vuelos Open - usar KPIs si están disponibles, sino calcular
    if kpis_open and kpis_open.get("flights_open") is not None:
        open_flights = int(kpis_open.get("flights_open", 0))
    elif kpis.get("flights_open") is not None:
        open_flights = int(kpis.get("flights_open", 0))
    else:
        # Fallback: calcular usando set para evitar duplicados
        covered_flight_ids = set()
        for t in selected_trips:
            covered_flight_ids.update(t.get("flights_covered", []))
        total_operative_flights = len([f for _, f in flights_df.iterrows() if not f.get('is_dh', False)])
        open_flights = total_operative_flights - len(covered_flight_ids)
    ws[f'A{row}'] = "Vuelos Sin Cubrir"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = open_flights
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = warning_fill if open_flights > 0 else good_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Horas DH (Deadhead) - total por piloto
    dh_hours = float(kpis_open.get("total_dh_hours", 0) or 0) if kpis_open else float(kpis.get("total_dh_hours", 0) or 0)
    ws[f'A{row}'] = "Horas DH (por piloto)"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = round(float(dh_hours), 1)
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Horas TAFB (Time Away From Base)
    tafb_hours = float(kpis_open.get("total_tafb_hours", 0) or 0) if kpis_open else float(kpis.get("total_tafb_hours", 0) or 0)
    ws[f'A{row}'] = "Horas TA (Time Away)"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = round(float(tafb_hours), 1)
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Costo Hotel (USD)
    hotel_usd = float(kpis_open.get("hotel_total_usd", 0) or 0) if kpis_open else float(kpis.get("hotel_total_usd", 0) or 0)
    ws[f'A{row}'] = "Costo Hotel (USD)"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = f"${hotel_usd:,.0f}"
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Viáticos (USD)
    viaticos_usd = float(kpis_open.get("viaticos_total_usd", 0) or 0) if kpis_open else float(kpis.get("viaticos_total_usd", 0) or 0)
    ws[f'A{row}'] = "Viaticos (USD)"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = f"${viaticos_usd:,.0f}"
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = value_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 1

    # Trips Parciales
    trips_partial_count = sum(1 for t in selected_trips if t.get('is_partial_trip'))
    ws[f'A{row}'] = "Trips Parciales"
    ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].border = border

    ws[f'B{row}'] = trips_partial_count
    ws[f'B{row}'].font = Font(name="Calibri", size=11)
    ws[f'B{row}'].fill = warning_fill if trips_partial_count > 0 else good_fill
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 2: DISTRIBUCIÓN POR TRIPULACIÓN
    # ============================================================================
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "DÍAS REQUERIDOS POR TIPO DE TRIPULACIÓN"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Headers
    ws[f'A{row}'] = "Tripulación"
    ws[f'B{row}'] = "Días (CAP)"
    ws[f'C{row}'] = "Días (COP)"
    ws[f'D{row}'] = "Días (CRP)"
    ws[f'E{row}'] = "Total Días"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 1

    # Datos por tripulación
    for crew_type in ['2P', '3P', '4P']:
        cap_days = man_days.get("cap_by_crew", {}).get(crew_type, 0)
        cop_days = man_days.get("cop_by_crew", {}).get(crew_type, 0)
        crp_days = man_days.get("crp_by_crew", {}).get(crew_type, 0)
        total_days = cap_days + cop_days + crp_days

        ws[f'A{row}'] = crew_type
        ws[f'B{row}'] = cap_days
        ws[f'C{row}'] = cop_days
        ws[f'D{row}'] = crp_days
        ws[f'E{row}'] = total_days

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1

    # Total
    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = man_days.get("cap", 0)
    ws[f'C{row}'] = man_days.get("cop", 0)
    ws[f'D{row}'] = man_days.get("crp", 0)
    ws[f'E{row}'] = man_days.get("cap", 0) + man_days.get("cop", 0) + man_days.get("crp", 0)
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 3: HORAS POR TRIPULACIÓN
    # ============================================================================
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "HORAS REQUERIDAS POR TIPO DE TRIPULACIÓN"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Headers
    ws[f'A{row}'] = "Tripulación"
    ws[f'B{row}'] = "Horas Block (CAP)"
    ws[f'C{row}'] = "Horas Block (COP)"
    ws[f'D{row}'] = "Horas Block (CRP)"
    ws[f'E{row}'] = "Total Block"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 1

    # Calcular horas por tripulación
    hours_by_crew = {'2P': {'cap': 0, 'cop': 0, 'crp': 0},
                     '3P': {'cap': 0, 'cop': 0, 'crp': 0},
                     '4P': {'cap': 0, 'cop': 0, 'crp': 0}}

    for trip in selected_trips:
        crew_raw = trip.get('trip_crew', '2P')
        # Normalizar: si ya es "2P"/"3P"/"4P" usar directo, si es número agregar "P"
        if isinstance(crew_raw, str) and crew_raw in hours_by_crew:
            crew_type = crew_raw
        else:
            crew_type = f"{crew_raw}P" if str(crew_raw) not in ['2P', '3P', '4P'] else str(crew_raw)
        if crew_type not in hours_by_crew:
            crew_type = '2P'  # Fallback
        block_hours = trip.get('block', 0)
        # Distribuir según rol (simplificado - asume distribución uniforme)
        hours_by_crew[crew_type]['cap'] += block_hours / 3
        hours_by_crew[crew_type]['cop'] += block_hours / 3
        hours_by_crew[crew_type]['crp'] += block_hours / 3

    for crew_type in ['2P', '3P', '4P']:
        ws[f'A{row}'] = crew_type
        ws[f'B{row}'] = round(float(hours_by_crew[crew_type]['cap']), 1)
        ws[f'C{row}'] = round(float(hours_by_crew[crew_type]['cop']), 1)
        ws[f'D{row}'] = round(float(hours_by_crew[crew_type]['crp']), 1)
        total_h = sum(hours_by_crew[crew_type].values())
        ws[f'E{row}'] = round(float(total_h), 1)

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1

    # Total
    total_cap_blk = kpis_open.get("cap_block_hours_total") if kpis_open else kpis.get("cap_block_hours_total", 0)
    total_cop_blk = kpis_open.get("cop_block_hours_total") if kpis_open else kpis.get("cop_block_hours_total", 0)
    total_crp_blk = kpis_open.get("crp_block_hours_total") if kpis_open else kpis.get("crp_block_hours_total", 0)

    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = round(float(total_cap_blk), 1)
    ws[f'C{row}'] = round(float(total_cop_blk), 1)
    ws[f'D{row}'] = round(float(total_crp_blk), 1)
    ws[f'E{row}'] = round(float(total_cap_blk + total_cop_blk + total_crp_blk), 1)
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 4: AERONAVES
    # ============================================================================
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "AERONAVES INVOLUCRADAS"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Obtener aeronaves únicas y sus horas
    aircraft_hours = {}
    for _, flight in flights_df.iterrows():
        tail = flight.get('tail', 'UNKNOWN')
        block = flight.get('blk_hours', 0)  # Corregido: usar blk_hours
        if tail not in aircraft_hours:
            aircraft_hours[tail] = 0
        aircraft_hours[tail] += block

    # Headers
    ws[f'A{row}'] = "Aeronave"
    ws[f'B{row}'] = "Horas Block"
    ws[f'C{row}'] = "% del Total"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 1

    total_aircraft_hours = sum(aircraft_hours.values())
    for tail, hours in sorted(aircraft_hours.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = tail
        ws[f'B{row}'] = round(float(hours), 1)
        ws[f'C{row}'] = round(hours/total_aircraft_hours*100, 1) if total_aircraft_hours > 0 else 0.0

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1

    # Total
    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = round(float(total_aircraft_hours), 1)
    ws[f'C{row}'] = "100%"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col}{row}'].fill = metric_fill
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 5: VUELOS OPEN (NO CUBIERTOS) POR PERÍODO
    # ============================================================================
    if open_time_rows:
        from calendar import monthrange

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "VUELOS OPEN (NO CUBIERTOS) POR PERÍODO"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = border
        ws.row_dimensions[row].height = 25
        row += 1

        # Clasificar vuelos por inicio/mitad/fin de mes
        open_inicio = []  # Días 1-5
        open_mitad = []   # Días 6-25
        open_fin = []     # Últimos 5 días del mes

        for orow in open_time_rows:
            dep = orow.get("dep_base")
            if dep is None:
                continue
            day = dep.day
            year = dep.year
            month = dep.month
            last_day = monthrange(year, month)[1]

            if day <= 5:
                open_inicio.append(orow)
            elif day >= last_day - 4:
                open_fin.append(orow)
            else:
                open_mitad.append(orow)

        total_open = len(open_time_rows)

        # Headers
        ws[f'A{row}'] = "Período"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "% del Total"
        ws[f'D{row}'] = "Rutas Principales"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
            ws[f'{col}{row}'].fill = metric_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 1

        # Función para obtener rutas principales
        def _get_top_routes(flights_list, top_n=3):
            from collections import Counter
            routes = Counter()
            for f in flights_list:
                route = f"{f.get('org', '?')}-{f.get('dst', '?')}"
                routes[route] += 1
            top = routes.most_common(top_n)
            return ", ".join([f"{r}({c})" for r, c in top]) if top else "-"

        # Inicio de mes (días 1-5)
        pct_inicio = (len(open_inicio) / total_open * 100) if total_open > 0 else 0
        ws[f'A{row}'] = "Inicio de mes (días 1-5)"
        ws[f'B{row}'] = len(open_inicio)
        ws[f'C{row}'] = round(float(pct_inicio), 1)
        ws[f'D{row}'] = _get_top_routes(open_inicio)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align if col != 'D' else left_align
            ws[f'{col}{row}'].border = border
        row += 1

        # Mitad de mes
        pct_mitad = (len(open_mitad) / total_open * 100) if total_open > 0 else 0
        ws[f'A{row}'] = "Mitad de mes (días 6-25)"
        ws[f'B{row}'] = len(open_mitad)
        ws[f'C{row}'] = round(float(pct_mitad), 1)
        ws[f'D{row}'] = _get_top_routes(open_mitad)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align if col != 'D' else left_align
            ws[f'{col}{row}'].border = border
        row += 1

        # Fin de mes (últimos 5 días)
        pct_fin = (len(open_fin) / total_open * 100) if total_open > 0 else 0
        ws[f'A{row}'] = "Fin de mes (últimos 5 días)"
        ws[f'B{row}'] = len(open_fin)
        ws[f'C{row}'] = round(float(pct_fin), 1)
        ws[f'D{row}'] = _get_top_routes(open_fin)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col}{row}'].fill = value_fill
            ws[f'{col}{row}'].alignment = center_align if col != 'D' else left_align
            ws[f'{col}{row}'].border = border
        row += 1

        # Total
        ws[f'A{row}'] = "TOTAL OPEN"
        ws[f'B{row}'] = total_open
        ws[f'C{row}'] = "100%"
        ws[f'D{row}'] = ""
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(name="Calibri", size=10, bold=True)
            ws[f'{col}{row}'].fill = warning_fill
            ws[f'{col}{row}'].alignment = center_align
            ws[f'{col}{row}'].border = border
        row += 2

    # ============================================================================
    # SECCIÓN 6: KPIs MENSUALES
    # ============================================================================
    if monthly_kpis and len(monthly_kpis) > 0:
        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}'] = "KPIs POR MES"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = border
        ws.row_dimensions[row].height = 25
        row += 1

        # Headers (sin Cobertura %)
        monthly_headers = [
            ("Mes", 15),
            ("Efic. Red", 10),
            ("Rotaciones", 11),
            ("Duties", 10),
            ("Duty Days", 11),
            ("Block (h)", 10),
            ("Duty (h)", 10),
            ("DH (h)", 10),
            ("Hotel ($)", 12),
            ("Viáticos ($)", 12),
            ("$/Blk Hr", 10),
            ("BOG %", 10),
        ]

        for idx, (header, width) in enumerate(monthly_headers):
            col_letter = get_column_letter(idx + 1)
            ws[f'{col_letter}{row}'] = header
            ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=9, bold=True)
            ws[f'{col_letter}{row}'].fill = metric_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = border
            ws.column_dimensions[col_letter].width = width
        row += 1

        # Data rows
        for m_data in monthly_kpis:
            ws[f'A{row}'] = m_data.get("month_name", "")
            ws[f'B{row}'] = round(float(m_data.get('avg_block_per_day', 0)), 2)
            ws[f'C{row}'] = m_data.get("trips", 0)
            ws[f'D{row}'] = m_data.get("duties", 0)
            ws[f'E{row}'] = m_data.get("duty_days", 0)
            ws[f'F{row}'] = round(float(m_data.get('block_hours_total', 0)), 1)
            ws[f'G{row}'] = round(float(m_data.get('duty_hours_total', 0)), 1)
            ws[f'H{row}'] = round(float(m_data.get('dh_hours', 0)), 1)
            ws[f'I{row}'] = f"${m_data.get('hotel_cost_usd', 0):,.0f}"
            ws[f'J{row}'] = f"${m_data.get('viaticos_usd', 0):,.0f}"
            ws[f'K{row}'] = round(float(m_data.get('direct_cost_per_block_hour', 0)), 2)
            ws[f'L{row}'] = round(float(m_data.get('bog_block_pct', 0)), 1)

            # Apply styles
            for col_idx in range(1, 13):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=9)
                ws[f'{col_letter}{row}'].fill = value_fill
                ws[f'{col_letter}{row}'].alignment = center_align
                ws[f'{col_letter}{row}'].border = border

            # Color efficiency cell based on value (green if >= 4.5)
            eff = m_data.get('avg_block_per_day', 0)
            ws[f'B{row}'].fill = good_fill if eff >= 4.5 else warning_fill
            row += 1

        # Totals row (sin Cobertura %)
        total_trips = sum(m.get('trips', 0) for m in monthly_kpis)
        total_duties = sum(m.get('duties', 0) for m in monthly_kpis)
        total_duty_days = sum(m.get('duty_days', 0) for m in monthly_kpis)
        total_block = sum(m.get('block_hours_total', 0) for m in monthly_kpis)
        total_duty_h = sum(m.get('duty_hours_total', 0) for m in monthly_kpis)
        total_dh = sum(m.get('dh_hours', 0) for m in monthly_kpis)
        total_hotel = sum(m.get('hotel_cost_usd', 0) for m in monthly_kpis)
        total_viaticos = sum(m.get('viaticos_usd', 0) for m in monthly_kpis)
        avg_cost_per_blk = ((total_hotel + total_viaticos) / max(1, total_block)) if total_block > 0 else 0
        avg_network_eff = (total_block / max(1, total_duty_days)) if total_duty_days > 0 else 0

        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = round(float(avg_network_eff), 2)
        ws[f'C{row}'] = total_trips
        ws[f'D{row}'] = total_duties
        ws[f'E{row}'] = total_duty_days
        ws[f'F{row}'] = round(float(total_block), 1)
        ws[f'G{row}'] = round(float(total_duty_h), 1)
        ws[f'H{row}'] = round(float(total_dh), 1)
        ws[f'I{row}'] = f"${total_hotel:,.0f}"
        ws[f'J{row}'] = f"${total_viaticos:,.0f}"
        ws[f'K{row}'] = round(float(avg_cost_per_blk), 2)
        ws[f'L{row}'] = ""

        for col_idx in range(1, 13):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=9, bold=True)
            ws[f'{col_letter}{row}'].fill = metric_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


def _create_monthly_kpis_executive(writer, monthly_kpis):
    """
    Crea una hoja ejecutiva de KPIs mensuales (posición 2, después del Dashboard).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Crear hoja en posición 2 (después de Dashboard que está en 0, y Parámetros en 1)
    ws = writer.book.create_sheet("KPIs_Mensuales", 2)

    # Estilos - Colores Avianca Cargo (Rojo)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    metric_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    value_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # TÍTULO PRINCIPAL
    ws.merge_cells('A1:N1')
    ws['A1'] = "KPIs MENSUALES - RESUMEN EJECUTIVO"
    ws['A1'].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # Fecha de generación
    ws.merge_cells('A2:N2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True)
    ws['A2'].alignment = center_align

    row = 4

    # ============================================================================
    # SECCIÓN 1: MÉTRICAS OPERATIVAS POR MES
    # ============================================================================
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "MÉTRICAS OPERATIVAS POR MES"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Headers
    headers = [
        ("Mes", 16), ("Efic. Red", 11), ("Rotaciones", 11), ("Duties", 10),
        ("Duty Days", 11), ("Block (h)", 11), ("Duty (h)", 11), ("DH (h)", 10),
        ("DH Vuelos", 10), ("A/C Changes", 11), ("Hotel Nights", 12),
        ("BOG Block", 11), ("MDE Block", 11), ("BOG %", 9)
    ]
    for idx, (header, width) in enumerate(headers):
        col_letter = get_column_letter(idx + 1)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border
        ws.column_dimensions[col_letter].width = width
    row += 1

    # Data rows
    for m in monthly_kpis:
        ws[f'A{row}'] = m.get("month_name", "")
        ws[f'B{row}'] = round(float(m.get('avg_block_per_day', 0)), 2)
        ws[f'C{row}'] = m.get("trips", 0)
        ws[f'D{row}'] = m.get("duties", 0)
        ws[f'E{row}'] = m.get("duty_days", 0)
        ws[f'F{row}'] = round(float(m.get('block_hours_total', 0)), 1)
        ws[f'G{row}'] = round(float(m.get('duty_hours_total', 0)), 1)
        ws[f'H{row}'] = round(float(m.get('dh_hours', 0)), 1)
        ws[f'I{row}'] = m.get("dh_flights", 0)
        ws[f'J{row}'] = m.get("ac_changes", 0)
        ws[f'K{row}'] = m.get("hotel_nights", 0)
        ws[f'L{row}'] = round(float(m.get('bog_block_hours', 0)), 1)
        ws[f'M{row}'] = round(float(m.get('mde_block_hours', 0)), 1)
        ws[f'N{row}'] = round(float(m.get('bog_block_pct', 0)), 1)

        for col_idx in range(1, 15):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col_letter}{row}'].fill = value_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = border

        # Color efficiency based on value
        eff = m.get('avg_block_per_day', 0)
        ws[f'B{row}'].fill = good_fill if eff >= 4.5 else warning_fill
        row += 1

    # Totals row
    total_trips = sum(m.get('trips', 0) for m in monthly_kpis)
    total_duties = sum(m.get('duties', 0) for m in monthly_kpis)
    total_duty_days = sum(m.get('duty_days', 0) for m in monthly_kpis)
    total_block = sum(m.get('block_hours_total', 0) for m in monthly_kpis)
    total_duty_h = sum(m.get('duty_hours_total', 0) for m in monthly_kpis)
    total_dh = sum(m.get('dh_hours', 0) for m in monthly_kpis)
    total_dh_flights = sum(m.get('dh_flights', 0) for m in monthly_kpis)
    total_ac = sum(m.get('ac_changes', 0) for m in monthly_kpis)
    total_hotel_nights = sum(m.get('hotel_nights', 0) for m in monthly_kpis)
    total_bog = sum(m.get('bog_block_hours', 0) for m in monthly_kpis)
    total_mde = sum(m.get('mde_block_hours', 0) for m in monthly_kpis)
    avg_eff = (total_block / max(1, total_duty_days)) if total_duty_days > 0 else 0
    bog_pct = (total_bog / max(1, total_block)) * 100 if total_block > 0 else 0

    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = round(float(avg_eff), 2)
    ws[f'C{row}'] = total_trips
    ws[f'D{row}'] = total_duties
    ws[f'E{row}'] = total_duty_days
    ws[f'F{row}'] = round(float(total_block), 1)
    ws[f'G{row}'] = round(float(total_duty_h), 1)
    ws[f'H{row}'] = round(float(total_dh), 1)
    ws[f'I{row}'] = total_dh_flights
    ws[f'J{row}'] = total_ac
    ws[f'K{row}'] = total_hotel_nights
    ws[f'L{row}'] = round(float(total_bog), 1)
    ws[f'M{row}'] = round(float(total_mde), 1)
    ws[f'N{row}'] = round(float(bog_pct), 1)

    for col_idx in range(1, 15):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 2: COSTOS POR MES
    # ============================================================================
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "COSTOS POR MES (USD)"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Cost headers
    cost_headers = [
        ("Mes", 16), ("Hotel ($)", 14), ("Viáticos ($)", 14), ("Costo Directo ($)", 15),
        ("$/Blk Hr Hotel", 13), ("$/Blk Hr Viát.", 13), ("$/Blk Hr Total", 13), ("Block (h)", 11)
    ]
    for idx, (header, width) in enumerate(cost_headers):
        col_letter = get_column_letter(idx + 1)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border
    row += 1

    # Cost data
    for m in monthly_kpis:
        ws[f'A{row}'] = m.get("month_name", "")
        ws[f'B{row}'] = f"${m.get('hotel_cost_usd', 0):,.0f}"
        ws[f'C{row}'] = f"${m.get('viaticos_usd', 0):,.0f}"
        ws[f'D{row}'] = f"${m.get('direct_cost_usd', 0):,.0f}"
        ws[f'E{row}'] = round(float(m.get('hotel_cost_per_block_hour', 0)), 2)
        ws[f'F{row}'] = round(float(m.get('viaticos_per_block_hour', 0)), 2)
        ws[f'G{row}'] = round(float(m.get('direct_cost_per_block_hour', 0)), 2)
        ws[f'H{row}'] = round(float(m.get('block_hours_total', 0)), 1)

        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col_letter}{row}'].fill = value_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = border
        row += 1

    # Cost totals
    total_hotel = sum(m.get('hotel_cost_usd', 0) for m in monthly_kpis)
    total_viaticos = sum(m.get('viaticos_usd', 0) for m in monthly_kpis)
    total_direct = total_hotel + total_viaticos
    avg_hotel_per_blk = (total_hotel / max(1, total_block)) if total_block > 0 else 0
    avg_viat_per_blk = (total_viaticos / max(1, total_block)) if total_block > 0 else 0
    avg_direct_per_blk = avg_hotel_per_blk + avg_viat_per_blk

    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = f"${total_hotel:,.0f}"
    ws[f'C{row}'] = f"${total_viaticos:,.0f}"
    ws[f'D{row}'] = f"${total_direct:,.0f}"
    ws[f'E{row}'] = round(float(avg_hotel_per_blk), 2)
    ws[f'F{row}'] = round(float(avg_viat_per_blk), 2)
    ws[f'G{row}'] = round(float(avg_direct_per_blk), 2)
    ws[f'H{row}'] = round(float(total_block), 1)

    for col_idx in range(1, 9):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border
    row += 2

    # ============================================================================
    # SECCIÓN 3: HORAS POR ROL Y MES
    # ============================================================================
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "HORAS BLOCK POR ROL Y MES"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = border
    ws.row_dimensions[row].height = 25
    row += 1

    # Role headers
    role_headers = [("Mes", 16), ("CAP Block", 12), ("COP Block", 12), ("CRP Block", 12),
                    ("CAP Duty", 12), ("COP Duty", 12), ("CRP Duty", 12)]
    for idx, (header, width) in enumerate(role_headers):
        col_letter = get_column_letter(idx + 1)
        ws[f'{col_letter}{row}'] = header
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border
    row += 1

    # Role data
    for m in monthly_kpis:
        ws[f'A{row}'] = m.get("month_name", "")
        ws[f'B{row}'] = round(float(m.get('cap_block_hours', 0)), 1)
        ws[f'C{row}'] = round(float(m.get('cop_block_hours', 0)), 1)
        ws[f'D{row}'] = round(float(m.get('crp_block_hours', 0)), 1)
        ws[f'E{row}'] = round(float(m.get('cap_duty_hours', 0)), 1)
        ws[f'F{row}'] = round(float(m.get('cop_duty_hours', 0)), 1)
        ws[f'G{row}'] = round(float(m.get('crp_duty_hours', 0)), 1)

        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10)
            ws[f'{col_letter}{row}'].fill = value_fill
            ws[f'{col_letter}{row}'].alignment = center_align
            ws[f'{col_letter}{row}'].border = border
        row += 1

    # Role totals
    total_cap_blk = sum(m.get('cap_block_hours', 0) for m in monthly_kpis)
    total_cop_blk = sum(m.get('cop_block_hours', 0) for m in monthly_kpis)
    total_crp_blk = sum(m.get('crp_block_hours', 0) for m in monthly_kpis)
    total_cap_dty = sum(m.get('cap_duty_hours', 0) for m in monthly_kpis)
    total_cop_dty = sum(m.get('cop_duty_hours', 0) for m in monthly_kpis)
    total_crp_dty = sum(m.get('crp_duty_hours', 0) for m in monthly_kpis)

    ws[f'A{row}'] = "TOTAL"
    ws[f'B{row}'] = round(float(total_cap_blk), 1)
    ws[f'C{row}'] = round(float(total_cop_blk), 1)
    ws[f'D{row}'] = round(float(total_crp_blk), 1)
    ws[f'E{row}'] = round(float(total_cap_dty), 1)
    ws[f'F{row}'] = round(float(total_cop_dty), 1)
    ws[f'G{row}'] = round(float(total_crp_dty), 1)

    for col_idx in range(1, 8):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'{col_letter}{row}'].fill = metric_fill
        ws[f'{col_letter}{row}'].alignment = center_align
        ws[f'{col_letter}{row}'].border = border


def _create_trips_gantt_chart(writer, selected_trips):
    """
    Crea una visualización tipo Gantt chart de los trips en el calendario.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from collections import defaultdict
    import random

    ws = writer.book.create_sheet("Trips_Gantt")

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Colores para trips (paleta variada)
    trip_colors = [
        "90EE90", "87CEEB", "FFD700", "FFA07A", "DDA0DD",
        "F0E68C", "98FB98", "AFEEEE", "DB7093", "F5DEB3",
        "E0BBE4", "C5E1A5", "FFF59D", "FFCCBC", "B2DFDB"
    ]

    # Recopilar información de trips
    trips_by_base = defaultdict(list)
    all_dates = set()

    for trip in selected_trips:
        base = trip.get("base", "UNK")
        trip_id = trip.get("id", "")
        chain = trip.get("chain", [])

        if not chain:
            continue

        # Obtener fechas del trip
        start_date = chain[0].get("start_base").date()
        end_date = chain[-1].get("end_base").date()

        # Recopilar todas las fechas
        current = start_date
        while current <= end_date:
            all_dates.add(current)
            current += timedelta(days=1)

        trips_by_base[base].append({
            "trip_id": trip_id,
            "start_date": start_date,
            "end_date": end_date,
            "days": (end_date - start_date).days + 1,
            "crew": trip.get("trip_crew", ""),
            "chain": chain,
        })

    if not all_dates:
        return

    # Ordenar fechas
    sorted_dates = sorted(all_dates)
    min_date = sorted_dates[0]
    max_date = sorted_dates[-1]

    # Calcular escala: cuántas horas por columna (ajustable)
    HOURS_PER_COLUMN = 2.0  # Cada columna representa 2 horas

    # Crear encabezado del calendario
    row = 1
    col = 2  # Columna inicial (después de Trip ID)

    # Escribir encabezado "Trip ID"
    ws.cell(row, 1, "Trip ID")
    ws.cell(row, 1).font = header_font
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).alignment = center_align
    ws.cell(row, 1).border = border

    # Crear columnas para cada fecha con referencia visual
    date_to_col = {}  # {date: columna_inicio}
    current_date = min_date

    while current_date <= max_date:
        # Cada día del calendario ocupa 12 columnas (24h / 2h por columna = 12 columnas)
        cols_per_day = int(24 / HOURS_PER_COLUMN)
        date_to_col[current_date] = col

        # Mergear columnas para el encabezado de fecha
        ws.merge_cells(start_row=row, start_column=col,
                      end_row=row, end_column=col + cols_per_day - 1)

        cell = ws.cell(row, col)
        cell.value = f"{current_date.strftime('%d/%m')}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

        # Ajustar ancho de columnas del día (37 píxeles ≈ 5.5 unidades)
        for c in range(col, col + cols_per_day):
            ws.column_dimensions[get_column_letter(c)].width = 5.5

        col += cols_per_day
        current_date += timedelta(days=1)

    # Segunda fila: días de la semana
    row = 2
    ws.cell(row, 1, "Base")
    ws.cell(row, 1).font = header_font
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).alignment = center_align
    ws.cell(row, 1).border = border

    cols_per_day = int(24 / HOURS_PER_COLUMN)
    for date, start_col in date_to_col.items():
        weekday = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"][date.weekday()]

        ws.merge_cells(start_row=row, start_column=start_col,
                      end_row=row, end_column=start_col + cols_per_day - 1)

        cell = ws.cell(row, start_col)
        cell.value = weekday
        cell.font = Font(name="Calibri", size=8)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Escribir trips por base
    row = 3
    for base in sorted(trips_by_base.keys()):
        for trip_info in trips_by_base[base]:
            # Columna 1: Trip ID con tipo de tripulación y horas de inicio/fin
            trip_id = trip_info["trip_id"]
            crew_type = trip_info.get("crew", "")

            # Obtener horas de inicio y fin del trip
            chain = trip_info.get("chain", [])
            if chain:
                start_time = chain[0].get("start_base")
                end_time = chain[-1].get("end_base")

                if start_time and end_time:
                    start_str = start_time.strftime("%H:%M")
                    end_str = end_time.strftime("%H:%M")
                    time_range = f"{start_str}-{end_str}"
                else:
                    time_range = ""
            else:
                time_range = ""

            # Construir etiqueta con Trip ID, tipo de tripulación y horario
            if crew_type and time_range:
                trip_label = f"{trip_id}\n{crew_type}\n{time_range}"
            elif crew_type:
                trip_label = f"{trip_id}\n{crew_type}"
            elif time_range:
                trip_label = f"{trip_id}\n{time_range}"
            else:
                trip_label = trip_id

            ws.cell(row, 1, trip_label)
            ws.cell(row, 1).alignment = center_align
            ws.cell(row, 1).border = border
            ws.cell(row, 1).font = Font(name="Calibri", size=7)

            # Asignar color aleatorio al trip
            color = random.choice(trip_colors)
            trip_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Procesar cada duty del trip con ancho proporcional a sus horas
            for duty in trip_info["chain"]:
                block_hours = duty.get("block", 0)

                # Obtener fecha de inicio del duty
                duty_start_dt = duty.get("start_base")
                if not duty_start_dt:
                    continue

                duty_start_date = duty_start_dt.date()

                # Calcular columna inicial basada en la fecha y hora del duty
                if duty_start_date not in date_to_col:
                    continue

                # Columna base del día
                day_start_col = date_to_col[duty_start_date]

                # Offset dentro del día basado en la hora de inicio
                hour_of_day = duty_start_dt.hour + duty_start_dt.minute / 60.0
                hour_offset = int(hour_of_day / HOURS_PER_COLUMN)

                col = day_start_col + hour_offset

                # Calcular número de celdas proporcional a las horas de block
                num_cells = max(1, int(round(block_hours / HOURS_PER_COLUMN)))

                # Obtener estaciones del duty
                flights = duty.get("flights", [])
                stations = []
                if flights:
                    for flt in flights:
                        org = flt.get("org", "")
                        dst = flt.get("dst", "")
                        if org and org not in stations:
                            stations.append(org)
                        if dst and dst not in stations:
                            stations.append(dst)

                route = "-".join(stations[:4]) if stations else ""
                duty_label = f"{fmt_dec(block_hours, 1)}h\n{route}"

                # Pintar y mergear celdas para el duty
                if num_cells == 1:
                    cell = ws.cell(row, col)
                    cell.value = duty_label
                    cell.fill = trip_fill
                    cell.border = border
                    cell.alignment = center_align
                    cell.font = Font(name="Calibri", size=7)
                else:
                    # Mergear celdas horizontalmente
                    for c in range(col, col + num_cells):
                        cell = ws.cell(row, c)
                        cell.fill = trip_fill
                        cell.border = border
                        cell.alignment = center_align

                    ws.merge_cells(
                        start_row=row, start_column=col,
                        end_row=row, end_column=col + num_cells - 1
                    )

                    # Escribir etiqueta en la primera celda mergeada
                    ws.cell(row, col).value = duty_label
                    ws.cell(row, col).font = Font(name="Calibri", size=7)

            row += 1

    # Ajustar ancho de primera columna (Trip ID, crew type, horario)
    ws.column_dimensions['A'].width = 18


def _create_pairings_old_format(writer, selected_trips):
    """
    Crea una hoja con los pairings en formato visual tipo calendario antiguo.
    Cada día ocupa 4 columnas. Por cada trip se muestran 5 filas:
      1. Encabezado del día (nombre + fecha) - sin salto de línea
      2. Ruta/asignación (negro negrita, fondo gris claro; rojo si es DH)
      3. Block Hours (2 celdas combinadas, izquierda si AM, derecha si PM)
      4. Hora inicio / Hora fin (sin combinar, salmón solo si cruza medianoche)
      5. Duty time H:MM (4 celdas combinadas)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from collections import OrderedDict

    ws = writer.book.create_sheet("Pairings_Formato_Old")

    # Estilos (fuente Arial)
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    route_font = Font(name="Arial", size=9, bold=True, color="000000")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    dh_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dh_route_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    bh_font = Font(name="Arial", size=9, bold=False, color="000000")
    time_font = Font(name="Arial", size=9, bold=False, color="000000")
    duty_font = Font(name="Arial", size=9, bold=False, color="000000")
    salmon_fill = PatternFill(start_color="FFC8A8", end_color="FFC8A8", fill_type="solid")
    # Estilos para filas de piloto individual
    ta_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    op_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pilot_label_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pilot_font = Font(name="Arial", size=8, bold=True, color="000000")
    pilot_label_font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
    pilot_ta_font = Font(name="Arial", size=8, bold=True, color="996600")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    center_align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hhmm_fmt = '[h]:mm'        # formato Excel para duraciones acumuladas
    time_fmt = 'h:mm'          # formato Excel para hora del día
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    weekday_en = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]

    COLS_PER_DAY = 4
    current_row = 1
    _utc_offset = timedelta(hours=UTC_OFFSET_HOURS)

    # Ordenar trips por fecha de inicio (ascendente)
    def _trip_start_date(t):
        ch = t.get("chain", [])
        pre = t.get("pre_dh")
        dates = []
        if ch and ch[0].get("start_base"):
            dates.append(ch[0]["start_base"])
        if pre and pre.get("legs"):
            for leg in pre["legs"]:
                dep_b = leg["dep_utc"] - _utc_offset
                dates.append(dep_b - timedelta(hours=1))
        return min(dates) if dates else datetime.max

    sorted_trips = sorted(selected_trips, key=_trip_start_date)

    for trip in sorted_trips:
        trip_id = trip.get("id", "")
        chain = trip.get("chain", [])
        pre_dh = trip.get("pre_dh")

        if not chain:
            continue

        trip_crew = trip.get("trip_crew", "2P")

        activities_by_date = OrderedDict()

        # Pre-DH legs
        if pre_dh and pre_dh.get('legs'):
            for leg in pre_dh['legs']:
                dep_base = leg['dep_utc'] - _utc_offset
                arr_base = leg['arr_utc'] - _utc_offset
                report_time = dep_base - timedelta(minutes=REPORT_MINUTES)
                end_time = arr_base + timedelta(minutes=DEBRIEF_MINUTES)
                # Ajuste traslado terrestre GRU-VCP para DH
                dh_org = str(leg.get('org', '')).upper()
                dh_dst = str(leg.get('dst', '')).upper()
                if dh_org in ('VCP', 'GRU') and dh_dst == 'BOG':
                    report_time = report_time - timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
                if dh_org == 'BOG' and dh_dst in ('VCP', 'GRU'):
                    end_time = end_time + timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
                blk_h = (leg['arr_utc'] - leg['dep_utc']).total_seconds() / 3600
                duty_h = (end_time - report_time).total_seconds() / 3600
                flt = str(leg.get('flt_num', '')).strip()
                leg_aln = leg.get('aln', 'AV')
                flt_str = f" {leg_aln} {flt}" if flt else ""
                route = f"DH{flt_str} {leg['org']}{leg['dst']}"
                d = report_time.date()
                activities_by_date.setdefault(d, []).append({
                    "route": route, "bh": blk_h, "start_dt": report_time,
                    "end_dt": end_time, "duty_h": duty_h, "is_dh": True,
                    "crew": trip_crew, "min_crew": trip_crew,
                })

        # Duties in chain
        for duty in chain:
            start_dt = duty.get("start_base")
            end_dt = duty.get("end_base")
            if not start_dt or not end_dt:
                continue
            flights = duty.get("flights", [])
            stops = [flights[0]['org']] + [f['dst'] for f in flights] if flights else []
            flt_nums = "/".join([str(f.get('flt_num', '')).strip() for f in flights if f.get('flt_num') is not None])
            route = f"{flt_nums} {''.join(stops)}" if flt_nums else "".join(stops)
            dh_legs = [f for f in flights if f.get("is_dh")]
            is_dh_duty = len(dh_legs) == len(flights) and len(flights) > 0
            if is_dh_duty and not flt_nums:
                route = f"DH {''.join(stops)}"
            duty_crew = trip_crew  # La tripulación completa vuela junta
            bh = float(duty.get("block", 0))
            duty_h = float(duty.get("duty_dur", 0))
            d = start_dt.date()
            activities_by_date.setdefault(d, []).append({
                "route": route, "bh": bh, "start_dt": start_dt,
                "end_dt": end_dt, "duty_h": duty_h, "is_dh": is_dh_duty,
                "crew": duty_crew,
                "min_crew": duty.get("min_crew", trip_crew),
            })

        # Rescue DH legs
        if trip.get('has_rescue_dh'):
            rescue_legs = trip.get('rescue_legs') or []
            for leg in rescue_legs:
                dep_base = leg['dep_utc'] - _utc_offset
                arr_base = leg['arr_utc'] - _utc_offset
                report_time = dep_base - timedelta(minutes=REPORT_MINUTES)
                end_time = arr_base + timedelta(minutes=DEBRIEF_MINUTES)
                # Ajuste traslado terrestre GRU-VCP para DH
                dh_org = str(leg.get('org', '')).upper()
                dh_dst = str(leg.get('dst', '')).upper()
                if dh_org in ('VCP', 'GRU') and dh_dst == 'BOG':
                    report_time = report_time - timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
                if dh_org == 'BOG' and dh_dst in ('VCP', 'GRU'):
                    end_time = end_time + timedelta(minutes=GROUND_TRANSFER_MINUTES_VCP_BOG)
                blk_h = (leg['arr_utc'] - leg['dep_utc']).total_seconds() / 3600
                duty_h = (end_time - report_time).total_seconds() / 3600
                flt = str(leg.get('flt_num', '')).strip()
                leg_aln = leg.get('aln', 'AV')
                flt_str = f" {leg_aln} {flt}" if flt else ""
                route = f"DH{flt_str} {leg['org']}{leg['dst']}"
                d = report_time.date()
                activities_by_date.setdefault(d, []).append({
                    "route": route, "bh": blk_h, "start_dt": report_time,
                    "end_dt": end_time, "duty_h": duty_h, "is_dh": True,
                    "crew": trip_crew, "min_crew": trip_crew,
                })

        if not activities_by_date:
            continue

        # Generar todos los días consecutivos del trip (incluyendo días sin actividad)
        all_act_dates = sorted(activities_by_date.keys())
        first_date = all_act_dates[0]
        last_date = all_act_dates[-1]

        last_day_acts = activities_by_date.get(last_date, [])
        if last_day_acts:
            last_end = last_day_acts[-1].get("end_dt")
            if last_end and last_end.date() > last_date:
                last_date = last_end.date()

        sorted_dates = []
        d_cursor = first_date
        while d_cursor <= last_date:
            sorted_dates.append(d_cursor)
            if d_cursor not in activities_by_date:
                activities_by_date[d_cursor] = []
            d_cursor += timedelta(days=1)
        num_days = len(sorted_dates)

        station_by_date = {}
        last_known_station = trip.get("base", "")
        for d in sorted_dates:
            acts = activities_by_date.get(d, [])
            if acts:
                last_act = acts[-1]
                end_dt = last_act.get("end_dt")
                found_dst = None
                for duty in chain:
                    if duty.get("start_base") and duty["start_base"].date() == d:
                        found_dst = duty.get("dst", "")
                if not found_dst and pre_dh and pre_dh.get('legs'):
                    for leg in pre_dh['legs']:
                        dep_b = leg['dep_utc'] - _utc_offset
                        if (dep_b - timedelta(hours=1)).date() == d:
                            found_dst = leg.get('dst', '')
                if not found_dst and trip.get('has_rescue_dh'):
                    for leg in (trip.get('rescue_legs') or []):
                        dep_b = leg['dep_utc'] - _utc_offset
                        if (dep_b - timedelta(hours=1)).date() == d:
                            found_dst = leg.get('dst', '')
                if found_dst:
                    last_known_station = found_dst
                station_by_date[d] = None
            else:
                station_by_date[d] = last_known_station

        midnight_days = set()
        for d in sorted_dates:
            acts = activities_by_date.get(d, [])
            for act in acts:
                end_dt = act.get("end_dt")
                if end_dt and end_dt.date() > d:
                    next_day = end_dt.date()
                    if next_day in set(sorted_dates) and not activities_by_date.get(next_day):
                        midnight_days.add(next_day)

        # Escribir encabezado del trip
        trip_header_cell = ws.cell(row=current_row, column=1)
        trip_header_cell.value = f"TRIP #{trip_id} | {trip.get('base', '')} | {trip.get('trip_crew', '')}"
        trip_header_cell.font = Font(name="Arial", size=10, bold=True, color="000080")
        trip_header_cell.alignment = center_align
        total_cols = num_days * COLS_PER_DAY
        if total_cols > 1:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=total_cols)
        current_row += 1

        # === FILA 1: Encabezados de día ===
        for day_idx, d in enumerate(sorted_dates):
            col_start = day_idx * COLS_PER_DAY + 1
            col_end = col_start + COLS_PER_DAY - 1
            ws.merge_cells(start_row=current_row, start_column=col_start,
                           end_row=current_row, end_column=col_end)
            cell = ws.cell(row=current_row, column=col_start)
            wd = weekday_en[d.weekday()]
            cell.value = f"{wd} {d.day}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            for c in range(col_start, col_end + 1):
                ws.cell(row=current_row, column=c).border = thin_border
        current_row += 1

        max_activities = max(1, max(len(activities_by_date.get(d, [])) for d in sorted_dates))

        station_font = Font(name="Arial", size=9, bold=False, color="0000FF")

        for act_idx in range(max_activities):
            # === FILA 2: Ruta/Asignación ===
            for day_idx, d in enumerate(sorted_dates):
                acts = activities_by_date.get(d, [])
                col_start = day_idx * COLS_PER_DAY + 1
                col_end = col_start + COLS_PER_DAY - 1
                ws.merge_cells(start_row=current_row, start_column=col_start,
                               end_row=current_row, end_column=col_end)
                cell = ws.cell(row=current_row, column=col_start)
                if act_idx < len(acts):
                    act = acts[act_idx]
                    crew_label = act.get("crew", "")
                    cell.value = f"{act['route']} {crew_label}".strip()
                    if act.get("is_dh"):
                        cell.font = dh_route_font
                        for c in range(col_start, col_end + 1):
                            ws.cell(row=current_row, column=c).fill = dh_fill
                    else:
                        cell.font = route_font
                        for c in range(col_start, col_end + 1):
                            ws.cell(row=current_row, column=c).fill = gray_fill
                elif act_idx == 0 and station_by_date.get(d):
                    if d in midnight_days:
                        cell.value = f"Duty {station_by_date[d]}"
                        cell.font = Font(name="Arial", size=9, bold=True, color="0000FF")
                        for c in range(col_start, col_end + 1):
                            ws.cell(row=current_row, column=c).fill = salmon_fill
                    else:
                        cell.value = station_by_date[d]
                        cell.font = station_font
                cell.alignment = center_align
                for c in range(col_start, col_end + 1):
                    ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

            # === FILA 3: Block Hours ===
            for day_idx, d in enumerate(sorted_dates):
                acts = activities_by_date.get(d, [])
                col_start = day_idx * COLS_PER_DAY + 1
                am_col1 = col_start
                am_col2 = col_start + 1
                pm_col1 = col_start + 2
                pm_col2 = col_start + 3

                if act_idx < len(acts):
                    act = acts[act_idx]
                    bh = act["bh"]
                    bh_val = bh / 24.0   # fracción de día para formato [h]:mm
                    is_pm = act["start_dt"].hour >= 12 if act["start_dt"] else False

                    if is_pm:
                        ws.merge_cells(start_row=current_row, start_column=am_col1,
                                       end_row=current_row, end_column=am_col2)
                        ws.merge_cells(start_row=current_row, start_column=pm_col1,
                                       end_row=current_row, end_column=pm_col2)
                        c_bh = ws.cell(row=current_row, column=pm_col1)
                        c_bh.value = bh_val
                        c_bh.number_format = hhmm_fmt
                        c_bh.font = bh_font
                        c_bh.alignment = center_align
                    else:
                        ws.merge_cells(start_row=current_row, start_column=am_col1,
                                       end_row=current_row, end_column=am_col2)
                        ws.merge_cells(start_row=current_row, start_column=pm_col1,
                                       end_row=current_row, end_column=pm_col2)
                        c_bh = ws.cell(row=current_row, column=am_col1)
                        c_bh.value = bh_val
                        c_bh.number_format = hhmm_fmt
                        c_bh.font = bh_font
                        c_bh.alignment = center_align
                else:
                    ws.merge_cells(start_row=current_row, start_column=am_col1,
                                   end_row=current_row, end_column=am_col2)
                    ws.merge_cells(start_row=current_row, start_column=pm_col1,
                                   end_row=current_row, end_column=pm_col2)

                for c in range(col_start, col_start + COLS_PER_DAY):
                    ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

            # === FILA 4: Horas inicio/fin ===
            for day_idx, d in enumerate(sorted_dates):
                acts = activities_by_date.get(d, [])
                col_start = day_idx * COLS_PER_DAY + 1
                am_col1 = col_start
                am_col2 = col_start + 1
                pm_col1 = col_start + 2
                pm_col2 = col_start + 3

                if act_idx < len(acts):
                    act = acts[act_idx]
                    start_dt = act["start_dt"]
                    end_dt = act["end_dt"]
                    start_val = (start_dt.hour * 60 + start_dt.minute) / 1440.0 if start_dt else None
                    end_val = (end_dt.hour * 60 + end_dt.minute) / 1440.0 if end_dt else None
                    is_pm_start = start_dt.hour >= 12 if start_dt else False
                    crosses_midnight = end_dt.date() > start_dt.date() if (start_dt and end_dt) else False

                    if is_pm_start:
                        c_s = ws.cell(row=current_row, column=pm_col1)
                        if start_val is not None:
                            c_s.value = start_val; c_s.number_format = time_fmt
                        c_s.font = time_font; c_s.alignment = center_align
                        c_e = ws.cell(row=current_row, column=pm_col2)
                        if end_val is not None:
                            c_e.value = end_val; c_e.number_format = time_fmt
                        c_e.font = time_font; c_e.alignment = center_align
                        if crosses_midnight:
                            ws.cell(row=current_row, column=pm_col2).fill = salmon_fill
                    else:
                        c_s = ws.cell(row=current_row, column=am_col1)
                        if start_val is not None:
                            c_s.value = start_val; c_s.number_format = time_fmt
                        c_s.font = time_font; c_s.alignment = center_align
                        c_e = ws.cell(row=current_row, column=am_col2)
                        if end_val is not None:
                            c_e.value = end_val; c_e.number_format = time_fmt
                        c_e.font = time_font; c_e.alignment = center_align
                        if crosses_midnight:
                            ws.cell(row=current_row, column=am_col2).fill = salmon_fill

                for c in range(col_start, col_start + COLS_PER_DAY):
                    ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

            # === FILA 5: Duty Time H:MM ===
            for day_idx, d in enumerate(sorted_dates):
                acts = activities_by_date.get(d, [])
                col_start = day_idx * COLS_PER_DAY + 1
                col_end = col_start + COLS_PER_DAY - 1
                ws.merge_cells(start_row=current_row, start_column=col_start,
                               end_row=current_row, end_column=col_end)
                cell = ws.cell(row=current_row, column=col_start)
                if act_idx < len(acts):
                    duty_h = acts[act_idx]["duty_h"]
                    cell.value = duty_h / 24.0
                    cell.number_format = hhmm_fmt
                cell.font = duty_font
                cell.alignment = center_align
                cell.border = thin_border
                for c in range(col_start, col_end + 1):
                    ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

        # === FILAS DE PILOTO INDIVIDUAL: 1 fila por piloto del trip ===
        crew_size = CREW_RANK.get(trip_crew, 2)
        if crew_size >= 2:
            pilots_list = _get_pilot_list(trip_crew)

            # Mapa: para cada día, cuál es el min_crew de la actividad
            day_min_crew = {}
            day_has_activity = {}
            for d in sorted_dates:
                acts = activities_by_date.get(d, [])
                if acts:
                    day_has_activity[d] = True
                    max_min = '2P'
                    for a in acts:
                        mc = a.get("min_crew", trip_crew)
                        if CREW_RANK.get(mc, 2) > CREW_RANK.get(max_min, 2):
                            max_min = mc
                    day_min_crew[d] = max_min
                else:
                    day_has_activity[d] = False
                    day_min_crew[d] = trip_crew

            for pid, role in pilots_list:
                for day_idx, d in enumerate(sorted_dates):
                    col_start = day_idx * COLS_PER_DAY + 1
                    col_end = col_start + COLS_PER_DAY - 1

                    if day_idx == 0:
                        label_cell = ws.cell(row=current_row, column=1)

                    ws.merge_cells(start_row=current_row, start_column=col_start,
                                   end_row=current_row, end_column=col_end)
                    cell = ws.cell(row=current_row, column=col_start)

                    if not day_has_activity.get(d, False):
                        cell.value = ""
                    else:
                        mc = day_min_crew.get(d, '2P')
                        ops = _ops_for_req(trip_crew, mc)
                        is_op = pid in ops

                        if is_op:
                            cell.value = f"{pid} OP"
                            cell.font = pilot_font
                            for c in range(col_start, col_end + 1):
                                ws.cell(row=current_row, column=c).fill = op_fill
                        else:
                            cell.value = f"{pid} TA"
                            cell.font = pilot_ta_font
                            for c in range(col_start, col_end + 1):
                                ws.cell(row=current_row, column=c).fill = ta_fill

                    cell.alignment = center_align
                    for c in range(col_start, col_end + 1):
                        ws.cell(row=current_row, column=c).border = thin_border
                current_row += 1

        # Fila vacía de separación entre trips
        current_row += 1

    # Ajustar ancho de columnas
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 7.5

    # Ajustar alto de filas
    for r in range(1, current_row):
        ws.row_dimensions[r].height = 24


def export_results_to_excel(
    output_path: str,
    flights_df: pd.DataFrame,
    trips: list,
    selected_trips: list,
    trip_legs_rows: list,
    trip_report_rows: list,
    pilot_lines_rows: list,
    pilot_lines_dates: list,
    daily_hours_rows: list,
    daily_pilots_excl_rows: list,
    viaticos_rows: list,
    kpis: dict,
    base_dist: dict,
    crew_dist: dict,
    man_days: dict,
    hotels_counter: Counter,
    hotels_by_month: Counter,
    duty_starts_rows: list,
    open_time_rows: list,
    pilot_hours_rows: list,
    pilots_by_day_rows: list,
    pilots_by_day_base_rows: list,
    plant_required_rows: list,
    dh_report_rows: list = None,
    kpis_open: dict = None,
    base_dist_open: dict = None,
    crew_dist_open: dict = None,
    man_days_open: dict = None,
    open_time_rows_open: list = None,
    monthly_kpis: list = None,
    base_analysis: dict = None,
    idle_windows: list = None,
):
    """
    Exporta resultados a Excel con múltiples pestañas.
    output_path: ruta del archivo .xlsx
    """
    output_path = str(output_path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # 1) Flights
    df_flights = flights_df.copy()
    df_flights = df_flights.sort_values(["tail", "dep_base"]).reset_index(drop=True)

    # 1b) Itinerario con cobertura - mapea vuelos a trips
    flight_to_trip = {}
    for t in selected_trips:
        trip_id = t.get("id", "")
        for fid in t.get("flights_covered", []):
            flight_to_trip[fid] = trip_id

    df_itinerario = flights_df.copy()
    df_itinerario["CUBIERTO"] = df_itinerario["id"].apply(lambda x: "SI" if x in flight_to_trip else "NO")
    df_itinerario["TRIP #"] = df_itinerario["id"].apply(lambda x: flight_to_trip.get(x, ""))

    # Agregar columnas RAZON y DESCRIPCION para vuelos no cubiertos
    def get_razon(flight_id):
        if flight_id in flight_to_trip:
            return ""  # Vuelo cubierto, no hay razón de exclusión
        exclusion = get_flight_exclusion(flight_id)
        return exclusion.get("razon", "")

    def get_descripcion(flight_id):
        if flight_id in flight_to_trip:
            return ""  # Vuelo cubierto, no hay descripción
        exclusion = get_flight_exclusion(flight_id)
        return exclusion.get("descripcion", "")

    df_itinerario["RAZON"] = df_itinerario["id"].apply(get_razon)
    df_itinerario["DESCRIPCION"] = df_itinerario["id"].apply(get_descripcion)

    # Reordenar columnas para que CUBIERTO, TRIP #, RAZON y DESCRIPCION estén al inicio
    cols = df_itinerario.columns.tolist()
    cols_new = ["CUBIERTO", "TRIP #", "RAZON", "DESCRIPCION"] + [c for c in cols if c not in ("CUBIERTO", "TRIP #", "RAZON", "DESCRIPCION")]
    df_itinerario = df_itinerario[cols_new]

    # DEBUG: Resumen de vuelos no cubiertos y sus razones
    if VERBOSE_OUTPUT:
        df_no_cubiertos = df_itinerario[df_itinerario["CUBIERTO"] == "NO"]
        if not df_no_cubiertos.empty:
            print("\n" + "="*80)
            print("  VUELOS NO CUBIERTOS - RESUMEN DE RAZONES")
            print("="*80)
            razones_count = df_no_cubiertos["RAZON"].value_counts()
            for razon, count in razones_count.items():
                print(f"  {razon}: {count} vuelos")
            print("-"*80)
            print("  Detalle de vuelos no cubiertos:")
            for _, row in df_no_cubiertos.iterrows():
                flt = row.get('flt', row.get('flt_num', ''))
                org = row.get('org', '')
                dst = row.get('dst', '')
                razon = row.get('RAZON', '')
                desc = row.get('DESCRIPCION', '')
                print(f"  - {flt} {org}-{dst}: [{razon}] {desc[:60]}...")
            print("="*80 + "\n")

    # 2) Trips (todos)
    def trip_to_row(t):
        return {
            "trip_id": t.get("id"),
            "base": t.get("base"),
            "trip_crew": t.get("trip_crew"),
            "days": t.get("days"),
            "block_hours": round(float(t.get("block", 0)), 2),
            "cost": round(float(t.get("cost", 0)), 2),
            "total_dh_hours": round(float(t.get("total_dh_hours", 0)), 2),
            "has_rescue_dh": bool(t.get("has_rescue_dh", False)),
            "rescue_from": t.get("rescue_from", ""),
            "final_rest_req_h": round(float(t.get("final_rest_req", 0)), 2) if t.get("final_rest_req") is not None else "",
            "flights_covered_n": len(t.get("flights_covered", set())) if t.get("flights_covered") is not None else "",
        }

    df_trips = pd.DataFrame([trip_to_row(t) for t in trips]).sort_values(["base", "trip_id"]).reset_index(drop=True)
    df_selected = pd.DataFrame([trip_to_row(t) for t in selected_trips]).sort_values(["base", "trip_id"]).reset_index(drop=True)

    # 2a) Trip Summary Report - Reporte consolidado de trips
    trip_summary_rows = []
    crew_sizes_map = {'2P': 2, '3P': 3, '4P': 4}
    for t in selected_trips:
        trip_id = t.get('id')
        chain = t.get('chain', [])
        pre_dh = t.get('pre_dh')

        # Fecha inicio/fin
        start_dt = chain[0]['start_base'] if chain else None
        if pre_dh and pre_dh.get('start_base'):
            start_dt = min(start_dt, pre_dh['start_base']) if start_dt else pre_dh['start_base']

        end_dt = chain[-1]['end_base'] if chain else None
        if t.get('has_rescue_dh'):
            rescue_arr = t.get('rescue_arr_base')
            if rescue_arr:
                end_dt = rescue_arr

        # Cantidad de pilotos
        trip_crew = t.get('trip_crew', '2P')
        num_pilots = crew_sizes_map.get(trip_crew, 2)

        # Block hours
        block_hours = float(t.get('block', 0))

        # Days
        days = t.get('days', 0)

        # Eficiencia (blk/days)
        efficiency = round(block_hours / days, 2) if days > 0 else 0

        # Secuencia completa de vuelos ejecutados (desde inicio hasta fin)
        flight_sequence = []

        # 1. Agregar DH de posicionamiento si existe
        if pre_dh:
            pre_legs = pre_dh.get('legs', [])
            if pre_legs:
                for leg in pre_legs:
                    org = leg.get('org', '')
                    dst = leg.get('dst', '')
                    if org and dst:
                        flight_sequence.append(f"[DH]{org}-{dst}")
            else:
                pre_from = pre_dh.get('from', '')
                pre_to = pre_dh.get('to', '')
                if pre_from and pre_to:
                    flight_sequence.append(f"[DH]{pre_from}-{pre_to}")

        # 2. Agregar vuelos operativos de cada duty
        for duty in chain:
            for f in duty.get('flights', []):
                org = f.get('org', '')
                dst = f.get('dst', '')
                is_dh = f.get('is_dh', False)
                if org and dst:
                    if is_dh:
                        flight_sequence.append(f"[DH]{org}-{dst}")
                    else:
                        flight_sequence.append(f"{org}-{dst}")

        # 3. Agregar DH de rescate si existe
        if t.get('has_rescue_dh'):
            rescue_legs = t.get('rescue_legs', [])
            if rescue_legs:
                for leg in rescue_legs:
                    org = leg.get('org', '')
                    dst = leg.get('dst', '')
                    if org and dst:
                        flight_sequence.append(f"[DH]{org}-{dst}")
            else:
                rescue_from = t.get('rescue_from', '')
                rescue_to = t.get('base', '')
                if rescue_from and rescue_to:
                    flight_sequence.append(f"[DH]{rescue_from}-{rescue_to}")

        flights_str = ' -> '.join(flight_sequence) if flight_sequence else ''

        # DH info
        has_dh = bool(pre_dh) or t.get('has_rescue_dh', False)
        dh_details = []
        if pre_dh:
            pre_legs = pre_dh.get('legs', [])
            if pre_legs:
                pre_route = '->'.join([leg.get('org', '') for leg in pre_legs] + [pre_legs[-1].get('dst', '')])
            else:
                pre_route = f"{pre_dh.get('from', '')}->{pre_dh.get('to', '')}"
            dh_details.append(f"POS: {pre_route}")
        if t.get('has_rescue_dh'):
            rescue_legs = t.get('rescue_legs', [])
            if rescue_legs:
                rescue_route = '->'.join([leg.get('org', '') for leg in rescue_legs] + [rescue_legs[-1].get('dst', '')])
            else:
                rescue_route = f"{t.get('rescue_from', '')}->{t.get('base', '')}"
            dh_details.append(f"RES: {rescue_route}")
        dh_str = '; '.join(dh_details) if dh_details else 'N/A'

        # DH hours
        total_dh_hours = float(t.get('total_dh_hours', 0))

        # Duty time total
        duty_hours = sum(float(d.get('duty', 0)) for d in chain)

        # Detectar si es trip parcial
        is_partial = t.get('is_partial_trip', False)
        is_third_pass = t.get('is_third_pass', False)
        final_station = t.get('final_station', t.get('base', ''))

        trip_summary_rows.append({
            'Trip #': trip_id,
            'Fecha Inicio': start_dt.strftime('%Y-%m-%d %H:%M') if start_dt else '',
            'Fecha Fin': end_dt.strftime('%Y-%m-%d %H:%M') if end_dt else '',
            'Base': t.get('base', ''),
            'Crew': trip_crew,
            'Pilotos': num_pilots,
            'Dias': days,
            'BLK (hrs)': round(block_hours, 2),
            'DH (hrs)': round(total_dh_hours, 2),
            'Duty (hrs)': round(duty_hours, 2),
            'Eficiencia (BLK/dia)': efficiency,
            'Secuencia Vuelos': flights_str,
            'Tiene DH': 'SI' if has_dh else 'NO',
            'Detalle DH': dh_str,
            'Es Parcial': 'SI' if is_partial else 'NO',
            'Estacion Final': final_station if is_partial else '',
        })

    df_trip_summary = pd.DataFrame(trip_summary_rows)
    if not df_trip_summary.empty:
        df_trip_summary = df_trip_summary.sort_values(['Base', 'Trip #']).reset_index(drop=True)

    # 2b) Trips de segunda pasada
    second_pass_rows = []
    for trip in selected_trips:
        if trip.get('is_second_pass'):
            chain = trip.get('chain', [])
            chain_route = " -> ".join([f"{d['org']}-{d['dst']}" for d in chain]) if chain else ""
            second_pass_rows.append({
                'Trip ID': f"TRIP #{trip['id']}",
                'Base': trip['base'],
                'Days': trip['days'],
                'Crew': f"{trip['trip_crew']}P",
                'Block Hours': round(float(trip['block']), 2),
                'Total DH Hours': round(float(trip.get('total_dh_hours', 0)), 2),
                'DH Ratio': round(float(trip.get('dh_ratio', 0)), 2),
                'Chain Length': trip.get('chain_length', 1),
                'Flights Covered': len(trip.get('flights_covered', [])),
                'Route': f"{trip.get('pre_dh', {}).get('from', '?')} -> DH -> {chain_route} -> DH -> {trip['base']}",
            })
    df_second_pass = pd.DataFrame(second_pass_rows) if second_pass_rows else pd.DataFrame()

    # 2c) Trips parciales (tercera pasada)
    partial_trips_rows = []
    for trip in selected_trips:
        if trip.get('is_partial_trip'):
            chain = trip.get('chain', [])
            chain_route = " -> ".join([f"{d['org']}-{d['dst']}" for d in chain]) if chain else ""
            partial_trips_rows.append({
                'Trip ID': f"TRIP #{trip['id']}",
                'Base': trip['base'],
                'Estacion Final': trip.get('final_station', ''),
                'Days': trip['days'],
                'Crew': trip.get('trip_crew', '2P'),
                'Block Hours': round(float(trip.get('block', 0)), 2),
                'DH Hours': round(float(trip.get('total_dh_hours', 0)), 2),
                'Chain Length': trip.get('chain_length', 1),
                'Flights Covered': len(trip.get('flights_covered', [])),
                'Route': f"{trip.get('pre_dh', {}).get('from', '?')} -> DH -> {chain_route} -> QUEDA EN {trip.get('final_station', '?')}",
            })
    df_partial_trips = pd.DataFrame(partial_trips_rows) if partial_trips_rows else pd.DataFrame()

    # 3) Legs / Duties dentro de trips seleccionados (detalle por día)
    df_legs = pd.DataFrame(trip_legs_rows)
    if not df_legs.empty:
        df_legs = df_legs.sort_values(["trip_id", "duty_start_base"]).reset_index(drop=True)

    
    # 3b) Trip report (formato como consola)
    df_trip_report = pd.DataFrame(trip_report_rows)
    if not df_trip_report.empty and "TRIP_ID" not in df_trip_report.columns:
        cur_trip = None
        trip_ids = []
        for _, row in df_trip_report.iterrows():
            ruta = row.get("RUTA/ACTIVIDAD", "")
            # Separator row (all empty)
            if all(str(v).strip() == "" for v in row.values):
                cur_trip = None
            elif isinstance(ruta, str) and ruta.startswith("TRIP #"):
                try:
                    cur_trip = ruta.split("TRIP #", 1)[1].split("|", 1)[0].strip()
                except Exception:
                    cur_trip = cur_trip
            trip_ids.append(cur_trip)
        df_trip_report["TRIP_ID"] = trip_ids
# 4) Hotels
    hotel_rows = []
    if hotels_by_month:
        for (stn, ym), v in sorted(hotels_by_month.items(), key=lambda x: (x[0][0], x[0][1])):
            hotel_rows.append({
                "station": stn,
                "month": ym,
                "nights_persona": int(v),
                "rate_usd": HOTEL_RATES_USD.get(stn),
                "total_usd": (int(v) * HOTEL_RATES_USD.get(stn, 0)),
            })
    else:
        for stn, v in hotels_counter.most_common():
            hotel_rows.append({
                "station": stn,
                "month": "",
                "nights_persona": int(v),
                "rate_usd": HOTEL_RATES_USD.get(stn),
                "total_usd": (int(v) * HOTEL_RATES_USD.get(stn, 0)),
            })
    df_hotels = pd.DataFrame(hotel_rows)

    # 5) Open time
    df_open = pd.DataFrame(open_time_rows)

    # 5a) Vuelos Open con formato de tarjeta para CAP y COP
    vuelos_open_rows = []
    mes_es = {1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
              7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"}
    for orow in open_time_rows:
        fid = orow.get("flight_id")
        flt_num = orow.get("flt_num", "")
        tail = orow.get("tail", "")
        org = orow.get("org", "")
        dst = orow.get("dst", "")
        dep_base = orow.get("dep_base")
        arr_base = orow.get("arr_base")
        blk = orow.get("blk_hours", 0)
        # Calcular duty (blk + report + debrief)
        duty_hrs = float(blk) + 1.5  # aproximación: 1h report + 0.5h debrief
        # Formato de hora
        dep_str = dep_base.strftime("%H:%M") if dep_base else ""
        arr_str = arr_base.strftime("%H:%M") if arr_base else ""
        fecha_str = f"{dep_base.day:02d}-{mes_es.get(dep_base.month, '')}" if dep_base else ""
        # Ruta completa (org-dst)
        rte = f"{org}-{dst}"
        # Viáticos simplificados para vuelo suelto (internacional sin pernocta)
        # Por vuelo individual: prima navegación (configurable)
        viat_cop = 0
        viat_usd = PRIMA_NAV_POR_VUELO_USD  # prima navegación básica por vuelo
        viat_str = f"COP {viat_cop} / USD {viat_usd}"
        # Tarjetas formateadas (una para cada posición)
        def _make_card(pos):
            return (
                f"FLT: {flt_num}\n"
                f"RTE: {rte}\n"
                f"ACFT: {tail}\n"
                f"TIME: {dep_str}-{arr_str}\n"
                f"BLK: {fmt_dec(float(blk), 1)}\n"
                f"DUTY: {fmt_dec(duty_hrs, 1)}\n"
                f"CREW: 2P {pos}\n"
                f"{fecha_str}\n\n"
                f"VIAT: {viat_str}"
            )
        # Agregar fila para CAP
        vuelos_open_rows.append({
            "position": "CAP",
            "fecha": dep_base.date() if dep_base else None,
            "fecha_str": fecha_str,
            "flight_id": fid,
            "flt_num": flt_num,
            "rte": rte,
            "tail": tail,
            "time": f"{dep_str}-{arr_str}",
            "blk": round(float(blk), 1),
            "duty": round(duty_hrs, 1),
            "crew": "2P CAP",
            "viat": viat_str,
            "card": _make_card("CAP"),
        })
        # Agregar fila para COP
        vuelos_open_rows.append({
            "position": "COP",
            "fecha": dep_base.date() if dep_base else None,
            "fecha_str": fecha_str,
            "flight_id": fid,
            "flt_num": flt_num,
            "rte": rte,
            "tail": tail,
            "time": f"{dep_str}-{arr_str}",
            "blk": round(float(blk), 1),
            "duty": round(duty_hrs, 1),
            "crew": "2P COP",
            "viat": viat_str,
            "card": _make_card("COP"),
        })
    df_vuelos_open = pd.DataFrame(vuelos_open_rows)
    if not df_vuelos_open.empty:
        df_vuelos_open = df_vuelos_open.sort_values(["position", "fecha", "flt_num"]).reset_index(drop=True)

    # 5b) Pilot hours (totales por rol)
    df_pilot_hours = pd.DataFrame(pilot_hours_rows)

    # 5c) Pilots required by day (global y por base)
    df_pilots_by_day = pd.DataFrame(pilots_by_day_rows)
    df_pilots_by_day_base = pd.DataFrame(pilots_by_day_base_rows)
    # 5d) Pilot lines (horizontal)
    df_pilot_lines = pd.DataFrame(pilot_lines_rows)
    if not df_pilot_lines.empty:
        cols = ["trip_number", "trip_days", "position"] + list(pilot_lines_dates or []) + ["VIATICOS_TOTAL"]
        cols = [c for c in cols if c in df_pilot_lines.columns]
        df_pilot_lines = df_pilot_lines.loc[:, cols]
    # 5e) Daily hours summary
    df_daily_hours = pd.DataFrame(daily_hours_rows)
    # 5f) Daily pilots excl (CAP/COP/CRP/DH/TA)
    df_daily_pilots_excl = pd.DataFrame(daily_pilots_excl_rows)
    # 5g) Viaticos detail
    df_viaticos = pd.DataFrame(viaticos_rows)
    # 5h) Planta requerida
    df_plant = pd.DataFrame(plant_required_rows)
    # 5i) Inicios de duty por estación
    df_duty_starts = pd.DataFrame(duty_starts_rows)


    # 6) KPIs (sin open)
    df_kpi = pd.DataFrame([
        {"metric": "system_name", "value": APP_NAME},
        {"metric": "coverage_pct", "value": kpis.get("coverage_pct")},
        {"metric": "network_eff_blk_per_day", "value": kpis.get("avg_network_eff")},
        {"metric": "cap_network_eff_blk_per_day", "value": kpis.get("cap_network_eff")},
        {"metric": "cop_network_eff_blk_per_day", "value": kpis.get("cop_network_eff")},
        {"metric": "crp_network_eff_blk_per_day", "value": kpis.get("crp_network_eff")},
        {"metric": "total_trips_selected", "value": kpis.get("total_trips")},
        {"metric": "total_block_hours", "value": kpis.get("total_blk")},
        {"metric": "total_trip_days", "value": kpis.get("total_days")},
        {"metric": "cap_block_hours_total", "value": kpis.get("cap_block_hours_total")},
        {"metric": "cop_block_hours_total", "value": kpis.get("cop_block_hours_total")},
        {"metric": "crp_block_hours_total", "value": kpis.get("crp_block_hours_total")},
        {"metric": "cap_duty_hours_total", "value": kpis.get("cap_duty_hours_total")},
        {"metric": "cop_duty_hours_total", "value": kpis.get("cop_duty_hours_total")},
        {"metric": "crp_duty_hours_total", "value": kpis.get("crp_duty_hours_total")},
        {"metric": "total_dh_hours_per_pilot", "value": kpis.get("total_dh_hours")},
        {"metric": "total_tafb_hours", "value": kpis.get("total_tafb_hours")},
        {"metric": "hotel_total_usd", "value": kpis.get("hotel_total_usd")},
        {"metric": "viaticos_total_usd", "value": kpis.get("viaticos_total_usd")},
        {"metric": "viaticos_total_cop", "value": kpis.get("viaticos_total_cop")},
    ])
    df_dist = pd.DataFrame([
        {"metric": "base_distribution", "value": str(base_dist)},
        {"metric": "pilot_base_distribution", "value": str(kpis.get("pilot_base_distribution", {}))},
        {"metric": "crew_distribution", "value": str(crew_dist)},
        {"metric": "man_days_cap", "value": man_days.get("cap")},
        {"metric": "man_days_cop", "value": man_days.get("cop")},
        {"metric": "man_days_crp", "value": man_days.get("crp")},
    ])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ============================================================================
        # DASHBOARD EJECUTIVO (Primera Hoja)
        # ============================================================================
        _create_executive_dashboard(writer, kpis, base_dist, crew_dist, man_days,
                                    selected_trips, flights_df, kpis_open, open_time_rows,
                                    monthly_kpis=monthly_kpis)

        # ============================================================================
        # PARÁMETROS DE CÁLCULO (Segunda Hoja)
        # ============================================================================
        _create_parameters_sheet(writer)

        # ============================================================================
        # GANTT CHART DE TRIPS (Visualización de calendario)
        # ============================================================================
        _create_trips_gantt_chart(writer, selected_trips)

        # ============================================================================
        # PAIRINGS FORMATO OLD (Calendario visual por trip)
        # ============================================================================
        _create_pairings_old_format(writer, selected_trips)

        # Resto de hojas
        df_flights.to_excel(writer, sheet_name=_safe_sheet_name("Flights"), index=False)
        df_itinerario.to_excel(writer, sheet_name=_safe_sheet_name("Itinerario_Cobertura"), index=False)
        df_trips.to_excel(writer, sheet_name=_safe_sheet_name("Trips_All"), index=False)
        df_selected.to_excel(writer, sheet_name=_safe_sheet_name("Trips_Selected"), index=False)
        if not df_trip_summary.empty:
            df_trip_summary.to_excel(writer, sheet_name=_safe_sheet_name("Trip_Summary"), index=False)
        if not df_second_pass.empty:
            df_second_pass.to_excel(writer, sheet_name=_safe_sheet_name("Second_Pass_Trips"), index=False)
        if not df_partial_trips.empty:
            df_partial_trips.to_excel(writer, sheet_name=_safe_sheet_name("Partial_Trips"), index=False)
        df_trip_report.to_excel(writer, sheet_name=_safe_sheet_name("Trip_Report"), index=False)
        df_legs.to_excel(writer, sheet_name=_safe_sheet_name("Trip_Legs_Raw"), index=False)
        df_hotels.to_excel(writer, sheet_name=_safe_sheet_name("Hotels"), index=False)
        df_open.to_excel(writer, sheet_name=_safe_sheet_name("OpenTime"), index=False)
        if not df_vuelos_open.empty:
            df_vuelos_open.to_excel(writer, sheet_name=_safe_sheet_name("Vuelos_Open"), index=False)
        df_kpi.to_excel(writer, sheet_name=_safe_sheet_name("KPIs"), index=False)
        df_dist.to_excel(writer, sheet_name=_safe_sheet_name("Distributions"), index=False)

        # Monthly KPIs sheet with executive style (positioned as 3rd sheet)
        if monthly_kpis:
            _create_monthly_kpis_executive(writer, monthly_kpis)

        # KPIs / Distributions con Open Tours (si aplica)
        if kpis_open:
            df_kpi_open = pd.DataFrame([
                {"metric": "system_name", "value": APP_NAME},
                {"metric": "coverage_pct", "value": kpis_open.get("coverage_pct")},
                {"metric": "network_eff_blk_per_day", "value": kpis_open.get("avg_network_eff")},
                {"metric": "cap_network_eff_blk_per_day", "value": kpis_open.get("cap_network_eff")},
                {"metric": "cop_network_eff_blk_per_day", "value": kpis_open.get("cop_network_eff")},
                {"metric": "crp_network_eff_blk_per_day", "value": kpis_open.get("crp_network_eff")},
                {"metric": "total_trips_selected", "value": kpis_open.get("total_trips")},
                {"metric": "total_block_hours", "value": kpis_open.get("total_blk")},
                {"metric": "total_trip_days", "value": kpis_open.get("total_days")},
                {"metric": "cap_block_hours_total", "value": kpis_open.get("cap_block_hours_total")},
                {"metric": "cop_block_hours_total", "value": kpis_open.get("cop_block_hours_total")},
                {"metric": "crp_block_hours_total", "value": kpis_open.get("crp_block_hours_total")},
                {"metric": "cap_duty_hours_total", "value": kpis_open.get("cap_duty_hours_total")},
                {"metric": "cop_duty_hours_total", "value": kpis_open.get("cop_duty_hours_total")},
                {"metric": "crp_duty_hours_total", "value": kpis_open.get("crp_duty_hours_total")},
                {"metric": "total_dh_hours_per_pilot", "value": kpis_open.get("total_dh_hours")},
                {"metric": "total_tafb_hours", "value": kpis_open.get("total_tafb_hours")},
                {"metric": "hotel_total_usd", "value": kpis_open.get("hotel_total_usd")},
                {"metric": "viaticos_total_usd", "value": kpis_open.get("viaticos_total_usd")},
                {"metric": "viaticos_total_cop", "value": kpis_open.get("viaticos_total_cop")},
            ])
            df_kpi_open.to_excel(writer, sheet_name=_safe_sheet_name("KPIs_Con_Open"), index=False)
        if base_dist_open is not None or crew_dist_open is not None or man_days_open is not None:
            df_dist_open = pd.DataFrame([
                {"metric": "base_distribution", "value": str(base_dist_open or {})},
                {"metric": "crew_distribution", "value": str(crew_dist_open or {})},
                {"metric": "man_days_cap", "value": (man_days_open or {}).get("cap")},
                {"metric": "man_days_cop", "value": (man_days_open or {}).get("cop")},
                {"metric": "man_days_crp", "value": (man_days_open or {}).get("crp")},
            ])
            df_dist_open.to_excel(writer, sheet_name=_safe_sheet_name("Distributions_Con_Open"), index=False)
        if open_time_rows_open is not None:
            df_open_open = pd.DataFrame(open_time_rows_open)
            df_open_open.to_excel(writer, sheet_name=_safe_sheet_name("OpenTime_Con_Open"), index=False)

        df_pilot_hours.to_excel(writer, sheet_name=_safe_sheet_name("Pilot_Hours"), index=False)
        df_pilots_by_day.to_excel(writer, sheet_name=_safe_sheet_name("Pilots_By_Day"), index=False)
        df_pilots_by_day_base.to_excel(writer, sheet_name=_safe_sheet_name("Pilots_By_Day_Base"), index=False)
        df_pilot_lines.to_excel(writer, sheet_name=_safe_sheet_name("Pilot_Lines"), index=False)
        df_daily_hours.to_excel(writer, sheet_name=_safe_sheet_name("Daily_Hours"), index=False)
        df_daily_pilots_excl.to_excel(writer, sheet_name=_safe_sheet_name("Pilots_By_Day_Excl"), index=False)
        df_viaticos.to_excel(writer, sheet_name=_safe_sheet_name("Viaticos_Detail"), index=False)
        df_plant.to_excel(writer, sheet_name=_safe_sheet_name("Plant_Required"), index=False)
        df_duty_starts.to_excel(writer, sheet_name=_safe_sheet_name("Duty_Starts_By_Station"), index=False)

        # DH Report por posición
        if dh_report_rows:
            df_dh_report = pd.DataFrame(dh_report_rows)
            # Ordenar por posición, fecha, trip_id
            if not df_dh_report.empty:
                sort_cols = [c for c in ['position', 'fecha', 'trip_id'] if c in df_dh_report.columns]
                if sort_cols:
                    df_dh_report = df_dh_report.sort_values(sort_cols).reset_index(drop=True)
                # Reordenar columnas para mejor lectura
                col_order = ['position', 'trip_id', 'dh_type', 'fecha', 'fecha_dia', 'flt_num',
                             'ruta', 'org', 'dst', 'hora_dep', 'hora_arr', 'blk_hours']
                col_order = [c for c in col_order if c in df_dh_report.columns]
                other_cols = [c for c in df_dh_report.columns if c not in col_order]
                df_dh_report = df_dh_report[col_order + other_cols]
            df_dh_report.to_excel(writer, sheet_name=_safe_sheet_name("DH_Report"), index=False)

        # Hoja Base_Analysis
        if base_analysis and base_analysis.get('bases'):
            ws_ba = writer.book.create_sheet("Base_Analysis")
            ba_bases = base_analysis['bases']
            headers = ['Base', 'Config %', 'Actual %', 'Natural %', 'Sugerida %',
                       'Eff Blk/Día', 'Trips', '2P', '3P', '4P', 'P-Days', 'Block(h)',
                       'DH Pos(h)', 'DH Res(h)']
            for c_idx, h in enumerate(headers, 1):
                ws_ba.cell(row=1, column=c_idx, value=h)
            for r_idx, b in enumerate(ba_bases, 2):
                s = base_analysis['base_stats'][b]
                ws_ba.cell(row=r_idx, column=1, value=b)
                ws_ba.cell(row=r_idx, column=2, value=round(base_analysis['configured_pct'].get(b, 0), 1))
                ws_ba.cell(row=r_idx, column=3, value=round(base_analysis['actual_pct'].get(b, 0), 1))
                ws_ba.cell(row=r_idx, column=4, value=round(base_analysis['natural_pct'].get(b, 0), 1))
                ws_ba.cell(row=r_idx, column=5, value=round(base_analysis['suggested_pct'].get(b, 0), 1))
                ws_ba.cell(row=r_idx, column=6, value=base_analysis['eff_by_base'].get(b, 0))
                ws_ba.cell(row=r_idx, column=7, value=s['trips'])
                ws_ba.cell(row=r_idx, column=8, value=s['crew_2p'])
                ws_ba.cell(row=r_idx, column=9, value=s['crew_3p'])
                ws_ba.cell(row=r_idx, column=10, value=s['crew_4p'])
                ws_ba.cell(row=r_idx, column=11, value=s['pilot_days'])
                ws_ba.cell(row=r_idx, column=12, value=round(s['block'], 1))
                ws_ba.cell(row=r_idx, column=13, value=round(s['dh_hours_positioning'], 1))
                ws_ba.cell(row=r_idx, column=14, value=round(s['dh_hours_rescue'], 1))
            r_sum = len(ba_bases) + 3
            ws_ba.cell(row=r_sum, column=1, value='DH Ahorro Estimado')
            ws_ba.cell(row=r_sum, column=2, value=base_analysis.get('dh_savings_estimate', 0))

        # Hoja Crew_Idle_Windows
        if idle_windows:
            ws_iw = writer.book.create_sheet("Crew_Idle_Windows")
            iw_headers = ['Trip', 'Base', 'Pilot', 'Role', 'Crew',
                          'Station', 'Window Start', 'Window End', 'Available Hrs']
            for c_idx, h in enumerate(iw_headers, 1):
                ws_iw.cell(row=1, column=c_idx, value=h)
            for r_idx, w in enumerate(idle_windows, 2):
                ws_iw.cell(row=r_idx, column=1, value=w.get('trip_id'))
                ws_iw.cell(row=r_idx, column=2, value=w.get('base'))
                ws_iw.cell(row=r_idx, column=3, value=w.get('pilot'))
                ws_iw.cell(row=r_idx, column=4, value=w.get('role'))
                ws_iw.cell(row=r_idx, column=5, value=w.get('trip_crew'))
                ws_iw.cell(row=r_idx, column=6, value=w.get('station'))
                ws_start = w.get('window_start')
                ws_end = w.get('window_end')
                ws_iw.cell(row=r_idx, column=7, value=ws_start.strftime('%Y-%m-%d %H:%M') if ws_start else '')
                ws_iw.cell(row=r_idx, column=8, value=ws_end.strftime('%Y-%m-%d %H:%M') if ws_end else '')
                ws_iw.cell(row=r_idx, column=9, value=w.get('available_hours', 0))

        # Ajustes de formato (freeze + auto ancho)
        wb = writer.book
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        def _apply_trip_colors(ws, trip_col_header, hide_trip_col=False):
            # Map trip id -> color
            header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            if trip_col_header not in header_map:
                return
            trip_col = header_map[trip_col_header]
            palette = [
                "FFF2CC", "D9E1F2", "E2F0D9", "FCE4D6",
                "E4DFEC", "DDEBF7", "F8CBAD", "C6E0B4",
            ]
            trip_color = {}
            for row in range(2, ws.max_row + 1):
                trip_id = ws.cell(row, trip_col).value
                if trip_id is None or trip_id == "":
                    continue
                if trip_id not in trip_color:
                    color = palette[len(trip_color) % len(palette)]
                    trip_color[trip_id] = PatternFill(start_color=color, end_color=color, fill_type="solid")
                fill = trip_color[trip_id]
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = fill
            if hide_trip_col:
                ws.column_dimensions[get_column_letter(trip_col)].hidden = True
        for sh in wb.sheetnames:
            ws = wb[sh]
            ws.freeze_panes = "A2"
            try:
                _autosize_worksheet_columns(ws)
            except Exception:
                pass
        # Colorear asignaciones por trip
        if "Trip_Report" in wb.sheetnames:
            _apply_trip_colors(wb["Trip_Report"], "TRIP_ID", hide_trip_col=True)
        if "Pilot_Lines" in wb.sheetnames:
            _apply_trip_colors(wb["Pilot_Lines"], "trip_number", hide_trip_col=False)
            from openpyxl.styles import Alignment
            ws_pl = wb["Pilot_Lines"]

            headers = [c.value for c in ws_pl[1]]
            skip_headers = {"trip_number", "position", "VIATICOS_TOTAL"}

            for col_idx, header in enumerate(headers, start=1):
                if header in skip_headers:
                    continue
                for row_idx in range(2, ws_pl.max_row + 1):
                    cell = ws_pl.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is None:
                        continue
                    sval = str(val)
                    if sval.strip() == "" or sval.strip().upper() == "DISPONIBLE":
                        continue
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                    try:
                        from openpyxl.styles import Font
                        cell.font = Font(bold=True)
                    except Exception:
                        pass

        # Colorear Itinerario_Cobertura: verde SI, rojo NO
        if "Itinerario_Cobertura" in wb.sheetnames:
            ws_itin = wb["Itinerario_Cobertura"]
            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # Encontrar columna CUBIERTO
            header_map = {cell.value: idx for idx, cell in enumerate(ws_itin[1], start=1)}
            cubierto_col = header_map.get("CUBIERTO", 1)
            for row in range(2, ws_itin.max_row + 1):
                cubierto_val = ws_itin.cell(row, cubierto_col).value
                if cubierto_val == "SI":
                    fill = fill_green
                else:
                    fill = fill_red
                for col in range(1, ws_itin.max_column + 1):
                    ws_itin.cell(row, col).fill = fill

        # Formatear Vuelos_Open con estilo de tarjeta
        if "Vuelos_Open" in wb.sheetnames:
            from openpyxl.styles import Alignment, Font
            ws_vo = wb["Vuelos_Open"]
            fill_beige = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            fill_cap = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # Verde claro para CAP
            fill_cop = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Azul claro para COP
            header_map = {cell.value: idx for idx, cell in enumerate(ws_vo[1], start=1)}
            card_col = header_map.get("card", None)
            position_col = header_map.get("position", None)
            for row in range(2, ws_vo.max_row + 1):
                pos_val = ws_vo.cell(row, position_col).value if position_col else ""
                if pos_val == "CAP":
                    fill = fill_cap
                elif pos_val == "COP":
                    fill = fill_cop
                else:
                    fill = fill_beige
                for col in range(1, ws_vo.max_column + 1):
                    cell = ws_vo.cell(row, col)
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            # Ajustar ancho de columna card
            if card_col:
                ws_vo.column_dimensions[get_column_letter(card_col)].width = 30
                # Ajustar altura de filas para mostrar el contenido
                for row in range(2, ws_vo.max_row + 1):
                    ws_vo.row_dimensions[row].height = 150

        # ── Formato numérico: estándar Excel (adapta a locale automáticamente) ─
        _NUM_FMT_INT = '#,##0'
        _NUM_FMT_DEC2 = '#,##0.00'
        _SKIP_SHEETS = {"Parametros_Calculo"}
        for sh in wb.sheetnames:
            if sh in _SKIP_SHEETS:
                continue
            ws = wb[sh]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                     min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, float):
                        abs_val = abs(cell.value)
                        if abs_val < 1 and abs_val > 0:
                            cell.number_format = _NUM_FMT_DEC2
                        elif abs_val == int(abs_val):
                            cell.number_format = _NUM_FMT_INT
                        else:
                            cell.number_format = _NUM_FMT_DEC2
                    elif isinstance(cell.value, int) and not isinstance(cell.value, bool):
                        cell.number_format = _NUM_FMT_INT

    return output_path


def _trip_calendar_span(trip, rules):
    """Calcula inicio/fin y días calendario consecutivos (incluye DH rescate si aplica)."""
    start_dt = trip['chain'][0]['start_base']
    pre_dh = trip.get('pre_dh')
    if pre_dh and pre_dh.get('start_base') is not None:
        start_dt = min(start_dt, pre_dh['start_base'])
    end_dt = trip['chain'][-1]['end_base']
    if trip.get('has_rescue_dh'):
        rescue_arr = trip.get('rescue_arr_base')
        if rescue_arr is not None:
            end_dt = rescue_arr
        else:
            rescue_from = trip.get('rescue_from')
            base = trip.get('base')
            try:
                dh_hours = rules.get_dh_time(rescue_from, base)
            except Exception:
                dh_hours = 0.0
            end_dt = end_dt + timedelta(hours=float(dh_hours or 0))
    days = (end_dt.date() - start_dt.date()).days + 1
    return start_dt, end_dt, days


def _ask_run_mode():
    print("\nModo de simulación:")
    print("  2P   = Solo 2P")
    print("  3P   = 2P y 3P")
    print("  3PONLY = Solo 3P")
    print("  4P   = 2P, 3P y 4P")
    print("  4P_MIN = 2P/3P y usar 4P solo si mejora cobertura")
    print("  AUTO = Ejecutar 2P / 3P / 4P y escoger el mejor")
    while True:
        choice = str(input("Ingrese 2P, 3P, 3PONLY, 4P, 4P_MIN o AUTO [AUTO]: ")).strip().upper()
        if choice == "":
            choice = "AUTO"
        if choice in {"2P", "3P", "3PONLY", "4P", "4P_MIN", "AUTO"}:
            if choice == "2P":
                return [("2P", {"2P"})]
            if choice == "3P":
                return [("3P", {"2P", "3P"})]
            if choice == "3PONLY":
                return [("3P_ONLY", {"3P"})]
            if choice == "4P":
                return [("4P", {"2P", "3P", "4P"})]
            if choice == "4P_MIN":
                return [("4P_MIN", {"2P", "3P", "4P"})]
            return [
                ("2P", {"2P"}),
                ("3P", {"2P", "3P"}),
                ("3P_ONLY", {"3P"}),
                ("4P", {"2P", "3P", "4P"}),
                ("4P_MIN", {"2P", "3P", "4P"}),
            ]
        print("Opción inválida. Intente de nuevo.")


# =====================================================================
# IDENTIFICACIÓN DE VENTANAS OCIOSAS (PILOTOS TA)
# =====================================================================
def _identify_idle_windows(selected_trips, rules):
    """Para cada trip 3P/4P, identifica duties donde algún piloto es TA (ocioso)."""
    windows = []
    for t in selected_trips:
        trip_crew = t.get('trip_crew', '2P')
        crew_size = CREW_RANK.get(trip_crew, 2)
        if crew_size <= 2:
            continue
        chain = t.get('chain', [])
        if not chain:
            continue
        pilots = _get_pilot_list(trip_crew)
        trip_id = t.get('id', 0)
        base = t.get('base', '')

        for pid, role in pilots:
            idle_start = None
            idle_station = None
            for i, duty in enumerate(chain):
                min_crew = duty.get('min_crew', trip_crew)
                ops = _ops_for_req(trip_crew, min_crew)
                if pid not in ops:
                    if idle_start is None:
                        idle_start = duty.get('start_base')
                        idle_station = duty.get('org', '')
                    idle_end = duty.get('end_base')
                    idle_station_end = duty.get('dst', '')
                    next_needs_pilot = True
                    if i + 1 < len(chain):
                        next_min = chain[i + 1].get('min_crew', trip_crew)
                        next_ops = _ops_for_req(trip_crew, next_min)
                        next_needs_pilot = (pid in next_ops)
                    if next_needs_pilot or i + 1 >= len(chain):
                        window_end = chain[i + 1].get('start_base') if i + 1 < len(chain) else None
                        available_hours = 0
                        if window_end and idle_start:
                            available_hours = (window_end - idle_start).total_seconds() / 3600
                        elif idle_end and idle_start:
                            available_hours = (idle_end - idle_start).total_seconds() / 3600
                        windows.append({
                            'trip_id': trip_id, 'base': base, 'pilot': pid,
                            'role': role, 'trip_crew': trip_crew,
                            'station': idle_station_end,
                            'window_start': idle_start,
                            'window_end': window_end or idle_end,
                            'available_hours': round(available_hours, 1),
                        })
                        idle_start = None
                        idle_station = None
                else:
                    idle_start = None
                    idle_station = None

    if windows:
        print(f"\n  " + "="*70)
        print(f"   VENTANAS OCIOSAS DE PILOTOS (TA)")
        print(f"  " + "="*70)
        print(f"  Total ventanas: {len(windows)}")
        total_hours = sum(w['available_hours'] for w in windows)
        print(f"  Total horas ociosas: {total_hours:.1f}h")
        by_crew = {}
        for w in windows:
            by_crew.setdefault(w['trip_crew'], []).append(w)
        for tc, ws_list in sorted(by_crew.items()):
            hrs = sum(w['available_hours'] for w in ws_list)
            print(f"  {tc}: {len(ws_list)} ventanas, {hrs:.1f}h ociosas")
        by_station = {}
        for w in windows:
            by_station.setdefault(w['station'], []).append(w)
        print(f"\n  {'Estación':<8} {'Ventanas':<10} {'Horas':<10}")
        print(f"  {'-'*8} {'-'*10} {'-'*10}")
        for stn in sorted(by_station, key=lambda s: -sum(w['available_hours'] for w in by_station[s])):
            ws_list = by_station[stn]
            hrs = sum(w['available_hours'] for w in ws_list)
            print(f"  {stn:<8} {len(ws_list):<10} {hrs:<10.1f}")
        print(f"  " + "="*70)
    return windows


# =====================================================================
# ANÁLISIS DE DISTRIBUCIÓN DE BASES
# =====================================================================
def _analyze_base_distribution(selected_trips, rules):
    """Analiza distribución natural/actual/sugerida de pilotos por base."""
    bases = sorted(rules.BASES)
    if len(bases) < 2:
        return None

    base_stats = {b: {
        'trips': 0, 'pilot_days': 0, 'block': 0.0, 'days': 0,
        'dh_hours_positioning': 0.0, 'dh_hours_rescue': 0.0,
        'duties': 0, 'crew_2p': 0, 'crew_3p': 0, 'crew_4p': 0
    } for b in bases}

    duty_base_affinity = []
    for t in selected_trips:
        base = t.get('base', '')
        if base not in base_stats:
            continue
        chain = t.get('chain', [])
        crew = t.get('trip_crew', '2P')
        crew_size = CREW_RANK.get(crew, 2)
        stats = base_stats[base]
        stats['trips'] += 1
        stats['pilot_days'] += crew_size * t.get('days', 0)
        stats['block'] += float(t.get('block', 0))
        stats['days'] += t.get('days', 0)
        stats['duties'] += len(chain)
        if crew == '2P': stats['crew_2p'] += 1
        elif crew == '3P': stats['crew_3p'] += 1
        else: stats['crew_4p'] += 1
        pre_dh = t.get('pre_dh')
        if pre_dh:
            stats['dh_hours_positioning'] += float(pre_dh.get('dh_hours', 0) or 0)
        rescue_hrs = float(t.get('rescue_dh_hours', 0) or 0)
        stats['dh_hours_rescue'] += rescue_hrs
        for d in chain:
            org = d.get('org', '')
            dst = d.get('dst', '')
            if org in bases:
                duty_base_affinity.append((d, org, 'starts_at_base'))
            elif dst in bases:
                duty_base_affinity.append((d, dst, 'ends_at_base'))
            else:
                duty_base_affinity.append((d, base, 'inherited'))

    natural_demand = {b: 0 for b in bases}
    for duty, best_base, reason in duty_base_affinity:
        natural_demand[best_base] += 1
    total_duties = sum(natural_demand.values())
    natural_pct = {b: (natural_demand[b] / total_duties * 100) if total_duties > 0 else 0 for b in bases}

    total_pd = sum(s['pilot_days'] for s in base_stats.values())
    actual_pct = {b: (base_stats[b]['pilot_days'] / total_pd * 100) if total_pd > 0 else 0 for b in bases}

    eff_by_base = {}
    for b in bases:
        s = base_stats[b]
        eff_by_base[b] = round(s['block'] / s['days'], 2) if s['days'] > 0 else 0

    total_dh_pos = sum(s['dh_hours_positioning'] for s in base_stats.values())
    total_dh_res = sum(s['dh_hours_rescue'] for s in base_stats.values())

    dh_savings_estimate = 0.0
    for t in selected_trips:
        base = t.get('base', '')
        chain = t.get('chain', [])
        pre_dh = t.get('pre_dh')
        if not chain or not pre_dh:
            continue
        first_org = chain[0].get('org', '')
        if first_org in bases and first_org != base:
            dh_savings_estimate += float(pre_dh.get('dh_hours', 0) or 0)

    configured_pct = {b: BASE_TARGETS.get(b, 0) * 100 for b in bases}
    suggested_pct = {}
    for b in bases:
        suggested_pct[b] = round(natural_pct[b] * 0.7 + actual_pct[b] * 0.3, 1)
    total_sug = sum(suggested_pct.values())
    if total_sug > 0:
        suggested_pct = {b: round(v / total_sug * 100, 1) for b, v in suggested_pct.items()}

    analysis = {
        'bases': bases, 'base_stats': base_stats,
        'natural_pct': natural_pct, 'actual_pct': actual_pct,
        'configured_pct': configured_pct, 'suggested_pct': suggested_pct,
        'eff_by_base': eff_by_base,
        'total_dh_positioning': round(total_dh_pos, 1),
        'total_dh_rescue': round(total_dh_res, 1),
        'dh_savings_estimate': round(dh_savings_estimate, 1),
    }

    print("\n  " + "="*70)
    print("   ANÁLISIS DE DISTRIBUCIÓN DE BASES")
    print("  " + "="*70)
    print(f"\n  {'Base':<6} {'Config':<10} {'Actual':<10} {'Natural':<10} {'Sugerida':<10} {'Eff Blk/Día':<12}")
    print(f"  {'-'*6} {'-'*10} {'-'*10} {'-'*10} {'-'*10} {'-'*12}")
    for b in bases:
        print(f"  {b:<6} {configured_pct[b]:>7.1f}%  {actual_pct[b]:>7.1f}%  {natural_pct[b]:>7.1f}%  {suggested_pct[b]:>7.1f}%  {eff_by_base[b]:>9.2f}")
    print(f"\n  DH Posicionamiento total: {total_dh_pos:.1f}h")
    print(f"  DH Rescate total:         {total_dh_res:.1f}h")
    if dh_savings_estimate > 0:
        print(f"  DH ahorro estimado: {dh_savings_estimate:.1f}h")
    print("  " + "="*70)
    return analysis


def solve_roster(
    flights_df,
    trips,
    target_distribution=None,
    export_excel=True,
    excel_output_path=None,
    rules=None,
    dh_table=None,
    coverage_first=False,
):
    if not trips:
        if VERBOSE_OUTPUT:
            print("No hay trips.")
        return None

    print("\r  [..] Optimizando .......                   ", end="", flush=True)
    if VERBOSE_OUTPUT:
        print(f"\n3. Ejecutando Optimizador (Con Balanceo y Restricciones)...")
    flight_ids = flights_df['id'].unique()
    flight_in_trips = {fid: [] for fid in flight_ids}
    for i, t in enumerate(trips):
        for fid in t['flights_covered']:
            if fid in flight_in_trips:
                flight_in_trips[fid].append(i)

    def _add_base_balance(prob, x_vars):
        if target_distribution and ENFORCE_BASE_BALANCE:
            total_trips_var = pulp.lpSum([x_vars[i] for i in range(len(trips))])
            for base, target in target_distribution.items():
                base_trips = pulp.lpSum([x_vars[i] for i in range(len(trips)) if trips[i]['base'] == base])
                prob += base_trips >= total_trips_var * (target - BASE_BALANCE_TOLERANCE)
                prob += base_trips <= total_trips_var * (target + BASE_BALANCE_TOLERANCE)

    crew_size_map = {"2P": 2, "3P": 3, "4P": 4}

    def _eff_expr(x_vars):
        return pulp.lpSum([
            (EFFICIENCY_DAY_WEIGHT * float(trips[i].get("days", 0) or 0) -
             EFFICIENCY_BLOCK_WEIGHT * float(trips[i].get("block", 0) or 0)) * x_vars[i]
            for i in range(len(trips))
        ])

    def _pilots_expr(x_vars):
        return pulp.lpSum([
            (PILOT_DAY_WEIGHT * crew_size_map.get(trips[i].get("trip_crew"), 0) *
             float(trips[i].get("days", 0) or 0)) * x_vars[i]
            for i in range(len(trips))
        ])

    def _cost_expr(x_vars):
        return pulp.lpSum([float(trips[i]['cost']) * x_vars[i] for i in range(len(trips))])

    def _objective_expr(x_vars):
        eff_expr = _eff_expr(x_vars)
        pilots_expr = _pilots_expr(x_vars)
        cost_expr = _cost_expr(x_vars)

        if OPTIMIZATION_OBJECTIVE == "EFFICIENCY":
            expr = eff_expr
        elif OPTIMIZATION_OBJECTIVE == "PILOTS":
            expr = pilots_expr
        elif OPTIMIZATION_OBJECTIVE == "PILOTS_EFF":
            expr = (PILOTS_EFF_WEIGHT_PILOTS * pilots_expr) + (PILOTS_EFF_WEIGHT_EFF * eff_expr)
        elif OPTIMIZATION_OBJECTIVE == "HYBRID":
            expr = (HYBRID_WEIGHT_EFF * eff_expr) + (HYBRID_WEIGHT_PILOTS * pilots_expr) + (HYBRID_WEIGHT_COST * cost_expr)
        else:
            expr = cost_expr

        if OPTIMIZATION_OBJECTIVE != "COST" and OBJECTIVE_TIEBREAKER_COST:
            expr += OBJECTIVE_TIEBREAKER_COST * cost_expr
        return expr

    # --- Fases de optimización ---
    best_cov = None
    best_pilots = None
    # Intentar HiGHS primero (más rápido), fallback a CBC
    try:
        _test_solver = pulp.HiGHS(msg=False, timeLimit=SOLVER_TIME_LIMIT_SECONDS)
        solver_cmd = _test_solver
        print("  [Solver] Usando HiGHS (API)")
    except Exception:
        solver_cmd = pulp.PULP_CBC_CMD(msg=False, timeLimit=SOLVER_TIME_LIMIT_SECONDS)
        print("  [Solver] HiGHS no disponible, usando CBC")

    x_final = None

    if ENABLE_LEXICOGRAPHIC:
        # Fase 1: maximizar cobertura
        prob1 = pulp.LpProblem("CrewScheduler_Coverage", pulp.LpMaximize)
        x1 = pulp.LpVariable.dicts("Trip", range(len(trips)), cat='Binary')
        y1 = pulp.LpVariable.dicts("Cover", flight_ids, cat='Binary')
        for fid in flight_ids:
            if flight_in_trips[fid]:
                prob1 += pulp.lpSum([x1[i] for i in flight_in_trips[fid]]) >= y1[fid]
            else:
                prob1 += y1[fid] == 0
        _add_base_balance(prob1, x1)
        prob1 += pulp.lpSum([y1[fid] for fid in flight_ids])
        try:
            prob1.solve(solver_cmd)
        except pulp.PulpSolverError as e:
            print(f"Solver error (cobertura): {e}")
            return None
        if pulp.LpStatus[prob1.status] != 'Optimal':
            if VERBOSE_OUTPUT:
                print(f"Solver status (cobertura): {pulp.LpStatus[prob1.status]}")
            return None
        best_cov = sum((y1[fid].value() or 0) for fid in flight_ids)
        if VERBOSE_OUTPUT:
            print(f">> Cobertura máxima alcanzada: {int(best_cov)} vuelos")

        # Fase 2: minimizar pilotos-día con cobertura fija
        prob2 = pulp.LpProblem("CrewScheduler_Pilots", pulp.LpMinimize)
        x2 = pulp.LpVariable.dicts("Trip", range(len(trips)), cat='Binary')
        y2 = pulp.LpVariable.dicts("Cover", flight_ids, cat='Binary')
        for fid in flight_ids:
            if flight_in_trips[fid]:
                prob2 += pulp.lpSum([x2[i] for i in flight_in_trips[fid]]) >= y2[fid]
            else:
                prob2 += y2[fid] == 0
        prob2 += pulp.lpSum([y2[fid] for fid in flight_ids]) >= best_cov
        _add_base_balance(prob2, x2)
        prob2 += _pilots_expr(x2)
        skip_eff = False
        try:
            prob2.solve(solver_cmd)
        except pulp.PulpSolverError as e:
            print(f"Solver error (pilotos): {e}")
            print("Usando solución de cobertura máxima (fase 1).")
            x_final = x1
            best_pilots = None
            skip_eff = True
        if not skip_eff and pulp.LpStatus[prob2.status] != 'Optimal':
            if VERBOSE_OUTPUT:
                print(f"Solver status (pilotos): {pulp.LpStatus[prob2.status]}")
            print("Usando solución de cobertura máxima (fase 1).")
            x_final = x1
            best_pilots = None
            skip_eff = True
        if not skip_eff:
            best_pilots = pulp.value(_pilots_expr(x2))

            # Fase 3: maximizar eficiencia con cobertura y pilotos-día fijos
            prob3 = pulp.LpProblem("CrewScheduler_Eff", pulp.LpMinimize)
            x3 = pulp.LpVariable.dicts("Trip", range(len(trips)), cat='Binary')
            y3 = pulp.LpVariable.dicts("Cover", flight_ids, cat='Binary')
            for fid in flight_ids:
                if flight_in_trips[fid]:
                    prob3 += pulp.lpSum([x3[i] for i in flight_in_trips[fid]]) >= y3[fid]
                else:
                    prob3 += y3[fid] == 0
            prob3 += pulp.lpSum([y3[fid] for fid in flight_ids]) >= best_cov
            if best_pilots is not None:
                prob3 += _pilots_expr(x3) <= best_pilots + PILOT_DAY_TOLERANCE
            _add_base_balance(prob3, x3)
            prob3 += _eff_expr(x3)
            try:
                prob3.solve(solver_cmd)
            except pulp.PulpSolverError as e:
                print(f"Solver error (eficiencia): {e}")
                x_final = x2
            if pulp.LpStatus[prob3.status] == 'Optimal':
                x_final = x3
            else:
                x_final = x2
    else:
        # --- Fase 1: maximizar cobertura (opcional) ---
        if coverage_first:
            prob1 = pulp.LpProblem("CrewScheduler_Coverage", pulp.LpMaximize)
            x1 = pulp.LpVariable.dicts("Trip", range(len(trips)), cat='Binary')
            y1 = pulp.LpVariable.dicts("Cover", flight_ids, cat='Binary')
            for fid in flight_ids:
                if flight_in_trips[fid]:
                    prob1 += pulp.lpSum([x1[i] for i in flight_in_trips[fid]]) >= y1[fid]
                else:
                    prob1 += y1[fid] == 0
            _add_base_balance(prob1, x1)
            prob1 += pulp.lpSum([y1[fid] for fid in flight_ids])
            prob1.solve(solver_cmd)
            if pulp.LpStatus[prob1.status] != 'Optimal':
                if VERBOSE_OUTPUT:
                    print(f"Solver status (cobertura): {pulp.LpStatus[prob1.status]}")
                return None
            best_cov = sum((y1[fid].value() or 0) for fid in flight_ids)
            if VERBOSE_OUTPUT:
                print(f">> Cobertura máxima alcanzada: {int(best_cov)} vuelos")

        # --- Fase 2: minimizar objetivo (con cobertura fijada) ---
        prob = pulp.LpProblem("CrewScheduler", pulp.LpMinimize)
        x = pulp.LpVariable.dicts("Trip", range(len(trips)), cat='Binary')
        if coverage_first:
            y = pulp.LpVariable.dicts("Cover", flight_ids, cat='Binary')
            for fid in flight_ids:
                if flight_in_trips[fid]:
                    prob += pulp.lpSum([x[i] for i in flight_in_trips[fid]]) >= y[fid]
                else:
                    prob += y[fid] == 0
            if best_cov is not None:
                prob += pulp.lpSum([y[fid] for fid in flight_ids]) >= best_cov
        else:
            for fid in flight_ids:
                if flight_in_trips[fid]:
                    prob += pulp.lpSum([x[i] for i in flight_in_trips[fid]]) >= 1
        _add_base_balance(prob, x)
        prob += _objective_expr(x)
        prob.solve(solver_cmd)

        if pulp.LpStatus[prob.status] != 'Optimal':
            if VERBOSE_OUTPUT:
                print(f"Solver status: {pulp.LpStatus[prob.status]}")
            return None
        x_final = x

    # =============================
    # Construcción de resultados
    # =============================
    selected_trips = [trips[i] for i in range(len(trips)) if (x_final[i].value() or 0) >= 0.9]
    print(f"\r  [OK] Optimización completada: {len(selected_trips)} rotaciones", flush=True)

    # Análisis de bases y ventanas ociosas
    base_analysis = _analyze_base_distribution(selected_trips, rules)
    idle_windows = _identify_idle_windows(selected_trips, rules)

    # Actualizar razones de exclusión: limpiar vuelos cubiertos
    covered_flight_ids = set()
    for t in selected_trips:
        for fid in t.get('flights_covered', []):
            covered_flight_ids.add(fid)

    # Limpiar razones para vuelos cubiertos (ya no están excluidos)
    for fid in covered_flight_ids:
        if fid in FLIGHT_EXCLUSION_REASONS:
            del FLIGHT_EXCLUSION_REASONS[fid]

    if VERBOSE_OUTPUT:
        print("\n" + "="*150)
        print(f"   REPORTE DETALLADO DE ASIGNACIÓN DE TRIPULACIONES")
        print("="*150)

    # Variables de agregación
    total_blk = 0
    total_days = 0
    covered = set()
    man_days = {
        'cap': 0, 'cop': 0, 'crp': 0,
        'cap_by_crew': {'2P': 0, '3P': 0, '4P': 0},
        'cop_by_crew': {'2P': 0, '3P': 0, '4P': 0},
        'crp_by_crew': {'2P': 0, '3P': 0, '4P': 0},
    }
    # Horas totales por rol (Block y Duty) — créditos por piloto
    cap_block_hours_total = 0.0
    cop_block_hours_total = 0.0
    crp_block_hours_total = 0.0
    cap_duty_hours_total = 0.0
    cop_duty_hours_total = 0.0
    crp_duty_hours_total = 0.0
    # Horas por mes y rol (para discriminación mensual en Pilot_Hours)
    cap_block_by_month = Counter()  # {(year, month): hours}
    cop_block_by_month = Counter()
    crp_block_by_month = Counter()
    cap_duty_by_month = Counter()
    cop_duty_by_month = Counter()
    crp_duty_by_month = Counter()
    # Horas DH y TAFB (Time Away From Base)
    total_dh_hours = 0.0
    total_tafb_hours = 0.0

    # Nuevas métricas AV KPIS
    total_duties = 0
    total_ac_changes = 0
    total_hotel_nights = 0
    total_dh_flights = 0
    # Distribución de duties por rangos de duty time
    duties_by_duty_time = {
        "00:00-04:00": 0,
        "04:01-08:00": 0,
        "08:01-10:00": 0,
        ">10:00": 0,
    }
    # Distribución de duties por rangos de block time
    duties_by_block_time = {
        "00:00-02:00": 0,
        "02:01-04:00": 0,
        "04:01-06:00": 0,
        "06:01-08:00": 0,
    }

    # ====================================================================
    # MONTHLY KPIs - Métricas mensuales para el Dashboard
    # ====================================================================
    # KPIs por mes: {(year, month): value}
    trips_by_month = Counter()
    duties_by_month = Counter()
    ac_changes_by_month = Counter()
    hotel_nights_by_month = Counter()
    dh_flights_by_month = Counter()
    dh_hours_by_month = Counter()
    tafb_hours_by_month = Counter()
    hotel_cost_by_month = Counter()
    viaticos_by_month = Counter()
    duty_days_by_month = Counter()
    block_hours_by_month = Counter()  # Total (sum of CAP+COP+CRP)
    duty_hours_by_month = Counter()   # Total duty hours
    flights_covered_by_month = Counter()
    flights_total_by_month = Counter()
    # Por base y mes: {(year, month, base): value}
    block_hours_by_month_base = Counter()
    duty_days_by_month_base = Counter()
    trips_by_month_base = Counter()

    # Pilotos requeridos por día (en hora base UTC-5)
    pilots_by_day = Counter()
    pilots_by_day_cap = Counter()
    pilots_by_day_cop = Counter()
    pilots_by_day_crp = Counter()
    # Inicios de duty por estación (personas)
    duty_starts_by_station = Counter()

    # Pilotos requeridos por día y por base (date, base)
    pilots_by_day_base_cap = Counter()
    pilots_by_day_base_cop = Counter()
    pilots_by_day_base_crp = Counter()
    # Pilotos requeridos por día (excluyente CAP/COP/CRP/DH/TA)
    daily_pilots_excl = {}
    lays = Counter()
    hotels_by_month = Counter()
    base_c = Counter()
    base_pilot_days = Counter()
    crew_c = Counter()

    if rules is None:
        rules = RuleEngine(dh_table=dh_table)

    # Calculate total flights by month for coverage calculation
    for _, flt in flights_df.iterrows():
        flt_date = flt['dep_utc'].date() if hasattr(flt['dep_utc'], 'date') else flt['dep_utc']
        if hasattr(flt_date, 'year'):
            flights_total_by_month[(flt_date.year, flt_date.month)] += 1

    # Para exportación: detalle por duty/leg
    trip_legs_rows = []
    # Para exportación: reporte de trips (tal cual formato consola)
    trip_report_rows = []
    # Para exportación: líneas por piloto (por día)
    pilot_lines_rows = []
    pilot_lines_dates = []
    pilot_lines_dates_set = set()
    # Para exportación: resumen diario de horas (CAP/COP/CRP/DH/TA)
    daily_hours = {}
    # Para exportación: viáticos
    viaticos_rows = []
    # Para exportación: reporte de DH por posición
    dh_report_rows = []

    weekday_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
    crew_sizes = {"2P": 2, "3P": 3, "4P": 4}
    def _op_counts(trip_crew, req):
        if trip_crew == "4P":
            if req == "4P":
                return {"CAP": 2, "COP": 2, "CRP": 0}
            if req == "3P":
                return {"CAP": 1, "COP": 2, "CRP": 0}
            return {"CAP": 1, "COP": 1, "CRP": 0}
        if trip_crew == "3P":
            if req == "3P":
                return {"CAP": 1, "COP": 1, "CRP": 1}
            return {"CAP": 1, "COP": 1, "CRP": 0}
        return {"CAP": 1, "COP": 1, "CRP": 0}
    def _fmt_base_dt(dt):
        wd = weekday_es[dt.weekday()]
        return f"{wd} {dt.strftime('%d-%b %H:%M')}"
    def _fmt_day(d):
        wd = weekday_es[d.weekday()]
        return f"{wd} {d.strftime('%d-%b')}"
    def _fmt_day_col(d):
        return d.strftime("%d-%b")

    def _add_dh_to_report(trip_id, trip_crew, dh_type, leg, dep_base, arr_base):
        """Agrega registros de DH al reporte por cada posición del trip."""
        flt_num = leg.get('flt_num', '')
        aln = leg.get('aln', 'AV')
        if flt_num and not any(str(flt_num).startswith(p) for p in ('AV', 'AM', 'LA', 'CM', 'AR')):
            flt_num = f"{aln}{flt_num}"
        org = leg.get('org', '')
        dst = leg.get('dst', '')
        blk_h = (leg.get('arr_utc', arr_base) - leg.get('dep_utc', dep_base)).total_seconds() / 3600 if leg.get('arr_utc') and leg.get('dep_utc') else (arr_base - dep_base).total_seconds() / 3600
        fecha_str = dep_base.strftime("%Y-%m-%d") if dep_base else ""
        fecha_dia = dep_base.strftime("%d-%b") if dep_base else ""
        hora_dep = dep_base.strftime("%H:%M") if dep_base else ""
        hora_arr = arr_base.strftime("%H:%M") if arr_base else ""

        # Determinar posiciones según trip_crew
        if trip_crew == '4P':
            positions = ['CAP1', 'CAP2', 'COP1', 'COP2']
        elif trip_crew == '3P':
            positions = ['CAP', 'COP', 'CRP']
        else:  # 2P
            positions = ['CAP', 'COP']

        for pos in positions:
            dh_report_rows.append({
                'trip_id': trip_id,
                'position': pos,
                'dh_type': dh_type,
                'fecha': fecha_str,
                'fecha_dia': fecha_dia,
                'flt_num': flt_num or 'DH',
                'ruta': f"{org}-{dst}",
                'org': org,
                'dst': dst,
                'hora_dep': hora_dep,
                'hora_arr': hora_arr,
                'blk_hours': round(blk_h, 2),
            })

    for t in selected_trips:
        trip_crew_size = crew_sizes.get(t.get('trip_crew'), 0)
        trip_start_dt, trip_end_dt, trip_days = _trip_calendar_span(t, rules)
        trip_start_date = trip_start_dt.date()
        trip_end_date = trip_end_dt.date()
        last_date = trip_start_date
        day_events = {}

        def _add_event(dt, kind, route, start_dt, end_dt, block_h, duty_h, req=None, viat=None, meta=None):
            d = dt.date()
            evt = {
                "time": dt,
                "kind": kind,
                "route": route,
                "start": start_dt,
                "end": end_dt,
                "block": block_h,
                "duty": duty_h,
                "req": req,
            }
            if viat:
                evt.update(viat)
            if meta:
                evt.update(meta)
            day_events.setdefault(d, []).append(evt)

        def _add_continuation(route_base, start_dt, end_dt):
            # Solo si toca el día calendario siguiente
            if end_dt.date() <= start_dt.date():
                return
            cur = start_dt.date() + timedelta(days=1)
            if cur > end_dt.date():
                return
            dt = datetime.combine(cur, datetime.min.time())
            day_events.setdefault(cur, []).append({
                "time": dt,
                "kind": "CONT",
                "route": f"CONT.. {route_base}",
                "start": dt,
                "end": dt,
                "block": 0.0,
                "duty": 0.0,
                "req": None,
            })
            # Línea CONT.. en Trip_Report
            d_date = _fmt_base_dt(dt)
            if VERBOSE_OUTPUT:
                print(f"{d_date:<12} {f'CONT.. {route_base}':<30} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} {'':<}")
            trip_report_rows.append({
                "FECHA": d_date,
                "RUTA/ACTIVIDAD": f"CONT.. {route_base}",
                "AVIÓN": "---",
                "REQ": "---",
                "ASG": "---",
                "BLK": "---",
                "MAX": "---",
                "DUTY": "---",
                "MAX.1": "---",
                "REST (Act/Req)": "---",
                "PERNOCTA": "",
            })

        def _emit_avl_until(target_date):
            nonlocal last_date
            cur = last_date + timedelta(days=1)
            while cur < target_date:
                dt = datetime.combine(cur, datetime.min.time())
                d_date = _fmt_base_dt(dt)
                route = "AVL - BLANCO"
                if VERBOSE_OUTPUT:
                    print(f"{d_date:<12} {route:<30} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} {'':<}")
                trip_report_rows.append({
                    "FECHA": d_date,
                    "RUTA/ACTIVIDAD": route,
                    "AVIÓN": "---",
                    "REQ": "---",
                    "ASG": "---",
                    "BLK": "---",
                    "MAX": "---",
                    "DUTY": "---",
                    "MAX.1": "---",
                    "REST (Act/Req)": "---",
                    "PERNOCTA": "",
                })
                last_date = cur
                cur += timedelta(days=1)

        total_blk += t['block']
        total_days += trip_days
        covered.update(t['flights_covered'])
        base_c[t['base']] += 1
        # Nota: total_dh_hours se calculará después desde dh_report_rows para garantizar consistencia
        # Calcular TAFB (Time Away From Base) - desde inicio hasta llegada a base
        trip_tafb = float(trip_days * 24)  # Aproximación en horas calendario
        total_tafb_hours += trip_tafb
        base_pilot_days[t['base']] += int(trip_crew_size * trip_days)
        crew_c[t['trip_crew']] += 1

        # ====================================================================
        # MONTHLY TRACKING - Trip level
        # ====================================================================
        trip_month_key = (trip_start_date.year, trip_start_date.month)
        trips_by_month[trip_month_key] += 1
        trips_by_month_base[(trip_start_date.year, trip_start_date.month, t['base'])] += 1
        duty_days_by_month[trip_month_key] += trip_days
        duty_days_by_month_base[(trip_start_date.year, trip_start_date.month, t['base'])] += trip_days
        tafb_hours_by_month[trip_month_key] += trip_tafb
        block_hours_by_month[trip_month_key] += t['block']
        block_hours_by_month_base[(trip_start_date.year, trip_start_date.month, t['base'])] += t['block']
        # Track covered flights by month
        for fid in t.get('flights_covered', []):
            # Get flight date from flights_df if available
            if fid in flights_df['id'].values:
                flt_row = flights_df[flights_df['id'] == fid].iloc[0]
                flt_date = flt_row['dep_utc'].date() if hasattr(flt_row['dep_utc'], 'date') else flt_row['dep_utc']
                if hasattr(flt_date, 'year'):
                    flights_covered_by_month[(flt_date.year, flt_date.month)] += 1

        # Recursos Humanos
        d = trip_days
        crew_type = t['trip_crew']
        if crew_type == '2P':
            man_days['cap'] += d
            man_days['cop'] += d
            man_days['cap_by_crew']['2P'] += d
            man_days['cop_by_crew']['2P'] += d
        elif crew_type == '3P':
            man_days['cap'] += d
            man_days['cop'] += d
            man_days['crp'] += d
            man_days['cap_by_crew']['3P'] += d
            man_days['cop_by_crew']['3P'] += d
            man_days['crp_by_crew']['3P'] += d
        elif crew_type == '4P':
            man_days['cap'] += d * 2
            man_days['cop'] += d * 2
            man_days['cap_by_crew']['4P'] += d * 2
            man_days['cop_by_crew']['4P'] += d * 2

        # Horas totales (Block y Duty) por rol — solo según REQ (sin contar TA)
        for dd in t.get('chain', []):
            req = dd.get('min_crew')
            blk = float(dd.get('block', 0) or 0)
            dty = float(dd.get('duty_dur', 0) or 0)
            dte = dd.get('start_base').date()
            month_key = (dte.year, dte.month)
            rec = daily_hours.setdefault(dte, {"date": dte, "cap_block": 0.0, "cop_block": 0.0, "crp_block": 0.0, "dh_hours": 0.0, "ta_hours": 0.0})
            if req == '2P':
                cap_block_hours_total += blk
                cop_block_hours_total += blk
                cap_duty_hours_total += dty
                cop_duty_hours_total += dty
                cap_block_by_month[month_key] += blk
                cop_block_by_month[month_key] += blk
                cap_duty_by_month[month_key] += dty
                cop_duty_by_month[month_key] += dty
                rec["cap_block"] += blk
                rec["cop_block"] += blk
            elif req == '3P':
                cap_block_hours_total += blk
                cop_block_hours_total += blk
                crp_block_hours_total += blk
                cap_duty_hours_total += dty
                cop_duty_hours_total += dty
                crp_duty_hours_total += dty
                cap_block_by_month[month_key] += blk
                cop_block_by_month[month_key] += blk
                crp_block_by_month[month_key] += blk
                cap_duty_by_month[month_key] += dty
                cop_duty_by_month[month_key] += dty
                crp_duty_by_month[month_key] += dty
                rec["cap_block"] += blk
                rec["cop_block"] += blk
                rec["crp_block"] += blk
            elif req == '4P':
                cap_block_hours_total += blk * 2
                cap_duty_hours_total += dty * 2
                cop_block_hours_total += blk * 2
                cop_duty_hours_total += dty * 2
                cap_block_by_month[month_key] += blk * 2
                cop_block_by_month[month_key] += blk * 2
                cap_duty_by_month[month_key] += dty * 2
                cop_duty_by_month[month_key] += dty * 2
                rec["cap_block"] += blk * 2
                rec["cop_block"] += blk * 2

            # TA hours (duty) por día
            req_size = crew_sizes.get(req, 0)
            ta_count = max(0, trip_crew_size - req_size)
            if ta_count > 0:
                rec["ta_hours"] += dty * ta_count

            # Acumular nuevas métricas AV KPIS
            total_duties += 1
            duties_by_month[month_key] += 1
            duty_hours_by_month[month_key] += dty

            # Distribución por duty time
            if dty <= 4.0:
                duties_by_duty_time["00:00-04:00"] += 1
            elif dty <= 8.0:
                duties_by_duty_time["04:01-08:00"] += 1
            elif dty <= 10.0:
                duties_by_duty_time["08:01-10:00"] += 1
            else:
                duties_by_duty_time[">10:00"] += 1

            # Distribución por block time
            if blk <= 2.0:
                duties_by_block_time["00:00-02:00"] += 1
            elif blk <= 4.0:
                duties_by_block_time["02:01-04:00"] += 1
            elif blk <= 6.0:
                duties_by_block_time["04:01-06:00"] += 1
            elif blk <= 8.0:
                duties_by_block_time["06:01-08:00"] += 1

            # Contar DH flights en el duty
            dh_count = len([f for f in dd.get('flights', []) if f.get('type') == 'DH'])
            total_dh_flights += dh_count
            dh_flights_by_month[month_key] += dh_count
            # Acumular horas DH por mes
            for f in dd.get('flights', []):
                if f.get('type') == 'DH':
                    dh_blk = float(f.get('blk_hours', 0) or 0)
                    dh_hours_by_month[month_key] += dh_blk

        # Acumular cambios de avión y noches de hotel por trip
        if t.get('tail_change', False):
            total_ac_changes += 1
            ac_changes_by_month[trip_month_key] += 1

        # Contar noches de hotel (layovers que no son en bases)
        for dd in t.get('chain', []):
            layover_station = dd.get('dst')
            if layover_station and layover_station not in DEFAULT_BASES:
                total_hotel_nights += 1
                # Track hotel nights by month
                layover_date = dd.get('end_base')
                if layover_date:
                    hotel_month_key = (layover_date.date().year, layover_date.date().month) if hasattr(layover_date, 'date') else (layover_date.year, layover_date.month)
                    hotel_nights_by_month[hotel_month_key] += 1

        # Multiplicadores por configuración
        if t['trip_crew'] == '2P':
            cap_n, cop_n, crp_n = 1, 1, 0
        elif t['trip_crew'] == '3P':
            cap_n, cop_n, crp_n = 1, 1, 1
        else:  # 4P
            cap_n, cop_n, crp_n = 2, 2, 0

        cur_date = trip_start_date
        while cur_date <= trip_end_date:
            pilots_by_day_cap[cur_date] += cap_n
            pilots_by_day_cop[cur_date] += cop_n
            pilots_by_day_crp[cur_date] += crp_n
            pilots_by_day[cur_date] += (cap_n + cop_n + crp_n)

            pilots_by_day_base_cap[(cur_date, t['base'])] += cap_n
            pilots_by_day_base_cop[(cur_date, t['base'])] += cop_n
            pilots_by_day_base_crp[(cur_date, t['base'])] += crp_n

            cur_date += timedelta(days=1)


        # Métricas del Trip Individual
        eff_metric = t['block'] / trip_days
        dh_info = f"{fmt_dec(t['total_dh_hours'], 1)}h DH" if t['total_dh_hours'] > 0 else "0h DH"
        rescue_flag = " [RESCATE]" if t['has_rescue_dh'] else ""
        second_pass_flag = " (2nd Pass)" if t.get('is_second_pass') else ""

        # --- Export (formato como consola) ---
        trip_report_rows.append({
            "FECHA": "",
            "RUTA/ACTIVIDAD": f"TRIP #{t['id']}{second_pass_flag} | BASE: {t['base']} | Config: {t['trip_crew']} | {trip_days} Días | Blk: {fmt_dec(t['block'], 1)}h | {dh_info}{rescue_flag}",
            "AVIÓN": "", "REQ": "", "ASG": "", "BLK": "", "MAX": "", "DUTY": "", "MAX.1": "", "REST (Act/Req)": "", "PERNOCTA": ""
        })
        trip_report_rows.append({
            "FECHA": "",
            "RUTA/ACTIVIDAD": f"Eficiencia Rotación: {fmt_dec(eff_metric, 2)} Block Hrs/Día",
            "AVIÓN": "", "REQ": "", "ASG": "", "BLK": "", "MAX": "", "DUTY": "", "MAX.1": "", "REST (Act/Req)": "", "PERNOCTA": ""
        })

        if VERBOSE_OUTPUT:
            print(f"\nTRIP #{t['id']} | BASE: {t['base']} | Config: {t['trip_crew']} | {trip_days} Días | Blk: {t['block']:.1f}h | {dh_info}{rescue_flag}")
            print(f"Eficiencia Rotación: {eff_metric:.2f} Block Hrs/Día")
            print("-" * 150)
            print(f"{'FECHA':<12} {'RUTA/ACTIVIDAD':<30} {'AVIÓN':<8} {'REQ':<4} {'ASG':<4} {'BLK':<5} {'MAX':<6} {'DUTY':<5} {'MAX':<6} {'REST (Act/Req)':<18} {'PERNOCTA'}")
            print("-" * 150)

        prev_end = None
        prev_blk = 0
        prev_dst = None

        pre_dh = t.get('pre_dh')
        if pre_dh and pre_dh.get('legs'):
            for leg in pre_dh['legs']:
                dep_base = leg['dep_utc'] - timedelta(hours=UTC_OFFSET_HOURS)
                arr_base = leg['arr_utc'] - timedelta(hours=UTC_OFFSET_HOURS)
                dh_report_base = dep_base - rules.REPORT_TIME
                dh_end_base = arr_base + rules.DEBRIEF
                dte = dh_report_base.date()
                blk_h = (leg['arr_utc'] - leg['dep_utc']).total_seconds() / 3600
                rec = daily_hours.setdefault(dte, {"date": dte, "cap_block": 0.0, "cop_block": 0.0, "crp_block": 0.0, "dh_hours": 0.0, "ta_hours": 0.0})
                rec["dh_hours"] += blk_h
                # Track DH hours by month
                dh_month_key = (dte.year, dte.month)
                dh_hours_by_month[dh_month_key] += blk_h
                dh_flights_by_month[dh_month_key] += 1
                _emit_avl_until(dh_report_base.date())
                flt = _fmt_flt_num(leg.get('flt_num', ''))
                aln = leg.get('aln', 'AV')
                route = f"DH {leg['org']}-{leg['dst']} [{aln}{flt}]" if flt else f"DH {leg['org']}-{leg['dst']}"
                time_info = f"{dh_report_base.strftime('%d-%b %H:%M')}-{dh_end_base.strftime('%H:%M')}"
                crew_info = f"({trip_crew_size} pilotos)" if trip_crew_size else ""
                dh_duty_h = (dh_end_base - dh_report_base).total_seconds() / 3600
                flt = _fmt_flt_num(leg.get('flt_num', ''))
                meta = {
                    "flt_nums": f"{aln}{flt}" if flt else "DH",
                    "tails": "---",
                    "crew": "DH",
                }
                _add_event(dh_report_base, "DH", f"DH {leg['org']}-{leg['dst']}", dh_report_base, dh_end_base, blk_h, dh_duty_h, meta=meta)
                # Agregar al reporte de DH
                _add_dh_to_report(t['id'], t.get('trip_crew', '2P'), 'PRE_DH', leg, dep_base, arr_base)
                if VERBOSE_OUTPUT:
                    print(f"{_fmt_base_dt(dh_report_base):<12} {f'{route} {time_info} {crew_info}'.strip():<30} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} {pre_dh.get('to','')}")
                trip_report_rows.append({
                    "FECHA": _fmt_base_dt(dh_report_base),
                    "RUTA/ACTIVIDAD": f"{route} {time_info} {crew_info}".strip(),
                    "AVIÓN": "---", "REQ": "---", "ASG": "---", "BLK": "---", "MAX": "---", "DUTY": "---", "MAX.1": "---", "REST (Act/Req)": "---", "PERNOCTA": pre_dh.get('to','')
                })
                if dh_end_base.date() > dh_report_base.date():
                    _add_continuation(f"DH {leg['org']}-{leg['dst']}", dh_report_base, dh_end_base)
                    last_date = dh_end_base.date()
                else:
                    last_date = dh_report_base.date()

        for i, duty in enumerate(t['chain']):
            # Línea de DH Transfer (solo para impresión)
            if prev_dst and prev_dst != duty['org']:
                if VERBOSE_OUTPUT:
                    print(f"{'':<12} >> DH TRANSFER: {prev_dst} -> {duty['org']:<12} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} {duty['org']}")
                trip_report_rows.append({
                    "FECHA": "",
                    "RUTA/ACTIVIDAD": f">> DH TRANSFER: {prev_dst} -> {duty['org']}",
                    "AVIÓN": "---", "REQ": "---", "ASG": "---", "BLK": "---", "MAX": "---", "DUTY": "---", "MAX.1": "---", "REST (Act/Req)": "---", "PERNOCTA": duty['org']
                })

            _emit_avl_until(duty['start_base'].date())
            d_date = _fmt_base_dt(duty['start_base'])
            duty_starts_by_station[(duty['start_base'].date(), duty['org'])] += trip_crew_size
            stops = [duty['flights'][0]['org']] + [f['dst'] for f in duty['flights']]
            route_base = "-".join(stops)
            route = route_base
            flt_nums = "/".join([str(f.get('flt_num', '')).strip() for f in duty['flights'] if f.get('flt_num') is not None])
            if flt_nums:
                route = f"{route} [{flt_nums}]"
            dh_legs = [f for f in duty.get("flights", []) if f.get("is_dh")]
            if dh_legs:
                dh_list = []
                for f in dh_legs:
                    flt = str(f.get("flt_num", "")).strip()
                    f_aln = f.get("aln", "AV")
                    if flt and not any(flt.startswith(p) for p in ('AV', 'AM', 'LA', 'CM', 'AR')):
                        flt = f"{f_aln}{flt}"
                    eta = ""
                    arr_base_dh = f.get("arr_base")
                    dep_base_dh = f.get("dep_base")
                    if hasattr(arr_base_dh, "strftime"):
                        eta = arr_base_dh.strftime("%H:%M")
                    part = f"{f.get('org')}-{f.get('dst')}"
                    if flt:
                        part += f" {flt}"
                    if eta:
                        part += f" {eta}"
                    dh_list.append(part)
                    # Agregar al reporte de DH
                    _add_dh_to_report(t['id'], t.get('trip_crew', '2P'), 'IN_DUTY', f, dep_base_dh, arr_base_dh)
                route = f"{route} +DH({', '.join(dh_list)})"
            req_crew = duty['min_crew']

            # Info Descanso
            rest_act = ""
            rest_req = ""
            rest_flag = ""
            rest_info = "INI"
            if prev_end:
                rest_act = (duty['start_base'] - prev_end).total_seconds() / 3600
                rest_req = rules.calculate_required_rest(prev_blk, is_base=False)
                rest_info = f"{fmt_dec(rest_act, 1)}h / {int(rest_req)}h"
                if (rest_act - rest_req) < 1.0:
                    rest_info += " (!)"
                    rest_flag = "LOW_MARGIN"

            # Contar Pernocta (Hotel)
            # Regla: Check-in hoteles es a las 15:00 (3 PM)
            # - Siempre: 1 noche base por pernoctar fuera de base
            # - Si llega antes de las 15:00: +1 noche adicional (early check-in)
            pernocta = duty['dst']
            if pernocta != t['base']:
                arrival_hour = duty['end_base'].hour
                night_date = duty['end_base'].date()
                night_key = night_date.strftime("%Y-%m")

                # Noche base: siempre se cuenta 1 noche por pernoctar
                lays[pernocta] += trip_crew_size
                hotels_by_month[(pernocta, night_key)] += trip_crew_size

                # Noche adicional si llega antes de las 15:00 (3 PM)
                # Necesita early check-in = habitación desde la noche anterior
                if arrival_hour < 15:
                    early_night = night_date - timedelta(days=1)
                    early_key = early_night.strftime("%Y-%m")
                    lays[pernocta] += trip_crew_size
                    hotels_by_month[(pernocta, early_key)] += trip_crew_size
            else:
                pernocta = "BASE (Fin)"

            # === Viáticos / primas ===
            stations = [duty['flights'][0]['org']] + [f['dst'] for f in duty['flights']]
            is_national = all(s in COLOMBIA_STATIONS for s in stations)
            outside_americas = any(s in OUTSIDE_AMERICAS_STATIONS for s in stations)

            # próximo reporte (para permanencia con descanso)
            next_start = None
            if i + 1 < len(t['chain']):
                next_start = t['chain'][i + 1]['start_base']
            elif t.get('has_rescue_dh'):
                rescue_legs = t.get('rescue_legs') or []
                if rescue_legs:
                    next_start = (rescue_legs[0]['dep_utc'] - timedelta(hours=UTC_OFFSET_HOURS)) - rules.REPORT_TIME
            if next_start is None:
                next_start = duty['end_base']

            rest_act = max(0.0, (next_start - duty['end_base']).total_seconds() / 3600)
            rest_req = rules.calculate_required_rest(duty['block'], is_base=False)
            has_rest = rest_act >= rest_req

            # Permanencia = tiempo entre finalizar duty e iniciar el siguiente (solo si termina fuera de base)
            if duty['dst'] != t['base']:
                perm_hours = (next_start - duty['end_base']).total_seconds() / 3600
                if perm_hours > 24:
                    perm_hours = 24.0
                if perm_hours < 0:
                    perm_hours = 0.0
            else:
                perm_hours = 0.0

            req = duty['min_crew']
            counts = {"2P": {"CAP":1,"COP":1,"CRP":0}, "3P": {"CAP":1,"COP":1,"CRP":1}, "4P": {"CAP":2,"COP":2,"CRP":0}}
            req_counts = counts.get(req, {"CAP":0,"COP":0,"CRP":0})
            req_total = req_counts["CAP"] + req_counts["COP"] + req_counts["CRP"]
            trip_counts = counts.get(t.get("trip_crew"), {"CAP":0,"COP":0,"CRP":0})
            trip_total = trip_counts["CAP"] + trip_counts["COP"] + trip_counts["CRP"]
            op_counts = _op_counts(t.get("trip_crew"), req)
            ta_count = max(0, trip_crew_size - (req_total))

            viat_cop = 0.0
            viat_usd = 0.0
            pn_cop = 0.0
            pn_usd = 0.0
            extra_usd = 0.0
            pc_cap_usd = 0.0  # Prima de comando para comandante(s)
            pc_cop_usd = 0.0  # Prima de comando para COP1 (solo 3P relief)

            if is_national:
                if 10 <= perm_hours <= 12:
                    viat_cop = 33412
                elif perm_hours > 12 and perm_hours <= 24 and has_rest:
                    viat_cop = 66773
                elif perm_hours < 12 and has_rest:
                    viat_cop = 33412

                if has_rest:
                    if perm_hours > 12:
                        pn_cop = 46266
                    else:
                        pn_cop = 25180
            else:
                if has_rest:
                    if perm_hours > 12:
                        # Viático CAP internacional con pernocta >12h
                        viat_usd = VIATICO_CAP_OUTSIDE_USD if outside_americas else VIATICO_CAP_AMERICAS_USD
                        pn_usd = PRIMA_NAV_CAP_OUTSIDE_USD if outside_americas else PRIMA_NAV_CAP_AMERICAS_USD
                    else:
                        # Viático COP internacional con pernocta <12h
                        viat_usd = VIATICO_COP_OUTSIDE_USD if outside_americas else VIATICO_COP_AMERICAS_USD
                        pn_usd = PRIMA_NAV_COP_USD

                same_day = duty['start_base'].date() == duty['end_base'].date()
                if same_day and not has_rest:
                    pn_usd += PRIMA_NAV_POR_VUELO_USD * len(duty['flights'])  # por vuelo
                    if trip_crew_size >= 3:
                        extra_usd = EXTRA_POSICIONAMIENTO_USD

                # Prima de comando internacional (usar parámetros configurables)
                trip_crew = t.get("trip_crew")
                if trip_crew == "2P":
                    # Tripulación sencilla internacional
                    pc_cap_usd = PRIMA_COMANDO_2P_CAP_USD
                elif trip_crew == "3P":
                    # Relief Pilot (1 CAP + 2 COP en práctica, o 1 CAP + 1 COP + 1 CRP)
                    pc_cap_usd = PRIMA_COMANDO_3P_CAP_USD
                    pc_cop_usd = PRIMA_COMANDO_3P_COP_USD
                elif trip_crew == "4P":
                    # Tripulación múltiple (2 CAP + 2 COP)
                    pc_cap_usd = PRIMA_COMANDO_4P_CAP1_USD + PRIMA_COMANDO_4P_CAP2_USD

            total_cop = viat_cop * trip_total + pn_cop * (op_counts["CAP"] + op_counts["COP"])
            total_usd = (
                viat_usd * trip_total
                + pn_usd * (op_counts["CAP"] + op_counts["COP"])
                + extra_usd * trip_total
                + pc_cap_usd  # Prima de comando para CAP(s)
                + pc_cop_usd  # Prima para COP1 en configuración 3P
            )

            viaticos_rows.append({
                "date": duty['start_base'].strftime("%Y-%m-%d"),
                "weekday": weekday_es[duty['start_base'].weekday()],
                "trip_id": t["id"],
                "duty_id": duty.get("id"),
                "route": "-".join(stations),
                "national": is_national,
                "outside_americas": outside_americas,
                "perm_hours": round(perm_hours, 2),
                "has_rest": has_rest,
                "REQ": req,
                "CAP": trip_counts["CAP"],
                "COP": trip_counts["COP"],
                "CRP": trip_counts["CRP"],
                "TA": ta_count,
                "viatico_cop_per_pilot": viat_cop,
                "viatico_usd_per_pilot": viat_usd,
                "prima_nav_cop_per_pilot": pn_cop,
                "prima_nav_usd_per_pilot": pn_usd,
                "extra_usd_per_pilot": extra_usd,
                "prima_comando_cap_usd": pc_cap_usd,
                "prima_comando_cop_usd": pc_cop_usd,
                "total_cop": round(total_cop, 2),
                "total_usd": round(total_usd, 2),
            })
            # Track viáticos by month
            viat_date = duty['start_base'].date() if hasattr(duty['start_base'], 'date') else duty['start_base']
            viat_month_key = (viat_date.year, viat_date.month)
            viaticos_by_month[viat_month_key] += total_usd

            lim = duty['limits']
            asg_crew = t['trip_crew']
            req_size = crew_sizes.get(req_crew, 0)
            asg_size = crew_sizes.get(asg_crew, 0)
            ta_count_asg = max(0, asg_size - req_size)
            if ta_count_asg > 0:
                asg_display = f"{req_crew}+{ta_count_asg}TA"
            else:
                asg_display = asg_crew

            tails = duty.get("tails") or [duty.get("tail")]
            tail_disp = "/".join([str(t) for t in tails if t is not None])
            meta = {
                "flt_nums": flt_nums,
                "tails": tail_disp,
                "crew": asg_display,
            }

            viat_payload = {
                "viat_cop": viat_cop,
                "viat_usd": viat_usd,
                "pn_cop": pn_cop,
                "pn_usd": pn_usd,
                "extra_usd": extra_usd,
                "pc_cap_usd": pc_cap_usd,
                "pc_cop_usd": pc_cop_usd,
                "is_national": is_national,
            }
            _add_event(
                duty['start_base'],
                "DUTY",
                route_base,
                duty['start_base'],
                duty['end_base'],
                float(duty['block']),
                float(duty['duty_dur']),
                req_crew,
                viat=viat_payload,
                meta=meta,
            )

            if duty.get("tail_change"):
                tail_disp = "/".join([str(t) for t in tails if t is not None])
            else:
                tail_disp = duty['tail']
            if VERBOSE_OUTPUT:
                print(f"{d_date:<12} {route:<30} {tail_disp:<8} {req_crew:<4} {asg_display:<4} {duty['block']:<5.1f} {lim['max_blk']:<6.1f} {duty['duty_dur']:<5.1f} {lim['max_duty']:<6.1f} {rest_info:<18} {pernocta}")
            trip_report_rows.append({
                "FECHA": d_date,
                "RUTA/ACTIVIDAD": route,
                "AVIÓN": tail_disp,
                "REQ": req_crew,
                "ASG": asg_display,
                "BLK": round(float(duty['block']), 1),
                "MAX": round(float(lim['max_blk']), 1) if lim.get('max_blk') is not None else "",
                "DUTY": round(float(duty['duty_dur']), 1),
                "MAX.1": round(float(lim['max_duty']), 1) if lim.get('max_duty') is not None else "",
                "REST (Act/Req)": rest_info,
                "PERNOCTA": pernocta,
            })

            # Row para exportación
            trip_legs_rows.append({
                "trip_id": t["id"],
                "duty_id": duty.get("id"),
                "duty_start_base": duty.get("start_base"),
                "duty_end_base": duty.get("end_base"),
                "route": route,
                "tail": duty.get("tail"),
                "org": duty.get("org"),
                "dst": duty.get("dst"),
                "req_crew": req_crew,
                "asg_crew": asg_display,
                "block_hours": duty.get("block"),
                "duty_hours": duty.get("duty_dur"),
                "max_blk": lim.get("max_blk"),
                "max_duty": lim.get("max_duty"),
                "rest_act_h": round(rest_act, 2) if rest_act != "" else "",
                "rest_req_h": round(rest_req, 2) if rest_req != "" else "",
                "rest_flag": rest_flag,
                "overnight": pernocta,
            })

            prev_end = duty['end_base']
            prev_blk = duty['block']
            prev_dst = duty['dst']
            if duty['end_base'].date() > duty['start_base'].date():
                _add_continuation(route_base, duty['start_base'], duty['end_base'])
                last_date = duty['end_base'].date()
            else:
                last_date = duty['start_base'].date()

        if t['has_rescue_dh']:
            if VERBOSE_OUTPUT:
                print(f"{'':<12} >> DH RESCATE: {t['rescue_from']} -> {t['base']:<12} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} BASE (Fin)")
            trip_report_rows.append({
                "FECHA": "",
                "RUTA/ACTIVIDAD": f">> DH RESCATE: {t['rescue_from']} -> {t['base']}",
                "AVIÓN": "---", "REQ": "---", "ASG": "---", "BLK": "---", "MAX": "---", "DUTY": "---", "MAX.1": "---", "REST (Act/Req)": "---", "PERNOCTA": "BASE (Fin)"
            })
            rescue_legs = t.get('rescue_legs') or []
            for leg in rescue_legs:
                dep_base = leg['dep_utc'] - timedelta(hours=UTC_OFFSET_HOURS)
                arr_base = leg['arr_utc'] - timedelta(hours=UTC_OFFSET_HOURS)
                dh_report_base = dep_base - rules.REPORT_TIME
                dh_end_base = arr_base + rules.DEBRIEF
                dte = dh_report_base.date()
                blk_h = (leg['arr_utc'] - leg['dep_utc']).total_seconds() / 3600
                rec = daily_hours.setdefault(dte, {"date": dte, "cap_block": 0.0, "cop_block": 0.0, "crp_block": 0.0, "dh_hours": 0.0, "ta_hours": 0.0})
                rec["dh_hours"] += blk_h
                # Track rescue DH hours by month
                rescue_dh_month_key = (dte.year, dte.month)
                dh_hours_by_month[rescue_dh_month_key] += blk_h
                dh_flights_by_month[rescue_dh_month_key] += 1
                _emit_avl_until(dh_report_base.date())
                flt = _fmt_flt_num(leg.get('flt_num', ''))
                aln = leg.get('aln', 'AV')
                route = f"DH {leg['org']}-{leg['dst']} [{aln}{flt}]" if flt else f"DH {leg['org']}-{leg['dst']}"
                time_info = f"{dh_report_base.strftime('%d-%b %H:%M')}-{dh_end_base.strftime('%H:%M')}"
                crew_info = f"({trip_crew_size} pilotos)" if trip_crew_size else ""
                dh_duty_h = (dh_end_base - dh_report_base).total_seconds() / 3600
                flt = _fmt_flt_num(leg.get('flt_num', ''))
                meta = {
                    "flt_nums": f"{aln}{flt}" if flt else "DH",
                    "tails": "---",
                    "crew": "DH",
                }
                _add_event(dh_report_base, "DH", f"DH {leg['org']}-{leg['dst']}", dh_report_base, dh_end_base, blk_h, dh_duty_h, meta=meta)
                # Agregar al reporte de DH
                _add_dh_to_report(t['id'], t.get('trip_crew', '2P'), 'RESCUE_DH', leg, dep_base, arr_base)
                if VERBOSE_OUTPUT:
                    print(f"{_fmt_base_dt(dh_report_base):<12} {f'{route} {time_info} {crew_info}'.strip():<30} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} BASE (Fin)")
                trip_report_rows.append({
                    "FECHA": _fmt_base_dt(dh_report_base),
                    "RUTA/ACTIVIDAD": f"{route} {time_info} {crew_info}".strip(),
                    "AVIÓN": "---", "REQ": "---", "ASG": "---", "BLK": "---", "MAX": "---", "DUTY": "---", "MAX.1": "---", "REST (Act/Req)": "---", "PERNOCTA": "BASE (Fin)"
                })
                if dh_end_base.date() > dh_report_base.date():
                    _add_continuation(f"DH {leg['org']}-{leg['dst']}", dh_report_base, dh_end_base)
                    last_date = dh_end_base.date()
                else:
                    last_date = dh_report_base.date()

        # Línea por piloto (por día)
        if trip_crew_size:

            if t.get('trip_crew') == '4P':
                pilots = [("CAP1", "COMANDANTE"), ("CAP2", "CAP"), ("COP1", "COP"), ("COP2", "COP")]
            elif t.get('trip_crew') == '3P':
                pilots = [("CAP", "COMANDANTE"), ("COP", "COP"), ("CRP", "CRP")]
            else:
                pilots = [("CAP", "COMANDANTE"), ("COP", "COP")]

            def _ops_for_req(req):
                if t.get('trip_crew') == '4P':
                    if req == '4P':
                        return {"CAP1","CAP2","COP1","COP2"}
                    if req == '3P':
                        return {"CAP1","COP1","COP2"}
                    return {"CAP1","COP1"}
                if t.get('trip_crew') == '3P':
                    if req == '3P':
                        return {"CAP","COP","CRP"}
                    return {"CAP","COP"}
                return {"CAP","COP"}

            dates = []
            cur = trip_start_date
            while cur <= trip_end_date:
                dates.append(cur)
                pilot_lines_dates_set.add(cur)
                cur += timedelta(days=1)

            # Encabezado horizontal (días)
            header_days = [_fmt_day_col(d) for d in dates]

            for pid, role in pilots:
                row = {
                    "trip_number": t["id"],
                    "trip_days": trip_days,
                    "position": pid,
                }
                activities = []
                sum_cop = 0.0
                sum_usd = 0.0
                # Para calcular Rest time desde el último duty/DH
                last_event_end = None

                for d in dates:
                    evs = sorted(day_events.get(d, []), key=lambda x: x["time"])
                    parts = []
                    day_cop = 0.0
                    day_usd = 0.0
                    # Detectar si hay DH + DUTY en el mismo día para calcular duty combinado
                    real_evs = [ev for ev in evs if ev["kind"] in ("DH", "DUTY")]
                    has_dh_and_duty = len(real_evs) > 1 and any(ev["kind"] == "DH" for ev in real_evs) and any(ev["kind"] == "DUTY" for ev in real_evs)
                    report_time_h = REPORT_MINUTES / 60.0
                    debrief_h = DEBRIEF_MINUTES / 60.0

                    # Precalcular duty combinado si hay DH + DUTY mismo día
                    # Duty combinado = (hora llegada último evento - hora salida primer evento) + 90 min presentación
                    combined_duty_h = None
                    dh_before_duty = False  # True si DH está antes del DUTY
                    if has_dh_and_duty:
                        dh_evs = [ev for ev in real_evs if ev["kind"] == "DH"]
                        duty_evs = [ev for ev in real_evs if ev["kind"] == "DUTY"]
                        if dh_evs and duty_evs:
                            # Determinar orden: DH antes de DUTY o DUTY antes de DH
                            first_dh = min(dh_evs, key=lambda x: x["start"])
                            first_duty = min(duty_evs, key=lambda x: x["start"])
                            last_dh = max(dh_evs, key=lambda x: x["end"])
                            last_duty = max(duty_evs, key=lambda x: x["end"])

                            if first_dh["start"] < first_duty["start"]:
                                # Caso 1: DH antes de DUTY (pre_dh + vuelo)
                                dh_before_duty = True
                                # dep = salida del DH (start_dh + report porque start = dep - report)
                                dep_first = first_dh["start"] + timedelta(minutes=REPORT_MINUTES)
                                # arr = llegada del vuelo (end_duty - debrief porque end = arr + debrief)
                                arr_last = last_duty["end"] - timedelta(minutes=DEBRIEF_MINUTES)
                            else:
                                # Caso 2: DUTY antes de DH (vuelo + rescue_dh)
                                dh_before_duty = False
                                # dep = salida del vuelo (start_duty + report)
                                dep_first = first_duty["start"] + timedelta(minutes=REPORT_MINUTES)
                                # arr = llegada del DH (end_dh - debrief)
                                arr_last = last_dh["end"] - timedelta(minutes=DEBRIEF_MINUTES)

                            # Duty combinado = (arr_último - dep_primero) + 90 min presentación
                            combined_duty_h = (arr_last - dep_first).total_seconds() / 3600 + report_time_h

                    for e in evs:
                        if e["kind"] == "CONT":
                            parts.append(e["route"])
                            continue
                        start = e["start"].strftime("%H:%M")
                        end = e["end"].strftime("%H:%M")
                        flt_nums = e.get("flt_nums") or ""
                        route_line = e.get("route") or ""
                        tails = e.get("tails") or "---"
                        crew_line = e.get("crew") or (e.get("req") or "")
                        # Calcular duty ajustado para DH+DUTY mismo día
                        duty_display = e['duty']
                        if has_dh_and_duty and combined_duty_h is not None:
                            if dh_before_duty:
                                # Caso 1: DH antes de DUTY - el DUTY muestra el combinado, DH muestra 0
                                if e["kind"] == "DH":
                                    duty_display = 0.0
                                elif e["kind"] == "DUTY":
                                    duty_display = combined_duty_h
                            else:
                                # Caso 2: DUTY antes de DH - el DH muestra el combinado, DUTY muestra 0
                                if e["kind"] == "DUTY":
                                    duty_display = 0.0
                                elif e["kind"] == "DH":
                                    duty_display = combined_duty_h
                        if e["kind"] == "DUTY":
                            ops = _ops_for_req(e.get("req"))
                            req_code = e.get("req") or ""
                            if pid in ops:
                                # Piloto operativo - mostrar req + posición
                                crew_line = f"{req_code} {pid}"
                            else:
                                # Piloto como TA (tripulante adicional)
                                crew_line = f"{req_code} TA"
                            # Viáticos base para todos los tripulantes
                            v_cop = float(e.get("viat_cop", 0.0) or 0.0)
                            v_usd = float(e.get("viat_usd", 0.0) or 0.0)
                            v_extra = float(e.get("extra_usd", 0.0) or 0.0)
                            day_cop += v_cop
                            day_usd += v_usd + v_extra
                            # Navegación solo para CAP/COP operativos
                            if pid in ops and (pid.startswith("CAP") or pid == "CAP" or pid.startswith("COP") or pid == "COP"):
                                day_cop += float(e.get("pn_cop", 0.0) or 0.0)
                                day_usd += float(e.get("pn_usd", 0.0) or 0.0)
                            # Prima de comando: CAP para comandantes, COP1 para copiloto más antiguo (3P)
                            if pid in ops and role == "COMANDANTE":
                                day_usd += float(e.get("pc_cap_usd", 0.0) or 0.0)
                            elif pid in ops and pid == "COP1":
                                day_usd += float(e.get("pc_cop_usd", 0.0) or 0.0)
                        elif e["kind"] == "DH":
                            crew_line = "DH"

                        # Calcular REST time desde el último duty/DH
                        rest_str = "---"
                        if last_event_end is not None and e["kind"] in ("DUTY", "DH"):
                            rest_hours = (e["start"] - last_event_end).total_seconds() / 3600
                            if rest_hours >= 0:
                                rest_str = f"{fmt_dec(rest_hours, 1)}h"

                        lines = [
                            f"FLT: {flt_nums}".strip(),
                            f"RTE: {route_line}".strip(),
                            f"ACFT: {tails}",
                            f"TIME: {start}-{end}",
                            f"BLK: {fmt_dec(e['block'], 1)}",
                            f"DUTY: {fmt_dec(duty_display, 1)}",
                            f"REST: {rest_str}",
                            f"CREW: {crew_line}",
                            f"{d.strftime('%d-%b').lower()}",
                        ]
                        parts.append("\n".join(lines))

                        # Actualizar last_event_end para el próximo cálculo de REST
                        if e["kind"] in ("DUTY", "DH"):
                            last_event_end = e["end"]
                    if parts:
                        if day_cop or day_usd:
                            parts.append(f"VIAT: COP {int(day_cop)} / USD {int(day_usd)}")
                        text = "\n\n".join(parts)
                    else:
                        text = "DISPONIBLE"
                    row[_fmt_day_col(d)] = text
                    activities.append(text)
                    sum_cop += day_cop
                    sum_usd += day_usd
                # Total viáticos al final del mes por piloto
                row["VIATICOS_TOTAL"] = f"COP {int(sum_cop)} / USD {int(sum_usd)}"
                # Solo exportar a Excel (no imprimir en consola)
                pilot_lines_rows.append(row)

        # Acumulado excluyente por día (CAP/COP/CRP/DH/TA)
        cur = trip_start_date
        while cur <= trip_end_date:
            evs = day_events.get(cur, [])
            duty_evs = [e for e in evs if e.get("kind") == "DUTY"]
            dh_evs = [e for e in evs if e.get("kind") == "DH"]
            rec = daily_pilots_excl.setdefault(cur, {"date": cur, "CAP": 0, "COP": 0, "CRP": 0, "DH": 0, "TA": 0})
            if duty_evs:
                req = max(duty_evs, key=lambda e: rules.CREW_RANK.get(e.get("req"), 0)).get("req")
                req_size = crew_sizes.get(req, 0)
                if req == "2P":
                    rec["CAP"] += 1
                    rec["COP"] += 1
                elif req == "3P":
                    rec["CAP"] += 1
                    rec["COP"] += 1
                    rec["CRP"] += 1
                elif req == "4P":
                    rec["CAP"] += 2
                    rec["COP"] += 2
                ta_count = max(0, trip_crew_size - req_size)
                rec["TA"] += ta_count
            elif dh_evs:
                rec["DH"] += trip_crew_size
            cur += timedelta(days=1)

        # Completar días sin asignación hasta fin del trip
        cur = last_date + timedelta(days=1)
        while cur <= trip_end_date:
            dt = datetime.combine(cur, datetime.min.time())
            d_date = _fmt_base_dt(dt)
            route = "AVL - BLANCO"
            if VERBOSE_OUTPUT:
                print(f"{d_date:<12} {route:<30} {'---':<8} {'---':<4} {'---':<4} {'---':<5} {'---':<6} {'---':<5} {'---':<6} {'---':<18} {'':<}")
            trip_report_rows.append({
                "FECHA": d_date,
                "RUTA/ACTIVIDAD": route,
                "AVIÓN": "---",
                "REQ": "---",
                "ASG": "---",
                "BLK": "---",
                "MAX": "---",
                "DUTY": "---",
                "MAX.1": "---",
                "REST (Act/Req)": "---",
                "PERNOCTA": "",
            })
            cur += timedelta(days=1)
        # Separador visual entre trips
        trip_report_rows.append({"FECHA":"","RUTA/ACTIVIDAD":"","AVIÓN":"","REQ":"","ASG":"","BLK":"","MAX":"","DUTY":"","MAX.1":"","REST (Act/Req)":"","PERNOCTA":""})

    # --- RESUMEN FINAL ---
    all_f = set(flight_ids)
    missing = all_f - covered
    cov_pct = (len(covered)/len(flight_ids))*100 if len(flight_ids) else 0
    avg_network_eff = (total_blk / total_days) if total_days > 0 else 0
    total_trips = sum(base_c.values())
    eff_cap = (cap_block_hours_total / man_days["cap"]) if man_days["cap"] > 0 else 0.0
    eff_cop = (cop_block_hours_total / man_days["cop"]) if man_days["cop"] > 0 else 0.0
    eff_crp = (crp_block_hours_total / man_days["crp"]) if man_days["crp"] > 0 else 0.0

    # Calcular KPIs de segunda pasada
    trips_second_pass = sum(1 for t in selected_trips if t.get('is_second_pass') and not t.get('is_partial_trip'))
    flights_covered_second_pass = sum(
        len(t.get('flights_covered', [])) for t in selected_trips if t.get('is_second_pass') and not t.get('is_partial_trip')
    )
    coverage_improvement = (flights_covered_second_pass / max(1, len(flight_ids))) * 100

    # Calcular KPIs de tercera pasada (trips parciales)
    trips_partial = sum(1 for t in selected_trips if t.get('is_partial_trip'))
    flights_covered_partial = sum(
        len(t.get('flights_covered', [])) for t in selected_trips if t.get('is_partial_trip')
    )
    coverage_partial = (flights_covered_partial / max(1, len(flight_ids))) * 100

    # Estaciones finales donde quedan crews (trips parciales)
    partial_final_stations = {}
    for t in selected_trips:
        if t.get('is_partial_trip'):
            station = t.get('final_station', 'UNKNOWN')
            partial_final_stations[station] = partial_final_stations.get(station, 0) + 1

    # Calcular métricas derivadas AV KPIS (parciales - falta hotel y viáticos)
    total_duty_hours = cap_duty_hours_total + cop_duty_hours_total + crp_duty_hours_total
    avg_block_per_day = (total_blk / total_days) if total_days > 0 else 0.0
    avg_block_per_duty = (total_blk / total_duties) if total_duties > 0 else 0.0
    avg_duty_per_duty = (total_duty_hours / total_duties) if total_duties > 0 else 0.0

    if VERBOSE_OUTPUT:
        print("\n" + "="*150)
        print(f"{APP_NAME} | DASHBOARD GERENCIAL")
        print("="*150)

        print(f"\n1. INDICADORES CLAVE (KPIs)")
        print(f"   > Cobertura de Vuelos:   {cov_pct:.1f}%")
        print(f"   > Eficiencia de Red:     {avg_network_eff:.2f} Block Hours / Día Promedio")
        print(f"   > Total Rotaciones:      {total_trips}")

        # KPIs de segunda pasada
        if trips_second_pass > 0:
            print(f"\n   SEGUNDA PASADA:")
            print(f"   > Trips de 2nd pass:     {trips_second_pass}")
            print(f"   > Vuelos cubiertos:      {flights_covered_second_pass}")
            print(f"   > Mejora de cobertura:   +{coverage_improvement:.1f}%")

        # KPIs de tercera pasada (trips parciales)
        if trips_partial > 0:
            print(f"\n   TERCERA PASADA (TRIPS PARCIALES):")
            print(f"   > Trips parciales:       {trips_partial}")
            print(f"   > Vuelos cubiertos:      {flights_covered_partial}")
            print(f"   > Cobertura adicional:   +{coverage_partial:.1f}%")
            print(f"   > Estaciones finales:    {partial_final_stations}")

        print(f"\n2. DISTRIBUCIÓN DE TRIPULACIONES")
        print(f"   > Por Base: {dict(base_c)}")
        print(f"   > Por Configuración: {dict(crew_c)}")

        print(f"\n3. RECURSOS HUMANOS REQUERIDOS (Días a Pagar)")
        print(f"   > Capitanes (CAP): {man_days['cap']} días")
        print(f"   > Copilotos (COP): {man_days['cop']} días")
        print(f"   > Relevos   (CRP): {man_days['crp']} días")

        print(f"   > Horas Block Totales CAP: {cap_block_hours_total:.1f} h")
        print(f"   > Horas Block Totales COP: {cop_block_hours_total:.1f} h")
        print(f"   > Horas Block Totales CRP: {crp_block_hours_total:.1f} h")
        print(f"   > Horas Duty  Totales CAP: {cap_duty_hours_total:.1f} h")
        print(f"   > Horas Duty  Totales COP: {cop_duty_hours_total:.1f} h")
        print(f"   > Horas Duty  Totales CRP: {crp_duty_hours_total:.1f} h")
        print(f"   > Eficiencia Red CAP: {eff_cap:.2f} Block Hrs/DÃ­a")
        print(f"   > Eficiencia Red COP: {eff_cop:.2f} Block Hrs/DÃ­a")
        print(f"   > Eficiencia Red CRP: {eff_crp:.2f} Block Hrs/DÃ­a")

        # Top 10 días por requerimiento total (útil para dimensionamiento)
        if pilots_by_day:
            top_days = sorted(pilots_by_day.items(), key=lambda x: x[1], reverse=True)[:10]
            print(f"\n   > Top días (pilotos requeridos):")
            for dte, tot in top_days:
                cap_n = pilots_by_day_cap.get(dte, 0)
                cop_n = pilots_by_day_cop.get(dte, 0)
                crp_n = pilots_by_day_crp.get(dte, 0)
                print(f"     - {dte.strftime('%Y-%m-%d')}: Total {tot} (CAP {cap_n}, COP {cop_n}, CRP {crp_n})")


        print(f"\n4. REQUERIMIENTOS DE ALOJAMIENTO (Top 10)")
    total_hotel_usd = 0.0
    for stn, count in lays.most_common(10):
        rate = HOTEL_RATES_USD.get(stn, 0)
        cost = count * rate
        total_hotel_usd += cost
        if VERBOSE_OUTPUT:
            if rate:
                print(f"   > {stn}: {count} noches-persona | USD {cost:.0f}")
            else:
                print(f"   > {stn}: {count} noches-persona | USD N/A")

    # Calculate hotel cost by month from hotels_by_month
    for (stn, ym), count in hotels_by_month.items():
        rate = HOTEL_RATES_USD.get(stn, 0)
        cost = count * rate
        year, month = int(ym.split('-')[0]), int(ym.split('-')[1])
        hotel_cost_by_month[(year, month)] += cost

    if VERBOSE_OUTPUT:
        print(f"\n5. VUELOS NO CUBIERTOS / OPEN TIME ({len(missing)})")
    open_time_rows = []
    if missing:
        miss_l = list(missing)
        miss_l.sort()
        for m in miss_l:
            r = flights_df[flights_df['id']==m].iloc[0]
            if VERBOSE_OUTPUT:
                print(f"   [OPEN] {r['flt_num']} {r['org']}-{r['dst']} ({r['dep_base'].strftime('%d-%b')}) {r['tail']}")
            open_time_rows.append({
                "flight_id": r["id"],
                "flt_num": r["flt_num"],
                "tail": r["tail"],
                "org": r["org"],
                "dst": r["dst"],
                "dep_base": r["dep_base"],
                "arr_base": r["arr_base"],
                "dep_utc": r["dep_utc"],
                "arr_utc": r["arr_utc"],
                "blk_hours": r["blk_hours"],
            })
    else:
        if VERBOSE_OUTPUT:
            print("   (Ninguno - Cobertura 100%)")
    if VERBOSE_OUTPUT:
        print("="*150)


    # ===================
    # Totales de Viáticos
    # ===================
    total_viaticos_usd = sum(float(v.get("total_usd", 0) or 0) for v in viaticos_rows)
    total_viaticos_cop = sum(float(v.get("total_cop", 0) or 0) for v in viaticos_rows)

    # ===================
    # Completar métricas derivadas AV KPIS (ahora que tenemos hotel y viáticos)
    # ===================
    hotel_cost_per_block_hour = (total_hotel_usd / total_blk) if total_blk > 0 else 0.0
    viaticos_per_block_hour = (total_viaticos_usd / total_blk) if total_blk > 0 else 0.0
    direct_cost_per_block_hour = hotel_cost_per_block_hour + viaticos_per_block_hour

    # ===================
    # Reportes adicionales (para Excel)
    # ===================
    # Horas por rol con discriminación mensual
    pilot_hours_rows = []

    # Recolectar todos los meses únicos
    all_months = set()
    for month_key in cap_block_by_month.keys():
        all_months.add(month_key)
    for month_key in cop_block_by_month.keys():
        all_months.add(month_key)
    for month_key in crp_block_by_month.keys():
        all_months.add(month_key)

    # Nombres de meses en español
    month_names = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    # Generar filas por mes y rol
    for year, month in sorted(all_months):
        month_str = f"{month_names[month]} {year}"

        # CAP
        pilot_hours_rows.append({
            "month": month_str,
            "role": "CAP",
            "block_hours": round(float(cap_block_by_month.get((year, month), 0)), 2),
            "duty_hours": round(float(cap_duty_by_month.get((year, month), 0)), 2),
        })

        # COP
        pilot_hours_rows.append({
            "month": month_str,
            "role": "COP",
            "block_hours": round(float(cop_block_by_month.get((year, month), 0)), 2),
            "duty_hours": round(float(cop_duty_by_month.get((year, month), 0)), 2),
        })

        # CRP
        pilot_hours_rows.append({
            "month": month_str,
            "role": "CRP",
            "block_hours": round(float(crp_block_by_month.get((year, month), 0)), 2),
            "duty_hours": round(float(crp_duty_by_month.get((year, month), 0)), 2),
        })

    # Agregar fila de totales al final
    pilot_hours_rows.append({
        "month": "TOTAL",
        "role": "CAP",
        "block_hours": round(float(cap_block_hours_total), 2),
        "duty_hours": round(float(cap_duty_hours_total), 2),
    })
    pilot_hours_rows.append({
        "month": "TOTAL",
        "role": "COP",
        "block_hours": round(float(cop_block_hours_total), 2),
        "duty_hours": round(float(cop_duty_hours_total), 2),
    })
    pilot_hours_rows.append({
        "month": "TOTAL",
        "role": "CRP",
        "block_hours": round(float(crp_block_hours_total), 2),
        "duty_hours": round(float(crp_duty_hours_total), 2),
    })

    # Pilotos requeridos por día (global)
    all_dates = sorted(set(list(pilots_by_day_cap.keys()) + list(pilots_by_day_cop.keys()) + list(pilots_by_day_crp.keys())))
    pilots_by_day_rows = []
    for dte in all_dates:
        cap_n = int(pilots_by_day_cap.get(dte, 0))
        cop_n = int(pilots_by_day_cop.get(dte, 0))
        crp_n = int(pilots_by_day_crp.get(dte, 0))
        pilots_by_day_rows.append({
            "date": dte.strftime("%Y-%m-%d"),
            "CAP_required": cap_n,
            "COP_required": cop_n,
            "CRP_required": crp_n,
            "Total_required": cap_n + cop_n + crp_n,
        })

    # Pilotos requeridos por día y base
    pilots_by_day_base_rows = []
    keys_base = set(list(pilots_by_day_base_cap.keys()) + list(pilots_by_day_base_cop.keys()) + list(pilots_by_day_base_crp.keys()))
    for (dte, base) in sorted(keys_base):
        cap_n = int(pilots_by_day_base_cap.get((dte, base), 0))
        cop_n = int(pilots_by_day_base_cop.get((dte, base), 0))
        crp_n = int(pilots_by_day_base_crp.get((dte, base), 0))
        pilots_by_day_base_rows.append({
            "date": dte.strftime("%Y-%m-%d"),
            "base": base,
            "CAP_required": cap_n,
            "COP_required": cop_n,
            "CRP_required": crp_n,
            "Total_required": cap_n + cop_n + crp_n,
        })

    # Resumen diario de horas (CAP/COP/CRP/DH/TA)
    daily_hours_rows = []
    for dte in sorted(daily_hours.keys()):
        rec = daily_hours[dte]
        daily_hours_rows.append({
            "date": dte.strftime("%Y-%m-%d"),
            "CAP_block_hours": round(float(rec.get("cap_block", 0.0)), 2),
            "COP_block_hours": round(float(rec.get("cop_block", 0.0)), 2),
            "CRP_block_hours": round(float(rec.get("crp_block", 0.0)), 2),
            "DH_block_hours": round(float(rec.get("dh_hours", 0.0)), 2),
            "TA_duty_hours": round(float(rec.get("ta_hours", 0.0)), 2),
        })

    # Pilotos requeridos por día (excluyente) con día de la semana
    daily_pilots_excl_rows = []
    for dte in sorted(daily_pilots_excl.keys()):
        rec = daily_pilots_excl[dte]
        daily_pilots_excl_rows.append({
            "date": dte.strftime("%Y-%m-%d"),
            "weekday": weekday_es[dte.weekday()],
            "CAP_required": int(rec.get("CAP", 0)),
            "COP_required": int(rec.get("COP", 0)),
            "CRP_required": int(rec.get("CRP", 0)),
            "DH_required": int(rec.get("DH", 0)),
            "TA_required": int(rec.get("TA", 0)),
            "Total_required": int(rec.get("CAP", 0) + rec.get("COP", 0) + rec.get("CRP", 0) + rec.get("DH", 0) + rec.get("TA", 0)),
        })

    # Planta requerida por mes (CAP/COP/CRP)
    plant_required_rows = []
    monthly = {}
    for dte, cap_n in pilots_by_day_cap.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly.setdefault(ym, {"cap": 0, "cop": 0, "crp": 0, "days_in_month": 0, "year": dte.year, "month": dte.month})
        rec["cap"] += cap_n
    for dte, cop_n in pilots_by_day_cop.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly.setdefault(ym, {"cap": 0, "cop": 0, "crp": 0, "days_in_month": 0, "year": dte.year, "month": dte.month})
        rec["cop"] += cop_n
    for dte, crp_n in pilots_by_day_crp.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly.setdefault(ym, {"cap": 0, "cop": 0, "crp": 0, "days_in_month": 0, "year": dte.year, "month": dte.month})
        rec["crp"] += crp_n

    monthly_base = {}
    for (dte, base), cap_n in pilots_by_day_base_cap.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly_base.setdefault((ym, base), {"cap": 0, "cop": 0, "crp": 0, "year": dte.year, "month": dte.month})
        rec["cap"] += cap_n
    for (dte, base), cop_n in pilots_by_day_base_cop.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly_base.setdefault((ym, base), {"cap": 0, "cop": 0, "crp": 0, "year": dte.year, "month": dte.month})
        rec["cop"] += cop_n
    for (dte, base), crp_n in pilots_by_day_base_crp.items():
        ym = dte.strftime("%Y-%m")
        rec = monthly_base.setdefault((ym, base), {"cap": 0, "cop": 0, "crp": 0, "year": dte.year, "month": dte.month})
        rec["crp"] += crp_n

    off_total = float(TRAINING_PILOTS + VACATION_PILOTS + ADMIN_PILOTS + DOCS_PILOTS + INCAP_PILOTS + UNION_PILOTS)
    for ym in sorted(monthly.keys()):
        rec = monthly[ym]
        year = rec["year"]
        month = rec["month"]
        try:
            import calendar
            days_in_month = calendar.monthrange(year, month)[1]
        except Exception:
            days_in_month = 30
        rec["days_in_month"] = days_in_month
        usable_days = max(1, days_in_month - FREE_DAYS_PER_MONTH)
        cap_dem = float(rec["cap"])
        cop_dem = float(rec["cop"])
        crp_dem = float(rec["crp"])
        total_dem = cap_dem + cop_dem + crp_dem
        cap_active = cap_dem / usable_days if usable_days > 0 else 0.0
        cop_active = cop_dem / usable_days if usable_days > 0 else 0.0
        crp_active = crp_dem / usable_days if usable_days > 0 else 0.0
        if total_dem > 0:
            cap_off = off_total * (cap_dem / total_dem)
            cop_off = off_total * (cop_dem / total_dem)
            crp_off = off_total * (crp_dem / total_dem)
        else:
            cap_off = cop_off = crp_off = 0.0
        import math
        reserve_factor = 1.0 + RESERVE_PCT
        cap_req = math.ceil((cap_active + cap_off) * reserve_factor)
        cop_req = math.ceil((cop_active + cop_off) * reserve_factor)
        crp_req = math.ceil((crp_active + crp_off) * reserve_factor)
        plant_required_rows.append({
            "month": ym,
            "base": "ALL",
            "days_in_month": days_in_month,
            "usable_days_per_pilot": usable_days,
            "cap_demand_days": round(cap_dem, 2),
            "cop_demand_days": round(cop_dem, 2),
            "crp_demand_days": round(crp_dem, 2),
            "off_total_pilots": off_total,
            "cap_off_alloc": round(cap_off, 2),
            "cop_off_alloc": round(cop_off, 2),
            "crp_off_alloc": round(crp_off, 2),
            "reserve_pct": RESERVE_PCT,
            "cap_required": cap_req,
            "cop_required": cop_req,
            "crp_required": crp_req,
            "total_required": cap_req + cop_req + crp_req,
        })

    for (ym, base) in sorted(monthly_base.keys()):
        rec = monthly_base[(ym, base)]
        year = rec["year"]
        month = rec["month"]
        try:
            import calendar
            days_in_month = calendar.monthrange(year, month)[1]
        except Exception:
            days_in_month = 30
        usable_days = max(1, days_in_month - FREE_DAYS_PER_MONTH)
        cap_dem = float(rec["cap"])
        cop_dem = float(rec["cop"])
        crp_dem = float(rec["crp"])
        total_dem = cap_dem + cop_dem + crp_dem
        cap_active = cap_dem / usable_days if usable_days > 0 else 0.0
        cop_active = cop_dem / usable_days if usable_days > 0 else 0.0
        crp_active = crp_dem / usable_days if usable_days > 0 else 0.0
        if total_dem > 0:
            cap_off = off_total * (cap_dem / total_dem)
            cop_off = off_total * (cop_dem / total_dem)
            crp_off = off_total * (crp_dem / total_dem)
        else:
            cap_off = cop_off = crp_off = 0.0
        import math
        reserve_factor = 1.0 + RESERVE_PCT
        cap_req = math.ceil((cap_active + cap_off) * reserve_factor)
        cop_req = math.ceil((cop_active + cop_off) * reserve_factor)
        crp_req = math.ceil((crp_active + crp_off) * reserve_factor)
        plant_required_rows.append({
            "month": ym,
            "base": base,
            "days_in_month": days_in_month,
            "usable_days_per_pilot": usable_days,
            "cap_demand_days": round(cap_dem, 2),
            "cop_demand_days": round(cop_dem, 2),
            "crp_demand_days": round(crp_dem, 2),
            "off_total_pilots": off_total,
            "cap_off_alloc": round(cap_off, 2),
            "cop_off_alloc": round(cop_off, 2),
            "crp_off_alloc": round(crp_off, 2),
            "reserve_pct": RESERVE_PCT,
            "cap_required": cap_req,
            "cop_required": cop_req,
            "crp_required": crp_req,
            "total_required": cap_req + cop_req + crp_req,
        })

    # Inicios de duty por estación (pivot días x estaciones)
    duty_starts_rows = []
    if duty_starts_by_station:
        stations = sorted({stn for (_, stn) in duty_starts_by_station.keys()})
        dates = sorted({d for (d, _) in duty_starts_by_station.keys()})
        for d in dates:
            row = {
                "date": d.strftime("%Y-%m-%d"),
                "weekday": weekday_es[d.weekday()],
            }
            for stn in stations:
                row[stn] = int(duty_starts_by_station.get((d, stn), 0))
            duty_starts_rows.append(row)

    # ====================================================================
    # Función auxiliar para construir KPIs mensuales
    # ====================================================================
    def _build_monthly_kpis(
        trips_by_month, duties_by_month, ac_changes_by_month,
        hotel_nights_by_month, dh_flights_by_month, dh_hours_by_month,
        tafb_hours_by_month, hotel_cost_by_month, viaticos_by_month,
        duty_days_by_month, block_hours_by_month, duty_hours_by_month,
        flights_covered_by_month, flights_total_by_month,
        block_hours_by_month_base, duty_days_by_month_base, trips_by_month_base,
        cap_block_by_month, cop_block_by_month, crp_block_by_month,
        cap_duty_by_month, cop_duty_by_month, crp_duty_by_month
    ):
        """Construye diccionario de KPIs por mes para el Dashboard."""
        month_names = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        # Recopilar todos los meses únicos
        all_months = set()
        for counter in [trips_by_month, duties_by_month, block_hours_by_month,
                       flights_total_by_month, cap_block_by_month]:
            all_months.update(counter.keys())

        monthly_data = []
        for year, month in sorted(all_months):
            month_str = f"{month_names[month]} {year}"
            ym_key = (year, month)

            # Block hours totales y por rol
            blk_total = float(block_hours_by_month.get(ym_key, 0))
            cap_blk = float(cap_block_by_month.get(ym_key, 0))
            cop_blk = float(cop_block_by_month.get(ym_key, 0))
            crp_blk = float(crp_block_by_month.get(ym_key, 0))

            # Duty hours totales y por rol
            dty_total = float(duty_hours_by_month.get(ym_key, 0))
            cap_dty = float(cap_duty_by_month.get(ym_key, 0))
            cop_dty = float(cop_duty_by_month.get(ym_key, 0))
            crp_dty = float(crp_duty_by_month.get(ym_key, 0))

            # Conteos
            trips_n = int(trips_by_month.get(ym_key, 0))
            duties_n = int(duties_by_month.get(ym_key, 0))
            duty_days_n = int(duty_days_by_month.get(ym_key, 0))
            ac_changes_n = int(ac_changes_by_month.get(ym_key, 0))
            hotel_nights_n = int(hotel_nights_by_month.get(ym_key, 0))
            dh_flights_n = int(dh_flights_by_month.get(ym_key, 0))
            dh_hours_n = float(dh_hours_by_month.get(ym_key, 0))
            tafb_hours_n = float(tafb_hours_by_month.get(ym_key, 0))

            # Costos
            hotel_cost_n = float(hotel_cost_by_month.get(ym_key, 0))
            viaticos_n = float(viaticos_by_month.get(ym_key, 0))
            direct_cost_n = hotel_cost_n + viaticos_n

            # Cobertura
            flights_total_n = int(flights_total_by_month.get(ym_key, 0))
            flights_covered_n = int(flights_covered_by_month.get(ym_key, 0))
            coverage_pct = (flights_covered_n / max(1, flights_total_n)) * 100 if flights_total_n > 0 else 0

            # Métricas derivadas
            avg_block_per_day = (blk_total / max(1, duty_days_n)) if duty_days_n > 0 else 0
            avg_block_per_duty = (blk_total / max(1, duties_n)) if duties_n > 0 else 0
            avg_duty_per_duty = (dty_total / max(1, duties_n)) if duties_n > 0 else 0
            hotel_cost_per_blk = (hotel_cost_n / max(1, blk_total)) if blk_total > 0 else 0
            viaticos_per_blk = (viaticos_n / max(1, blk_total)) if blk_total > 0 else 0
            direct_cost_per_blk = hotel_cost_per_blk + viaticos_per_blk

            # Por base
            bog_blk = float(block_hours_by_month_base.get((year, month, 'BOG'), 0))
            mde_blk = float(block_hours_by_month_base.get((year, month, 'MDE'), 0))
            bog_days = int(duty_days_by_month_base.get((year, month, 'BOG'), 0))
            mde_days = int(duty_days_by_month_base.get((year, month, 'MDE'), 0))
            bog_trips = int(trips_by_month_base.get((year, month, 'BOG'), 0))
            mde_trips = int(trips_by_month_base.get((year, month, 'MDE'), 0))

            bog_blk_pct = (bog_blk / max(1, blk_total)) * 100 if blk_total > 0 else 0
            mde_blk_pct = (mde_blk / max(1, blk_total)) * 100 if blk_total > 0 else 0

            monthly_data.append({
                "year": year,
                "month": month,
                "month_name": month_str,
                # Cobertura
                "flights_total": flights_total_n,
                "flights_covered": flights_covered_n,
                "coverage_pct": round(coverage_pct, 2),
                # Conteos
                "trips": trips_n,
                "duties": duties_n,
                "duty_days": duty_days_n,
                "ac_changes": ac_changes_n,
                "hotel_nights": hotel_nights_n,
                "dh_flights": dh_flights_n,
                # Horas
                "block_hours_total": round(blk_total, 2),
                "duty_hours_total": round(dty_total, 2),
                "dh_hours": round(dh_hours_n, 2),
                "tafb_hours": round(tafb_hours_n, 2),
                "cap_block_hours": round(cap_blk, 2),
                "cop_block_hours": round(cop_blk, 2),
                "crp_block_hours": round(crp_blk, 2),
                "cap_duty_hours": round(cap_dty, 2),
                "cop_duty_hours": round(cop_dty, 2),
                "crp_duty_hours": round(crp_dty, 2),
                # Costos
                "hotel_cost_usd": round(hotel_cost_n, 2),
                "viaticos_usd": round(viaticos_n, 2),
                "direct_cost_usd": round(direct_cost_n, 2),
                # Métricas derivadas
                "avg_block_per_day": round(avg_block_per_day, 2),
                "avg_block_per_duty": round(avg_block_per_duty, 2),
                "avg_duty_per_duty": round(avg_duty_per_duty, 2),
                "hotel_cost_per_block_hour": round(hotel_cost_per_blk, 2),
                "viaticos_per_block_hour": round(viaticos_per_blk, 2),
                "direct_cost_per_block_hour": round(direct_cost_per_blk, 2),
                # Por base
                "bog_block_hours": round(bog_blk, 2),
                "mde_block_hours": round(mde_blk, 2),
                "bog_block_pct": round(bog_blk_pct, 2),
                "mde_block_pct": round(mde_blk_pct, 2),
                "bog_duty_days": bog_days,
                "mde_duty_days": mde_days,
                "bog_trips": bog_trips,
                "mde_trips": mde_trips,
            })

        return monthly_data

    # Calcular KPIs de segunda pasada
    trips_second_pass = sum(1 for t in selected_trips if t.get('is_second_pass'))
    flights_covered_second_pass = sum(
        len(t.get('flights_covered', [])) for t in selected_trips if t.get('is_second_pass')
    )
    coverage_improvement = (flights_covered_second_pass / max(1, len(flight_ids))) * 100 if flight_ids else 0

    # Calcular total_dh_hours desde dh_report_rows para garantizar consistencia con el reporte
    total_dh_hours = sum(float(row.get('blk_hours', 0) or 0) for row in dh_report_rows)

    results = {
        "selected_trips": selected_trips,
        "trip_report_rows": trip_report_rows,
        "pilot_lines_rows": pilot_lines_rows,
        "pilot_lines_dates_set": pilot_lines_dates_set,
        "trip_legs_rows": trip_legs_rows,
        "kpis": {
            "coverage_pct": round(cov_pct, 2),
            "total_flights": len(flight_ids),
            "flights_covered": len(covered),
            "flights_open": len(missing),
            "avg_network_eff": round(avg_network_eff, 4),
            "total_trips": int(total_trips),
            "total_blk": round(float(total_blk), 2),
            "total_days": int(total_days),
            "cap_block_hours_total": round(float(cap_block_hours_total), 2),
            "cop_block_hours_total": round(float(cop_block_hours_total), 2),
            "crp_block_hours_total": round(float(crp_block_hours_total), 2),
            "cap_duty_hours_total": round(float(cap_duty_hours_total), 2),
            "cop_duty_hours_total": round(float(cop_duty_hours_total), 2),
            "crp_duty_hours_total": round(float(crp_duty_hours_total), 2),
            "total_dh_hours": round(float(total_dh_hours), 2),
            "total_tafb_hours": round(float(total_tafb_hours), 2),
            "cap_network_eff": round(float(eff_cap), 4),
            "cop_network_eff": round(float(eff_cop), 4),
            "crp_network_eff": round(float(eff_crp), 4),
            "hotel_total_usd": round(float(total_hotel_usd), 2),
            "viaticos_total_usd": round(float(total_viaticos_usd), 2),
            "viaticos_total_cop": round(float(total_viaticos_cop), 2),
            "pilot_base_distribution": dict(base_pilot_days),
            "trips_second_pass": int(trips_second_pass),
            "flights_covered_second_pass": int(flights_covered_second_pass),
            "coverage_improvement": round(coverage_improvement, 2),
            # Métricas de tercera pasada (trips parciales)
            "trips_partial": int(trips_partial),
            "flights_covered_partial": int(flights_covered_partial),
            "coverage_partial": round(coverage_partial, 2),
            "partial_final_stations": partial_final_stations,
            # Nuevas métricas AV KPIS
            "total_duties": int(total_duties),
            "total_duty_hours": round(float(total_duty_hours), 2),
            "total_ac_changes": int(total_ac_changes),
            "total_hotel_nights": int(total_hotel_nights),
            "total_dh_flights": int(total_dh_flights),
            "avg_block_per_day": round(float(avg_block_per_day), 2),
            "avg_block_per_duty": round(float(avg_block_per_duty), 2),
            "avg_duty_per_duty": round(float(avg_duty_per_duty), 2),
            "hotel_cost_per_block_hour": round(float(hotel_cost_per_block_hour), 2),
            "viaticos_per_block_hour": round(float(viaticos_per_block_hour), 2),
            "direct_cost_per_block_hour": round(float(direct_cost_per_block_hour), 2),
            # Distribuciones por rangos
            "duties_duty_00_04": int(duties_by_duty_time["00:00-04:00"]),
            "duties_duty_04_08": int(duties_by_duty_time["04:01-08:00"]),
            "duties_duty_08_10": int(duties_by_duty_time["08:01-10:00"]),
            "duties_duty_gt_10": int(duties_by_duty_time[">10:00"]),
            "duties_block_00_02": int(duties_by_block_time["00:00-02:00"]),
            "duties_block_02_04": int(duties_by_block_time["02:01-04:00"]),
            "duties_block_04_06": int(duties_by_block_time["04:01-06:00"]),
            "duties_block_06_08": int(duties_by_block_time["06:01-08:00"]),
        },
        "base_distribution": dict(base_c),
        "crew_distribution": dict(crew_c),
        "man_days": dict(man_days),
        "hotels": lays,
        "hotels_by_month": hotels_by_month,
        "missing_flights": missing,
        "open_time_rows": open_time_rows,
        "pilot_hours_rows": pilot_hours_rows,
        "pilots_by_day_rows": pilots_by_day_rows,
        "pilots_by_day_base_rows": pilots_by_day_base_rows,
        "plant_required_rows": plant_required_rows,
        "daily_hours_rows": daily_hours_rows,
        "daily_pilots_excl_rows": daily_pilots_excl_rows,
        "viaticos_rows": viaticos_rows,
        "duty_starts_rows": duty_starts_rows,
        "dh_report_rows": dh_report_rows,
        "base_analysis": base_analysis,
        "idle_windows": idle_windows,
        # ====================================================================
        # Monthly KPIs - Datos mensuales para el Dashboard
        # ====================================================================
        "monthly_kpis": _build_monthly_kpis(
            trips_by_month, duties_by_month, ac_changes_by_month,
            hotel_nights_by_month, dh_flights_by_month, dh_hours_by_month,
            tafb_hours_by_month, hotel_cost_by_month, viaticos_by_month,
            duty_days_by_month, block_hours_by_month, duty_hours_by_month,
            flights_covered_by_month, flights_total_by_month,
            block_hours_by_month_base, duty_days_by_month_base, trips_by_month_base,
            cap_block_by_month, cop_block_by_month, crp_block_by_month,
            cap_duty_by_month, cop_duty_by_month, crp_duty_by_month
        ),
    }

    # =============================
    # Exportación a Excel
    # =============================
    if export_excel:
        if excel_output_path is None:
            # Por defecto: mismo folder del itinerario (si existe), o cwd
            base_dir = os.path.dirname(os.path.abspath(__file__))
            try:
                # Si el DF viene del load_schedule, no traemos aquí FILE_PATH; el usuario puede pasar excel_output_path.
                pass
            except Exception:
                pass
            fname = f"Resultados_{APP_NAME_FILE}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            excel_output_path = os.path.join(base_dir, fname)

        pilot_lines_dates = [_fmt_day_col(d) for d in sorted(pilot_lines_dates_set)]
        out = export_results_to_excel(
            output_path=excel_output_path,
            flights_df=flights_df,
            trips=trips,
            selected_trips=selected_trips,
            trip_legs_rows=trip_legs_rows,
            trip_report_rows=trip_report_rows,
            pilot_lines_rows=pilot_lines_rows,
            pilot_lines_dates=pilot_lines_dates,
            daily_hours_rows=results.get("daily_hours_rows", []),
            daily_pilots_excl_rows=results.get("daily_pilots_excl_rows", []),
            viaticos_rows=results.get("viaticos_rows", []),
            kpis=results["kpis"],
            base_dist=results["base_distribution"],
            crew_dist=results["crew_distribution"],
            man_days=results["man_days"],
            hotels_counter=lays,
            hotels_by_month=results.get("hotels_by_month", Counter()),
            duty_starts_rows=results.get("duty_starts_rows", []),
            open_time_rows=open_time_rows,
            pilot_hours_rows=results.get("pilot_hours_rows", []),
            pilots_by_day_rows=results.get("pilots_by_day_rows", []),
            pilots_by_day_base_rows=results.get("pilots_by_day_base_rows", []),
            plant_required_rows=results.get("plant_required_rows", []),
            dh_report_rows=results.get("dh_report_rows", []),
            monthly_kpis=results.get("monthly_kpis", []),
            base_analysis=results.get("base_analysis"),
            idle_windows=results.get("idle_windows"),
        )
        print(f"\n✅ Exportado a Excel: {out}")
        try:
            os.startfile(out)
        except Exception:
            pass

    return results


def _merge_dict_counts(a, b):
    out = Counter()
    if a:
        out.update(a)
    if b:
        out.update(b)
    return dict(out)


def _merge_man_days(a, b):
    a = a or {}
    b = b or {}
    result = {}
    # Merge top-level numeric keys (cap, cop, crp)
    for k in ['cap', 'cop', 'crp']:
        result[k] = (a.get(k, 0) or 0) + (b.get(k, 0) or 0)
    # Merge nested dictionaries (cap_by_crew, cop_by_crew, crp_by_crew)
    for k in ['cap_by_crew', 'cop_by_crew', 'crp_by_crew']:
        a_dict = a.get(k, {}) or {}
        b_dict = b.get(k, {}) or {}
        result[k] = {
            '2P': (a_dict.get('2P', 0) or 0) + (b_dict.get('2P', 0) or 0),
            '3P': (a_dict.get('3P', 0) or 0) + (b_dict.get('3P', 0) or 0),
            '4P': (a_dict.get('4P', 0) or 0) + (b_dict.get('4P', 0) or 0),
        }
    return result


def _aggregate_pilots_by_day(rows):
    total = Counter()
    cap = Counter()
    cop = Counter()
    crp = Counter()
    for row in rows or []:
        try:
            d = datetime.strptime(str(row.get("date")), "%Y-%m-%d").date()
        except Exception:
            continue
        cap_n = int(row.get("CAP_required", 0) or 0)
        cop_n = int(row.get("COP_required", 0) or 0)
        crp_n = int(row.get("CRP_required", 0) or 0)
        total[d] += cap_n + cop_n + crp_n
        cap[d] += cap_n
        cop[d] += cop_n
        crp[d] += crp_n
    return total, cap, cop, crp


def _build_open_time_rows(flights_df, missing_set):
    rows = []
    if not missing_set:
        return rows
    miss_l = list(missing_set)
    miss_l.sort()
    for m in miss_l:
        r = flights_df[flights_df['id'] == m].iloc[0]
        rows.append({
            "flight_id": r["id"],
            "flt_num": r["flt_num"],
            "tail": r["tail"],
            "org": r["org"],
            "dst": r["dst"],
            "dep_base": r["dep_base"],
            "arr_base": r["arr_base"],
            "dep_utc": r["dep_utc"],
            "arr_utc": r["arr_utc"],
            "blk_hours": r["blk_hours"],
        })
    return rows


def _hotel_total_top10_usd(hotels_counter):
    total_hotel_usd = 0.0
    for stn, count in (hotels_counter or Counter()).most_common(10):
        rate = HOTEL_RATES_USD.get(stn, 0)
        total_hotel_usd += float(count) * float(rate or 0)
    return round(total_hotel_usd, 2)


def _combine_results_with_open(res_main, res_open, flights_df, open_df):
    missing_main = set(res_main.get("missing_flights") or [])
    missing_open = set(res_open.get("missing_flights") or [])
    open_ids = set(open_df["id"].tolist())
    covered_open = open_ids - missing_open
    combined_missing = missing_main - covered_open

    k1 = res_main.get("kpis", {}) or {}
    k2 = res_open.get("kpis", {}) or {}

    total_blk = float(k1.get("total_blk", 0) or 0) + float(k2.get("total_blk", 0) or 0)
    total_days = int(k1.get("total_days", 0) or 0) + int(k2.get("total_days", 0) or 0)
    total_trips = int(k1.get("total_trips", 0) or 0) + int(k2.get("total_trips", 0) or 0)

    cap_block_total = float(k1.get("cap_block_hours_total", 0) or 0) + float(k2.get("cap_block_hours_total", 0) or 0)
    cop_block_total = float(k1.get("cop_block_hours_total", 0) or 0) + float(k2.get("cop_block_hours_total", 0) or 0)
    crp_block_total = float(k1.get("crp_block_hours_total", 0) or 0) + float(k2.get("crp_block_hours_total", 0) or 0)
    cap_duty_total = float(k1.get("cap_duty_hours_total", 0) or 0) + float(k2.get("cap_duty_hours_total", 0) or 0)
    cop_duty_total = float(k1.get("cop_duty_hours_total", 0) or 0) + float(k2.get("cop_duty_hours_total", 0) or 0)
    crp_duty_total = float(k1.get("crp_duty_hours_total", 0) or 0) + float(k2.get("crp_duty_hours_total", 0) or 0)
    # DH y TAFB totales
    total_dh_hours = float(k1.get("total_dh_hours", 0) or 0) + float(k2.get("total_dh_hours", 0) or 0)
    total_tafb_hours = float(k1.get("total_tafb_hours", 0) or 0) + float(k2.get("total_tafb_hours", 0) or 0)

    man_days = _merge_man_days(res_main.get("man_days"), res_open.get("man_days"))
    eff_cap = (cap_block_total / man_days.get("cap", 0)) if man_days.get("cap", 0) > 0 else 0.0
    eff_cop = (cop_block_total / man_days.get("cop", 0)) if man_days.get("cop", 0) > 0 else 0.0
    eff_crp = (crp_block_total / man_days.get("crp", 0)) if man_days.get("crp", 0) > 0 else 0.0
    avg_network_eff = (total_blk / total_days) if total_days > 0 else 0.0

    all_f = set(flights_df['id'].unique())
    cov_pct = (len(all_f) - len(combined_missing)) / len(all_f) * 100 if len(all_f) else 0.0

    hotels_counter = Counter()
    hotels_counter.update(res_main.get("hotels", Counter()))
    hotels_counter.update(res_open.get("hotels", Counter()))

    base_dist = _merge_dict_counts(res_main.get("base_distribution"), res_open.get("base_distribution"))
    crew_dist = _merge_dict_counts(res_main.get("crew_distribution"), res_open.get("crew_distribution"))

    open_time_rows_open = _build_open_time_rows(flights_df, combined_missing)

    pilots_total_main, cap_main, cop_main, crp_main = _aggregate_pilots_by_day(res_main.get("pilots_by_day_rows"))
    pilots_total_open, cap_open, cop_open, crp_open = _aggregate_pilots_by_day(res_open.get("pilots_by_day_rows"))
    pilots_total = pilots_total_main + pilots_total_open
    pilots_cap = cap_main + cap_open
    pilots_cop = cop_main + cop_open
    pilots_crp = crp_main + crp_open

    # Calculate second pass metrics for combined results
    all_selected_trips = (res_main.get("selected_trips", []) or []) + (res_open.get("selected_trips", []) or [])
    trips_second_pass_combined = sum(1 for t in all_selected_trips if t.get('is_second_pass'))
    flights_covered_second_pass_combined = sum(
        len(t.get('flights_covered', [])) for t in all_selected_trips if t.get('is_second_pass')
    )
    coverage_improvement_combined = (flights_covered_second_pass_combined / max(1, len(all_f))) * 100

    kpis_open = {
        "coverage_pct": round(cov_pct, 2),
        "avg_network_eff": round(avg_network_eff, 4),
        "total_trips": int(total_trips),
        "total_blk": round(total_blk, 2),
        "total_days": int(total_days),
        "cap_block_hours_total": round(cap_block_total, 2),
        "cop_block_hours_total": round(cop_block_total, 2),
        "crp_block_hours_total": round(crp_block_total, 2),
        "cap_duty_hours_total": round(cap_duty_total, 2),
        "cop_duty_hours_total": round(cop_duty_total, 2),
        "crp_duty_hours_total": round(crp_duty_total, 2),
        "total_dh_hours": round(total_dh_hours, 2),
        "total_tafb_hours": round(total_tafb_hours, 2),
        "cap_network_eff": round(eff_cap, 4),
        "cop_network_eff": round(eff_cop, 4),
        "crp_network_eff": round(eff_crp, 4),
        "hotel_total_usd": _hotel_total_top10_usd(hotels_counter),
        "viaticos_total_usd": round(
            float(k1.get("viaticos_total_usd", 0) or 0) + float(k2.get("viaticos_total_usd", 0) or 0), 2
        ),
        "viaticos_total_cop": round(
            float(k1.get("viaticos_total_cop", 0) or 0) + float(k2.get("viaticos_total_cop", 0) or 0), 2
        ),
        "trips_second_pass": int(trips_second_pass_combined),
        "flights_covered_second_pass": int(flights_covered_second_pass_combined),
        "coverage_improvement": round(coverage_improvement_combined, 2),
    }

    combined = {
        "kpis": kpis_open,
        "base_distribution": dict(base_dist),
        "crew_distribution": dict(crew_dist),
        "man_days": man_days,
        "hotels": hotels_counter,
        "missing_flights": combined_missing,
        "open_time_rows": open_time_rows_open,
        "pilots_by_day_total": pilots_total,
        "pilots_by_day_cap": pilots_cap,
        "pilots_by_day_cop": pilots_cop,
        "pilots_by_day_crp": pilots_crp,
    }
    return combined


def _merge_date_keyed_rows(rows_main, rows_extra, date_col="date"):
    """Merge dos listas de dicts keyed por fecha, sumando columnas numéricas."""
    if not rows_extra:
        return list(rows_main or [])
    if not rows_main:
        return list(rows_extra or [])
    merged = {}
    for row in list(rows_main) + list(rows_extra):
        key = row.get(date_col, "")
        if key not in merged:
            merged[key] = dict(row)
        else:
            existing = merged[key]
            for k, v in row.items():
                if k == date_col:
                    continue
                if isinstance(v, (int, float)):
                    existing[k] = (existing.get(k, 0) or 0) + v
                elif k not in existing or not existing[k]:
                    existing[k] = v
    return [merged[k] for k in sorted(merged.keys())]


def _merge_date_base_keyed_rows(rows_main, rows_extra):
    """Merge dos listas de dicts keyed por (date, base), sumando columnas numéricas."""
    if not rows_extra:
        return list(rows_main or [])
    if not rows_main:
        return list(rows_extra or [])
    merged = {}
    for row in list(rows_main) + list(rows_extra):
        key = (row.get("date", ""), row.get("base", ""))
        if key not in merged:
            merged[key] = dict(row)
        else:
            existing = merged[key]
            for k, v in row.items():
                if k in ("date", "base", "weekday"):
                    continue
                if isinstance(v, (int, float)):
                    existing[k] = (existing.get(k, 0) or 0) + v
                elif k not in existing or not existing[k]:
                    existing[k] = v
    return [merged[k] for k in sorted(merged.keys())]


def _rebuild_plant_required(pilots_by_day_rows, pilots_by_day_base_rows):
    """Reconstruye plant_required_rows a partir de pilots_by_day mergeados."""
    plant_rows = []
    # Agregar por mes (ALL)
    monthly = {}
    for row in pilots_by_day_rows or []:
        try:
            d = datetime.strptime(str(row.get("date")), "%Y-%m-%d").date()
        except Exception:
            continue
        ym = d.strftime("%Y-%m")
        rec = monthly.setdefault(ym, {"cap": 0, "cop": 0, "crp": 0, "year": d.year, "month": d.month})
        rec["cap"] += int(row.get("CAP_required", 0) or 0)
        rec["cop"] += int(row.get("COP_required", 0) or 0)
        rec["crp"] += int(row.get("CRP_required", 0) or 0)

    monthly_base = {}
    for row in pilots_by_day_base_rows or []:
        try:
            d = datetime.strptime(str(row.get("date")), "%Y-%m-%d").date()
        except Exception:
            continue
        ym = d.strftime("%Y-%m")
        base = row.get("base", "")
        rec = monthly_base.setdefault((ym, base), {"cap": 0, "cop": 0, "crp": 0, "year": d.year, "month": d.month})
        rec["cap"] += int(row.get("CAP_required", 0) or 0)
        rec["cop"] += int(row.get("COP_required", 0) or 0)
        rec["crp"] += int(row.get("CRP_required", 0) or 0)

    off_total = float(TRAINING_PILOTS + VACATION_PILOTS + ADMIN_PILOTS + DOCS_PILOTS + INCAP_PILOTS + UNION_PILOTS)
    import calendar as _cal
    import math as _math
    reserve_factor = 1.0 + RESERVE_PCT

    for ym in sorted(monthly.keys()):
        rec = monthly[ym]
        days_in_month = _cal.monthrange(rec["year"], rec["month"])[1]
        usable_days = max(1, days_in_month - FREE_DAYS_PER_MONTH)
        cap_dem, cop_dem, crp_dem = float(rec["cap"]), float(rec["cop"]), float(rec["crp"])
        total_dem = cap_dem + cop_dem + crp_dem
        cap_active = cap_dem / usable_days
        cop_active = cop_dem / usable_days
        crp_active = crp_dem / usable_days
        if total_dem > 0:
            cap_off = off_total * (cap_dem / total_dem)
            cop_off = off_total * (cop_dem / total_dem)
            crp_off = off_total * (crp_dem / total_dem)
        else:
            cap_off = cop_off = crp_off = 0.0
        cap_req = _math.ceil((cap_active + cap_off) * reserve_factor)
        cop_req = _math.ceil((cop_active + cop_off) * reserve_factor)
        crp_req = _math.ceil((crp_active + crp_off) * reserve_factor)
        plant_rows.append({
            "month": ym, "base": "ALL", "days_in_month": days_in_month,
            "usable_days_per_pilot": usable_days,
            "cap_demand_days": round(cap_dem, 2), "cop_demand_days": round(cop_dem, 2),
            "crp_demand_days": round(crp_dem, 2), "off_total_pilots": off_total,
            "cap_off_alloc": round(cap_off, 2), "cop_off_alloc": round(cop_off, 2),
            "crp_off_alloc": round(crp_off, 2), "reserve_pct": RESERVE_PCT,
            "cap_required": cap_req, "cop_required": cop_req, "crp_required": crp_req,
            "total_required": cap_req + cop_req + crp_req,
        })

    for (ym, base) in sorted(monthly_base.keys()):
        rec = monthly_base[(ym, base)]
        days_in_month = _cal.monthrange(rec["year"], rec["month"])[1]
        usable_days = max(1, days_in_month - FREE_DAYS_PER_MONTH)
        cap_dem, cop_dem, crp_dem = float(rec["cap"]), float(rec["cop"]), float(rec["crp"])
        total_dem = cap_dem + cop_dem + crp_dem
        cap_active = cap_dem / usable_days
        cop_active = cop_dem / usable_days
        crp_active = crp_dem / usable_days
        if total_dem > 0:
            cap_off = off_total * (cap_dem / total_dem)
            cop_off = off_total * (cop_dem / total_dem)
            crp_off = off_total * (crp_dem / total_dem)
        else:
            cap_off = cop_off = crp_off = 0.0
        cap_req = _math.ceil((cap_active + cap_off) * reserve_factor)
        cop_req = _math.ceil((cop_active + cop_off) * reserve_factor)
        crp_req = _math.ceil((crp_active + crp_off) * reserve_factor)
        plant_rows.append({
            "month": ym, "base": base, "days_in_month": days_in_month,
            "usable_days_per_pilot": usable_days,
            "cap_demand_days": round(cap_dem, 2), "cop_demand_days": round(cop_dem, 2),
            "crp_demand_days": round(crp_dem, 2), "off_total_pilots": off_total,
            "cap_off_alloc": round(cap_off, 2), "cop_off_alloc": round(cop_off, 2),
            "crp_off_alloc": round(crp_off, 2), "reserve_pct": RESERVE_PCT,
            "cap_required": cap_req, "cop_required": cop_req, "crp_required": crp_req,
            "total_required": cap_req + cop_req + crp_req,
        })
    return plant_rows


def _merge_all_pass_results(res_main, res_open=None, res_solo=None):
    """Merge los datos agregados de todas las pasadas (main + open tours + solo DH).
    Retorna un dict con todas las rows/counters/kpis combinados listos para exportar."""

    results = [r for r in [res_main, res_open, res_solo] if r]
    if len(results) <= 1:
        return res_main  # Nada que mergear

    # --- Rows date-keyed: sumar por fecha ---
    merged_daily_hours = list(res_main.get("daily_hours_rows") or [])
    merged_pilots_by_day = list(res_main.get("pilots_by_day_rows") or [])
    merged_pilots_by_day_base = list(res_main.get("pilots_by_day_base_rows") or [])
    merged_daily_pilots_excl = list(res_main.get("daily_pilots_excl_rows") or [])
    merged_duty_starts = list(res_main.get("duty_starts_rows") or [])

    for extra in [res_open, res_solo]:
        if not extra:
            continue
        merged_daily_hours = _merge_date_keyed_rows(merged_daily_hours, extra.get("daily_hours_rows"))
        merged_pilots_by_day = _merge_date_keyed_rows(merged_pilots_by_day, extra.get("pilots_by_day_rows"))
        merged_pilots_by_day_base = _merge_date_base_keyed_rows(merged_pilots_by_day_base, extra.get("pilots_by_day_base_rows"))
        merged_daily_pilots_excl = _merge_date_keyed_rows(merged_daily_pilots_excl, extra.get("daily_pilots_excl_rows"))
        merged_duty_starts = _merge_date_keyed_rows(merged_duty_starts, extra.get("duty_starts_rows"))

    # --- Rows list: concatenar ---
    merged_viaticos = list(res_main.get("viaticos_rows") or [])
    merged_pilot_hours = list(res_main.get("pilot_hours_rows") or [])
    for extra in [res_open, res_solo]:
        if not extra:
            continue
        merged_viaticos.extend(extra.get("viaticos_rows") or [])
        merged_pilot_hours.extend(extra.get("pilot_hours_rows") or [])

    # --- Counters: sumar ---
    merged_hotels = Counter(res_main.get("hotels") or {})
    merged_hotels_by_month = Counter(res_main.get("hotels_by_month") or {})
    for extra in [res_open, res_solo]:
        if not extra:
            continue
        merged_hotels.update(extra.get("hotels") or {})
        merged_hotels_by_month.update(extra.get("hotels_by_month") or {})

    # --- Dicts: merge ---
    merged_base_dist = dict(res_main.get("base_distribution") or {})
    merged_crew_dist = dict(res_main.get("crew_distribution") or {})
    merged_man_days = dict(res_main.get("man_days") or {})
    for extra in [res_open, res_solo]:
        if not extra:
            continue
        merged_base_dist = _merge_dict_counts(merged_base_dist, extra.get("base_distribution"))
        merged_crew_dist = _merge_dict_counts(merged_crew_dist, extra.get("crew_distribution"))
        merged_man_days = _merge_man_days(merged_man_days, extra.get("man_days"))

    # --- Plant required: recalcular desde pilots_by_day mergeados ---
    merged_plant = _rebuild_plant_required(merged_pilots_by_day, merged_pilots_by_day_base)

    # --- KPIs: recalcular sumas ---
    k_main = res_main.get("kpis") or {}
    merged_kpis = dict(k_main)
    for extra in [res_open, res_solo]:
        if not extra:
            continue
        k_extra = extra.get("kpis") or {}
        for key in ["total_blk", "total_days", "total_trips",
                     "cap_block_hours_total", "cop_block_hours_total", "crp_block_hours_total",
                     "cap_duty_hours_total", "cop_duty_hours_total", "crp_duty_hours_total",
                     "total_dh_hours", "total_tafb_hours",
                     "viaticos_total_usd", "viaticos_total_cop"]:
            merged_kpis[key] = round(float(merged_kpis.get(key, 0) or 0) + float(k_extra.get(key, 0) or 0), 2)
    # Recalcular métricas derivadas
    total_days = merged_kpis.get("total_days", 0) or 1
    merged_kpis["avg_network_eff"] = round(float(merged_kpis.get("total_blk", 0)) / max(1, total_days), 4)
    cap_days = merged_man_days.get("cap", 0) or 1
    cop_days = merged_man_days.get("cop", 0) or 1
    crp_days = merged_man_days.get("crp", 0) or 1
    merged_kpis["cap_network_eff"] = round(float(merged_kpis.get("cap_block_hours_total", 0)) / max(1, cap_days), 4)
    merged_kpis["cop_network_eff"] = round(float(merged_kpis.get("cop_block_hours_total", 0)) / max(1, cop_days), 4)
    merged_kpis["crp_network_eff"] = round(float(merged_kpis.get("crp_block_hours_total", 0)) / max(1, crp_days), 4)
    merged_kpis["hotel_total_usd"] = _hotel_total_top10_usd(merged_hotels)

    # --- Monthly KPIs: mergear por (year, month) ---
    main_monthly = res_main.get("monthly_kpis") or []
    extra_monthlies = []
    for extra in [res_open, res_solo]:
        if extra and extra.get("monthly_kpis"):
            extra_monthlies.extend(extra["monthly_kpis"])
    if extra_monthlies:
        monthly_by_key = {}
        for mk in main_monthly:
            key = (mk.get("year"), mk.get("month"))
            monthly_by_key[key] = dict(mk)
        # Campos aditivos (se suman), los derivados se recalculan después
        _ADDITIVE_MONTHLY = {
            "flights_total", "flights_covered",
            "trips", "duties", "duty_days", "ac_changes", "hotel_nights", "dh_flights",
            "block_hours_total", "duty_hours_total", "dh_hours", "tafb_hours",
            "cap_block_hours", "cop_block_hours", "crp_block_hours",
            "cap_duty_hours", "cop_duty_hours", "crp_duty_hours",
            "hotel_cost_usd", "viaticos_usd", "direct_cost_usd",
            "bog_block_hours", "mde_block_hours",
            "bog_duty_days", "mde_duty_days", "bog_trips", "mde_trips",
        }
        for mk in extra_monthlies:
            key = (mk.get("year"), mk.get("month"))
            if key not in monthly_by_key:
                monthly_by_key[key] = dict(mk)
            else:
                existing = monthly_by_key[key]
                for k, v in mk.items():
                    if k in ("year", "month", "month_name"):
                        continue
                    if k in _ADDITIVE_MONTHLY and isinstance(v, (int, float)):
                        existing[k] = (existing.get(k, 0) or 0) + v
        # Recalcular métricas derivadas de cada mes
        for key, m in monthly_by_key.items():
            blk = m.get("block_hours_total", 0) or 0
            dty = m.get("duty_hours_total", 0) or 0
            dd = m.get("duty_days", 0) or 1
            du = m.get("duties", 0) or 1
            ft = m.get("flights_total", 0) or 1
            fc = m.get("flights_covered", 0) or 0
            hc = m.get("hotel_cost_usd", 0) or 0
            vu = m.get("viaticos_usd", 0) or 0
            m["avg_block_per_day"] = round(blk / max(1, dd), 2)
            m["avg_block_per_duty"] = round(blk / max(1, du), 2)
            m["avg_duty_per_duty"] = round(dty / max(1, du), 2)
            m["hotel_cost_per_block_hour"] = round(hc / max(1, blk), 2)
            m["viaticos_per_block_hour"] = round(vu / max(1, blk), 2)
            m["direct_cost_per_block_hour"] = round((hc + vu) / max(1, blk), 2)
            m["coverage_pct"] = round((fc / max(1, ft)) * 100, 2)
            bog_blk = m.get("bog_block_hours", 0) or 0
            mde_blk = m.get("mde_block_hours", 0) or 0
            m["bog_block_pct"] = round((bog_blk / max(1, blk)) * 100, 2)
            m["mde_block_pct"] = round((mde_blk / max(1, blk)) * 100, 2)
        merged_monthly = [monthly_by_key[k] for k in sorted(monthly_by_key.keys())]
    else:
        merged_monthly = main_monthly

    return {
        "daily_hours_rows": merged_daily_hours,
        "pilots_by_day_rows": merged_pilots_by_day,
        "pilots_by_day_base_rows": merged_pilots_by_day_base,
        "daily_pilots_excl_rows": merged_daily_pilots_excl,
        "duty_starts_rows": merged_duty_starts,
        "viaticos_rows": merged_viaticos,
        "pilot_hours_rows": merged_pilot_hours,
        "hotels": merged_hotels,
        "hotels_by_month": merged_hotels_by_month,
        "base_distribution": merged_base_dist,
        "crew_distribution": merged_crew_dist,
        "man_days": merged_man_days,
        "plant_required_rows": merged_plant,
        "kpis": merged_kpis,
        "monthly_kpis": merged_monthly,
        # Estos ya se manejan aparte (trip_report, pilot_lines, dh_report, open_time)
    }


def _print_dashboard_from_summary(summary, title_suffix):
    if not summary:
        return
    kpis = summary.get("kpis", {}) or {}
    base_dist = summary.get("base_distribution", {}) or {}
    crew_dist = summary.get("crew_distribution", {}) or {}
    man_days = summary.get("man_days", {}) or {}
    hotels = summary.get("hotels", Counter()) or Counter()
    missing = summary.get("missing_flights", set()) or set()
    pilots_total = summary.get("pilots_by_day_total", Counter()) or Counter()
    pilots_cap = summary.get("pilots_by_day_cap", Counter()) or Counter()
    pilots_cop = summary.get("pilots_by_day_cop", Counter()) or Counter()
    pilots_crp = summary.get("pilots_by_day_crp", Counter()) or Counter()

    # Resumen ejecutivo limpio
    total_pilot_days = man_days.get('cap', 0) + man_days.get('cop', 0) + man_days.get('crp', 0)
    hotel_usd = float(kpis.get('hotel_total_usd', 0) or 0)
    viaticos_usd = float(kpis.get('viaticos_total_usd', 0) or 0)
    dh_hours = float(kpis.get('total_dh_hours', 0) or 0)

    ui.summary(f"RESULTADOS {title_suffix}", {
        "Cobertura": f"{float(kpis.get('coverage_pct', 0) or 0):.1f}%",
        "Eficiencia de Red": f"{float(kpis.get('avg_network_eff', 0) or 0):.2f} Blk/Día",
        "Total Rotaciones": int(kpis.get('total_trips', 0) or 0),
        "Días Piloto (CAP/COP/CRP)": f"{man_days.get('cap', 0)} / {man_days.get('cop', 0)} / {man_days.get('crp', 0)} = {total_pilot_days}",
        "Horas DH": f"{dh_hours:.1f}",
        "Costo Hotel": f"USD ${hotel_usd:,.0f}",
        "Viáticos": f"USD ${viaticos_usd:,.0f}",
        "Vuelos Open": len(missing),
    })

    if VERBOSE_OUTPUT:
        print(f"\n{'-'*60}")
        print(f"DETALLES ADICIONALES")
        print(f"{'-'*60}")

        print(f"\nDISTRIBUCION DE TRIPULACIONES")
        print(f"   Por Base: {dict(base_dist)}")
        print(f"   Por Configuracion: {dict(crew_dist)}")

        print(f"\nHORAS POR ROL")
        print(f"   CAP Block: {float(kpis.get('cap_block_hours_total', 0) or 0):.1f}h | Duty: {float(kpis.get('cap_duty_hours_total', 0) or 0):.1f}h | Eff: {float(kpis.get('cap_network_eff', 0) or 0):.2f}")
        print(f"   COP Block: {float(kpis.get('cop_block_hours_total', 0) or 0):.1f}h | Duty: {float(kpis.get('cop_duty_hours_total', 0) or 0):.1f}h | Eff: {float(kpis.get('cop_network_eff', 0) or 0):.2f}")
        print(f"   CRP Block: {float(kpis.get('crp_block_hours_total', 0) or 0):.1f}h | Duty: {float(kpis.get('crp_duty_hours_total', 0) or 0):.1f}h | Eff: {float(kpis.get('crp_network_eff', 0) or 0):.2f}")

        if pilots_total:
            top_days = sorted(pilots_total.items(), key=lambda x: x[1], reverse=True)[:5]
            print(f"\nTOP DIAS (pilotos requeridos):")
            for dte, tot in top_days:
                cap_n = pilots_cap.get(dte, 0)
                cop_n = pilots_cop.get(dte, 0)
                crp_n = pilots_crp.get(dte, 0)
                print(f"   {dte.strftime('%Y-%m-%d')}: Total {tot} (CAP {cap_n}, COP {cop_n}, CRP {crp_n})")

        print(f"\nALOJAMIENTO (Top 5)")
        for stn, count in hotels.most_common(5):
            rate = HOTEL_RATES_USD.get(stn, 0)
            cost = count * rate
            print(f"   {stn}: {count} noches | USD {cost:.0f}")

        if missing:
            open_rows = summary.get("open_time_rows") or []
            print(f"\nVUELOS OPEN ({len(missing)})")
            for r in open_rows[:10]:
                try:
                    dep_str = r["dep_base"].strftime("%d-%b")
                except Exception:
                    dep_str = ""
                print(f"   {r.get('flt_num')} {r.get('org')}-{r.get('dst')} ({dep_str}) {r.get('tail')}")
            if len(open_rows) > 10:
                print(f"   ... y {len(open_rows) - 10} más")


def _find_dh_options(flights_by_org, times_by_org, target_stations, time_start, time_end, direction="arrive"):
    """Busca vuelos DH en una ventana de tiempo.
    direction='arrive': DH que llegan a target_stations entre time_start y time_end (posicionamiento)
    direction='depart': DH que salen de target_stations entre time_start y time_end (rescate)
    """
    options = []
    if direction == "arrive":
        for src_org, src_flights in flights_by_org.items():
            src_times = times_by_org.get(src_org, [])
            if not src_times:
                continue
            start_idx = bisect_left(src_times, time_start)
            for fi in range(start_idx, len(src_flights)):
                f = src_flights[fi]
                if f["dep_utc"] > time_end:
                    break
                if f["dst"] in target_stations and f["arr_utc"] <= time_end:
                    options.append(f)
    else:  # depart
        for dep_org in target_stations:
            src_flights = flights_by_org.get(dep_org, [])
            src_times = times_by_org.get(dep_org, [])
            if not src_times:
                continue
            start_idx = bisect_left(src_times, time_start)
            for fi in range(start_idx, len(src_flights)):
                f = src_flights[fi]
                if f["dep_utc"] > time_end:
                    break
                options.append(f)
    return options


def _print_dh_list(options, ref_time, is_positioning=True):
    """Imprime lista numerada de opciones DH."""
    if is_positioning:
        label_time = "MARGEN"
        options.sort(key=lambda x: x["arr_utc"], reverse=True)
    else:
        label_time = "ESPERA"
        options.sort(key=lambda x: x["dep_utc"])

    print(f"    {'#':>3}  {'VUELO':<12} {'RUTA':<12} {'FECHA':<12} {'DEP(UTC)':<10} {'ARR(UTC)':<10} {'BLK':>5}  {label_time}")
    print("    " + "-" * 80)
    for j, f in enumerate(options[:30], 1):
        fnum = f.get("flt_num", "DH")[:10]
        ruta = f"{f['org']}-{f['dst']}"
        fecha = f["dep_utc"].strftime("%d-%b")
        dep_t = f["dep_utc"].strftime("%H:%M")
        arr_t = f["arr_utc"].strftime("%H:%M")
        blk = f"{f.get('blk', 0):.1f}"
        if is_positioning:
            delta_h = (ref_time - f["arr_utc"]).total_seconds() / 3600
        else:
            delta_h = (f["dep_utc"] - ref_time).total_seconds() / 3600
        print(f"    {j:>3}  {fnum:<12} {ruta:<12} {fecha:<12} {dep_t:<10} {arr_t:<10} {blk:>5}  {delta_h:>5.1f}h")
    if len(options) > 30:
        print(f"    ... y {len(options) - 30} más")


def _interactive_dh_explorer(flights_df, dh_index, combined_summary, res_main):
    """Menú interactivo para emparejar vuelos sin cubrir con DH y crear trips manuales."""
    missing_ids = set()
    if combined_summary and "missing_flights" in combined_summary:
        missing_ids = set(combined_summary.get("missing_flights") or set())
    if not missing_ids and res_main:
        missing_ids = set(res_main.get("missing_flights") or set())
    if not missing_ids:
        return

    missing_df = flights_df[flights_df["id"].isin(missing_ids)].copy()
    if missing_df.empty:
        return

    missing_df = missing_df.sort_values(["dep_utc"]).reset_index(drop=True)
    flights_by_org = dh_index.get("flights_by_org", {}) if dh_index else {}
    times_by_org = dh_index.get("times_by_org", {}) if dh_index else {}
    search_window_h = 48

    # Trips manuales creados por el usuario
    manual_trips = []
    covered_ids = set()

    print("\n" + "=" * 70)
    print("  CONSTRUCTOR MANUAL DE TRIPS — Vuelos sin cubrir")
    print("=" * 70)

    while True:
        # Recalcular lista excluyendo vuelos ya emparejados
        remaining = missing_df[~missing_df["id"].isin(covered_ids)].reset_index(drop=True)
        if remaining.empty:
            print("\n  Todos los vuelos han sido emparejados.")
            break

        print(f"\n  {'#':>3}  {'VUELO':<12} {'RUTA':<12} {'FECHA':<12} {'DEP(UTC)':<10} {'ARR(UTC)':<10} {'BLK':>5}  {'RAZON'}")
        print("  " + "-" * 85)
        for i, (_, row) in enumerate(remaining.iterrows(), 1):
            flt = str(row.get("flt_num", ""))[:10]
            ruta = f"{row['org']}-{row['dst']}"
            fecha = row["dep_utc"].strftime("%d-%b") if pd.notna(row.get("dep_utc")) else ""
            dep_t = row["dep_utc"].strftime("%H:%M") if pd.notna(row.get("dep_utc")) else ""
            arr_t = row["arr_utc"].strftime("%H:%M") if pd.notna(row.get("arr_utc")) else ""
            blk = f"{row.get('blk_hours', 0):.1f}"
            reason_info = get_flight_exclusion(row["id"])
            razon = str(reason_info.get("razon", ""))[:20]
            print(f"  {i:>3}  {flt:<12} {ruta:<12} {fecha:<12} {dep_t:<10} {arr_t:<10} {blk:>5}  {razon}")

        print(f"\n  Pendientes: {len(remaining)}  |  Trips creados: {len(manual_trips)}")
        print("  Ingrese # para emparejar, 'g' para guardar trips, 'q' para salir.")

        try:
            choice = input("\n  Vuelo #: ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if choice.lower() in ('q', 'quit', 'salir'):
            break
        if choice.lower() in ('g', 'guardar'):
            if manual_trips:
                _export_manual_trips(manual_trips, flights_df)
            else:
                print("  No hay trips para guardar.")
            continue
        try:
            idx = int(choice)
            if idx < 1 or idx > len(remaining):
                print(f"  Número inválido. Ingrese 1-{len(remaining)}")
                continue
        except ValueError:
            print("  Entrada inválida.")
            continue

        row = remaining.iloc[idx - 1]
        flt_id = row["id"]
        flt_num = str(row.get("flt_num", ""))
        org = row["org"]
        dst = row["dst"]
        dep_utc = row["dep_utc"]
        arr_utc = row["arr_utc"]
        blk_hours = float(row.get("blk_hours", 0) or 0)

        print(f"\n  {'=' * 60}")
        print(f"  VUELO: {flt_num}  |  {org} -> {dst}")
        print(f"  DEP: {dep_utc.strftime('%d-%b %H:%M')} UTC  |  ARR: {arr_utc.strftime('%d-%b %H:%M')} UTC")
        reason_info = get_flight_exclusion(flt_id)
        if reason_info.get("razon"):
            print(f"  RAZON: {reason_info['razon']}")
            if reason_info.get("descripcion"):
                print(f"  DETALLE: {reason_info['descripcion']}")
        print(f"  {'=' * 60}")

        # ── PASO 1: DH POSICIONAMIENTO ──
        print(f"\n  >> DH POSICIONAMIENTO (llegar a {org} antes de {dep_utc.strftime('%d-%b %H:%M')} UTC)")
        window_start = dep_utc - timedelta(hours=search_window_h)
        pos_options = _find_dh_options(flights_by_org, times_by_org,
                                       _equiv_stations(org), window_start, dep_utc, "arrive")
        selected_pos = None
        if pos_options:
            _print_dh_list(pos_options, dep_utc, is_positioning=True)
            try:
                pos_choice = input(f"\n    Seleccione DH posicionamiento (1-{min(len(pos_options),30)}), 's' para omitir: ").strip()
            except (EOFError, KeyboardInterrupt):
                break
            if pos_choice.lower() not in ('s', 'skip', ''):
                try:
                    pi = int(pos_choice)
                    if 1 <= pi <= min(len(pos_options), 30):
                        pos_options.sort(key=lambda x: x["arr_utc"], reverse=True)
                        selected_pos = pos_options[pi - 1]
                        print(f"    >> Seleccionado: {selected_pos.get('flt_num','DH')} {selected_pos['org']}-{selected_pos['dst']} "
                              f"{selected_pos['dep_utc'].strftime('%d-%b %H:%M')}")
                except ValueError:
                    pass
        else:
            print(f"    Sin opciones de posicionamiento en ventana de {search_window_h}h")
            try:
                input("    Presione Enter para continuar con DH rescate...")
            except (EOFError, KeyboardInterrupt):
                break

        # ── PASO 2: DH RESCATE ──
        print(f"\n  >> DH RESCATE (salir de {dst} despues de {arr_utc.strftime('%d-%b %H:%M')} UTC)")
        window_end = arr_utc + timedelta(hours=search_window_h)
        rescue_options = _find_dh_options(flights_by_org, times_by_org,
                                          _equiv_stations(dst), arr_utc, window_end, "depart")
        selected_rescue = None
        if rescue_options:
            _print_dh_list(rescue_options, arr_utc, is_positioning=False)
            try:
                res_choice = input(f"\n    Seleccione DH rescate (1-{min(len(rescue_options),30)}), 's' para omitir: ").strip()
            except (EOFError, KeyboardInterrupt):
                break
            if res_choice.lower() not in ('s', 'skip', ''):
                try:
                    ri = int(res_choice)
                    if 1 <= ri <= min(len(rescue_options), 30):
                        rescue_options.sort(key=lambda x: x["dep_utc"])
                        selected_rescue = rescue_options[ri - 1]
                        print(f"    >> Seleccionado: {selected_rescue.get('flt_num','DH')} {selected_rescue['org']}-{selected_rescue['dst']} "
                              f"{selected_rescue['dep_utc'].strftime('%d-%b %H:%M')}")
                except ValueError:
                    pass
        else:
            print(f"    Sin opciones de rescate en ventana de {search_window_h}h")

        # ── PASO 3: CONFIRMAR TRIP ──
        print(f"\n  {'─' * 60}")
        print(f"  TRIP PROPUESTO:")
        trip_legs = []
        if selected_pos:
            trip_legs.append(f"    DH POS:  {selected_pos.get('flt_num','DH'):>8}  {selected_pos['org']}-{selected_pos['dst']}  "
                           f"{selected_pos['dep_utc'].strftime('%d-%b %H:%M')}-{selected_pos['arr_utc'].strftime('%H:%M')}")
        trip_legs.append(f"    CARGO:   {flt_num:>8}  {org}-{dst}  "
                        f"{dep_utc.strftime('%d-%b %H:%M')}-{arr_utc.strftime('%H:%M')}")
        if selected_rescue:
            trip_legs.append(f"    DH RES:  {selected_rescue.get('flt_num','DH'):>8}  {selected_rescue['org']}-{selected_rescue['dst']}  "
                           f"{selected_rescue['dep_utc'].strftime('%d-%b %H:%M')}-{selected_rescue['arr_utc'].strftime('%H:%M')}")
        for leg in trip_legs:
            print(leg)
        print(f"  {'─' * 60}")

        try:
            confirm = input("  Confirmar trip? (s/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            break
        if confirm in ('s', 'si', 'y', 'yes'):
            trip_data = {
                "trip_id": len(manual_trips) + 1,
                "flight_id": flt_id,
                "flight_num": flt_num,
                "flight_org": org,
                "flight_dst": dst,
                "flight_dep_utc": dep_utc,
                "flight_arr_utc": arr_utc,
                "flight_blk": blk_hours,
                "dh_pos": selected_pos,
                "dh_rescue": selected_rescue,
            }
            manual_trips.append(trip_data)
            covered_ids.add(flt_id)
            print(f"  >> Trip #{trip_data['trip_id']} creado. Vuelo {flt_num} emparejado.")
        else:
            print("  >> Trip descartado.")

    # Guardar al salir si hay trips pendientes
    if manual_trips:
        try:
            save = input(f"\n  Hay {len(manual_trips)} trips manuales. Guardar en Excel? (s/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            save = 's'
        if save in ('s', 'si', 'y', 'yes', ''):
            _export_manual_trips(manual_trips, flights_df)


def _export_manual_trips(manual_trips, flights_df):
    """Exporta los trips creados manualmente a un archivo Excel."""
    rows = []
    for t in manual_trips:
        trip_id = t["trip_id"]
        # Leg DH posicionamiento
        if t["dh_pos"]:
            p = t["dh_pos"]
            rows.append({
                "TRIP": trip_id,
                "LEG": 1,
                "TIPO": "DH POS",
                "VUELO": p.get("flt_num", "DH"),
                "ORIGEN": p["org"],
                "DESTINO": p["dst"],
                "DEP_UTC": p["dep_utc"],
                "ARR_UTC": p["arr_utc"],
                "BLK": round(p.get("blk", 0), 1),
            })
        # Leg cargo
        leg_num = 2 if t["dh_pos"] else 1
        rows.append({
            "TRIP": trip_id,
            "LEG": leg_num,
            "TIPO": "CARGO",
            "VUELO": t["flight_num"],
            "ORIGEN": t["flight_org"],
            "DESTINO": t["flight_dst"],
            "DEP_UTC": t["flight_dep_utc"],
            "ARR_UTC": t["flight_arr_utc"],
            "BLK": round(t["flight_blk"], 1),
        })
        # Leg DH rescate
        if t["dh_rescue"]:
            r = t["dh_rescue"]
            rows.append({
                "TRIP": trip_id,
                "LEG": leg_num + 1,
                "TIPO": "DH RES",
                "VUELO": r.get("flt_num", "DH"),
                "ORIGEN": r["org"],
                "DESTINO": r["dst"],
                "DEP_UTC": r["dep_utc"],
                "ARR_UTC": r["arr_utc"],
                "BLK": round(r.get("blk", 0), 1),
            })

    out_df = pd.DataFrame(rows)
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            f"Trips_Manuales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    out_df.to_excel(out_path, index=False, sheet_name="Trips Manuales")
    print(f"\n  [OK] Trips manuales exportados: {out_path}")
    print(f"       {len(manual_trips)} trips, {len(rows)} legs totales")
    try:
        os.startfile(out_path)
    except Exception:
        pass


if __name__ == "__main__":
    def _pick_file(title, initial_dir):
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            path = filedialog.askopenfilename(
                title=title,
                initialdir=initial_dir,
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            root.destroy()
            return path
        except Exception:
            return ""

    def _pick_dir(title, initial_dir):
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            path = filedialog.askdirectory(title=title, initialdir=initial_dir)
            root.destroy()
            return path
        except Exception:
            return ""

    # Limpiar consola y mostrar encabezado
    ui.clear_console()
    ui.header()

    scenarios = _ask_run_mode()

    ui.section("SELECCIÓN DE ARCHIVOS")
    schedule_path = _pick_file("Seleccione el archivo de vuelos (itinerario carguero)", PROJECT_ROOT)
    if not schedule_path:
        ui.error("Archivo de vuelos no seleccionado. Abortando.")
        sys.exit(1)
    ui.status(f"Itinerario: {os.path.basename(schedule_path)}", done=True)

    dh_path = _pick_file("Seleccione el archivo de DH (itinerario pasajeros)", PROJECT_ROOT)
    if not dh_path:
        ui.error("Archivo de DH no seleccionado. Abortando.")
        sys.exit(1)
    ui.status(f"DH: {os.path.basename(dh_path)}", done=True)

    output_dir = _pick_dir("Seleccione la carpeta de salida de resultados", PROJECT_ROOT)
    if not output_dir:
        ui.error("Carpeta de salida no seleccionada. Abortando.")
        sys.exit(1)
    ui.status(f"Salida: {output_dir}", done=True)

    ui.section("CARGANDO DATOS")
    base_rules = RuleEngine()
    df = load_schedule(schedule_path, tails_filter=TARGET_TAILS)
    df_duty = df
    df_dh_sched = pd.DataFrame()
    load_dh_for_duty = False
    if ALLOW_DH_IN_DUTY and (DH_IN_DUTY_MODE or "").strip().upper() != "NONE":
        load_dh_for_duty = True
    if ENABLE_OPEN_TOURS and OPEN_TOUR_INCLUDE_DH_FLIGHTS:
        load_dh_for_duty = True
    if load_dh_for_duty:
        dh_flights = _load_dh_flights(dh_path)
        if dh_flights:
            before_cnt = len(dh_flights)
            dh_flights = _filter_dh_flights_for_duty(dh_flights, df)
            after_cnt = len(dh_flights)
            if before_cnt != after_cnt and VERBOSE_OUTPUT:
                print(f"DH para duty filtrados: {before_cnt} -> {after_cnt}")
        df_dh_sched = _dh_flights_to_schedule_df(dh_flights)
        if not df_dh_sched.empty and (ALLOW_DH_IN_DUTY and (DH_IN_DUTY_MODE or "").strip().upper() != "NONE"):
            df_duty = pd.concat([df, df_dh_sched], ignore_index=True)
    cargo_dh = _cargo_flights_to_dh(df) if not df.empty else []
    dh_table, dh_index = load_dh_data(dh_path, base_rules, extra_flights=cargo_dh)

    ui.status(f"Vuelos cargados: {len(df)}", done=True)
    ui.status(f"Aviones: {', '.join(df['tail'].unique()) if not df.empty else 'N/A'}", done=True)

    if not df.empty:
        ui.section("OPTIMIZACIÓN")
        best = None
        results_by_scn = []
        total_scenarios = len(scenarios)

        # Definir etapas de progreso
        _stages = [
            "Generando duties y trips",
            "Optimizando solver",
            "Open Tours",
            "Solo DH",
            "Exportando resultados",
        ]
        progress = StageProgress(_stages)

        for idx, (label, allowed_crews) in enumerate(scenarios, 1):
            res = None
            used_allowed = allowed_crews
            open_allowed = allowed_crews
            if label == "4P_MIN":
                allowed_main = set(allowed_crews) - {"4P"}
                if not allowed_main:
                    allowed_main = {"2P", "3P"}
                rules_main = RuleEngine(dh_table=dh_table, dh_index=dh_index, allowed_crews=allowed_main)
                progress.advance()
                with _suppress_stdout():
                    trips_main = generate_trips(df_duty, rules_main)
                progress.advance()
                with _suppress_stdout():
                    res = solve_roster(df, trips_main, target_distribution=BASE_TARGETS, export_excel=False, rules=rules_main)
                used_allowed = allowed_main
                open_allowed = set(allowed_main) | {"4P"}
            else:
                rules = RuleEngine(dh_table=dh_table, dh_index=dh_index, allowed_crews=allowed_crews)
                progress.advance()
                with _suppress_stdout():
                    all_trips = generate_trips(df_duty, rules)
                progress.advance()
                with _suppress_stdout():
                    res = solve_roster(df, all_trips, target_distribution=BASE_TARGETS, export_excel=False, rules=rules)
                progress.advance()
            results_by_scn.append((label, res, used_allowed, open_allowed))
            if not res:
                continue
            cov = float(res["kpis"].get("coverage_pct", 0) or 0)
            eff = float(res["kpis"].get("avg_network_eff", 0) or 0)
            md = res.get("man_days") or {}
            pilot_days = float(md.get("cap", 0) + md.get("cop", 0) + md.get("crp", 0))
            score = (-cov, pilot_days, -eff)
            if best is None or score < best["score"]:
                best = {"label": label, "allowed": used_allowed, "open_allowed": open_allowed, "score": score}

        if total_scenarios > 1:
            ui.section("COMPARACIÓN DE ESCENARIOS")
            results_map = {label: (res, allowed, open_allowed) for label, res, allowed, open_allowed in results_by_scn}
            for label, res, _allowed, _open_allowed in results_by_scn:
                if not res:
                    ui.result(label, "sin solución")
                    continue
                cov = float(res['kpis'].get('coverage_pct', 0) or 0)
                eff = float(res['kpis'].get('avg_network_eff', 0) or 0)
                ui.result(label, f"Cobertura {cov:.1f}% | Eficiencia {eff:.2f}")

        if best:
            if total_scenarios > 1:
                ui.section("MEJOR ESCENARIO SELECCIONADO")
                cov_best = -best['score'][0]
                ui.success(f"{best['label']}: Cobertura {cov_best:.1f}%")
                best_res_tuple = results_map.get(best["label"])
                best_res = best_res_tuple[0] if best_res_tuple else None
                if best_res and VERBOSE_OUTPUT:
                    best_dist = best_res["kpis"].get("pilot_base_distribution", {})
                    print(f"  Distribución de pilotos (pilot-days): {best_dist}")

            if total_scenarios == 1:
                label_s, res_s, allowed_s, open_s = results_by_scn[0]
                if not res_s:
                    ui.error("No se pudo generar la solución final.")
                    sys.exit(1)
                res_main = res_s
                if label_s == "4P_MIN":
                    rules = rules_main
            else:
                rules = RuleEngine(dh_table=dh_table, dh_index=dh_index, allowed_crews=best["allowed"])
                with _suppress_stdout():
                    all_trips = generate_trips(df_duty, rules)
                    res_main = solve_roster(df, all_trips, target_distribution=BASE_TARGETS, export_excel=False, rules=rules)
                if not res_main:
                    ui.error("No se pudo generar la solución final.")
                    sys.exit(1)

            trip_report_rows = list(res_main.get("trip_report_rows") or [])
            pilot_lines_rows = list(res_main.get("pilot_lines_rows") or [])
            pilot_lines_dates_set = set(res_main.get("pilot_lines_dates_set") or set())
            trip_legs_rows = list(res_main.get("trip_legs_rows") or [])

            # Acumular todos los selected_trips de todas las pasadas
            all_selected_trips = list(res_main.get("selected_trips") or [])

            combined_summary = None
            res_open = None
            res_solo = None

            # Open tours para vuelos no cubiertos (anexar a Trip_Report y Pilot_Lines)
            if ENABLE_OPEN_TOURS and res_main:
                missing = list(res_main.get("missing_flights") or [])
                if not missing:
                    # Sin vuelos pendientes — avanzar etapas OT y SD
                    progress.advance()
                    progress.advance()
                else:
                    progress.advance(f"{len(missing)} vuelos sin cubrir")
                    open_df = df[df["id"].isin(missing)].copy()
                    if not open_df.empty:
                        rules_tour = RuleEngine(dh_table=dh_table, dh_index=dh_index, allowed_crews={"2P", "3P", "4P"}, allow_same_day_duty=OPEN_TOUR_ALLOW_SAME_DAY)
                        rules_tour.MAX_TRIP_DAYS = OPEN_TOUR_MAX_DAYS
                        rules_tour.MAX_DH_HOURS = OPEN_TOUR_MAX_DH_HOURS
                        rules_tour.MAX_DH_LEGS = OPEN_TOUR_MAX_DH_LEGS
                        open_df_duty = open_df
                        if OPEN_TOUR_INCLUDE_DH_FLIGHTS and not df_dh_sched.empty:
                            open_df_duty = pd.concat([open_df, df_dh_sched], ignore_index=True)
                        with _suppress_stdout():
                            open_trips = generate_trips(
                                open_df_duty, rules_tour,
                                max_dh_hours_per_trip=OPEN_TOUR_MAX_DH_HOURS_PER_TRIP,
                                max_dh_ratio=OPEN_TOUR_MAX_DH_RATIO,
                                allow_dh_only_duties=OPEN_TOUR_INCLUDE_DH_FLIGHTS
                            )

                        with _suppress_stdout():
                            res_open = solve_roster(open_df, open_trips, target_distribution=None, export_excel=False, rules=rules_tour, coverage_first=True)
                        if res_open:
                            # Prefijo para identificar open tours
                            def _prefix_trip_line(s):
                                if isinstance(s, str) and s.startswith("TRIP #"):
                                    try:
                                        rest = s.split("TRIP #", 1)[1]
                                        tid, tail = rest.split("|", 1)
                                        tid = tid.strip()
                                        return f"TRIP #OT-{tid} |{tail}"
                                    except Exception:
                                        return s
                                return s

                            open_trip_report = []
                            for row in res_open.get("trip_report_rows") or []:
                                new_row = dict(row)
                                new_row["RUTA/ACTIVIDAD"] = _prefix_trip_line(new_row.get("RUTA/ACTIVIDAD", ""))
                                open_trip_report.append(new_row)

                            if open_trip_report:
                                trip_report_rows.append({"FECHA":"","RUTA/ACTIVIDAD":"OPEN TOURS","AVIÓN":"","REQ":"","ASG":"","BLK":"","MAX":"","DUTY":"","MAX.1":"","REST (Act/Req)":"","PERNOCTA":""})
                                trip_report_rows.extend(open_trip_report)

                            # Acumular trips y legs de Open Tours
                            all_selected_trips.extend(res_open.get("selected_trips") or [])
                            trip_legs_rows.extend(res_open.get("trip_legs_rows") or [])

                            open_pilot_lines = []
                            for row in res_open.get("pilot_lines_rows") or []:
                                new_row = dict(row)
                                if "trip_number" in new_row:
                                    new_row["trip_number"] = f"OT-{new_row['trip_number']}"
                                open_pilot_lines.append(new_row)
                            pilot_lines_rows.extend(open_pilot_lines)

                            pilot_lines_dates_set.update(res_open.get("pilot_lines_dates_set") or set())

                            # Combinar dh_report_rows de Open Tours con prefijo OT-
                            open_dh_rows = []
                            for row in res_open.get("dh_report_rows") or []:
                                new_row = dict(row)
                                if "trip_id" in new_row:
                                    new_row["trip_id"] = f"OT-{new_row['trip_id']}"
                                open_dh_rows.append(new_row)
                            main_dh_rows = res_main.get("dh_report_rows", [])
                            combined_dh_rows = list(main_dh_rows) + open_dh_rows
                            res_main["dh_report_rows"] = combined_dh_rows

                            combined_summary = _combine_results_with_open(res_main, res_open, df, open_df)

                            # Recalcular total_dh_hours desde combined_dh_rows para garantizar consistencia
                            combined_dh_total = sum(float(row.get('blk_hours', 0) or 0) for row in combined_dh_rows)
                            if combined_summary and "kpis" in combined_summary:
                                combined_summary["kpis"]["total_dh_hours"] = round(combined_dh_total, 2)

                        # ── Pase Individual Solo-DH ─────────────────────────────────────
                        if ENABLE_SOLO_DH_PASS:
                            if res_open:
                                still_missing_solo = list(res_open.get("missing_flights") or [])
                            else:
                                still_missing_solo = list(missing)
                            if not still_missing_solo:
                                progress.advance()
                            else:
                                progress.advance(f"{len(still_missing_solo)} vuelos")
                                solo_df = df[df["id"].isin(still_missing_solo)].copy()
                                if not solo_df.empty:
                                    with _suppress_stdout():
                                        solo_trips, rules_solo = cover_individual_flights_with_dh(
                                            still_missing_solo, df, dh_table, dh_index
                                        )
                                    if solo_trips:
                                        with _suppress_stdout():
                                            res_solo = solve_roster(
                                                solo_df, solo_trips,
                                                target_distribution=None,
                                                export_excel=False,
                                                rules=rules_solo,
                                                coverage_first=True,
                                            )
                                        if res_solo:
                                            def _prefix_solo_line(s):
                                                if isinstance(s, str) and s.startswith("TRIP #"):
                                                    try:
                                                        rest = s.split("TRIP #", 1)[1]
                                                        tid, tail = rest.split("|", 1)
                                                        return f"TRIP #SD-{tid.strip()} |{tail}"
                                                    except Exception:
                                                        return s
                                                return s

                                            solo_trip_report = []
                                            for row in res_solo.get("trip_report_rows") or []:
                                                new_row = dict(row)
                                                new_row["RUTA/ACTIVIDAD"] = _prefix_solo_line(new_row.get("RUTA/ACTIVIDAD", ""))
                                                solo_trip_report.append(new_row)
                                            if solo_trip_report:
                                                trip_report_rows.append({"FECHA": "", "RUTA/ACTIVIDAD": "PASE SOLO-DH", "AVIÓN": "", "REQ": "", "ASG": "", "BLK": "", "MAX": "", "DUTY": "", "MAX.1": "", "REST (Act/Req)": "", "PERNOCTA": ""})
                                                trip_report_rows.extend(solo_trip_report)

                                            # Acumular trips y legs de Solo DH
                                            all_selected_trips.extend(res_solo.get("selected_trips") or [])
                                            trip_legs_rows.extend(res_solo.get("trip_legs_rows") or [])

                                            solo_pilot_lines = []
                                            for row in res_solo.get("pilot_lines_rows") or []:
                                                new_row = dict(row)
                                                if "trip_number" in new_row:
                                                    new_row["trip_number"] = f"SD-{new_row['trip_number']}"
                                                solo_pilot_lines.append(new_row)
                                            pilot_lines_rows.extend(solo_pilot_lines)
                                            pilot_lines_dates_set.update(res_solo.get("pilot_lines_dates_set") or set())

                                            solo_dh_rows = []
                                            for row in res_solo.get("dh_report_rows") or []:
                                                new_row = dict(row)
                                                if "trip_id" in new_row:
                                                    new_row["trip_id"] = f"SD-{new_row['trip_id']}"
                                                solo_dh_rows.append(new_row)
                                            existing_dh_rows = list(res_main.get("dh_report_rows") or [])
                                            res_main["dh_report_rows"] = existing_dh_rows + solo_dh_rows

                                            # Actualizar cobertura en combined_summary
                                            solo_missing_ids = set(res_solo.get("missing_flights") or [])
                                            solo_covered_count = len(still_missing_solo) - len(solo_missing_ids)
                                            solo_trip_count = len(res_solo.get("selected_trips") or [])
                                            total_f_count = len(df["id"].unique())
                                            solo_covered_ids = set(still_missing_solo) - solo_missing_ids
                                            if combined_summary and "kpis" in combined_summary:
                                                prev_missing = set(combined_summary.get("missing_flights") or set())
                                                combined_summary["missing_flights"] = prev_missing - solo_covered_ids
                                                prev_cov_pct = float(combined_summary["kpis"].get("coverage_pct", 0) or 0)
                                                prev_cov_count = round(prev_cov_pct * total_f_count / 100)
                                                new_cov_count = min(total_f_count, prev_cov_count + solo_covered_count)
                                                combined_summary["kpis"]["coverage_pct"] = round(new_cov_count / total_f_count * 100, 2) if total_f_count else 0
                                                combined_summary["kpis"]["flights_covered"] = new_cov_count
                                                combined_summary["kpis"]["flights_open"] = len(combined_summary["missing_flights"])
                                                combined_summary["kpis"]["trips_solo_dh"] = solo_trip_count
                                                combined_summary["kpis"]["flights_covered_solo_dh"] = solo_covered_count
                                            else:
                                                combined_summary = _combine_results_with_open(res_main, res_solo, df, solo_df)
                                                if combined_summary and "kpis" in combined_summary:
                                                    combined_summary["kpis"]["trips_solo_dh"] = solo_trip_count
                                                    combined_summary["kpis"]["flights_covered_solo_dh"] = solo_covered_count

                        else:
                            # Solo DH deshabilitado
                            progress.advance()
            else:
                # Open Tours deshabilitado — avanzar etapas OT y SD
                progress.advance()
                progress.advance()

            # Etapa final: exportación
            progress.done("Optimización completada")
            print()

            # ── Preguntar si desea crear trips manuales antes de exportar ──
            try:
                _manual = input("\n  Desea crear trips manuales para vuelos sin cubrir? (s/n) [n]: ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                _manual = 'n'
            if _manual in ('s', 'si', 'y', 'yes'):
                _interactive_dh_explorer(df, dh_index, combined_summary, res_main)

            default_name = f"Resultados_{APP_NAME_FILE}_{best['label']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            try:
                custom_name = input(f"\n  Nombre del archivo [{default_name}]: ").strip()
            except (EOFError, KeyboardInterrupt):
                custom_name = ""
            if custom_name:
                if not custom_name.lower().endswith(".xlsx"):
                    custom_name += ".xlsx"
                OUTPUT_XLSX = os.path.join(output_dir, custom_name)
            else:
                OUTPUT_XLSX = os.path.join(output_dir, default_name)
            pilot_lines_dates = [d.strftime("%d-%b") for d in sorted(pilot_lines_dates_set)]

            # Mergear datos de todas las pasadas para que las hojas reflejen el cálculo completo
            merged = _merge_all_pass_results(res_main, res_open, res_solo)

            export_results_to_excel(
                output_path=OUTPUT_XLSX,
                flights_df=df,
                trips=all_trips,
                selected_trips=all_selected_trips,
                trip_legs_rows=trip_legs_rows,
                trip_report_rows=trip_report_rows,
                pilot_lines_rows=pilot_lines_rows,
                pilot_lines_dates=pilot_lines_dates,
                daily_hours_rows=merged.get("daily_hours_rows", []),
                daily_pilots_excl_rows=merged.get("daily_pilots_excl_rows", []),
                viaticos_rows=merged.get("viaticos_rows", []),
                kpis=merged.get("kpis", {}),
                base_dist=merged.get("base_distribution", {}),
                crew_dist=merged.get("crew_distribution", {}),
                man_days=merged.get("man_days", {}),
                hotels_counter=merged.get("hotels", Counter()),
                hotels_by_month=merged.get("hotels_by_month", Counter()),
                duty_starts_rows=merged.get("duty_starts_rows", []),
                open_time_rows=_build_open_time_rows(df, (combined_summary or {}).get("missing_flights") or res_main.get("missing_flights") or set()),
                pilot_hours_rows=merged.get("pilot_hours_rows", []),
                pilots_by_day_rows=merged.get("pilots_by_day_rows", []),
                pilots_by_day_base_rows=merged.get("pilots_by_day_base_rows", []),
                plant_required_rows=merged.get("plant_required_rows", []),
                dh_report_rows=res_main.get("dh_report_rows", []),
                kpis_open=(combined_summary or {}).get("kpis"),
                base_dist_open=(combined_summary or {}).get("base_distribution"),
                crew_dist_open=(combined_summary or {}).get("crew_distribution"),
                man_days_open=(combined_summary or {}).get("man_days"),
                open_time_rows_open=(combined_summary or {}).get("open_time_rows"),
                monthly_kpis=merged.get("monthly_kpis", []),
                base_analysis=res_main.get("base_analysis"),
                idle_windows=res_main.get("idle_windows"),
            )
            # Mostrar resumen ejecutivo
            if combined_summary:
                _print_dashboard_from_summary(combined_summary, "(CON OPEN TOURS)")
            else:
                main_summary = {
                    "kpis": res_main.get("kpis", {}),
                    "base_distribution": res_main.get("base_distribution", {}),
                    "crew_distribution": res_main.get("crew_distribution", {}),
                    "man_days": res_main.get("man_days", {}),
                    "hotels": res_main.get("hotels", Counter()),
                    "missing_flights": res_main.get("missing_flights", set()),
                }
                _print_dashboard_from_summary(main_summary, "FINAL")

            ui.success(f"Archivo generado: {os.path.basename(OUTPUT_XLSX)}")
            try:
                os.startfile(OUTPUT_XLSX)
            except Exception:
                pass
